import tkinter as tk # A. [GUI 기본 도구] 컴퓨터 화면에 창, 버튼, 입력창 등을 만드는 파이썬 표준 GUI 라이브러리입니다.
from tkinter import ttk, messagebox, filedialog # B. [GUI 확장 도구] ttk=더 예쁜 위젯, messagebox=알림/확인 팝업창, filedialog=파일 열기/저장 대화상자를 제공합니다.
import threading # C. [멀티스레드] 여러 작업을 동시에 실행할 수 있게 해주는 도구입니다. API 호출이나 파일 다운로드 중에도 화면이 멈추지 않도록 백그라운드 작업에 사용됩니다.
import time # D. [시간 도구] 현재 시각 확인, 일정 시간 대기(sleep), 경과 시간 측정 등 시간 관련 기능을 제공합니다.
import os # E. [파일/폴더 도구] 파일 존재 여부 확인, 폴더 경로 조합, 파일 삭제 등 운영체제의 파일 시스템을 다루는 도구입니다.
import sys # F. [시스템 정보 도구] 실행 중인 파이썬 환경 정보, 실행 파일 경로, 표준 에러 출력 등 시스템 수준의 정보와 설정에 접근합니다.
from cryptography.fernet import Fernet # G. [암호화 도구] 대칭키 암호화(AES 기반) 라이브러리입니다. 인증키를 파일에 저장할 때 평문 노출을 막기 위해 암호화/복호화에 사용됩니다.
MASTER_KEY = b'u7_K-5D4fR9zP2mN8xL1qJ6vH3sB0tG9wE8rT7yU4iA=' # H. [암호화 마스터키] Fernet 암호화에 사용되는 고정 키입니다. 이 키로 인증키를 암호화하여 key.cfg 파일에 저장하고, 프로그램 시작 시 같은 키로 복호화하여 불러옵니다.
cipher_suite = Fernet(MASTER_KEY) # I. [암호화 엔진 생성] MASTER_KEY를 사용하는 Fernet 암호화/복호화 객체를 생성합니다. cipher_suite.encrypt()로 암호화, cipher_suite.decrypt()로 복호화합니다.
import webbrowser # J. [웹 브라우저 도구] 운영체제의 기본 웹 브라우저를 열어 URL로 이동하거나, mailto: 링크로 이메일 작성 창을 띄우는 데 사용됩니다.
import platform # K. [플랫폼 판별 도구] 현재 실행 중인 운영체제가 Windows인지 macOS(Darwin)인지 Linux인지 판별합니다. OS별로 글꼴, 버튼 스타일, 커서 모양 등을 다르게 적용할 때 사용됩니다.
from datetime import datetime, timedelta # L. [날짜/시간 도구] datetime=현재 날짜·시각 조회 및 문자열 변환, timedelta=날짜·시간 간의 덧셈/뺄셈 연산에 사용됩니다. 영업일 계산, 자정 초기화, POS 일시정지 시각 계산 등에 핵심적으로 쓰입니다.
import xml.etree.ElementTree as ET # M. [XML 파서] 서울시 버스 API 응답은 XML 형식입니다. 이 모듈로 XML 텍스트를 파싱하여 headerCd, plainNo, arrmsg 등 필요한 데이터를 추출합니다.
from urllib.parse import unquote # N. [URL 디코딩 도구] 인증키에 포함된 URL 인코딩 문자(%XX)를 원래 문자로 변환합니다. API 호출 시 serviceKey 파라미터에 디코딩된 키를 전달해야 정상 인증됩니다.

# O. [운영체제 판별] platform.system() 호출 결과를 전역 변수에 저장합니다.
#    "Windows" → 윈도우, "Darwin" → macOS, "Linux" → 리눅스.
#    이후 글꼴 설정, 버튼 위젯 선택, 커서 모양, 경로 처리 등 OS별 분기에 반복 사용됩니다.
CURRENT_OS = platform.system()

# P. [OS별 글꼴 및 글자 크기 설정] 운영체제에 따라 화면에 표시할 글꼴 이름과 크기를 다르게 지정합니다.
#    윈도우는 "맑은 고딕"/"돋움"/"Consolas"를, macOS는 "AppleGothic"/"Apple SD Gothic Neo"/"Menlo"를 사용합니다.
#    macOS는 같은 pt 수치에서 글자가 작게 렌더링되므로 1.4배 확대한 크기를 적용합니다.
#    SZ_L=제목, SZ_M=중간, SZ_S=일반, SZ_XS=작은, SZ_XXS=아주 작은 글자에 각각 사용됩니다.
if CURRENT_OS == "Windows":
    FONT_MAIN = "맑은 고딕" # a. [윈도우 주 글꼴] 버튼, 제목, 안내문구 등 대부분의 UI 텍스트에 사용됩니다.
    FONT_SUB = "돋움" # b. [윈도우 보조 글꼴] 부제목, 제작자 정보, 링크 텍스트 등 보조 영역에 사용됩니다.
    FONT_MONO = "Consolas" # c. [윈도우 고정폭 글꼴] 로그 창, 인증키 입력창, 숫자 카운터 등 글자 폭이 일정해야 하는 곳에 사용됩니다.
    SZ_L = 19 # d. [큰 글자 크기] 화면 상단 안내 제목("정류소와 노선을 선택하고...")에 사용됩니다.
    SZ_M = 11 # e. [중간 글자 크기] 정류소 이름 라벨, 표 제목 등에 사용됩니다.
    SZ_S = 9 # f. [일반 글자 크기] 버튼 텍스트, 갱신주기 라벨, 카운터 텍스트 등 대부분의 UI 요소에 사용됩니다.
    SZ_XS = 8 # g. [작은 글자 크기] 제작자 정보, 데이터 출처 링크, 안내 문구 등에 사용됩니다.
    SZ_XXS = 7 # h. [아주 작은 글자 크기] 인증키 입력창 내부 텍스트, 기록 삭제 버튼 등에 사용됩니다.
else:
    FONT_MAIN = "AppleGothic" # i. [macOS 주 글꼴] 맥에서 한글을 정상 표시하는 기본 고딕 글꼴입니다.
    FONT_SUB = "Apple SD Gothic Neo" # j. [macOS 보조 글꼴] 맥 전용 산돌 고딕 네오 글꼴로, 더 세밀한 한글 표현이 필요한 곳에 사용됩니다.
    FONT_MONO = "Menlo" # k. [macOS 고정폭 글꼴] 맥의 기본 고정폭 글꼴로, 로그와 코드 형태의 텍스트 표시에 사용됩니다.
    SZ_L = int(15 * 1.4) # l. [macOS 큰 글자] 윈도우 19pt 대신 15×1.4=21pt를 적용합니다. macOS의 글자 렌더링이 윈도우보다 작으므로 보정합니다.
    SZ_M = int(11 * 1.4)
    SZ_S = int(9 * 1.4)
    SZ_XS = int(8 * 1.4)
    SZ_XXS = int(7 * 1.4)

# Q. [외부 라이브러리 로드] 프로그램 실행에 필수적인 외부 패키지들을 불러옵니다.
#    하나라도 설치되어 있지 않으면 ImportError가 발생하며, except 블록에서 사용자에게
#    pip install 명령어를 안내하는 팝업을 띄운 뒤 프로그램을 종료합니다.
try:
    import requests # a. [HTTP 통신] 서울시 버스 API 서버에 GET 요청을 보내고 XML 응답을 받아오는 라이브러리입니다.
    import pandas as pd # b. [데이터 처리] 도착 기록을 표(DataFrame) 형태로 관리하고, 엑셀 파일(.xlsx)로 읽기/쓰기할 때 사용됩니다.
    from selenium import webdriver # c. [브라우저 자동화] 서울 데이터 광장에서 최신 버스 노선 엑셀 파일을 자동으로 다운로드하기 위해 크롬 브라우저를 프로그래밍으로 제어합니다.
    from selenium.webdriver.chrome.service import Service # d. [크롬 드라이버 서비스] ChromeDriver 실행 파일의 경로와 프로세스를 관리합니다.
    from selenium.webdriver.chrome.options import Options # e. [크롬 옵션] headless 모드(화면 없이 실행), 다운로드 경로 설정, 봇 감지 우회 등 크롬 브라우저의 실행 옵션을 설정합니다.
    from selenium.webdriver.common.by import By # f. [요소 탐색 방법] XPATH, CSS_SELECTOR 등 웹 페이지 요소를 찾는 기준을 지정합니다.
    from selenium.webdriver.support.ui import WebDriverWait # g. [명시적 대기] 특정 요소가 나타날 때까지 최대 N초간 기다리는 기능입니다. 페이지 로딩이 완료되기 전에 클릭하는 오류를 방지합니다.
    from selenium.webdriver.support import expected_conditions as EC # h. [대기 조건] "요소가 클릭 가능해질 때까지", "요소가 보일 때까지" 등 WebDriverWait의 대기 조건을 정의합니다.
    from webdriver_manager.chrome import ChromeDriverManager # i. [드라이버 자동 설치] 현재 설치된 크롬 버전에 맞는 ChromeDriver를 자동으로 다운로드하고 경로를 반환합니다. 수동 설치 없이 크롬 자동화를 사용할 수 있게 합니다.
except ImportError as e:
    # j. [라이브러리 미설치 오류 처리] 필수 패키지가 없으면 숨겨진 임시 창을 만들고,
    #    사용자에게 pip install 명령어를 안내하는 오류 팝업을 표시한 뒤 프로그램을 종료합니다.
    root = tk.Tk()
    root.withdraw()
    msg = f"필수 라이브러리가 없습니다.\n아래 명령어를 터미널(CMD)에 실행해 주세요.\n\npip install requests pandas openpyxl cryptography selenium webdriver-manager\n\n에러 내용: {e}"
    messagebox.showerror("실행 오류", msg)
    exit()


# ══════════════════════════════════════════════════════════════════════════════
# 1. [MacButton 클래스] macOS 전용 커스텀 버튼 위젯
# ──────────────────────────────────────────────────────────────────────────────
# macOS의 tkinter 기본 tk.Button은 배경색(bg) 변경이 불가능한 제한이 있습니다.
# 이 클래스는 tk.Frame(배경판) + tk.Label(텍스트)을 조합하여
# 배경색·전경색·호버 효과·비활성 스타일을 완전히 제어할 수 있는 버튼을 구현합니다.
# tk.Button과 동일한 인터페이스(configure, cget, pack, state 등)를 제공하므로
# 코드 전체에서 AnyButton 변수를 통해 OS별로 자동 선택됩니다.
# 윈도우/리눅스에서는 기본 tk.Button을, macOS에서만 이 MacButton을 사용합니다.
# ══════════════════════════════════════════════════════════════════════════════
class MacButton(tk.Frame):
    """tk.Button drop-in replacement with full bg/fg control on macOS."""

    # 1-1. [생성자] MacButton 인스턴스를 생성합니다.
    #   parent      : 이 버튼이 배치될 부모 위젯
    #   text        : 버튼에 표시할 텍스트
    #   command     : 클릭 시 실행할 콜백 함수
    #   bg/fg       : 기본 배경색/전경색
    #   activebackground/activeforeground : 마우스 호버 시 배경색/전경색
    #   disabledforeground : 비활성 상태의 전경색
    #   state       : "normal" 또는 "disabled" — 초기 활성 상태
    #   padx/pady   : Label 내부 여백 (macOS 버튼 크기 조절에 사용)
    #   width/height: Label의 문자 단위 너비/높이
    #   나머지 파라미터는 tk.Button 호환성을 위해 받되 일부는 무시됩니다.
    def __init__(self, parent, text="", command=None,
                 bg="#272727", fg="#f1f1f1",
                 activebackground=None, activeforeground=None,
                 disabledforeground="#555555",
                 font=None, cursor="pointinghand",
                 state="normal",
                 padx=4, pady=2,
                 width=0, height=0,
                 relief="flat", bd=0,
                 highlightthickness=0, highlightbackground=None,
                 overrelief=None,
                 **_ignored):
        # a. [Frame 옵션 구성] 배경색, 테두리 스타일 등 Frame 수준의 옵션을 딕셔너리로 준비합니다.
        _frm = dict(bg=bg, relief=relief, bd=bd,
                    highlightthickness=highlightthickness)
        if highlightbackground:
            _frm["highlightbackground"] = highlightbackground
        super().__init__(parent, **_frm)

        # b. [내부 상태 저장] 정상/호버/비활성 시 사용할 색상과 상태값을 인스턴스 변수에 보관합니다.
        self._norm_bg  = bg # c. 정상 상태 배경색
        self._norm_fg  = fg # d. 정상 상태 전경색
        self._act_bg   = activebackground  or bg # e. 호버 시 배경색 (미지정 시 정상 배경색 사용)
        self._act_fg   = activeforeground  or fg # f. 호버 시 전경색
        self._dis_fg   = disabledforeground # g. 비활성 상태 전경색
        self._state    = state # h. 현재 상태 ("normal" 또는 "disabled")
        self._command  = command # i. 클릭 시 실행할 함수
        self._hand_cur = cursor # j. 정상 상태의 커서 모양 (macOS: "pointinghand")

        # k. [초기 외관 설정] state에 따라 커서와 전경색을 결정합니다.
        _en  = (state == "normal")
        _cur = cursor if _en else "arrow"
        _fg  = fg     if _en else disabledforeground

        # l. [내부 Label 생성] 실제로 텍스트를 표시하는 Label 위젯을 Frame 안에 배치합니다.
        #    이 Label이 버튼의 시각적 본체 역할을 합니다.
        lbl_kw = dict(bg=bg, fg=_fg, padx=padx, pady=pady, cursor=_cur)
        if font:   lbl_kw["font"]   = font
        if width:  lbl_kw["width"]  = width
        if height: lbl_kw["height"] = height
        self._lbl = tk.Label(self, text=text, **lbl_kw)
        self._lbl.pack(fill="both", expand=True)

        # m. [마우스 이벤트 바인딩] Frame과 Label 양쪽 모두에 마우스 이벤트를 연결합니다.
        #    두 위젯 모두 바인딩해야 Label 위에서도, Frame 테두리 위에서도 이벤트가 발생합니다.
        for w in (self, self._lbl):
            w.bind("<Enter>",           self._on_enter,   "+")
            w.bind("<Leave>",           self._on_leave,   "+")
            w.bind("<Button-1>",        self._on_press,   "+")
            w.bind("<ButtonRelease-1>", self._on_release, "+")

    # 1-2. [마우스 진입] 마우스가 버튼 위로 올라오면 호버 색상(activebackground/activeforeground)을 적용합니다.
    #   비활성 상태에서는 색상 변경 없이 무시합니다.
    def _on_enter(self, _e):
        if self._state == "normal":
            self._lbl.configure(bg=self._act_bg, fg=self._act_fg)
            tk.Frame.configure(self, bg=self._act_bg)

    # 1-3. [마우스 이탈] 마우스가 버튼 밖으로 나가면 원래 색상(정상 또는 비활성)으로 복원합니다.
    def _on_leave(self, _e):
        _fg = self._norm_fg if self._state == "normal" else self._dis_fg
        self._lbl.configure(bg=self._norm_bg, fg=_fg)
        tk.Frame.configure(self, bg=self._norm_bg)

    # 1-4. [마우스 누름] 클릭이 시작되면 _pressed 플래그를 True로 설정합니다.
    #   이 플래그는 _on_release에서 실제 클릭 완료 여부를 판단하는 데 사용됩니다.
    def _on_press(self, _e):
        self._pressed = True

    # 1-5. [마우스 놓음] 누름 상태에서 버튼을 놓으면 command 콜백을 실행합니다.
    #   _pressed 플래그와 state="normal" 조건을 모두 만족해야 실행되며,
    #   버튼 바깥에서 놓는 경우(_pressed가 False인 경우)에는 실행되지 않습니다.
    def _on_release(self, _e):
        if getattr(self, "_pressed", False) and self._state == "normal":
            self._pressed = False
            if self._command:
                self._command()

    # 1-6. [속성 변경] tk.Button의 configure/config와 동일한 인터페이스를 제공합니다.
    #   command, state, text, bg, fg, font, padx, pady, width, height, cursor 등
    #   주요 속성을 키워드 인자로 받아 내부 Label과 Frame에 반영합니다.
    #   tk.Button에서만 사용되는 속성(overrelief 등)은 호환성을 위해 받되 무시합니다.
    def configure(self, **kw):
        if "command" in kw:
            self._command = kw.pop("command")

        # a. [상태 변경] state가 바뀌면 커서와 전경색을 즉시 갱신합니다.
        if "state" in kw:
            self._state = kw.pop("state")
            _en  = (self._state == "normal")
            _cur = self._hand_cur if _en else "arrow"
            _fg  = self._norm_fg  if _en else self._dis_fg
            self._lbl.configure(fg=_fg, bg=self._norm_bg, cursor=_cur)
            tk.Frame.configure(self, bg=self._norm_bg, cursor=_cur)

        if "text" in kw:
            self._lbl.configure(text=kw.pop("text"))

        # b. [배경색 변경] bg 또는 background 키를 처리합니다. 정상 배경색과 호버 배경색을 동시에 갱신합니다.
        _bg = kw.pop("bg", kw.pop("background", None))
        if _bg is not None:
            self._norm_bg = _bg
            self._act_bg  = _bg
            _fg = self._norm_fg if self._state == "normal" else self._dis_fg
            self._lbl.configure(bg=_bg, fg=_fg)
            tk.Frame.configure(self, bg=_bg)

        # c. [전경색 변경] fg 또는 foreground 키를 처리합니다. 정상 상태일 때만 Label에 즉시 반영합니다.
        _fg = kw.pop("fg", kw.pop("foreground", None))
        if _fg is not None:
            self._norm_fg = _fg
            if self._state == "normal":
                self._lbl.configure(fg=_fg)

        if "activebackground" in kw: self._act_bg = kw.pop("activebackground")
        if "activeforeground" in kw: self._act_fg = kw.pop("activeforeground")
        if "disabledforeground" in kw: self._dis_fg = kw.pop("disabledforeground")

        if "font"   in kw: self._lbl.configure(font=kw.pop("font"))
        if "padx"   in kw: self._lbl.configure(padx=kw.pop("padx"))
        if "pady"   in kw: self._lbl.configure(pady=kw.pop("pady"))
        if "width"  in kw: self._lbl.configure(width=kw.pop("width"))
        if "height" in kw: self._lbl.configure(height=kw.pop("height"))
        if "cursor" in kw:
            _c = kw.pop("cursor")
            self._hand_cur = _c
            if self._state == "normal":
                self._lbl.configure(cursor=_c)
                tk.Frame.configure(self, cursor=_c)

        # d. [Frame 전용 속성] highlightthickness, relief 등은 Frame에 직접 적용합니다.
        for _k in ["highlightthickness", "highlightbackground", "relief", "bd"]:
            if _k in kw:
                try: tk.Frame.configure(self, **{_k: kw.pop(_k)})
                except: kw.pop(_k, None)
        # e. [무시 속성] tk.Button에만 존재하는 overrelief는 MacButton에서 의미 없으므로 조용히 제거합니다.
        for _k in ["overrelief"]:
            kw.pop(_k, None)
        if kw:
            try: tk.Frame.configure(self, **kw)
            except: pass

    config = configure # f. [별칭] config는 configure의 별칭입니다. tkinter 관례에 따라 두 이름 모두 사용 가능합니다.

    # 1-7. [크기 고정] 현재 렌더링된 크기를 그대로 고정합니다.
    #   "자동 기록 시작" 버튼이 "중지"로 텍스트가 바뀔 때 버튼 크기가 줄어들지 않도록
    #   setup_ui()에서 초기 렌더링 후 한 번 호출합니다.
    #   pack_propagate(False)로 내부 Label 크기 변화가 Frame에 전파되지 않게 합니다.
    def lock_size(self):
        self.update_idletasks()
        w = self.winfo_reqwidth()
        h = self.winfo_reqheight()
        tk.Frame.configure(self, width=w, height=h)
        self.pack_propagate(False)
        self._size_locked = True

    # 1-8. [속성 조회] tk.Button의 cget과 동일한 인터페이스를 제공합니다.
    #   state, text, bg, fg 등 주요 속성을 문자열 키로 조회할 수 있습니다.
    def cget(self, key):
        if key == "state":  return self._state
        if key == "text":   return self._lbl.cget("text")
        if key in ("bg", "background"):  return self._norm_bg
        if key in ("fg", "foreground"):  return self._norm_fg
        try: return tk.Frame.cget(self, key)
        except: return ""


# R. [OS별 버튼 위젯 선택] macOS에서는 MacButton을, 윈도우/리눅스에서는 기본 tk.Button을 사용합니다.
#    이후 코드에서 AnyButton(...)으로 버튼을 생성하면 OS에 맞는 위젯이 자동 선택됩니다.
AnyButton = MacButton if CURRENT_OS == "Darwin" else tk.Button


# S. [공공누리 저작권 이미지 Base64 인코딩 데이터]
#    공공누리(KOGL) 저작권 마크 PNG 이미지를 Base64 문자열로 코드에 직접 내장했습니다.
#    별도의 이미지 파일 없이 프로그램 내부에서 바로 디코딩하여 화면에 표시할 수 있습니다.
#    setup_ui()에서 이 문자열을 디코딩하여 상단 정보 영역에 공공누리 마크로 표시합니다.
_GONGGONG_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAJUAAAA2CAYAAADOKtsPAAAvwUlEQVR42u2deZiU1ZX/P+9S+9LV+0bTTUODyCIgm6LgvhDQcWSMEqKiiCFG1F80Ehd0HkdlXKIxo1nGJGKMRB7HjaCRoIIhbIKAtIKAiNA0Te9dXXu9y++PqnupbtEmoDP/5OWpp5vuqne599xzvud7vve0Ylu2/edly3jmmWfYvHkzhmGgqiq2bcuv/zyO/1BVFdM0SSaTFBQUoKoqDoeDVCpFJBJBUZSv/bxpmrhcLkzTBMCyLGzbxul0kkqlcKraCd2frany3OJeLctC0zQ0TZO/+6rDsiy8Xi/9+vXjnnvu4TvTpqFs/XCLfdddd/Hee+8xZMgQRo4cicvlIp1O43Q6sSzrn5ZxAkfvCVu1ahXhcJhx48ZRW1tLOp0+pnNYloWu6/I8qqpmDPIE58fIGpC4T6fTCUA6nZbX/LojGo2yc+dOPvvsM+rq6nj22Wfhpnk/tD0ut33mpDPsFX95204nU7ZtWnYiFrdty/7n60RfpiW///yzvfZ555xrV5ZX2L/972dtM230+XnLMHueK+d8tmXblmWd2Euc37R6XMtIpY/p+SzDtN9ftdq+4LzzbRXFvuq7V9r6ihUr8Pl8XHXVVUyZMgXd4QDAkf1q88/wdyKHoihg26Ao2LZNV1cXLS0tGe+laX2Or6KqPTyW8FC28FCaeoL3d+QbRVGwTJN0Oo2qqmjHMP+KqnLm5Ml897vfpaGhgXfffRc9HA4zYMAAxo4di9PlIpVM4nQ6UTXtyKD88zgB0JKdOdtGURTcbjculwvtWMc3O6e2ZaHlhCLTNNEdjhNf8jlGLwzdlb03bBtF7eP+rMxnTz31VFwuF4FAAF3TNJxOp/RMTpcrE1NTqcyD6+o/DeNEbMowM8agKFiWRTqdJpVKYRhGZk76MItYdwR/IICiqpiGIY1QdzgyxnDivlQarWma2LaNI4urTNNE6SMRUBWFZCKR8XKWhWEY6OJBE4lED4vVdR1FVf8Z/E40+9N1bMtCUVVUVcXlcuF2uzPGlUzicLu+9vP+QABsm3Xr1rFv3z4Mw8CyLFwuF7FYDGcfQLqvw8p6SwHKRdY3ePBgThk1qs/5t7P3Yts2FRUVNDU1HfFUAvXnxkrbsrC1f4a/E/MDGSwlRjGRSJBOp9E0DafL1aenUlB46aWXuP/++zlw4ACBQIBYLIbH48mmb+aJZafZGzMMA7fbLamOSy65hIceeoiBQ+q+3iizYTgajdLR0YGqqpngZhkGmqJkwKRloQjwpyoYKDgtQIUYBiom7pQNmk5KU3GgEo+n8XocWd7CQFXBdszsQGqk0klUlwcTMLFxoqHaoNpgWIZMWwV9oeaAU9u2JW8i/i9CQI/UGkgmk7iy4ds0zcxnjsHVplMpHE4npmEcwS22nbmWpsprplIpnE5nxsXL9yE9fCqZlPDBMs0MLrXtDCDPSf0F1MC2UfvAVIlolOd+9zvCnZ2UFBWhKApBvx8jGwpVMgA7ZaQxTZPCwkL2NzTg8Xhwu90UlRSzf/9+LMvC7/VlqAPLwuFwZAB/dnwty0JRFLxuN0UFBbz3zjv85c03uXnIrX16YkAauW3b9Ok7nVGDmE8hjYIvZqBrbqIu2GR1sH7vLvaGEzRHO4gqJkkrRcBSGBEs5fSyAYwtq8ENBF06Whr0eBSCHpJYpLBw2Rq6rktjyAWtYiIVRUHTNFJZjCe+V1X1SxxKrrcV3IumqH0CVU3TsATpBxnj0jSZeQmjdjqdJBIJ3G43QOY+UGTG7HS5pDGJROdEj1Qq1YPrEmMkxkZBwTRNfD4fhmGwd+9eRo0Zw8yZM5kwYQIOl5O2tjZ+/etf89e3VxAKhXBkxyYej+N0OuW5BNlt2zamaUrc948efQdkn43TNPDiBqebnakE/7VtHa/s2YYj4GN/woR0GlRQUVEti782HeKXez+hpKiYef1GMnXASQxRAIcHYglUr4aiOLA1hXg8ztq1a/H5fGiahmEYktUXhhMOhwkGgwwbNgyv19vDoITXaGpqYufOnTidzgzYdDiIxWK4HM4+GW9d16mrqyOYl0fz4cPs3r0bTdNIp9PUDRlMWVkZsVgMr9eLw+GQVQan0wmWTbiri40bN5KXl0cqlWL48OHkhUIY6XSf5GFfRzKZlJOb68Fzs8B0Oo0Vs3G5XNTW1nLnnXdy8cUXZxZFdp0WFxejorBy5Urys/fpdrt7RAdVVSVjL877rRhVCgunrYMKf2r+nCfq/87mcCumSwcjDqoDPA6wLSxsLM0JKZVUysTsDvPQxnfZ0LCXG0eO56xQEUrSgSMch6CDOCZv/nk5CxYskN5HDJyu68RiMTRNw+12o6oqyWSS2tpaZs2axdSpUykoKEDXdSzD5JWX/4f/+I//kPjQNE0cDkefqy0cDjNy5Eh+/vOfc3IgwFNPPcUf/vCHTOYVi/H/bv8xd955p5wA6QGFJ1IU1qxZw5w5c1BVFY/Hw8KFC5k1a9Y3kqGJSRYeShi0KKMZponX6yWeTJBMJjnnnHOYOnUqKEomg3dkIsHQoUO56qqrWLlypQT7vasluSW5EynP9ckXOOM2UV3n4f2bmLf1z2yMH8Z0uXCpLjBsBqkqE/0hpgSKOMOdzxjNS4WughYnkmikNaDyUuun/HTb2yxt24vtcoDuh4SJQmZ1dXZ29kgWhIF5vV4Z7hKJBKlUit27d3PHHXdw9dVXs2PHDsmtpNNpWltbZW0tmUySTCbl4H3Vy+Px0NnZmZk4VZWG7PV68fl8PPfcc7z99tuoqko6nZaTmTvwmqZlvKLLRSqVkpOVSia/EfI010OJ+xbXFqUbXdexbZva2lrIpvmOrNf2eDyk02lOPvlkSkpKsCxLjrUId73PK6DGtxP+3G5eaPyEJ3eso9NMQjAEze1Uaz5mnDSRM0tLqMkroUzzolkKh6w4q8MNLGv6hPcb99CdMLGCLjZ0H+TRrX+jclw+Z3jywVJwYaMCmqLg93ppbW3Ftm0ikQjhzk5ZgxKEoWmahEIhVK+XDzdt4rZbbuGPf/wjJeXluJ1OQsEgDk2js70dwzBIJ5PE+5jYVCpFSUnJkQHR9R5pe2trK08//TQjR46ksrLyCH8jJtvOTIBt2+i6TiKRyPI7aga0n6CnysVRwgB6G10ikUDVM9AhEon0wJfi/Q6Hg+7ubqLRqHwGkXT0Pq+45rdmVO8raX6z/m80K2l0Xx580cbUgScxf9xkznTl4zTMDLdl2aAqBDQPtUV1XOarYb/nILM3L2dXKk7C62RzZzu//3QbFePOotZWUIyUBIRdXV0AzJ07l9NOO01OlJiscDjMmjVrePvtt+nq6qKgoID6+nqWLVvG7NmzSSQSckILCwu58sorGTlyJM4sqP66Yq3H46Guri6T+WYzIkVRMAyD/Px81q9fz7PPPsuPf/xj/H5/j2xUeA9haJqmZTyVbffMEk+EQD3KhAtvKK4rMOT69evpaG8nv6BALpJoNIrP5+Nvf/sbhw8fprK8XI6XuD8B1nuUl47z6POJ/7BtHfuSMSjwYsQTzKoZxdPjLySoQNjswKnlZdJqwwJFxTbA0iDgcVBbU8N/1HyPn750mE8cGrbbwfLdnzC+vJqZ/QYQMDODEggE5KoZNWoU5190UWaFWxaoKmY6jeZ0cskllzBo0CCefPJJuru7cbvdvPbaa1w/dy5er1e6cdu2GTduXOY8fc8YlmXJbE2oAAQfpCrgcrlYsmQJkydP5uyzz0bTtAyeMQx0TZeGKCZYYK1vAlOJhZWbDYvrCQ+UTqdJp9P4/X4aGhro7OwkGAxmDF9R8Pl8JJNJ6uvrCQQCpNNpvF4vqVTqKw3paF7xmL1rWFUgbRNUdWxMUhqYQALY1drO6sZGkh4PxCIM9Pu4dPRIggpgglvLA1MFVEyXTkKzSDoNLDWO00oSUpNMt0P85NRz0IMqKN20uFK8d2g3sRTYLheOLHFmWRbJZLKnfifLnWkOB7ZlYQHnnHce1QMGYFgWpm0TjcexLYuUYWAButNJMp0mkUphmSZ2X/8UUHVNZkmJRAKPy0U8GkUFkvEE+Xkh2lpaWfTQw7QcbkZFwUxnDCpX3mLleDrBNn9TXqo3pyQ8lUWGS4tGo7g8Hn71m99Q0a8y80zZul06ncblcnH//fdz1VVX0draKuU0YhHmJgDiOn1pqb7SqJy6A1tV6E7GM24UG9u0SQA7Is18lm4hqnfhC/m4uvYUpvrLwTCxXCniJOmyoqQVC82yccct3Gkdt+FBMVykOzMA85IhI5iiFBPEj+rS2LBvB+FkFCWWoQScTueXwC+WdaSEniUeVU3D7/fL+pmqqhmeJes1bNuWbLXD4UDVtB6r+6teuQBVkIIAXq+XvLw8wuEwgUCALVu28Itf/CJj6LougbjgsLTs9QzDwEineygMvq1DkKDBYJA5c+YwZsyYTIktJ0QKsrWktJQZM2Zw+umn09HRcdyYqe/SlJXJnpJ2lmCzFHRVIQ580HEAf3cUJWUTiCQI6Y6M0sKhYSsqCpCn+tAVlUOqyRrCPHVwO/O2ruTaXau5PbmLD5ItpIApFbW4uiMYzhj7QknWJZuwVbsHW527ysm6biPLlbiy2Gjnzp20tbXh8/kkl2JmeS2hqDRNk66uLuKxmMQOX/eSzH3OylRVFb/fzzXXXENJSQmJRAKfz8fSpUtZ9sYbEq+YWV4tlUpJjysIxf+VgrVtE4/H6d+/P5dffjker7cHQSreIziniRMnctZZZ0mv+m0cumJYKLqG7cgATNVUQIUWLcbW5i8I52toLg9tXTFe2f8J/bwFXOSswOvS8Tp1PkxG2HrgAH85UM/7rfs4TBLcHlQU2GWwp9ti/rSruHzwqXzcuJc3EvuIxbr5pHUfkaKBmYnJckqhUCjzoKaZMapsNV5M9tatW3nhhRdobGykoKCA1tZWRo8eLd21w+GQ6fPzzz/PihUr+nThXr+PGTNmcPbZZ6NoGfIvHo9LL3D22Wfj9/tZtGgRuq5z+PBhnnnmGU499VQqstmgw+HA7XbLgqyUtvwvSLEFfzdu3LhMFiv4q5wkwchWCMT9jB07loqKCsLh8DeSSHwZqGcBpq7r2MqR0rWuKgyq7A+fdqMkwcLF+tZmDm9ayX87/KQwafVCc8qgq7uTeKSbEo+Xy6qHcXH/oZziLqQwBVvjDVQoboaacHv1eH5QcyE79u/hZH8xPiOT2YnYHo1GefXVV9m+fbs0FKfbzd69e4lEIjQ2NrJv3z5M0ySVSlFcXHyEZATi8bjUgW/ZsoVUKtXnAMSTCYYOHcrkyZPRNV0apzAuXdf53ve+x+rVq3n77bcpLi7mo48+4re//S333nuvDM2CgjAMg0QiIXmvb9uwbNvG7/czYcKEHuEsl9sSpLJtZsb5lFNOobq6mk2bNn07RqWqKnbaQLFsUNRMPmhBpeLh/Orh1Ect1rXuI04S0+FgZ7qNna5uSBgQtkDX0JJwRkF/rhw8mulVJ9HfUiAJYJJfdjJegINd/PuV16IEPJx77rn0+/53UYOKXFF+v59YLMbGjRtZt26dJBkVTSMajcowGQgE0HUdp9PJzJkzOeWUU3rU+oSbF0y83Yen0hx6j7qjkICIMJxKpQjl53PTTTexdetWWd5YvHgxEydO5PwLLkBVVSKRCHl5ebLU9L91mKZJaWkplZWVmWdQFNLp9JGidXYshAFquk5RaSl+v/9buyddURTcio6aMrGBlGLhNCFgqaQ27uKx6lF8VFLLex2fsyfRxWexNsJpC5/tpJ/uolp1c9boUVxYO4xaFJxGRvdl+jIZpA+bFAppO8YXkUM07Wli18cfYRS4+X83/FDG9nQ6LTkrgYtUVSWd3U0iPJFt29TV1XHhhRdy1913Z9x7Oi1VlbquE4/HGThwYMYA+wDL3mCAyspKubI1TSOZTGLbNj6fT2KuKWedxaxZs3j66afJy8ujo6ODxx9/nNGjRxMIBCSl4XA4pBbtf+OwLAu/3y+L3F9VYkmn0zg0vQd98G0BdT2VTuNxudAtEflU0FXUqM3nK9ez4JV53HPLT7hp9EjMqiEknTp2LI3L4SRQXkqJ7sAFBAElLVRfYKmgKRouA9At3nxnBXvaGgmVFdC+v5WSsmKwjpQg4vE4Xq+X6upqCgsLZXFTzRqJ1+ulsLCQmpoaLrvsMqpraqQiIHc7kQiBCxYs4OKpU/sOPzlyWTNtSA8nito+nw8BEW6//XY2bNjAmjVrKC0tZffu3SxZsoTq6mpZGxSlHCl/+ZaNS3BjXq8XRVWxsky58FYiEgjOCtsG25bk8rdiVBoG3VYCy+NCt1VU0wQdEo4klaMG0PLLJm76yXxKSkqYM2cOP73rLjmZlmmi5nAbiuPIAOq2jQMFy0rjsh387k8vUe4txug2KKodyIiaIaBnVpqoXXWGw9x77bX864zLexRNc6UZIm3PpK4atmlJSkFkZAJo99CG9UEuqqqKlk3Fhac0DAPTtkBVsEwbr9/HnT9dwK7rr6e9owO/38+LL7zAuHHj8Hk82NnVbyuZhaVoKicqnVU0jUQqha0oON1uuru7cblcqIqCadsomsbBQ4dYt2EDJWVlRyS9uk5SScgyl4AFLpeLXbt2cbilBSXL/otFKcYil9w97vCXm1paloVKZrWOGDGCU089lfr6eqLRKEuXLuXSSy/l5GHD6OrszMg7cmQZuRV1uSHV4eC/f/Urdu/ejWVZhMNhzrvgfIYPH56ZNNOUVX9d1/F6vT0EX19lCDJtzuG3tByvZWUlvH3tW8zdQKmr2pdkNwLsC4wyatQofvCDH/DYY4+RTqc5ePAgnZ2dJJNJPD4f4XBYjoncSXOCeipRcQiHw3LyE4kEWpa3SyQSLFq0iDfeeEMWtqXx5YyBnS0daZrGrl270HUdh8NBIpHAMIweBfxM2dd9AkA9697FQAjvMXjwYMaPH8/HH3+M0+mkqamJu+66i9/85jeUlJYSjUTwBfw9iLjcCrhpmry74q88+eSTxONxYokE+fn5XHbZZUcmPgcgi3JDD6P5Ci7lCElqS3wgiEtVVeXAHEv4ERmoMB4hxhNZce4EB4NBrr76atasWcP69evxe31Sa6U7HPh8PqnpMgyjB445nkMYk7g3kWH6fD6pDQ+HwyiKwrvvvivLL7IGmYMTvV6vJIelHDmLyeT9ZhdPrtriuDyVmFAATdex7CMDfN111/H222/T3t6Oy+Vi06ZNXH311Tz66KOcfPLJPdy7rulkN4txYP9+tm7dyv0LFxKJRHC5XMSTSS699FLOO++8HqFNeCmtl7v9OnJO/i6LiQzDIBqNomkaJSUlR37fh6MwLVPiMVV3yPKEGI/cVS52bBcXFzN79mw+++wzwu0d5Ofn093dTVtbGwUFBfKzDocjU2g/QZGex+PhtttuY9KkSfJ+vF4vzc3NAIRCIZLJZA9PnSt9DgaDcpGZpkk0GiUYDEonkEqlCIVCtLW1EQgEaG9vZ9GiRaSPs0yjC6/Um9cQ3NGAAQO4//77ufXWW4lEIoRCIbZu3crMmTO5+OKLufzyywkGg+Tl5aGqKl1dXezYsYM33niD5cuXkxcI4PF42LdvH5dedhm33HKLzOZyNei5pZLc+lNvCe3XeRuPx4PH48GyLFavXs3hw4f7xAWKlvFqo0aN4qwpZ8nwIjLJo7HTANOmTWP79u089bMnULJFWzNHa/VNsdXifoYNG8apY8dKPX0sGmXLli18/PHHTJ06lQkTJuB0uXjxj3+kqKiIs846iw0bNlBbW8u2bduorKyktbWVQYMGMfKUU2htaSGRSBAKhdiwYQPjxo1j4MCBsgj+7LPP0t3dfXxGJTyG1HdnRfyKomRcvw3TL7mEzs5OHnroIZLJpJRT/OEPf+CF55+nrq6Oqqoq4vE4+/bto7m5GdM0cTudxGIx0uk0F198MbfddhvVNTU9cIuo+wnDyjXw3qHraBhL4cimB/H7trY2XnrpJeLxOI4+0uaUlSnp3HTTTZw15SzptUWYyb2m0KcLLzBnzhy2fLCJFStW4HK5cLlcdHV19QinKsoJG5XYgQPIPXmvvPIKixcvJpFI8PHHH/Pggw+SSqV45plnqKio4NChQzQ0NJCXl8cjjzzC+eefz7Zt2zj//POZM2cOy5YtIxwOc/a55/Dq668xfOQISktLJaWTMtI9MuPjCn+5K1EhB7xnPcbMmTOpq6vj3nvvZdeuXaRSqQw3Y5rs3buXjz/+WComhRpRZB3Tp0/n3//93ykpL8/sOMnudTMMg1gshmEYElQK8H40DqW3t+hdWM0V11mWhc/n65OnCrpdsigrykGmaWZkL1kvKdJzAVxFWCktLWXu3Lns37+fxsZGNIejBw6zc/ZRnqhhCayTTqVobGzkr3/9Kw0NDQwaNIhPPvmEl19+mWnTplFRUYHX6+XnP/85kyZNwu/3Ew6H8Xq9MiExDIMDBw5gWRbBYBCHwyHHWyxsUSQ/bqAuJqR3NmXbNqqiyi1GEydO5MUXX+Stt97i97//PZ9//jlWdr+YAI2iwJmfn89JJ53E7bffzmmnnYaSreo7sx1lHA4HqqpSWVnJSSedREFBAeFIhMrKyh4g82hAOze0JOMJXO43gwYNYty4cVIvJMjLdF/KT8ukrq6O2tpaVE3j5JNP5rTTTiORSFBRUUFJSYks26hZubFYOLZtc8HFF7Nr1y5ef/11nG43TqeTioqKHlnXicqJhfJBeKqSkpKMsjWdZs+ePSiKQl5eHkVFRXJRXHTRRXg8HpLJJIFAAFVVCQaDFBYW0tDQwObNmwmHwwysG4Rt25K8Ffp+kTQdd/gTvEQut5MrBlMUBUXLGF9xaQlXX3sN0y+9hKamJt544w06Ojpob2/Htm1KSkoYOnQoo0ePZuDAgT3SUuGhxKpTVZXR48by57+8Jf8vJquHYuFrDpcnQzpOv/QSvjN9Wg9sdqyMtFiZpmUy6+rvc9X3ZspV7fP5ehi31+v9knHfdMt8fvCjm3osBKllt+wvKSvl/R2DF0ulUui6noUimc95vF4GDx7MqlWrJEAvKSnhnXfe4YorriAQCLB9+3aam5vZuXMneXl5RKNRPB4PjY2NbNu2jblz51JZWcnSpUvBsknE4ng9XrBsUokkik3f29tyniGXS9SPZaUc7fv8/Hzcbjd33nlnD+4ol0w81pQ59/O5nztWsCvCVO8mbb3P91VG1ZuryyUCj9WbHG1PXm9O7XjUlMLbhMNhKQXSNI1hw4bx/e9/n1AoRDqdZvz48RQWFlKY9VaVlZVs376dCRMmUFdXR1FRkTTAoqIi3B4P2DbV1dW89dZbRxaLouDJcoXHUpDnKCFe7/szPQclV1Lh8Xi+ZES9PdyxGMQ3cfQ2jH9EOvJNnO8feQ5pXMeAufLy8qiurmbp0qUMGDCAsrIykskkkyZNYsqUKTidTrkbKRKJyI0PBQUFnHHGGcRiMSoqKujq6sLn89HZ2Uk0GqW5uZlktsPP2LFjCYfD7N27l1AoxEcffcSuXbuYNGnScakm9GMdgFxPILdc55RQektfTzSlzj3Xsbzvq7zrsXiGr9ql8k08xwmDdF1n7ty53HPPPcyePTvD98XjUqQoFrTITAUe6r1bOx6PSyFjPB6XtI4I8bFYTJ7P7XYzfvx4LrnkkuPUUx3jBH9VKBENsnrzM717IByLJ/xHjOlYDO1Yw/D/heEc6zVj0Shnn3MOzz33HJ999hm6ruPz+WhtbZWGJUotoqIgdlfnlqH8fj+dnZ0EAoGMDj8bVkV2a9s2RUVFHD58GI/Hw4QJE/Bm8eQ3blRf0o4fhXTMxVK9mfJva7D7ev+xht9v6sj1drkeLnd3zZe83zHcn8fjwbYsBg4cSN3gwXLrl6pmsvLcpiJWto4qtVSiBCaahYB8v5HrDLJYzTAMTh07Vr7vWMJzb6CeaRryD0yeKGGIDCuX3xKerHfHlv8rL9EbHB/Lq7eB/COgWjma0RzHe770GVXtocIQDUVUTSPS3f2lhetwOmUrSLExQ1GUjJH0wn6qKB7bNrrDgdvjkVzd8XamPiajyuUqcrMs8TWXi+m9P/9YeJreN567d+9Y9URHe//R9sr19TpRPHXUz3wDi8a2LJwuF/FYTF7HSKfxBwJyw4ads6UqnUrJHdKmYaBkZT1Wdue0LdSt2fOKNkjC0MRn/pEtWrnPros+3Tt27JBxVpB9ototdDi5+CkWi+Hz+eiORmUIFCDQ7XZLklBMutvtpqOjg7y8PNm5TwBLcb3c4raotuvZ6+WWdnJrgkJhkduMQ7j/WCwm700oF4QeSywAXddlPwe/3y/BqmgCZmW5ItG5rqioiJaWlh4teBRFkWK53Gvpuo6W/Vk6naatrY1kPE7A56OpsZEd9fXYikI0GiUUCkkuSaT+6XRaPr+maXR3d8vOMrKXgtOZKUdlFaddXV2EQiH5M9u2SaVSUiIjWjf5/X6ppM0lm8U4SrXK10QUuY/A6WTXzp3Eo1HyAgGU6sp+tqIoVFVVSYGbkG6IGpZY8aK+JrZ+J5NJtCwzLupTBQUFdHV1ycERLQlFxiFuxOVyEYlE8Pv9pFIpWW+zbZtYLIbb7c6UCrL3ZNs2gUBAttYR50wmk4RCoR4DJq4pJlNVVcnyi00JLpdLqhqExkg09ujRSzyrKxKT1ZEV54n7VbJGIQCwYL+DwSDR7IZUsRhSqZTsTlxVVZUhVnMWqcfjIZFISL2+w+FAyRq4UBJ0d3fj8/lkcb8zHMbj8UjZSiKR+NJ4a5pGOByW5RjxPF6vF6fTSXt7Oz6fT/Zb8Pv9EsS7s7bQOxrkLiAhpzlw4ADpdBplaN1gW+iwhfBLy242EHyUYMXFdqBcojOVFXfleoxIJCK1OyK9FWmw0OmIc4pMRExGrldKpVLY2d4Ign8RLtnr9RKLxeQAxONxaeymacpsSKTO0WhUZkLiZ0LyYpomeXl5dOdgFLGrJpXl5dJZICsMW9yf6BojVBKifCL0SkpO8iI8nWh8H41GsYBgMChrjclkEp/PRzQazRTbs+MudFXied1uN5FIBHe2X1cymZTXFdEmFxvJ58mOv+gGE41GZclM7A8U13A6nZg58Odorbno4eD1emXNVRlQ1d/OTT2FNQuLz8VUYmvU+PHjGTlyJPF4HCtHYLd27VoaGxu5/vrrpcfJrSW9+uqr9OvXj+HDh7Nu3ToikYjUTwsPNGjQIDmJ0WgUXVW59NJL2bRpEx9++KEMXyLkiNBw0kknyYkRCoPGxkba29spLy+nvLxcGnsymSSRSHDo0CEMw6CiokL2Ni8tLaWpqYna2lo6Ojro6OqivLxcGkskEqGwsBDTNGltbcUwDOrq6qTRBINBWltbaWlpyXjELCTw+/14vV46OjrweDyoqkppaSnxZJLdu3fj8XgIBoNUVlbK8WpsbCQZj8uQXF1djWVZHDhwAEVROOecc1j57ruyk96YMWPIy8vDMAx27NjBwYMHZRirqKjAzPay2rlzJ8OHDycQCPDRRx9RVVVFIpHAmVWViL9IsXHjRlKJhBzjXMiRi1lzsa0sKOfySbliL3EygSmEhxg1ahRTp07NrMzsiUaPGcMt8+fz6quvcvrppxMMBikoKCAejxMKhQiFQixdupThw4ezcOFC5s+fz8qVK2VtzTAMioqKuPXWW2VP7lgsxn333cf06dPp6uqSRiUeQlEUurq6mDRpEj/84Q8JhUL4/X7a29upq6tj1apVXHPNNcybN48LL7xQYrFUKsXYsWM599xzaWlpYcqUKQwaNIi8vDw2b97Mli1bmDFjBu+88w6KpnH++edz2mmn0dDQwJ///GeCwSBFRUU8/PDD3HDDDfzkJz9h1apVFBcXE4vFqKmp4Qc/+AHbtm3LNBfJdvs977zz6N+/P59++inbt29n5MiRpAyDJ554gpqaGm677TZqa2tl85ENGzbw9C9+QSwWY9asWVx66aWk02nuv/9+tm3bxsKFC3nnvfeoqqrigQceIBwO09jYSF5eHvPnz+f3v/89ixcvZsyYMdxzzz2MGDGCLVu2cM899zBx4kTOPPNM7rvvPqZMmUJRURH9+vWjvr6e+vp6HnroIa688kqam5qOSpv0DoUioqmqmgHqArDmAmZFUeQGAvFmEUYee+wxfvazn3HLLbewYuVKdF3nueeeIxwO03Cwge9///sUFhYyc+ZMnn/+ea666irmzZsna1jBYJCamhpGjx4t43o4HGbdunX86le/oqGhgalTp7JgwQKpWxfuXdyb6HZSUFDAmjVr+Pvf/05tbS2lpaVs3LiRhQsXyvc++eSTPPjgg1xxxRU0NjayatUqPvjgA7xeL93d3SxbtoxFixZhGAZbtmyhpqZGaov+tno1XV1d/Nu//Rsvvvgizz77LD/60Y8YM2YMc+fOpaWlhY0bN/LII49QWlqKoijMnj2bjo4OGXKSWW8UCAQ4/fTTWb58OWvXrmXo0KE0HDxIa2srN910EwUFBUyYMIFINMKUyVN4/vnn2bBuHZs3b+bKK6/k2muvZcaMGdx9993MmDFDzk9ZWRlVVVWsWLGCZcuWUV5ezqmnnio9aHl5OStXrmT58uWEQiH69+8vQ2NXVxeRSATDMBg+fDhFRUU0NTVRXVPD5Zdfzu9/+1sJKYTH7M0ACBsS0ETPFaPldvsV1ieKq+FwmHQ6zciRIykuLsa2bRYuXMjIUaPYsmULHR0dlJeXc8W/XUE4HGbChAncddddhEIhfD4f9fX11NXV4fP5yAuFeOCBB6RIPx6Ps2TJEjZu3MisWbPo378/5eXlMgQKT5a7yyN3AdTW1qJpGldccQWDBg0C4NChQ6TTaUaMGIGqqhQWFnLDDTfQ0NBAd3c3W7ZsQVEUPB4PH3/yMV1dXeTn51NUVMSkSZMoLi6WevUZM2awY8cOLrvsMj755BMJlDs6Opg4cSL5+fk888wzcst7Y2Mj06ZN4y9/+QuHDh4EoKWlhba2Nvx+P9FolIMHD2LbNrNnz+ajjz7i2Wef5Y477mDJkiUSTL///vts27aN/v37s3//frZt28ZJJ53EaaedJjePaprGypUrefzxx3nggQdIJpOMGDGCPXv28OCDD2KaJocOHeK6664jEAjwxRdfEA6H6d+/P06nk7q6OmbNmsXatWtZuXIlnZ2dVFVV0dLcLHeDH42zE/9PZ/uair/qZdt2xlMJkJzLPeVao2j+1dzcTF1dHU8//TRdXV043G4uueQSLrzwQgn0amtr8Xq9mYl3Ornhhhvktqcf/ehHvPDCC/x0wQIKCgrkTfj9fnbt2kV+fj5nn302Bw8e5PXXX+fTTz+lpaVFatiFhxIrQwzqjBkzJBju6uri9NNPx+PxEIvFOPfcc+Vmz+3bt6NmMVpTUxOTJ09m06ZNTJ82ndGjR1NTU8NLL73E4sWLufbaa/F6vdx88834fD6uu+46xo4dK3HLihUrWLt2LWVlZUQiEWpra6murmbjxo2Ew2G540WI9txuN5MnT6a2ro4zzzyT9evXY1kWa9eupaCggO7ubpYuXcrgwYPliv/ggw9wuVz4/X7pBcvLy2ltbZU4tF+/flLt6fF4mDFjhjTKl19+mf/6r/9i9OjRdHR08Mgjj3D99dczfvx4IpEI8Xic7u5umWlWVlYyYsQICgoKaGlp4Z133pEJjoBEvVsaiaRGLHhVVTN6KuGResdI8dXlcsnMY/Xq1Vx00UVMnz6dyZMnY2W9yEsvvcT69esJh8OMHj2af/3Xf6WmpgaXy0V9fT1LliwhEonw8MMPEwgEMqrLYJB4NEoikeCKGTPobG+nraWFro4OPt6+neLiYq677joKCwtlfaq3KjQSifD6669z2WWXMWXKFAoKCrAsi3fffZc333yTPXv20L9/f2688UaGDRuG3++nq6uLlStX8uKLLzJkyBCeeOIJ3nrrLVavXs19993Hc889J8PXU08+yd13383DDz4oB3Ls2LE8//zzNDc1UV9fz9y5c2U4LisrY8yYMdx4440cOnQId1YlOn/+fKZNm8YF557LT37yExYtWsTevXvZt3cvjQ0NXH755YwcOVJ6qDPOOIOKsjLq6+t58803cblc3HvvvYwePZq9e/eyf/9+nE4nzU1NPLJoEZMnT+a0006jrKyMVCrF2rVrWbt2LevXr+eiCy4Ay6K9tRWfx8PGjRtpaGhg1qxZ7N+3j5899hjXXXcdhw4dIhmP09XRIfm0ZDzeo+4rDKl3i2yRsVuWdWx6qpaWFsrLy0kmk7S2tjJ//nyuvvpqfv3rX7P7s884//zz+c///E8uv/xy9uzZw3333Uc0GmXJkiUoisLMmTN56KGHOPvss7n22mtpbW1l6NChvPLKK9y1YAF///vfJQ0wb948Bg8ezI9//GN0Xc/wQlnCNJdYzOXKhg0bxjXXXMNdd93FqlWrKCoq4qmnnsLlcjFnzhzuvvtuioqKmDdvHvv27eP000/nT3/6E+vXrycYDPLaa6/xy1/+EkVRuPfee6moqKCgoEC2JHr11VdlZ71UKsXcuXNlD1Kfz0d1dTU7d+6U3FBzc7MkUi3LoqqqilAoxB133MG+ffu4+eab+Zd/+RdGjx5Nd3c3TU1NPProozz++ONomsabb77JsGHDWL58OatXr8bj8TB37lxuvPFGNm7cyPPPP080GqW4uFjyWj/84Q9ZvHgxTzzxBIFAgIcffhin08mKFSt49NFHuf/++1m8eDFbt25lyZIlTJ8+XXJriUSCDz74QN7voEGDmDBhwrHpqY6noGyaJsFgENM06ezsxOfzUVRURCKRoLGxkdbWVvbv34+aFX8J6UQikWDnzp2SligrK5PVdZfLJfuhi+ZlqVQKh8PBggUL+Pzzz+Vf8zz55JP805/+RF5enuRTxO/E91988QXbt2/nO9/5DkOGDCEYDNLW1sb777+PpmmsWrWKmTNncu2119Lc3MyoUaNYvnw5e/fu5cCBA6xatUoWxzdv3szevXsZMGAAhmGQl5fHJ598InFmMplkzZo1eDweCgoK8Pv9tLW1sXv3bundFfhTGN6hQ4dYsGCB3KFjGAYvv/yyNAixG7qjo0NSK3v37pV0jG3bNDc3c+utt8oNpP369ePDDz+UVMbatWupqqriggsuIBQK0dDQwBdffIHL5aK9vZ2bb75ZbtgQbH19fT35+fnU19ezdetWfD4fTU1NjBs3jiFDhuB2u+Wew3+oXDWwuuZri2wul4ukqAtlu5uMGDGC2bNnM3ToUKzsz//nf/6H1157jc7OTqZMmcK0adMYOnQopmnS1tbG7373O95//32p3aysrGTevHkszmrdRdwWxuLI2USwcOFCli1bxnvvvSd3MAsSz+l00tXVRWFhIRMmTKCkpIR0Os3GjRvZs2ePlNL269ePsWPHUl5ezs6dO9m+fTsdHR2yP1YuVtB1ndLSUhKJBM3NzZL7EX/9qn///qRSKZqamigqKaG6ulpu0BQM++eff05HRweOLA4RZSnDMORfuKisrKSjowPDMOju7qaqqgqn00ljYyNlZWW0t7fLaoKoWhQXF0vMNmTIEPbs2UM4HKasrIyJEydSlP1TIx988AEfffQRHo9HZvYA3d3dqKpKIBCgqqqKTZs2UVVVlWkSF4/LXqEDBgzg008/7dmJ+ShUwlH1bH0ZlWidk7vDtbm5mYKCAvr374+iaUQiEZqbm6XHcTqd+P1+QqEQAO3t7bS0tEhJq/B+kUgEn8dDV1cXeXl5EiOJnRwyRdV12apH1BaFa04kEgSDQVkmEu0eI5EIRUVFRCIRyYK3t7dLplvXdYLBoOy1LrKXQCBAPB6ns7NTUigi48zPz5c93ePxOIWFhUTjcVmWyt0YAWRKUFnyUFAzovwh+liJJiCCc8stWwm9lEhUxCLr7u7G6/XKslAwGCQejxOJRGRNNpVKUVhYKK8liG1FUeT4iD7ykUgEj8cjgXhnZyeubD/W4zEqrSAUuv/rjEow7GKCNU0jFArh8Xj44osviESjxGIxWaTM3cbe0dFBLBajra2NAQMG0NjYSDAYBDJdXhRFwZ0lOUXdqrCwkO7u7i/VHcW1o9GoHCi3200gEODQoUPk5+fLARXbp9rb22XHPZ/PJ0Ouw+GguLiYtrY22X3Ptm1ZC0wkEpSWlkojE2y605n5Oy9Cn9/d3Y2m65RnW0iL3SgCHpimiStbeBYgX5wzFosRCoXIy8ujpaVFpuXCI1VWVsqanCClRUnE7/fL6kAgEKChoUH2m6qoqCCRbS8gDDh3K7t4hry8PFkvFTuvhQfVNI28vDza29ulIR6toPxVKo7/D9MIwVcT9HmyAAAAAElFTkSuQmCC"
)

# ══════════════════════════════════════════════════════════════════════════════
# 2. [SeoulBusArrivalRecorder 클래스] 서울 버스 도착 기록 프로그램의 전체 기능을 담은 메인 클래스
# ──────────────────────────────────────────────────────────────────────────────
# 이 클래스 하나에 프로그램의 모든 기능이 포함됩니다:
#   - UI 레이아웃 구성 (상단 안내, 컨트롤 바, 실시간 현황 표, 도착 기록 표, 로그 창)
#   - 라이트/다크 테마 전환
#   - 인증키 암호화 저장/로드 및 서버 검증
#   - 서울시 버스 공공 API 11종 호출 (도착 조회, 위치 조회, 정류소 검색 등)
#   - 자동/수동 모니터링 루프 (설정 주기마다 반복 갱신)
#   - 버스 도착 판정 및 엑셀 자동/수동 저장
#   - API 호출 통계 관리 (오늘/어제, 메인키/백업키 구분)
#   - 자정 자동 초기화, POS 일시정지/재개, 영업일 완결 파일 자동 생성
# __init__에서 모든 상태 변수를 초기화하고, setup_ui()로 화면을 그린 뒤,
# root.mainloop()에 의해 사용자 이벤트를 처리하는 구조입니다.
# ══════════════════════════════════════════════════════════════════════════════
class SeoulBusArrivalRecorder:

    # 2-1. [__init__ 함수] 프로그램이 처음 실행될 때 호출되는 생성자 함수입니다.
    #      창 제목·크기 설정, 실행 경로 탐색, 아이콘 적용, 모든 상태 변수 초기화,
    #      저장된 인증키 로드, UI 구성, 자정 초기화 예약, 엑셀 다운로드 예약까지
    #      프로그램 시작에 필요한 모든 준비를 한 번에 수행합니다.
    def __init__(self, root):
        # 2-1-a. [루트 창 설정] tkinter 메인 창 객체를 저장하고 제목, 초기 크기, 최소 크기를 지정합니다.
        self.root = root
        self.root.title("서울버스 정류소 듀얼 도착기록 프로그램 v1.3.90")
        self.root.geometry("1200x800")
        self.root.minsize(960, 400)

        # 2-1-b. [ttk 테마 설정] ttk 위젯(Treeview, Scrollbar 등)의 기본 외관을 'clam' 테마로 지정합니다.
        #        clam 테마는 배경색·전경색 커스터마이징이 가장 유연한 테마입니다.
        self.style = ttk.Style()
        try:
            self.style.theme_use('clam')
        except:
            pass

        # 2-1-c. [실행 경로 탐색] 프로그램이 .exe(PyInstaller 빌드)인지 .py(소스 코드)인지 판별하여
        #        리소스 디렉토리(아이콘 등)와 작업 디렉토리(엑셀·키 파일 저장)를 각각 결정합니다.
        #        resource_dir : 아이콘, 내장 리소스가 위치하는 폴더
        #        current_dir  : key.cfg, 엑셀 파일이 저장되는 실제 작업 폴더
        if getattr(sys, 'frozen', False):
            # 2-1-c-i.  PyInstaller로 빌드된 실행파일(.exe/.app)일 때의 경로 설정입니다.
            #           _MEIPASS는 PyInstaller가 임시로 풀어놓는 내부 리소스 폴더 경로입니다.
            self.resource_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(sys.executable)))
            base_exe_path = os.path.abspath(sys.executable)
            if CURRENT_OS == "Darwin":
                # 2-1-c-ii. macOS .app 번들은 실행파일이 MyApp.app/Contents/MacOS/ 안에 있으므로
                #           3단계 상위 폴더로 올라가야 사용자가 접근 가능한 작업 폴더에 도달합니다.
                self.current_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(base_exe_path))))
            else:
                self.current_dir = os.path.dirname(base_exe_path)
        else:
            # 2-1-c-iii. 소스 코드(.py)로 직접 실행 중일 때는 스크립트 파일 위치가 곧 작업 폴더입니다.
            self.resource_dir = os.path.dirname(os.path.abspath(__file__))
            self.current_dir = self.resource_dir

        # 2-1-d. [파일 경로 사전 정의] 인증키 파일(key.cfg)과 아이콘 파일의 절대 경로를 미리 조합해 둡니다.
        self.key_file_path = os.path.join(self.current_dir, "key.cfg")
        icon_path_png = os.path.join(self.resource_dir, "icon.png")
        icon_path_ico = os.path.join(self.resource_dir, "icon.ico")
        icon_path_icns = os.path.join(self.resource_dir, "icon.icns")

        # 2-1-e. [윈도우 작업표시줄 아이콘 설정] 윈도우에서는 ctypes로 AppUserModelID를 지정해야
        #        작업 표시줄에 프로그램 고유 아이콘이 정상적으로 표시됩니다.
        if CURRENT_OS == "Windows":
            import ctypes
            try:
                myappid = 'Bus_Arrival_Recoder(v1.3.88)'
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
            except: pass

        # 2-1-f. [창 아이콘 적용] PNG 아이콘이 있으면 창 제목 바에 적용하고,
        #        윈도우에서는 추가로 ICO 형식 아이콘도 적용합니다.
        try:
            if os.path.exists(icon_path_png):
                img = tk.PhotoImage(file=icon_path_png)
                self.root.iconphoto(True, img)
            if CURRENT_OS == "Windows" and os.path.exists(icon_path_ico):
                self.root.iconbitmap(icon_path_ico)
        except Exception as e:
            print(f"아이콘 설정 실패: {e}")

        # 2-1-g. [상단 컨테이너] 화면 최상단에 안내 문구·제작자 정보·컨트롤 바가 배치될 프레임입니다.
        self.top_container = tk.Frame(self.root)
        self.top_container.pack(side="top", fill="x")

        # 2-1-h. [인증키 변수] 메인 인증키와 백업 인증키를 화면 위젯과 연동하기 위한 tkinter 문자열 변수입니다.
        #        key_locked는 인증키가 서버 검증을 통과해 확정된 상태인지를 나타냅니다.
        self.service_key_var = tk.StringVar(value="")
        self.backup_key_var = tk.StringVar(value="")
        self.key_locked = False

        # 2-1-i. [검색 버튼·확정키 저장소] 정류소 검색 버튼 참조 리스트와, 검증 완료된 최종 인증키를 보관합니다.
        self.btn_searches = []
        self.final_main_key = ""
        self.final_backup_key = ""

        # 2-1-j. [정류소·갱신 상태 변수] 정류소1·2의 ARS 번호, 갱신 주기(초), 모니터링 ON/OFF 플래그입니다.
        self.ars_ids = [tk.StringVar(), tk.StringVar()]
        self.refresh_interval_var = tk.StringVar(value="20")
        self.is_monitoring = False

        # 2-1-k. [자동 저장 설정] 자동 저장 대상 엑셀 파일 경로와 저장 허용 플래그입니다.
        #        can_auto_save가 False이면 파일 생성 오류로 인해 저장이 불가능한 상태입니다.
        self.auto_save_path = None
        self.can_auto_save = True

        # 2-1-l. [폴백 API 플래그] 주 API가 한도를 초과했을 때 대체 API를 사용할지 여부를 나타냅니다.
        self.use_fallback_api = False
        self.use_fallback_pos_api = False

        # 2-1-m. [핵심 데이터 저장소] 프로그램 실행 중 수집·관리되는 모든 핵심 데이터를 초기화합니다.
        #        target_st_info  : 정류소1·2에 대한 설정 정보(ID, 이름, 선택된 노선 목록, ord_map 등)
        #        recorded_data   : 도착 판정된 모든 기록을 누적 저장하는 리스트 (엑셀 저장 원본)
        #        last_arrival_logs : 정류소별로 차량번호→마지막 기록 시각을 저장해 5분 내 중복을 방지
        #        route_corp_map  : 노선ID → 운수사명 매핑 (getRouteInfo API 응답에서 수집)
        #        rid_to_rnm      : 노선ID → 노선번호 매핑 (로그 표시용)
        #        excel_multi_corp_map : 노선번호 → 운수사명 집합 (서울시 엑셀에서 수집, 공동배차 표시용)
        #        temp_pos_data   : 한 갱신 사이클 내에서 정류소1·2 간 POS 응답을 공유하는 임시 캐시
        self.target_st_info = [{}, {}]
        self.recorded_data = []
        self.last_arrival_logs = [{}, {}]
        self.route_corp_map = {}
        self.rid_to_rnm = {}
        self.excel_multi_corp_map = {}
        self.temp_pos_data = {}

        # 2-1-n. [차량번호 캐시] ARR1/ARR2 응답에서 받은 차량번호(plainNo1, plainNo2)를 저장합니다.
        #        키: (정류소ID, 노선ID) 튜플 / 값: (plainNo1, plainNo2) 튜플
        #        SINF 폴백 시 VID API 호출 없이 캐시된 번호판을 바로 사용할 수 있습니다.
        self.veh_cache = {}

        # 2-1-o. [POS 일시정지 테이블] 차량이 없는 노선의 위치 조회를 첫차 시각 전까지 건너뛰기 위한 예약 테이블입니다.
        #        키: 노선ID / 값: 정지 해제 시각(datetime) — 이 시각이 되면 다시 호출을 허용합니다.
        #        pos_resume_logged : POS 재개 로그를 한 번만 출력하기 위한 중복 방지 집합입니다.
        self.pos_suspend_until = {}
        self.pos_resume_logged = set()

        # 2-1-p. [날짜 추적] 자정을 넘겼는지 감지하여 일시정지 테이블과 API 카운터를 초기화하는 데 사용합니다.
        self._last_date = datetime.now().date()

        # 2-1-q. [SINF 로그 날짜 추적] 실시간 현황 예비2 API(SINF) 보조 호출 안내 로그를 정류소별로 하루 1회만 출력하기 위한 딕셔너리입니다.
        #        키: 정류소 인덱스(0 또는 1) / 값: 마지막 로그 출력 날짜(date)
        self._sinf_logged_date = {}

        # 2-1-r. [macOS 버튼 활성 플래그] macOS에서 state="disabled"가 클릭을 완전히 차단하지 못하는 버그를 우회합니다.
        #        각 버튼 핸들러 진입 시 이 플래그를 먼저 확인하여 비활성 상태의 클릭을 무시합니다.
        self._btn_active = {'toggle': False, 'manual': False}

        # 2-1-s. [노선 전체 정류소 캐시] 정류소 검색창에서 노선별 전체 정류소 목록(getStaionByRoute)을
        #        한 번만 호출하고 결과를 저장해 둡니다. confirm_selection에서 재사용되며,
        #        정류소가 바뀔 때마다 초기화됩니다.
        #        _strt_cache     : {노선ID: XML 루트 요소}
        #        _strt_ord_cache : {노선ID: 해당 정류소의 순번(seq/staOrd 문자열)}
        self._strt_cache = {}
        self._strt_ord_cache = {}

        # 2-1-t. [영업일 완결 파일 추적] 이미 완결 파일을 저장한 영업일 날짜를 기억하여 중복 저장을 방지합니다.
        #        {"YYYY-MM-DD", ...} 형태의 집합입니다.
        self._completed_dates_saved = set()

        # 2-1-u. [스레드 안전 잠금] 엑셀 파일 쓰기와 데이터 갱신이 동시에 실행되지 않도록 막는 잠금장치입니다.
        #        _save_lock    : perform_auto_save() 진입을 한 스레드만 허용합니다.
        #        _refresh_lock : refresh_data() 진입을 한 스레드만 허용합니다.
        #        두 잠금 모두 자동 모니터링 스레드와 수동 갱신 스레드가 동시에 실행될 때
        #        Race Condition으로 인한 데이터 누락·역전을 방지합니다.
        self._save_lock = threading.Lock()
        self._refresh_lock = threading.Lock()

        # 2-1-v. [이중 저장 보호] 마지막으로 엑셀에 성공적으로 저장된 시점의 recorded_data 크기입니다.
        #        refresh_data() 사이클 종료 시 현재 크기와 비교해 미저장 기록이 있으면
        #        자동으로 재저장을 시도합니다(구제 저장).
        self._saved_record_count = 0

        # 2-1-w. [API 호출 통계] 오늘 호출한 API 종류별 횟수를 기록합니다.
        #        api_stats          : 메인키+백업키 합산 통계 (메인 화면 카운터에 표시)
        #        api_stats_by_key   : 메인키/백업키를 구분한 개별 통계 (API 상세 창에 표시)
        #        api_stats_yesterday, api_stats_by_key_yesterday : 자정 초기화 직전에 복사된 어제 스냅샷
        self.api_stats = {
            "ARR1": 0, "ARR2": 0, "SINF": 0, "POS1": 0, "POS2": 0, "SCNM": 0, "SCID": 0, "RINF": 0, "SLST": 0, "VID": 0, "VLD": 0
        }
        self.api_stats_by_key = {
            "main": {k: 0 for k in self.api_stats},
            "back": {k: 0 for k in self.api_stats},
        }
        self.api_stats_yesterday = {k: 0 for k in self.api_stats}
        self.api_stats_by_key_yesterday = {
            "main": {k: 0 for k in self.api_stats},
            "back": {k: 0 for k in self.api_stats},
        }

        # 2-1-x. [API 주소 모음] 서울시 공공 버스 API 11종의 엔드포인트 URL을 사전(딕셔너리)으로 관리합니다.
        #        fetch_api()에서 실제 호출에 사용하고, 통계 창에서 URL 표시에도 사용됩니다.
        self.api_urls = {
            "ARR1": "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRoute",
            "ARR2": "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRouteAll",
            "POS1": "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRtid",
            "SINF": "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
            "POS2": "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRouteSt",
            "SCNM": "http://ws.bus.go.kr/api/rest/stationinfo/getStationByName",
            "SCID": "http://ws.bus.go.kr/api/rest/stationinfo/getRouteByStation",
            "RINF": "http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo",
            "SLST": "http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute",
            "VID": "http://ws.bus.go.kr/api/rest/buspos/getBusPosByVehId",
            "VLD": "http://ws.bus.go.kr/api/rest/busRouteInfo/getBusRouteList"
        }

        # 2-1-y. [기타 UI 상태] 통계 팝업 창 참조와 다크모드 플래그를 초기화합니다.
        self.stats_win = None
        self.is_dark_mode = False

        # 2-1-z. [초기화 마무리 순서] 아래 5단계를 순서대로 실행하여 프로그램 시작을 완료합니다.
        #        ① 저장된 인증키 파일(key.cfg)이 있으면 불러와 자동 검증 시작
        #        ② 전체 UI 레이아웃 구성 (setup_ui)
        #        ③ 창 닫기(X) 버튼에 종료 확인 팝업 연결
        #        ④ 다음 자정(00:00)에 API 카운터를 초기화하는 타이머 예약
        #        ⑤ 서울시 버스 노선 엑셀 파일을 백그라운드에서 자동 다운로드 시작
        self.load_saved_key()
        self.setup_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_window_close)
        self.root.after(200, self._schedule_midnight_reset)
        self.root.after(500, self.start_excel_download_thread)

    # 2-2. [_on_window_close 함수] 사용자가 창의 X 버튼을 클릭했을 때 호출됩니다.
    #      실수로 종료하는 것을 막기 위해 "정말 종료하시겠습니까?" 확인 팝업을 띄우고,
    #      '예'를 선택한 경우에만 프로그램을 완전히 종료(destroy)합니다.
    def _on_window_close(self):
        if messagebox.askyesno("종료 확인", "정말 종료하시겠습니까?"):
            self.root.destroy()

    # 2-3. [_schedule_midnight_reset 함수] 다음 자정(00:00:00)까지 남은 시간을 밀리초 단위로 계산하여
    #      _do_midnight_reset 함수를 tkinter 타이머(root.after)로 예약합니다.
    #      _do_midnight_reset 내부에서 이 함수를 다시 호출하므로 매일 자정마다 반복 실행됩니다.
    def _schedule_midnight_reset(self):
        now = datetime.now()
        from datetime import timedelta
        next_midnight = (now + timedelta(days=1)).replace(
            hour=0, minute=0, second=0, microsecond=0)
        ms_left = int((next_midnight - now).total_seconds() * 1000)
        # 2-3-a. 자정 직후 0ms가 나오면 무한 재귀가 될 수 있으므로 최소 1초를 보장합니다.
        ms_left = max(ms_left, 1000)
        self.root.after(ms_left, self._do_midnight_reset)

    # 2-4. [_do_midnight_reset 함수] 매일 자정(00:00)에 자동 실행되는 초기화 함수입니다.
    #      ① 오늘 통계를 '어제 스냅샷'으로 복사 (통계 창 '어제의 합계' 탭에 표시됨)
    #      ② 오늘 카운터를 전부 0으로 초기화
    #      ③ SINF 보조 로그 날짜 기록 초기화 (다음 날 첫 호출 시 안내 로그 재출력)
    #      ④ UI 카운터 갱신 및 로그 기록
    #      ⑤ 다음 날 자정을 위해 _schedule_midnight_reset()을 재호출 (필수: 없으면 첫날만 작동)
    def _do_midnight_reset(self):
        self.api_stats_yesterday = dict(self.api_stats)
        self.api_stats_by_key_yesterday = {
            slot: dict(vals) for slot, vals in self.api_stats_by_key.items()
        }
        for k in self.api_stats:
            self.api_stats[k] = 0
        for slot in self.api_stats_by_key.values():
            for k in slot:
                slot[k] = 0
        self._sinf_logged_date.clear()
        self.update_api_counter_ui()
        self.log("📅 [자정] API 호출 카운터를 초기화했습니다. 어제의 합계는 통계 창에서 확인할 수 있습니다.")
        # 2-4-a. 다음 날 자정을 위한 재예약입니다. 이 줄이 없으면 첫날만 실행되고 이후 작동하지 않습니다.
        self._schedule_midnight_reset()

    # 2-5. [get_theme 함수] 현재 테마 모드(라이트/다크)에 따른 전체 색상 팔레트를 딕셔너리로 반환합니다.
    #      배경색, 전경색, 버튼 색상, 토글 스위치 색상, 스크롤바 색상, 입력창 테두리 색상 등
    #      프로그램 전체에서 사용하는 모든 색상 값이 이 한 곳에 정의되어 있습니다.
    #      apply_theme(), get_btn_style(), setup_ui() 등 거의 모든 UI 함수에서 호출됩니다.
    def get_theme(self):
        if self.is_dark_mode:
            return {
                # 2-5-a. [다크 모드 팔레트] 어두운 배경에 밝은 글씨 기반의 색상 체계입니다.
                "bg_root":      "#0f0f0f",      # 메인 창 배경 (가장 어두운 검정)
                "bg_panel":     "#212121",      # 상단 패널·상태 바 배경
                "bg_card":      "#272727",      # 컨트롤 바·버튼 영역 배경
                "bg_entry":     "#121212",      # 입력창 배경
                "bg_entry_ro":  "#1a1a1a",      # 읽기 전용 입력창 배경
                "fg_main":      "#f1f1f1",      # 주 텍스트 색상 (밝은 흰색)
                "fg_sub":       "#aaaaaa",      # 보조 텍스트 (회색)
                "fg_accent":    "#3ea6ff",      # 강조 색상 (파란색, "자동 기록 시작" 등)
                "fg_warn":      "#ff4e45",      # 경고 색상 (빨간색)
                "fg_link":      "#5ebe5e",      # 외부 링크 색상 (초록색)
                "fg_email":     "#3ea6ff",      # 이메일 링크 색상
                "fg_st1":       "#3ea6ff",      # 정류소1 제목 색상
                "fg_st2":       "#ff4e45",      # 정류소2 제목 색상
                "fg_counter":   "#aaaaaa",      # 카운터 라벨 색상
                "fg_counter_v": "#f1f1f1",      # 카운터 숫자 색상
                "bg_log":       "#0f0f0f",      # 로그 창 배경
                "fg_log":       "#dfe6e9",      # 로그 텍스트 색상
                "bg_tree":      "#212121",      # Treeview 표 배경
                "fg_tree":      "#f1f1f1",      # Treeview 표 텍스트
                "bg_tree_sel":  "#3f3f3f",      # Treeview 선택 행 배경
                "border":       "#3f3f3f",      # 테두리·구분선 색상
                "sash":         "#3f3f3f",      # PanedWindow 분할선 색상
                "btn_normal_bg":      "#272727", # 일반 버튼 배경
                "btn_normal_fg":      "#f1f1f1", # 일반 버튼 텍스트
                "btn_normal_active":  "#3f3f3f", # 일반 버튼 호버 배경
                "btn_normal_border":  "#3f3f3f", # 일반 버튼 테두리
                "btn_outline_bg":     "#272727", # 아웃라인 버튼 배경
                "btn_outline_fg":     "#3ea6ff", # 아웃라인 버튼 텍스트 (강조 파란색)
                "btn_outline_border": "#3ea6ff", # 아웃라인 버튼 테두리
                "btn_disabled_bg":    "#1a1a1a", # 비활성 버튼 배경
                "btn_disabled_fg":    "#555555", # 비활성 버튼 텍스트
                "btn_disabled_border":"#2a2a2a", # 비활성 버튼 테두리
                "toggle_track_off":  "#3f3f3f",  # 토글 스위치 OFF 트랙 색상
                "toggle_track_on":   "#3ea6ff",  # 토글 스위치 ON 트랙 색상
                "toggle_thumb":      "#f1f1f1",  # 토글 스위치 동그라미(thumb) 색상
                "toggle_label":      "#aaaaaa",  # 토글 라벨 텍스트 색상
                "toggle_bg":         "#272727",  # 토글 영역 배경
                "sb_bg":      "#3f3f3f",         # 스크롤바 thumb 배경
                "sb_trough":  "#1a1a1a",         # 스크롤바 trough(홈) 배경
                "sb_arrow":   "#aaaaaa",         # 스크롤바 화살표 색상
                "entry_hl":       "#48484a",     # 입력창 비포커스 테두리 (macOS용)
                "entry_hl_focus": "#0a84ff",     # 입력창 포커스 테두리 (macOS용)
            }
        else:
            return {
                # 2-5-b. [라이트 모드 팔레트] 밝은 배경에 어두운 글씨 기반의 색상 체계입니다.
                "bg_root":      "#ffffff",
                "bg_panel":     "#ecf0f1",
                "bg_card":      "#f1f2f6",
                "bg_entry":     "#ffffff",
                "bg_entry_ro":  "#f0f0f0",
                "fg_main":      "#2d3436",
                "fg_sub":       "#7f8c8d",
                "fg_accent":    "#0984e3",
                "fg_warn":      "#e74c3c",
                "fg_link":      "#006400",
                "fg_email":     "#2980b9",
                "fg_st1":       "#0984e3",
                "fg_st2":       "#d63031",
                "fg_counter":   "#636e72",
                "fg_counter_v": "#2d3436",
                "bg_log":       "#2d3436",
                "fg_log":       "#dfe6e9",
                "bg_tree":      "#ffffff",
                "fg_tree":      "#000000",
                "bg_tree_sel":  "#0078d7",
                "border":       "#cccccc",
                "sash":         "#cccccc",
                "btn_normal_bg":      "#2ecc71",
                "btn_normal_fg":      "#ffffff",
                "btn_normal_active":  "#27ae60",
                "btn_normal_border":  "#1e8449",
                "btn_outline_bg":     "#eafaf1",
                "btn_outline_fg":     "#27ae60",
                "btn_outline_border": "#27ae60",
                "btn_disabled_bg":    "#e8e8e8",
                "btn_disabled_fg":    "#a0a0a0",
                "btn_disabled_border":"#c0c0c0",
                "toggle_track_off":  "#bbbbbb",
                "toggle_track_on":   "#2ecc71",
                "toggle_thumb":      "#ffffff",
                "toggle_label":      "#7f8c8d",
                "toggle_bg":         "#f1f2f6",
                "sb_bg":      "#c0c0c0",
                "sb_trough":  "#f0f0f0",
                "sb_arrow":   "#606060",
                "entry_hl":       "#c7c7c7",
                "entry_hl_focus": "#007aff",
            }

    # 2-6. [apply_theme 함수] is_dark_mode 상태에 맞춰 프로그램 전체 위젯의 색상을 일괄 갱신합니다.
    #      setup_ui()에서 _tw 딕셔너리에 등록해 둔 모든 위젯을 카테고리별로 순회하며
    #      배경색·전경색·글꼴 등을 현재 테마 팔레트로 다시 칠합니다.
    #      라이트↔다크 토글 클릭 시(_on_toggle_theme) 호출됩니다.
    def apply_theme(self):
        T = self.get_theme()
        # 2-6-a. [루트·컨테이너 배경] 메인 창과 상단 컨테이너의 배경색을 갱신합니다.
        self.root.configure(bg=T["bg_root"])
        self.top_container.configure(bg=T["bg_panel"])

        # 2-6-b. [상단 상태 바 프레임] 안내 문구, 제작자 정보 영역의 배경색을 갱신합니다.
        for w in self._tw.get("status_frames", []):
            try: w.configure(bg=T["bg_panel"])
            except: pass
        # 2-6-c. [상단 상태 바 라벨] 제목, 이메일, 링크 등 개별 라벨의 색상을 역할별로 갱신합니다.
        for key, w in self._tw.get("status_labels", {}).items():
            try:
                if key == "title_text":  w.configure(bg=T["bg_panel"], fg=T["fg_main"])
                elif key == "title_text3":  w.configure(bg=T["bg_panel"], fg=T["fg_main"])
                elif key == "title_blue":w.configure(bg=T["bg_panel"], fg=T["fg_accent"])
                elif key == "maker":     w.configure(bg=T["bg_panel"], fg=T["fg_sub"])
                elif key == "email":     w.configure(bg=T["bg_panel"], fg=T["fg_email"])
                elif key == "link1":     w.configure(bg=T["bg_panel"], fg=T["fg_link"])
                elif key == "link2":     w.configure(bg=T["bg_panel"], fg="#7fb3ff" if self.is_dark_mode else "#000080")
                elif key == "maker_bracket1": w.configure(bg=T["bg_panel"], fg=T["fg_sub"])
                elif key == "maker_bracket2": w.configure(bg=T["bg_panel"], fg=T["fg_sub"])
            except: pass

        # 2-6-d. [컨트롤 바 프레임] 버튼·입력창이 모인 컨트롤 바 영역의 배경색을 갱신합니다.
        for w in self._tw.get("ctrl_frames", []):
            try: w.configure(bg=T["bg_card"])
            except: pass
        # 2-6-e. [자동 저장 상태 라벨] "자동 기록 중..." 또는 안내 문구의 색상을 갱신합니다.
        try: self.lbl_auto_save_status.configure(bg=T["bg_panel"], fg=T["fg_warn"])
        except: pass
        # 2-6-f. [갱신 주기 관련 라벨·입력창] "갱신주기(초):" 라벨과 입력창의 색상을 갱신합니다.
        for w in self._tw.get("interval_labels", []):
            try: w.configure(bg=T["bg_card"], fg=T["fg_main"])
            except: pass
        try:
            self.entry_refresh_interval.configure(
                bg=T["bg_entry"], fg=T["fg_main"],
                insertbackground=T["fg_main"],
                readonlybackground=T["bg_entry_ro"])
        except: pass
        # 2-6-g. [인증키 관련 라벨·프레임] 인증키 영역의 배경색을 갱신합니다.
        for w in self._tw.get("key_labels", []):
            try: w.configure(bg=T["bg_card"], fg=T["fg_main"])
            except: pass
        for w in self._tw.get("key_frames", []):
            try: w.configure(bg=T["bg_card"])
            except: pass

        # 2-6-h. [API 카운터 영역] 실시간현황/도착기록/기타 숫자판의 배경·텍스트 색상을 갱신합니다.
        try: self.api_stats_container.configure(bg=T["bg_card"])
        except: pass
        try: self._tw["stats_outer"].configure(bg=T["bg_card"])
        except: pass
        for lbl in self._tw.get("stat_key_labels", []):
            try: lbl.configure(bg=T["bg_card"], fg=T["fg_counter"])
            except: pass
        for key, lbl in self.stat_value_labels.items():
            try: lbl.configure(bg=T["bg_card"], fg=T["fg_counter_v"])
            except: pass

        # 2-6-i. [버튼 일괄 재스타일] 등록된 모든 버튼의 색상을 현재 테마에 맞게 다시 적용합니다.
        self._restyle_all_buttons(T)

        # 2-6-j. [분할 패널 구분선] PanedWindow의 sash(드래그 구분선) 색상을 갱신합니다.
        for pw in [self.super_paned, self.main_paned]:
            try: pw.configure(bg=T["sash"])
            except: pass

        # 2-6-k. [Treeview 표 영역 프레임] 실시간 현황·도착 기록 표를 감싸는 프레임 배경색을 갱신합니다.
        for w in self._tw.get("tree_frames", []):
            try: w.configure(bg=T["bg_root"])
            except: pass

        # 2-6-l. [정류소 제목·ARS 입력창] 정류소 이름 라벨과 ARS 번호 표시 입력창의 색상을 갱신합니다.
        for lbl in self.lbl_st_names:
            try: lbl.configure(fg=T["fg_st1"], bg=T["bg_root"])
            except: pass
        for lbl in self.lbl_hist_titles:
            try: lbl.configure(fg=T["fg_st2"], bg=T["bg_root"])
            except: pass
        for w in self._tw.get("ars_entries", []):
            try: w.configure(readonlybackground=T["bg_entry_ro"], fg=T["fg_main"])
            except: pass

        # 2-6-m. [로그 창] 하단 로그 텍스트 영역과 감싸는 프레임의 색상을 갱신합니다.
        try:
            self.txt_log.configure(bg=T["bg_log"], fg=T["fg_log"])
            self._tw["log_outer"].configure(bg=T["bg_root"])
        except: pass

        # 2-6-n. [Treeview·스크롤바·입력창·토글 스타일] ttk 위젯 스타일과 토글 스위치를 재적용합니다.
        self._apply_treeview_style(T)
        self._apply_scrollbar_style(T)
        if CURRENT_OS == "Darwin":
            self._mac_style_all_entries(T)
        self._draw_toggle()

    # 2-7. [_restyle_all_buttons 함수] _tw["buttons"]에 등록된 모든 버튼의 색상을 현재 테마에 맞게 일괄 재적용합니다.
    #      각 버튼은 (위젯, [타입문자열]) 형태로 저장되어 있으며, 타입(normal/disabled/outline)에 따라
    #      get_btn_style()을 호출하여 새 색상 딕셔너리를 얻어 configure()합니다.
    def _restyle_all_buttons(self, T):
        for btn, type_cell in self._tw.get("buttons", []):
            try:
                style = self.get_btn_style(type_cell[0])
                style.pop("state", None)
                btn.configure(**style)
            except: pass

    # 2-8. [_apply_treeview_style 함수] ttk.Treeview 위젯(표)의 본문·헤더·선택행 색상을
    #      현재 테마 팔레트에 맞게 ttk.Style로 설정합니다.
    #      다크 모드에서는 어두운 배경에 밝은 글씨, 라이트 모드에서는 흰 배경에 검은 글씨를 적용합니다.
    def _apply_treeview_style(self, T):
        sty = ttk.Style()
        if self.is_dark_mode:
            sty.configure("Treeview",
                background=T["bg_tree"], foreground=T["fg_tree"],
                fieldbackground=T["bg_tree"], rowheight=22)
            sty.configure("Treeview.Heading",
                background=T["bg_card"], foreground=T["fg_main"], relief="flat")
            sty.map("Treeview",
                background=[("selected", T["bg_tree_sel"])],
                foreground=[("selected", T["fg_main"])])
            sty.map("Treeview.Heading",
                background=[("active", T["border"])])
        else:
            sty.configure("Treeview",
                background="#ffffff", foreground="#000000",
                fieldbackground="#ffffff", rowheight=22)
            sty.configure("Treeview.Heading",
                background="#f0f0f0", foreground="#000000", relief="raised")
            sty.map("Treeview",
                background=[("selected", "#0078d7")],
                foreground=[("selected", "#ffffff")])
            sty.map("Treeview.Heading",
                background=[("active", "#e0e0e0")])

    # 2-9. [_apply_scrollbar_style 함수] "App.Vertical.TScrollbar" 스타일을 현재 테마에 맞게 설정합니다.
    #      모든 ttk.Scrollbar가 이 스타일을 공유하므로 한 번 호출하면 프로그램 전체에 적용됩니다.
    def _apply_scrollbar_style(self, T):
        sty = ttk.Style()
        sty.configure("App.Vertical.TScrollbar",
            background=T["sb_bg"],
            troughcolor=T["sb_trough"],
            arrowcolor=T["sb_arrow"],
            bordercolor=T["sb_trough"],
            darkcolor=T["sb_bg"],
            lightcolor=T["sb_bg"],
            relief="flat",
        )
        sty.map("App.Vertical.TScrollbar",
            background=[
                ("active",   T["border"]),
                ("disabled", T["sb_trough"]),
            ])

    # 2-10. [_mac_style_entry 함수] macOS에서 tk.Entry 위젯에 네이티브 스타일(얇은 테두리, 포커스 시 파란 테두리)을
    #       적용합니다. relief="flat" + highlightthickness=1로 macOS TextField 느낌을 구현합니다.
    def _mac_style_entry(self, entry, T):
        try:
            entry.configure(
                relief="flat",
                highlightthickness=1,
                highlightbackground=T.get("entry_hl", "#c7c7c7"),
                highlightcolor=T.get("entry_hl_focus", "#007aff"),
            )
        except Exception:
            pass

    # 2-11. [_mac_style_all_entries 함수] 현재 존재하는 모든 tk.Entry 위젯에 macOS 스타일을 일괄 적용합니다.
    #       갱신주기 입력창과 ARS 번호 표시 입력창이 대상입니다.
    def _mac_style_all_entries(self, T):
        targets = []
        try: targets.append(self.entry_refresh_interval)
        except: pass
        for e in self._tw.get("ars_entries", []):
            targets.append(e)
        for e in targets:
            self._mac_style_entry(e, T)

    # 2-12. [_build_toggle 함수] 라이트/다크 테마 전환용 토글 스위치를 Canvas 위젯으로 생성합니다.
    #       50×24px 크기의 둥근 트랙 안에 원형 thumb과 이모티콘(☀/🌙)을 그립니다.
    #       클릭 이벤트를 _on_toggle_theme에 연결합니다.
    def _build_toggle(self, parent):
        T = self.get_theme()
        TRACK_W, TRACK_H = 50, 24
        self._toggle_canvas = tk.Canvas(
            parent, width=TRACK_W + 6, height=TRACK_H + 6,
            highlightthickness=0,
            cursor="pointinghand" if CURRENT_OS == "Darwin" else "hand2"
        )
        self._toggle_canvas.pack(side="left", padx=(6, 6))
        self._toggle_canvas.bind("<Button-1>", self._on_toggle_theme)
        self._draw_toggle()

    # 2-13. [_draw_toggle 함수] 토글 스위치의 그래픽을 현재 테마 상태에 맞게 Canvas에 다시 그립니다.
    #       라이트 모드: 왼쪽 해 아이콘 + 오른쪽 흰 thumb
    #       다크 모드:  왼쪽 흰 thumb + 오른쪽 달 아이콘
    def _draw_toggle(self):
        if not hasattr(self, "_toggle_canvas"):
            return
        T = self.get_theme()
        c = self._toggle_canvas
        TRACK_W, TRACK_H, THUMB_R = 50, 24, 10
        ox, oy = 3, 3
        c.delete("all")
        c.configure(bg=T["bg_card"])
        r = TRACK_H // 2
        thumb_cy = oy + r
        import math

        # ── 공통: 검정 둥근 트랙 ──
        track_color = "#1a1a1a" if self.is_dark_mode else "#1a1a1a"
        c.create_arc(ox, oy, ox + TRACK_H, oy + TRACK_H,
                     start=90, extent=180, fill=track_color, outline=track_color)
        c.create_arc(ox + TRACK_W - TRACK_H, oy, ox + TRACK_W, oy + TRACK_H,
                     start=270, extent=180, fill=track_color, outline=track_color)
        c.create_rectangle(ox + r, oy, ox + TRACK_W - r, oy + TRACK_H,
                           fill=track_color, outline=track_color)

        icon_color = "#ffffff"

        if self.is_dark_mode:
            # ════════════════════════════════
            # 다크 모드: 왼쪽 thumb + 오른쪽 달
            # ════════════════════════════════

            # ── 왼쪽 흰 thumb ──
            thumb_cx = ox + THUMB_R + 3
            c.create_oval(thumb_cx - THUMB_R, thumb_cy - THUMB_R,
                          thumb_cx + THUMB_R, thumb_cy + THUMB_R,
                          fill="#ffffff", outline="#ffffff")

            # ── 오른쪽 초승달 ──
            moon_cx = ox + TRACK_W - r
            moon_cy = thumb_cy
            moon_r = 6
            c.create_oval(moon_cx - moon_r, moon_cy - moon_r,
                          moon_cx + moon_r, moon_cy + moon_r,
                          fill=icon_color, outline=icon_color)
            # 트랙색 원을 겹쳐 초승달 형태
            cut_r = moon_r - 1
            c.create_oval(moon_cx - cut_r + 4, moon_cy - cut_r - 1,
                          moon_cx + cut_r + 4, moon_cy + cut_r - 1,
                          fill=track_color, outline=track_color)

        else:
            # ════════════════════════════════
            # 라이트 모드: 왼쪽 해 + 오른쪽 thumb
            # ════════════════════════════════

            # ── 왼쪽 해 아이콘 (선 스타일) ──
            sun_cx = ox + r
            sun_cy = thumb_cy
            # 중심 원 (얇은 윤곽선)
            sr = 3.0
            c.create_oval(sun_cx - sr, sun_cy - sr,
                          sun_cx + sr, sun_cy + sr,
                          fill=icon_color, outline=icon_color)
            # 8방향 직선 광선
            ray_inner = 4.5
            ray_outer = 7.5
            for i in range(8):
                angle = math.radians(i * 45)
                x1 = sun_cx + ray_inner * math.cos(angle)
                y1 = sun_cy + ray_inner * math.sin(angle)
                x2 = sun_cx + ray_outer * math.cos(angle)
                y2 = sun_cy + ray_outer * math.sin(angle)
                c.create_line(x1, y1, x2, y2,
                              fill=icon_color, width=1.5, capstyle="round")

            # ── 오른쪽 흰 thumb ──
            thumb_cx = ox + TRACK_W - THUMB_R - 3
            c.create_oval(thumb_cx - THUMB_R, thumb_cy - THUMB_R,
                          thumb_cx + THUMB_R, thumb_cy + THUMB_R,
                          fill="#ffffff", outline="#ffffff")

        try:
            if hasattr(self, "_toggle_label"):
                self._toggle_label.configure(text="", bg=T["bg_card"])
        except: pass

    # 2-14. [_on_toggle_theme 함수] 토글 스위치 클릭 시 호출됩니다.
    #       is_dark_mode 플래그를 반전시킨 뒤 apply_theme()을 호출하여 전체 UI 색상을 즉시 전환합니다.
    def _on_toggle_theme(self, event=None):
        self.is_dark_mode = not self.is_dark_mode
        self.apply_theme()

    # 2-15. [setup_ui 함수] 프로그램의 전체 화면 레이아웃을 구성하는 핵심 함수입니다.
    #       상단 안내 문구 → 컨트롤 바(버튼·입력창) → 실시간 현황 표 → 도착 기록 표 → 로그 창
    #       순서로 위젯을 생성·배치하며, 테마 변경 시 색상을 다시 칠할 수 있도록
    #       모든 위젯을 _tw 딕셔너리에 카테고리별로 등록합니다.
    def setup_ui(self):
        # 2-15-a. [위젯 추적 저장소 초기화] apply_theme()에서 색상을 재적용할 때 참조할
        #         모든 위젯을 카테고리별로 저장하는 딕셔너리입니다.
        self._tw = {
            "status_frames": [], "status_labels": {}, "ctrl_frames": [],
            "interval_labels": [], "key_labels": [], "key_frames": [],
            "stat_key_labels": [], "tree_frames": [], "ars_entries": [],
            "buttons": [],
        }

        T = self.get_theme()

        # 2-15-b. [상단 상태 바] 화면 최상단에 프로그램 사용 안내 문구와 제작자 정보를 표시하는 영역입니다.
        frame_status = tk.Frame(self.top_container, pady=2, bg=T["bg_panel"])
        frame_status.pack(fill="x", side="top")
        self._tw["status_frames"].append(frame_status)

        # 2-15-c. [좌측 안내 영역] "정류소와 노선을 선택하고 자동 기록 시작 버튼을 누르세요." 문구를 배치합니다.
        frame_left_status = tk.Frame(frame_status, bg=T["bg_panel"])
        frame_left_status.pack(side="left", padx=10, pady=2)
        self._tw["status_frames"].append(frame_left_status)
        self.lbl_info_container = tk.Frame(frame_left_status, bg=T["bg_panel"])
        self.lbl_info_container.pack(anchor="w", pady=(2, 0))
        self._tw["status_frames"].append(self.lbl_info_container)

        # 2-15-d. [안내 문구 3분할] "정류소와 노선을 선택하고 " + "자동 기록 시작" + " 버튼을 누르세요."
        #         가운데 "자동 기록 시작"만 강조 색상(fg_accent)으로 표시합니다.
        lbl_t1 = tk.Label(self.lbl_info_container, text="정류소와 노선을 선택하고 ",
                          font=(FONT_MAIN, SZ_L, "bold"), fg=T["fg_main"], bg=T["bg_panel"])
        lbl_t1.pack(side="left")
        self._tw["status_labels"]["title_text"] = lbl_t1
        lbl_t2 = tk.Label(self.lbl_info_container, text="자동 기록 시작",
                          font=(FONT_MAIN, SZ_L, "bold"), fg=T["fg_accent"], bg=T["bg_panel"])
        lbl_t2.pack(side="left")
        self._tw["status_labels"]["title_blue"] = lbl_t2
        lbl_t3 = tk.Label(self.lbl_info_container, text=" 버튼을 누르세요.",
                          font=(FONT_MAIN, SZ_L, "bold"), fg=T["fg_main"], bg=T["bg_panel"])
        lbl_t3.pack(side="left")
        self._tw["status_labels"]["title_text3"] = lbl_t3

        # 2-15-e. [자동 저장 안내 라벨] 자동 기록 시작 전에는 안내 문구를, 시작 후에는 저장 파일명을 표시합니다.
        self.lbl_auto_save_status = tk.Label(
            frame_left_status,
            text=" ※ 자동 기록 시작 버튼을 작동시키면 도착 기록이 엑셀파일로 자동 저장됩니다.",
            font=(FONT_SUB, SZ_XS, "bold"), fg=T["fg_warn"], bg=T["bg_panel"]
        )
        self.lbl_auto_save_status.pack(anchor="w", pady=(0, 2))

        # 2-15-f. [우측 정보 영역] 제작자 이메일, 데이터 출처 링크, 공공누리 마크를 우측에 배치합니다.
        frame_right_status = tk.Frame(frame_status, bg=T["bg_panel"])
        frame_right_status.pack(side="right", padx=10, pady=2)
        self._tw["status_frames"].append(frame_right_status)

        frame_right_horiz = tk.Frame(frame_right_status, bg=T["bg_panel"])
        frame_right_horiz.pack(side="top", anchor="e")
        self._tw["status_frames"].append(frame_right_horiz)

        frame_info_block = tk.Frame(frame_right_horiz, bg=T["bg_panel"])
        frame_info_block.pack(side="left", anchor="e")
        self._tw["status_frames"].append(frame_info_block)

        # 2-15-g. [제작자 행] "[ 만든이 : 박 국 환 (이메일) ]" 형태의 라벨을 배치합니다.
        frame_right_top = tk.Frame(frame_info_block, bg=T["bg_panel"])
        frame_right_top.pack(side="top", anchor="e")
        self._tw["status_frames"].append(frame_right_top)

        lbl_m1 = tk.Label(frame_right_top, text=" [ 만든이 : 박 국 환 (",
                           font=(FONT_SUB, SZ_XS), fg=T["fg_sub"], bg=T["bg_panel"])
        lbl_m1.pack(side="left", padx=(5, 0))
        self._tw["status_labels"]["maker_bracket1"] = lbl_m1
        lbl_email = tk.Label(frame_right_top, text="ggoyong2@naver.com",
                             font=(FONT_SUB, SZ_XS, "underline"), fg=T["fg_email"],
                             bg=T["bg_panel"], cursor="hand2")
        lbl_email.pack(side="left")
        lbl_email.bind("<Button-1>", lambda e: self.send_email_link("ggoyong2@naver.com"))
        self._tw["status_labels"]["email"] = lbl_email
        lbl_m2 = tk.Label(frame_right_top, text=") ]",
                           font=(FONT_SUB, SZ_XS), fg=T["fg_sub"], bg=T["bg_panel"])
        lbl_m2.pack(side="left")
        self._tw["status_labels"]["maker_bracket2"] = lbl_m2

        # 2-15-h. [데이터 출처 링크] 공공데이터포털과 서울 데이터 광장 링크를 클릭 가능한 라벨로 배치합니다.
        frame_sources = tk.Frame(frame_info_block, bg=T["bg_panel"])
        frame_sources.pack(side="top", anchor="e")
        self._tw["status_frames"].append(frame_sources)
        lbl_link1 = tk.Label(frame_sources,
                             text="공공데이터포털 Open API (https://www.data.go.kr)",
                             font=(FONT_SUB, SZ_XS), fg=T["fg_link"], bg=T["bg_panel"], cursor="hand2")
        lbl_link1.pack(side="top", anchor="e")
        lbl_link1.bind("<Button-1>", lambda e: self.open_link("https://www.data.go.kr"))
        self._tw["status_labels"]["link1"] = lbl_link1
        lbl_link2 = tk.Label(frame_sources,
                             text="서울시 버스운행노선 정보 (https://data.seoul.go.kr)",
                             font=(FONT_SUB, SZ_XS), fg="#000080", bg=T["bg_panel"], cursor="hand2")
        lbl_link2.pack(side="top", anchor="e")
        lbl_link2.bind("<Button-1>", lambda e: self.open_link("https://data.seoul.go.kr/dataList/OA-15066/F/1/datasetView.do"))
        self._tw["status_labels"]["link2"] = lbl_link2

        # 2-15-i. [공공누리 마크 이미지] Base64로 인코딩된 공공누리 저작권 마크 PNG를 디코딩하여 표시합니다.
        #         Pillow(PIL)가 있으면 ImageTk로, 없으면 tkinter 기본 PhotoImage로 표시합니다.
        try:
            import base64 as _b64s, io as _ios
            _img_data_s = _b64s.b64decode(_GONGGONG_B64)
            try:
                from PIL import Image as _PILs, ImageTk as _PILTks
                _pil_s = _PILs.open(_ios.BytesIO(_img_data_s))
                self._gonggong_photo_status = _PILTks.PhotoImage(_pil_s)
            except ImportError:
                self._gonggong_photo_status = tk.PhotoImage(data=_GONGGONG_B64)
            lbl_gonggong_s = tk.Label(
                frame_right_horiz, image=self._gonggong_photo_status,
                bg=T["bg_panel"], bd=0, relief="flat"
            )
            lbl_gonggong_s.pack(side="left", padx=(10, 0), anchor="center")
            self._tw["status_frames"].append(lbl_gonggong_s)
        except Exception:
            pass

        # 2-15-j. [컨트롤 바] 인증키 입력, 갱신주기, 자동기록시작, 수동갱신, 엑셀저장, API현황 버튼,
        #         카운터, 라이트/다크 토글이 한 줄에 배치되는 핵심 조작 패널입니다.
        self.frame_ctrl_master = tk.Frame(self.top_container, pady=4, bg=T["bg_card"])
        self.frame_ctrl_master.pack(fill="x", side="top")
        self._tw["ctrl_frames"].append(self.frame_ctrl_master)

        frame_main_content = tk.Frame(self.frame_ctrl_master, bg=T["bg_card"])
        frame_main_content.pack(side="left", fill="x", expand=True)
        self._tw["ctrl_frames"].append(frame_main_content)

        PAD = 5
        # 2-15-k. [버튼 공통 패딩] 윈도우: ipadx=6, ipady=4 / macOS: 0 (MacButton이 자체 패딩 사용)
        BTN_IPADY = 0 if CURRENT_OS == "Darwin" else 4
        BTN_IPADX = 0 if CURRENT_OS == "Darwin" else 6

        # 2-15-l. [컨트롤 바 레이아웃] 좌측=[인증키 버튼] / 중앙=[갱신주기+시작+수동+저장] / 우측=[API현황+카운터+토글]
        frame_row2 = tk.Frame(frame_main_content, bg=T["bg_card"])
        frame_row2.pack(fill="x", side="top", pady=(2, 1))
        self._tw["ctrl_frames"].append(frame_row2)

        # 2-15-m. [좌측: 인증키 버튼] 인증키 입력/변경 버튼을 좌측에 배치합니다.
        left_r2 = tk.Frame(frame_row2, bg=T["bg_card"])
        left_r2.pack(side="left", padx=(5, 0), fill="y")
        self._tw["key_frames"].append(left_r2)

        BTN_H = 1
        self.btn_key_manage = AnyButton(
            left_r2, text="인증키 입력",
            command=self.open_key_dialog,
            width=0 if CURRENT_OS == "Darwin" else 10,
            height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_key_manage.pack(side="left", padx=(0, 6), fill="y",
                                 ipadx=BTN_IPADX, ipady=BTN_IPADY)
        self._tw["buttons"].append((self.btn_key_manage, ["normal"]))

        # 2-15-n. [중앙: 갱신주기 + 자동기록시작 + 수동갱신 + 다른이름저장]
        mid_r2 = tk.Frame(frame_row2, bg=T["bg_card"])
        mid_r2.pack(side="left", fill="y")
        self._tw["ctrl_frames"].append(mid_r2)

        lbl_ival = tk.Label(mid_r2, text="갱신주기(초):", bg=T["bg_card"],
                            font=(FONT_MAIN, SZ_S, "bold"), fg=T["fg_main"])
        lbl_ival.pack(side="left", padx=(0, 2))
        self._tw["interval_labels"].append(lbl_ival)

        self.entry_refresh_interval = tk.Entry(
            mid_r2, textvariable=self.refresh_interval_var,
            width=4, bg=T["bg_entry"], fg=T["fg_main"],
            insertbackground=T["fg_main"], font=(FONT_MONO, SZ_S),
            readonlybackground=T["bg_entry_ro"]
        )
        self.entry_refresh_interval.pack(side="left", padx=(0, PAD))
        if CURRENT_OS == "Darwin":
            self._mac_style_entry(self.entry_refresh_interval, T)

        self.btn_toggle = AnyButton(
            mid_r2, text="자동 기록 시작",
            command=self._on_toggle_monitoring,
            width=0 if CURRENT_OS == "Darwin" else 11,
            height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_toggle.pack(side="left", padx=(0, PAD),
                             ipadx=BTN_IPADX, ipady=BTN_IPADY)
        self._tw["buttons"].append((self.btn_toggle, ["normal"]))

        self.btn_manual = AnyButton(
            mid_r2, text="수동 갱신",
            command=self.manual_refresh,
            width=0 if CURRENT_OS == "Darwin" else 7,
            height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_manual.pack(side="left", padx=(0, PAD),
                             ipadx=BTN_IPADX, ipady=BTN_IPADY)
        self._tw["buttons"].append((self.btn_manual, ["normal"]))

        self.btn_save_excel = AnyButton(
            mid_r2, text="다른 이름으로 엑셀 저장",
            command=self.save_to_excel, height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_save_excel.pack(side="left", padx=(0, 8),
                                 ipadx=BTN_IPADX, ipady=BTN_IPADY)
        self._tw["buttons"].append((self.btn_save_excel, ["normal"]))

        # 2-15-o. [우측: API 호출 현황 버튼 + 카운터 + 토글]
        right_r2 = tk.Frame(frame_row2, bg=T["bg_card"])
        right_r2.pack(side="right", padx=(0, 5), fill="y")
        self._tw["ctrl_frames"].append(right_r2)

        self.btn_api_stats = AnyButton(
            right_r2, text="API 호출 현황",
            command=self.open_api_stats_window,
            width=0 if CURRENT_OS == "Darwin" else 11,
            height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_api_stats.pack(side="left", padx=(0, 4), fill="y",
                                ipadx=BTN_IPADX, ipady=BTN_IPADY)
        self._tw["buttons"].append((self.btn_api_stats, ["normal"]))

        # 2-15-p. [카운터 영역] 실시간 현황/도착 기록/기타 3개의 API 호출 합계를 숫자로 표시합니다.
        stats_outer = tk.Frame(right_r2, bg=T["bg_card"])
        stats_outer.pack(side="left", fill="y")
        stats_outer.grid_rowconfigure(0, weight=1)
        stats_outer.grid_rowconfigure(1, weight=0)
        stats_outer.grid_rowconfigure(2, weight=1)
        stats_outer.grid_columnconfigure(0, weight=1)
        self._tw["stats_outer"] = stats_outer

        self.api_stats_container = tk.Frame(stats_outer, bg=T["bg_card"])
        self.api_stats_container.grid(row=1, column=0, sticky="")
        self.stat_value_labels = {}
        stat_layout_3 = [
            ("실시간 현황", "REALTIME"),
            ("도착 기록",   "ARRIVAL"),
            ("기타",        "OTHER"),
        ]
        for c, (label, key) in enumerate(stat_layout_3):
            lbl_key = tk.Label(
                self.api_stats_container,
                text=label, font=(FONT_MONO, SZ_S, "bold"),
                fg=T["fg_counter"], bg=T["bg_card"], anchor="w"
            )
            lbl_key.grid(row=0, column=c*2, padx=(4, 1), pady=2, sticky="w")
            self._tw["stat_key_labels"].append(lbl_key)
            val_lbl = tk.Label(
                self.api_stats_container,
                text="0", font=(FONT_MONO, SZ_S, "bold"),
                fg=T["fg_counter_v"], bg=T["bg_card"], width=5, anchor="w"
            )
            val_lbl.grid(row=0, column=c*2+1, padx=(0, 4), pady=2, sticky="w")
            self.stat_value_labels[key] = val_lbl

        # 2-15-q. [토글 스위치 배치] 카운터 우측에 라이트/다크 전환 토글 스위치를 배치합니다.
        #         세로 중앙 정렬을 위해 toggle_inner 프레임으로 감쌉니다.
        toggle_wrapper = tk.Frame(right_r2, bg=T["bg_card"])
        toggle_wrapper.pack(side="left", fill="y", padx=(4, 0))
        self._tw["ctrl_frames"].append(toggle_wrapper)
        toggle_inner = tk.Frame(toggle_wrapper, bg=T["bg_card"])
        toggle_inner.pack(anchor="center", expand=True, fill="y")
        self._tw["ctrl_frames"].append(toggle_inner)
        self._build_toggle(toggle_inner)

        # 2-15-r. [초기 상태 반영] 카운터 숫자와 버튼 활성/비활성 상태를 현재 상태에 맞게 즉시 갱신합니다.
        self.update_api_counter_ui()
        self.update_button_states()

        # 2-15-s. [인증키 로드 후 버튼 표시] 저장된 인증키가 이미 검증 완료 상태이면
        #         "인증키 입력" 버튼을 "인증키 변경"(outline 스타일)으로 바꿉니다.
        if self.key_locked:
            _ok = self.get_btn_style("outline")
            _ok.pop("state", None)
            self.btn_key_manage.config(text="인증키 변경", **_ok)
            self._update_btn_tracking(self.btn_key_manage, "outline")

        # 2-15-t. [분할 패널 생성] 상단(실시간 현황 표 + 도착 기록 표)과 하단(로그 창)을
        #         세로 드래그로 크기 조절할 수 있는 PanedWindow를 생성합니다.
        self.super_paned = tk.PanedWindow(self.root, orient=tk.VERTICAL,
                                          sashrelief=tk.RAISED, sashwidth=6)
        self.super_paned.pack(fill="both", expand=True)
        self.main_paned = tk.PanedWindow(self.super_paned, orient=tk.VERTICAL,
                                         sashrelief=tk.RAISED, sashwidth=6)
        self.setup_trees()
        self.super_paned.add(self.main_paned, stretch="always", minsize=220)

        # 2-15-u. [로그 창] 화면 하단에 작업 내역을 실시간으로 기록하는 텍스트 영역입니다.
        #         state="disabled"로 사용자가 직접 편집할 수 없게 하고,
        #         스크롤바를 연결하여 이전 로그를 확인할 수 있습니다.
        log_outer_frame = tk.Frame(self.super_paned, padx=10, pady=5, bg=T["bg_root"])
        self._tw["log_outer"] = log_outer_frame
        log_scroll_frame = tk.Frame(log_outer_frame, bg=T["bg_root"])
        log_scroll_frame.pack(fill="both", expand=True)
        self.txt_log = tk.Text(log_scroll_frame, height=10, bg=T["bg_log"], fg=T["fg_log"],
                               font=(FONT_MONO, SZ_S), state="disabled")
        log_scrollbar = ttk.Scrollbar(log_scroll_frame, orient="vertical", command=self.txt_log.yview, style="App.Vertical.TScrollbar")
        self.txt_log.configure(yscrollcommand=log_scrollbar.set)
        self.txt_log.pack(side="left", fill="both", expand=True)
        log_scrollbar.pack(side="right", fill="y")
        self.super_paned.add(log_outer_frame, minsize=60, height=180)

        # 2-15-v. [초기 스타일 적용] 스크롤바 색상을 현재 테마에 맞게 즉시 적용합니다.
        self._apply_scrollbar_style(self.get_theme())

        # 2-15-w. [macOS 버튼 크기 고정] macOS에서 "자동 기록 시작" → "중지"로 텍스트가 바뀔 때
        #         버튼 크기가 줄어들지 않도록 초기 렌더링 크기를 lock_size()로 고정합니다.
        if CURRENT_OS == "Darwin":
            self.root.after_idle(self.btn_toggle.lock_size)

    # 2-16. [setup_trees 함수] 실시간 현황 표 2개와 도착 기록 표 2개를 생성하여 main_paned에 배치합니다.
    #       각 표는 정류소1·정류소2용으로 나란히 배치되며, 제목 라벨·ARS 번호 표시·검색 버튼·
    #       Treeview 표·스크롤바로 구성됩니다.
    def setup_trees(self):
        T = self.get_theme()
        # 2-16-a. [실시간 현황 표 영역] 정류소1·2 각각의 도착 예정 시간을 보여주는 Treeview 2개를 생성합니다.
        #         열 구성: 노선 / 1번차량 / 도착정보 / 2번차량 / 도착정보
        frame_rt_container = tk.Frame(self.main_paned, bg=T["bg_root"])
        self._tw["tree_frames"].append(frame_rt_container)
        rt_inner = tk.Frame(frame_rt_container, bg=T["bg_root"])
        rt_inner.pack(fill="both", expand=True)
        self._tw["tree_frames"].append(rt_inner)
        rt_inner.grid_columnconfigure(0, weight=1); rt_inner.grid_columnconfigure(1, weight=1)
        rt_inner.grid_rowconfigure(0, weight=1)

        self.btn_searches = []
        self.trees_rt = []; self.lbl_st_names = []

        for i in range(2):
            f = tk.Frame(rt_inner, bg=T["bg_root"])
            f.grid(row=0, column=i, sticky="nsew", padx=2)
            self._tw["tree_frames"].append(f)
            header = tk.Frame(f, bg=T["bg_root"])
            header.pack(fill="x", pady=2)
            self._tw["tree_frames"].append(header)
            inner_header = tk.Frame(header, bg=T["bg_root"])
            inner_header.pack(anchor="center")
            self._tw["tree_frames"].append(inner_header)

            lbl = tk.Label(inner_header, text=f"[정류소 {i+1}] 실시간 현황",
                           fg=T["fg_st1"], bg=T["bg_root"],
                           font=(FONT_SUB, SZ_M, "bold"))
            lbl.pack(side="left")
            self.lbl_st_names.append(lbl)

            ars_entry = tk.Entry(inner_header, textvariable=self.ars_ids[i],
                                 width=10, state="readonly",
                                 readonlybackground=T["bg_entry_ro"],
                                 fg=T["fg_main"], font=(FONT_MONO, SZ_M, "bold"))
            ars_entry.pack(side="left", padx=5)
            if CURRENT_OS == "Darwin":
                self._mac_style_entry(ars_entry, T)
            self._tw["ars_entries"].append(ars_entry)

            # 2-16-b. [검색 버튼] 인증키가 확정되기 전에는 비활성(disabled) 상태이며,
            #         인증키 검증 성공 후 활성화됩니다.
            _s_search = self.get_btn_style("disabled")
            _s_search.pop("state", None)
            btn_search = AnyButton(inner_header, text="검색",
                                   command=lambda idx=i: self.open_search_window(idx),
                                   width=0 if CURRENT_OS == "Darwin" else 5,
                                   state="disabled", **_s_search)
            btn_search.pack(side="left", padx=(4, 0),
                            ipady=0 if CURRENT_OS == "Darwin" else 1)
            self.btn_searches.append(btn_search)
            self._tw["buttons"].append((btn_search, ["disabled"]))

            tree_frame = tk.Frame(f, bg=T["bg_root"])
            tree_frame.pack(fill="both", expand=True)
            self._tw["tree_frames"].append(tree_frame)
            cols = ("route", "bus1_no", "bus1_msg", "bus2_no", "bus2_msg")
            tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
            for col in cols: tree.heading(col, text=self.get_col_name(col))
            tree.column("route", width=69); tree.column("bus1_no", width=86)
            tree.column("bus1_msg", width=130); tree.column("bus2_no", width=86)
            tree.column("bus2_msg", width=129)
            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview, style="App.Vertical.TScrollbar")
            tree.configure(yscrollcommand=vsb.set)
            tree.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")
            self.trees_rt.append(tree)
        self.main_paned.add(frame_rt_container, minsize=100)

        # 2-16-c. [도착 기록 표 영역] 실제로 도착한 버스의 기록을 누적 표시하는 Treeview 2개를 생성합니다.
        #         열 구성: 데이터 시각 / 노선 / 차량번호 / 운수사명 / 상태
        frame_hist_container = tk.Frame(self.main_paned, bg=T["bg_root"])
        self._tw["tree_frames"].append(frame_hist_container)
        hist_inner = tk.Frame(frame_hist_container, bg=T["bg_root"])
        hist_inner.pack(fill="both", expand=True)
        self._tw["tree_frames"].append(hist_inner)
        hist_inner.grid_columnconfigure(0, weight=1); hist_inner.grid_columnconfigure(1, weight=1)
        hist_inner.grid_rowconfigure(0, weight=1)

        self.trees_hist = []; self.lbl_hist_titles = []
        for i in range(2):
            f = tk.Frame(hist_inner, bg=T["bg_root"])
            f.grid(row=0, column=i, sticky="nsew", padx=2)
            self._tw["tree_frames"].append(f)
            header = tk.Frame(f, bg=T["bg_root"])
            header.pack(fill="x", pady=2)
            self._tw["tree_frames"].append(header)
            inner_header = tk.Frame(header, bg=T["bg_root"])
            inner_header.pack(anchor="center")
            self._tw["tree_frames"].append(inner_header)

            lbl = tk.Label(inner_header, text=f"[정류소 {i+1}] 도착 기록",
                           fg=T["fg_st2"], bg=T["bg_root"],
                           font=(FONT_SUB, SZ_M, "bold"))
            lbl.pack(side="left")
            self.lbl_hist_titles.append(lbl)

            _s_del = self.get_btn_style("normal", font_size=SZ_XXS)
            btn_del = AnyButton(inner_header, text="기록 삭제",
                                command=lambda idx=i: self.clear_history(idx),
                                **_s_del)
            btn_del.pack(side="left", padx=10,
                         ipady=0 if CURRENT_OS == "Darwin" else 1)
            self._tw["buttons"].append((btn_del, ["normal"]))

            tree_frame = tk.Frame(f, bg=T["bg_root"])
            tree_frame.pack(fill="both", expand=True)
            self._tw["tree_frames"].append(tree_frame)
            cols = ("data_time", "route", "veh_no", "corp", "status")
            tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
            for col in cols: tree.heading(col, text=self.get_col_name(col))
            tree.column("data_time", width=135); tree.column("route", width=63)
            tree.column("veh_no", width=94); tree.column("corp", width=135)
            tree.column("status", width=73)
            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview, style="App.Vertical.TScrollbar")
            tree.configure(yscrollcommand=vsb.set)
            tree.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")
            self.trees_hist.append(tree)
        self.main_paned.add(frame_hist_container, minsize=100)

        # 2-16-d. [초기 표 스타일 적용] 방금 생성된 표들에 현재 테마의 Treeview 스타일을 즉시 적용합니다.
        self._apply_treeview_style(T)

    # 2-17. [get_btn_style 함수] 버튼 타입(normal/outline/disabled)과 현재 테마에 따라
    #       버튼에 적용할 색상·글꼴·커서·테두리 등의 스타일 딕셔너리를 반환합니다.
    #       macOS와 윈도우/리눅스의 스타일을 각각 다르게 처리하며,
    #       구버전 색상 코드(#28a745 등)도 자동으로 타입명으로 변환합니다.
    def get_btn_style(self, btn_type="normal", font_size=None):
        # 2-17-a. [구버전 호환] 이전 버전에서 색상 코드로 버튼 타입을 지정하던 방식을
        #         "normal"/"outline" 문자열로 자동 변환합니다.
        _color_map = {
            "#28a745": "normal", "#007bff": "normal", "#8e44ad": "normal",
            "#f1c40f": "normal", "#2ecc71": "normal",
            "#d63031": "outline", "#7f8c8d": "outline",
        }
        if btn_type.startswith("#"):
            btn_type = _color_map.get(btn_type, "normal")

        target_sz = font_size if font_size else SZ_S
        T = self.get_theme()
        CURSOR_HAND = "pointinghand" if CURRENT_OS == "Darwin" else "hand2"

        if CURRENT_OS == "Darwin":
            # 2-17-b. [macOS 버튼 스타일] MacButton은 배경색을 직접 제어하므로 얇은 테두리(1px) +
            #         정밀한 색상 값으로 macOS 네이티브 느낌의 버튼을 구현합니다.
            if not self.is_dark_mode:
                _bdr_n = "#c7c7c7"
                _bdr_o = "#065fd4"
                _bdr_d = "#d0d0d0"
                if btn_type == "disabled":
                    return {
                        "bg": "#e5e5e5", "fg": "#888888",
                        "activebackground": "#e5e5e5", "activeforeground": "#888888",
                        "disabledforeground": "#888888",
                        "font": (FONT_MAIN, target_sz, "normal"), "cursor": "arrow",
                        "relief": "solid", "bd": 0,
                        "highlightthickness": 1, "highlightbackground": _bdr_d,
                        "padx": 2, "pady": 0, "state": "normal",
                    }
                elif btn_type == "outline":
                    return {
                        "bg": "#f2f2f2", "fg": "#065fd4",
                        "activebackground": "#e0e0e0", "activeforeground": "#065fd4",
                        "font": (FONT_MAIN, target_sz, "bold"), "cursor": CURSOR_HAND,
                        "relief": "solid", "bd": 0,
                        "highlightthickness": 1, "highlightbackground": _bdr_o,
                        "padx": 2, "pady": 0,
                    }
                else:
                    return {
                        "bg": "#f2f2f2", "fg": "#0f0f0f",
                        "activebackground": "#e0e0e0", "activeforeground": "#0f0f0f",
                        "font": (FONT_MAIN, target_sz, "bold"), "cursor": CURSOR_HAND,
                        "relief": "solid", "bd": 0,
                        "highlightthickness": 1, "highlightbackground": _bdr_n,
                        "padx": 2, "pady": 0,
                    }
            else:
                _bdr_n = "#555555"
                _bdr_o = "#3ea6ff"
                _bdr_d = "#333333"
                if btn_type == "disabled":
                    return {
                        "bg": "#1a1a1a", "fg": "#555555",
                        "activebackground": "#1a1a1a", "activeforeground": "#555555",
                        "disabledforeground": "#555555",
                        "font": (FONT_MAIN, target_sz, "normal"), "cursor": "arrow",
                        "relief": "solid", "bd": 0,
                        "highlightthickness": 1, "highlightbackground": _bdr_d,
                        "padx": 2, "pady": 0, "state": "normal",
                    }
                elif btn_type == "outline":
                    return {
                        "bg": "#272727", "fg": "#3ea6ff",
                        "activebackground": "#3f3f3f", "activeforeground": "#3ea6ff",
                        "font": (FONT_MAIN, target_sz, "bold"), "cursor": CURSOR_HAND,
                        "relief": "solid", "bd": 0,
                        "highlightthickness": 1, "highlightbackground": _bdr_o,
                        "padx": 2, "pady": 0,
                    }
                else:
                    return {
                        "bg": "#272727", "fg": "#f1f1f1",
                        "activebackground": "#3f3f3f", "activeforeground": "#f1f1f1",
                        "font": (FONT_MAIN, target_sz, "bold"), "cursor": CURSOR_HAND,
                        "relief": "solid", "bd": 0,
                        "highlightthickness": 1, "highlightbackground": _bdr_n,
                        "padx": 2, "pady": 0,
                    }
        else:
            # 2-17-c. [윈도우/리눅스 버튼 스타일] tk.Button의 relief, bd 설정으로 윈도우 네이티브 느낌을 구현합니다.
            if btn_type == "disabled":
                return {
                    "bg": T["btn_disabled_bg"],
                    "fg": T["btn_disabled_fg"],
                    "font": (FONT_MAIN, target_sz, "normal"),
                    "relief": "groove", "bd": 2,
                    "highlightbackground": T["btn_disabled_border"],
                    "highlightthickness": 1,
                    "cursor": "arrow",
                    "state": "disabled",
                    "disabledforeground": T["btn_disabled_fg"],
                }
            elif btn_type == "outline":
                return {
                    "bg": T["btn_outline_bg"],
                    "fg": T["btn_outline_fg"],
                    "activebackground": T["btn_outline_bg"],
                    "activeforeground": T["btn_outline_fg"],
                    "font": (FONT_MAIN, target_sz, "bold"),
                    "relief": "ridge", "bd": 2,
                    "highlightbackground": T["btn_outline_border"],
                    "highlightthickness": 1,
                    "cursor": CURSOR_HAND,
                    "overrelief": "solid",
                }
            else:
                return {
                    "bg": T["btn_normal_bg"],
                    "fg": T["btn_normal_fg"],
                    "activebackground": T["btn_normal_active"],
                    "activeforeground": T["btn_normal_fg"],
                    "font": (FONT_MAIN, target_sz, "bold"),
                    "relief": "raised", "bd": 2,
                    "highlightbackground": T["btn_normal_border"],
                    "highlightthickness": 1,
                    "cursor": CURSOR_HAND,
                    "overrelief": "solid",
                }

    # 2-18. [get_col_name 함수] Treeview 열 코드(route, bus1_no 등)를 화면에 표시할 한글 이름으로 변환합니다.
    def get_col_name(self, code):
        mapping = {"route": "노선", "bus1_no": "1번차량", "bus1_msg": "도착정보", "bus2_no": "2번차량", "bus2_msg": "도착정보", "data_time": "데이터 시각", "veh_no": "차량번호", "corp": "운수사명", "status": "상태"}
        return mapping.get(code, code)

    # 2-19. [open_link 함수] 운영체제의 기본 웹 브라우저로 지정된 URL을 엽니다.
    def open_link(self, url):
        webbrowser.open_new(url)

    # 2-20. [send_email_link 함수] 운영체제의 기본 이메일 클라이언트로 새 메일 작성 창을 띄웁니다.
    def send_email_link(self, email):
        webbrowser.open_new(f"mailto:{email}")

    # 2-21. [prevent_column_resize 함수] Treeview 표의 열 구분선(separator)을 드래그하여
    #       열 너비를 변경하는 것을 차단합니다. "break"를 반환하면 이벤트가 전파되지 않습니다.
    def prevent_column_resize(self, event):
        if event.widget.identify_region(event.x, event.y) == "separator":
            return "break"

    # 2-22. [open_api_stats_window 함수] API 호출 상세 통계 팝업 창을 엽니다.
    #       '오늘의 합계'와 '어제의 합계' 두 개의 탭으로 구성되며,
    #       각 탭에는 API 종류별 메인키/백업키 개별 호출 횟수와 합계를 표로 표시합니다.
    #       창이 이미 열려있으면 새 창을 만들지 않고 기존 창을 맨 앞으로 가져옵니다.
    #       1초마다 자동으로 최신 통계를 갱신하며, 창을 닫으면 갱신 루프가 자동 중단됩니다.
    def open_api_stats_window(self):
        # 2-22-a. [중복 방지] 이미 열려있으면 기존 창을 최상위로 올립니다.
        if self.stats_win is not None and self.stats_win.winfo_exists():
            self.stats_win.lift()
            self.stats_win.focus_force()
            return

        T = self.get_theme()

        self.stats_win = tk.Toplevel(self.root)
        self.stats_win.title("API 호출 상세 현황")
        self.stats_win.geometry("785x460")
        self.stats_win.minsize(735, 420)
        self.stats_win.transient(self.root)
        self.stats_win.protocol("WM_DELETE_WINDOW", self.on_stats_win_close)
        self.stats_win.configure(bg=T["bg_panel"])

        lbl_title = tk.Label(self.stats_win,
                 text="API 호출 상세 통계",
                 font=(FONT_MAIN, SZ_M, "bold"),
                 bg=T["bg_panel"], fg=T["fg_main"], pady=8)
        lbl_title.pack(side="top", fill="x")

        # 2-22-b. [탭 스타일 설정] 오늘/어제 탭의 배경·전경·선택 색상을 현재 테마에 맞게 설정합니다.
        sty = ttk.Style()
        if self.is_dark_mode:
            sty.configure("StatsNB.TNotebook",
                          background=T["bg_panel"], borderwidth=0)
            sty.configure("StatsNB.TNotebook.Tab",
                          background=T["bg_card"], foreground=T["fg_sub"],
                          padding=[8, 3])
            sty.map("StatsNB.TNotebook.Tab",
                    background=[("selected", T["bg_tree_sel"])],
                    foreground=[("selected", T["fg_main"])])
        else:
            sty.configure("StatsNB.TNotebook",
                          background="#ecf0f1", borderwidth=0)
            sty.configure("StatsNB.TNotebook.Tab",
                          background="#dfe6e9", foreground="#2d3436",
                          padding=[8, 3])
            sty.map("StatsNB.TNotebook.Tab",
                    background=[("selected", "#ffffff")],
                    foreground=[("selected", "#2d3436")])

        notebook = ttk.Notebook(self.stats_win, style="StatsNB.TNotebook")
        notebook.pack(fill="both", expand=True, padx=6, pady=(0, 6))

        # 2-22-c. [표 생성 헬퍼] 통계 탭 안에 Treeview + Scrollbar 조합을 만드는 내부 함수입니다.
        def make_tree(parent):
            parent.configure(bg=T["bg_panel"])
            cols = ("abbr", "url", "main", "back", "total")
            t = ttk.Treeview(parent, columns=cols, show="headings",
                             style="Stats.Treeview")
            t.heading("abbr",  text="사용 API")
            t.heading("url",   text="API 실제 호출 주소")
            t.heading("main",  text="메인키")
            t.heading("back",  text="백업키")
            t.heading("total", text="합계")
            t.column("abbr",  width=149, anchor="w",      stretch=False)
            t.column("url",   width=395, anchor="w",      stretch=True)
            t.column("main",  width=60,  anchor="center", stretch=False)
            t.column("back",  width=60,  anchor="center", stretch=False)
            t.column("total", width=65,  anchor="center", stretch=False)
            t.bind('<Button-1>', self.prevent_column_resize)
            sb = ttk.Scrollbar(parent, orient="vertical", command=t.yview, style="App.Vertical.TScrollbar")
            t.configure(yscrollcommand=sb.set)
            t.pack(side="left", fill="both", expand=True, padx=(4, 0), pady=4)
            sb.pack(side="right", fill="y", pady=4, padx=(0, 4))
            return t

        # 2-22-d. [통계 전용 Treeview 스타일] 메인 화면의 표와 독립적인 스타일을 적용합니다.
        if self.is_dark_mode:
            sty.configure("Stats.Treeview",
                background=T["bg_tree"], foreground=T["fg_tree"],
                fieldbackground=T["bg_tree"], rowheight=22)
            sty.configure("Stats.Treeview.Heading",
                background=T["bg_card"], foreground=T["fg_main"], relief="flat")
            sty.map("Stats.Treeview",
                background=[("selected", T["bg_tree_sel"])],
                foreground=[("selected", T["fg_main"])])
            sty.map("Stats.Treeview.Heading",
                background=[("active", T["border"])])
        else:
            sty.configure("Stats.Treeview",
                background="#ffffff", foreground="#000000",
                fieldbackground="#ffffff", rowheight=22)
            sty.configure("Stats.Treeview.Heading",
                background="#f0f0f0", foreground="#000000", relief="raised")
            sty.map("Stats.Treeview",
                background=[("selected", "#0078d7")],
                foreground=[("selected", "#ffffff")])
            sty.map("Stats.Treeview.Heading",
                background=[("active", "#e0e0e0")])

        tab_today = tk.Frame(notebook, bg=T["bg_panel"])
        notebook.add(tab_today, text="  오늘의 합계  ")
        tree_today = make_tree(tab_today)

        tab_yest = tk.Frame(notebook, bg=T["bg_panel"])
        notebook.add(tab_yest, text="  어제의 합계  ")
        tree_yest = make_tree(tab_yest)

        # 2-22-e. [API 이름 매핑] 내부 약어(ARR1 등)를 사용자가 읽기 쉬운 한글 이름으로 변환합니다.
        _api_name_map = {
            "ARR1": "실시간 현황 API",
            "ARR2": "실시간 현황 예비 API",
            "SINF": "실시간 현황 예비2 API",
            "POS1": "도착 기록 API",
            "POS2": "도착 기록 예비 API",
            "SCNM": "정류소명 검색 API",
            "SCID": "정류소번호 검색 API",
            "RINF": "노선기본정보 API",
            "SLST": "노선전체정류소 API",
            "VID":  "차량번호확인 API",
            "VLD":  "인증키검증용 API",
        }

        # 2-22-f. [표 데이터 채우기] API 통계 딕셔너리를 읽어 표의 각 행에 데이터를 삽입합니다.
        #         마지막 행에는 전체 합계를 표시합니다.
        def fill_tree(t, stats_dict, by_key_dict, label_row):
            for item in t.get_children():
                t.delete(item)
            total_all = total_main = total_back = 0
            for key, count in stats_dict.items():
                url   = self.api_urls.get(key, "-")
                m_cnt = by_key_dict["main"].get(key, 0)
                b_cnt = by_key_dict["back"].get(key, 0)
                api_label = _api_name_map.get(key, key)
                t.insert("", "end", values=(api_label, url, f"{m_cnt}", f"{b_cnt}", f"{count}"))
                total_all  += count
                total_main += m_cnt
                total_back += b_cnt
            t.insert("", "end", values=(" ", " ", " ", " ", " "))
            t.insert("", "end", values=(label_row, " - ",
                                        f"{total_main}", f"{total_back}", f"{total_all}"))

        # 2-22-g. [1초 주기 자동 갱신] 창이 열려있는 동안 1초마다 오늘·어제 탭을 최신 데이터로 갱신합니다.
        #         창이 닫히면(winfo_exists 실패) 루프가 자동 종료됩니다.
        def update_stats_loop():
            if not self.stats_win or not self.stats_win.winfo_exists():
                return
            fill_tree(tree_today, self.api_stats,
                      self.api_stats_by_key, "합계")
            fill_tree(tree_yest, self.api_stats_yesterday,
                      self.api_stats_by_key_yesterday, "합계")
            self.stats_win.after(1000, update_stats_loop)

        update_stats_loop()

    # 2-23. [on_stats_win_close 함수] 통계 창을 닫을 때 Toplevel 위젯을 파괴하고
    #       stats_win 참조를 None으로 초기화하여 다음 열기 시 새 창을 생성할 수 있게 합니다.
    def on_stats_win_close(self):
        if self.stats_win:
            self.stats_win.destroy()
            self.stats_win = None

    # 2-24. [update_button_states 함수] 현재 프로그램 상태(정류소 설정 여부, 모니터링 ON/OFF)에 따라
    #       자동기록시작/수동갱신 버튼의 텍스트·활성 상태·연결 함수를 업데이트합니다.
    #       모니터링 중이면 시작 버튼을 "중지"(outline)로, 아니면 "자동 기록 시작"(normal)으로 표시합니다.
    def update_button_states(self):
        # 2-24-a. 정류소1 또는 정류소2 중 하나라도 노선이 설정되어 있는지 확인합니다.
        has_info = (bool(self.target_st_info[0].get('routes')) or
                    bool(self.target_st_info[1].get('routes')))

        # 2-24-b. [버튼 설정 헬퍼] 버튼 1개의 타입·활성 여부·연결 함수·텍스트를 한 번에 설정합니다.
        def apply_btn(btn, btn_type, is_active, cmd, text=None):
            effective_type = btn_type if is_active else "disabled"
            style = self.get_btn_style(effective_type)
            if text:
                style["text"] = text
            style.pop("state", None)
            self._update_btn_tracking(btn, effective_type)
            if CURRENT_OS == "Darwin":
                # 2-24-c. [macOS 활성 플래그 동기화] MacButton의 state="disabled"가 클릭을 완전히
                #         차단하지 못하는 버그를 우회하기 위해 별도 플래그를 관리합니다.
                if btn is self.btn_toggle:
                    self._btn_active['toggle'] = is_active
                elif btn is self.btn_manual:
                    self._btn_active['manual'] = is_active
                btn.config(state=("normal" if is_active else "disabled"),
                           command=(cmd if is_active else lambda: None), **style)
            else:
                btn.config(state=("normal" if is_active else "disabled"),
                           command=(cmd if is_active else lambda: None), **style)

        if self.is_monitoring:
            # 2-24-d. [모니터링 중] 시작 버튼 → "중지"(outline), 수동 갱신 → 활성화
            apply_btn(self.btn_toggle, "outline", True,
                      self.stop_monitoring, text="중지")
            apply_btn(self.btn_manual, "normal", True, self.manual_refresh)
        else:
            # 2-24-e. [대기 중] 시작 버튼 → 정류소 설정 여부에 따라 활성/비활성
            apply_btn(self.btn_toggle, "normal", has_info,
                      self.start_monitoring, text="자동 기록 시작")
            apply_btn(self.btn_manual, "normal", has_info, self.manual_refresh)

    # 2-25. [_on_toggle_monitoring 함수] 자동기록시작/중지 통합 버튼 클릭 시 호출됩니다.
    #       현재 모니터링 상태에 따라 start_monitoring() 또는 stop_monitoring()으로 분기합니다.
    def _on_toggle_monitoring(self):
        # 2-25-a. [macOS 비활성 클릭 차단] disabled 상태에서도 클릭이 발생하는 macOS 버그를 플래그로 차단합니다.
        if CURRENT_OS == "Darwin" and not self._btn_active.get('toggle', True):
            return
        if self.is_monitoring:
            self.stop_monitoring()
        else:
            self.start_monitoring()

    # 2-26. [_trigger_counter_update 함수] 어떤 스레드에서 호출해도 안전하게 메인 화면의
    #       API 호출 카운터 숫자를 갱신합니다. root.after(0, ...)은 tkinter 이벤트 루프에
    #       갱신을 위임하므로 스레드 안전합니다.
    def _trigger_counter_update(self):
        self.root.after(0, self.update_api_counter_ui)

    # 2-27. [_update_btn_tracking 함수] _tw["buttons"] 리스트에서 해당 버튼의 현재 타입 문자열을 갱신합니다.
    #       apply_theme()에서 버튼 색상을 재적용할 때 이 타입을 참조하므로,
    #       버튼 스타일이 바뀔 때마다 반드시 호출해야 합니다.
    def _update_btn_tracking(self, btn, type_str):
        for item in self._tw.get("buttons", []):
            if item[0] is btn:
                item[1][0] = type_str
                break

    # 2-28. [update_api_counter_ui 함수] 화면 상단 카운터의 '실시간 현황' / '도착 기록' / '기타'
    #       숫자를 현재 api_stats에서 계산하여 라벨에 표시합니다.
    #       실시간 현황 = ARR1 + ARR2 + SINF, 도착 기록 = POS1 + POS2, 기타 = 나머지 전부
    def update_api_counter_ui(self):
        s = self.api_stats
        realtime = s.get("ARR1",0) + s.get("ARR2",0) + s.get("SINF",0)
        arrival  = s.get("POS1",0) + s.get("POS2",0)
        other    = s.get("SCNM",0) + s.get("SCID",0) + s.get("RINF",0) + s.get("SLST",0) + s.get("VID",0) + s.get("VLD",0)
        for key, val_lbl in self.stat_value_labels.items():
            if key == "REALTIME": val_lbl.config(text=str(realtime))
            elif key == "ARRIVAL": val_lbl.config(text=str(arrival))
            elif key == "OTHER":   val_lbl.config(text=str(other))

    # 2-29. [load_saved_key 함수] 프로그램 시작 시 key.cfg 파일이 존재하면 암호화된 인증키를 복호화하여
    #       메인/백업 인증키 변수에 로드하고, 자동으로 서버 유효성 검사를 시작합니다.
    #       파일 형식: MASTER_KEY로 Fernet 암호화된 "메인키|백업키" 문자열
    def load_saved_key(self):
        if os.path.exists(self.key_file_path):
            try:
                with open(self.key_file_path, "rb") as f:
                    encrypted_data = f.read()
                    if encrypted_data:
                        # 2-29-a. [복호화] cipher_suite(Fernet)으로 암호문을 평문으로 변환합니다.
                        decrypted_data = cipher_suite.decrypt(encrypted_data).decode('utf-8')
                        main_k, back_k = "", ""
                        # 2-29-b. "|" 구분자로 메인키와 백업키를 분리합니다.
                        if "|" in decrypted_data:
                            main_k, back_k = decrypted_data.split("|", 1)
                        else:
                            main_k = decrypted_data

                        self.service_key_var.set(main_k)
                        self.backup_key_var.set(back_k)

                        if main_k.strip() or back_k.strip():
                            # 2-29-c. UI(로그 창)가 준비될 때까지 짧은 지연 후 자동 검증을 시작합니다.
                            self.root.after(100, lambda: self.log("🚀 저장된 인증키를 불러왔습니다. 유효성 검사를 시작합니다..."))
                            self.root.after(150, self.toggle_key_lock)

            except Exception as e:
                print(f"인증키 복호화 실패: {e}")

    # 2-30. [save_key_to_file 함수] 현재 메인/백업 인증키를 "메인키|백업키" 형태로 합친 뒤
    #       Fernet(MASTER_KEY)으로 암호화하여 key.cfg 파일에 저장합니다.
    def save_key_to_file(self):
        main_k = self.service_key_var.get().strip()
        back_k = self.backup_key_var.get().strip()

        combined_data = f"{main_k}|{back_k}"

        try:
            encrypted_data = cipher_suite.encrypt(combined_data.encode('utf-8'))
            with open(self.key_file_path, "wb") as f:
                f.write(encrypted_data)
            self.log("🔒 메인/백업 인증키가 암호화되어 안전하게 저장되었습니다.")
        except Exception as e:
            self.log(f"⚠ 인증키 암호화 저장 실패: {e}")
            
    # 2-31. [open_key_dialog 함수] 인증키를 새로 입력하거나 변경하는 모달(grab_set) 팝업 창을 띄웁니다.
    #       메인 인증키(64자리 필수)와 백업 인증키(0 또는 64자리)를 입력받고,
    #       64자리 조건을 만족해야 '인증키 입력' 버튼이 활성화됩니다.
    #       버튼 클릭 시 _validate_and_apply_keys()로 서버 검증을 수행하고,
    #       성공하면 key.cfg에 저장 후 창을 닫습니다.
    def open_key_dialog(self):
        T = self.get_theme()

        key_win = tk.Toplevel(self.root)
        key_win.title("인증키 입력")
        # 2-31-a. [모달 설정] grab_set()으로 이 창이 열려있는 동안 메인 창 조작을 차단합니다.
        key_win.grab_set()
        key_win.resizable(False, False)
        key_win.configure(bg=T["bg_card"])

        pad_f = tk.Frame(key_win, bg=T["bg_card"], padx=18, pady=16)
        pad_f.pack(fill="both", expand=True)

        # 2-31-b. [임시 변수] 확정 전까지 service_key_var에 직접 쓰지 않고 임시 변수로 관리합니다.
        #         검증 실패 시 원래 키를 보존하기 위함입니다.
        tmp_main = tk.StringVar(value=self.service_key_var.get())
        tmp_back = tk.StringVar(value=self.backup_key_var.get())

        # 2-31-c. [입력 행 생성 헬퍼] 라벨 + 비밀번호 입력창(show="*") 한 쌍을 만드는 내부 함수입니다.
        def _make_row(parent, label_text, var):
            row = tk.Frame(parent, bg=T["bg_card"])
            row.pack(fill="x", pady=4)
            tk.Label(row, text=label_text, font=(FONT_MAIN, SZ_S),
                     bg=T["bg_card"], fg=T["fg_main"], width=10, anchor="e").pack(side="left")
            ent = tk.Entry(row, textvariable=var, show="*",
                           width=68, bg=T["bg_entry"], fg=T["fg_main"],
                           insertbackground=T["fg_main"], font=(FONT_MONO, SZ_XXS),
                           readonlybackground=T["bg_entry_ro"])
            ent.pack(side="left", padx=(6, 0))
            if CURRENT_OS == "Darwin":
                self._mac_style_entry(ent, T)
            return ent

        ent_main = _make_row(pad_f, "메인 인증키 :", tmp_main)
        ent_back = _make_row(pad_f, "백업 인증키 :", tmp_back)

        # 2-31-d. [안내 문구] 메인 64자리 필수, 백업 0 또는 64자리 선택 안내를 표시합니다.
        lbl_hint = tk.Label(pad_f,
            text="  ※ 메인 인증키는 64자리 필수, 백업 인증키는 선택(0 또는 64자리)입니다.",
            font=(FONT_SUB, SZ_XS), fg=T["fg_warn"], bg=T["bg_card"])
        lbl_hint.pack(anchor="w", pady=(2, 8))

        # 2-31-e. [버튼 영역] '인증키 입력'(초기 비활성)과 '취소' 버튼을 배치합니다.
        btn_row = tk.Frame(pad_f, bg=T["bg_card"])
        btn_row.pack(fill="x")

        _dis_style = self.get_btn_style("disabled")
        _dis_style.pop("state", None)
        btn_apply = AnyButton(btn_row, text="인증키 입력",
                              command=lambda: _do_apply(),
                              state="disabled",
                              **_dis_style)
        btn_apply.pack(side="left", padx=(0, 8),
                       ipady=0 if CURRENT_OS == "Darwin" else 4)

        AnyButton(btn_row, text="취소",
                  command=key_win.destroy,
                  **self.get_btn_style("normal")).pack(
            side="left", ipady=0 if CURRENT_OS == "Darwin" else 4)

        # 2-31-f. [64자리 검증 → 버튼 활성/비활성] 입력창 내용이 변경될 때마다 호출되어
        #         메인키 64자리 + 백업키 0 또는 64자리 조건을 만족하면 버튼을 활성화합니다.
        def _check_apply_btn(*_):
            ml = len(tmp_main.get().strip())
            bl = len(tmp_back.get().strip())
            ok = (ml == 64) and (bl == 0 or bl == 64)
            if ok:
                _s = self.get_btn_style("normal"); _s.pop("state", None)
                btn_apply.config(state="normal", **_s)
            else:
                _s = self.get_btn_style("disabled"); _s.pop("state", None)
                btn_apply.config(state="disabled", **_s)

        # 2-31-g. [실시간 감시] tkinter 변수의 trace_add를 사용하여 키 입력이 바뀔 때마다
        #         _check_apply_btn을 자동 호출합니다.
        tmp_main.trace_add("write", _check_apply_btn)
        tmp_back.trace_add("write", _check_apply_btn)
        _check_apply_btn()

        # 2-31-h. [인증키 입력 버튼 로직] 서버 검증을 실행하고 성공 시 창을 닫습니다.
        def _do_apply():
            main_input = tmp_main.get().strip()
            back_input = tmp_back.get().strip()
            if not main_input:
                messagebox.showwarning("알림", "메인 인증키를 입력해주세요.", parent=key_win)
                return
            # 2-31-h-i. 검증 중 버튼 중복 클릭을 방지하기 위해 즉시 비활성화합니다.
            _ds2 = self.get_btn_style("disabled"); _ds2.pop("state", None)
            btn_apply.config(state="disabled", **_ds2)
            key_win.update()

            # 2-31-h-ii. [검증 실패 콜백] 해당 키 변수를 빈 문자열로 초기화합니다.
            def _on_bad_key(var_obj):
                var_obj.set("")

            success = self._validate_and_apply_keys(
                main_input, back_input,
                on_bad_main=lambda: _on_bad_key(tmp_main),
                on_bad_back=lambda: _on_bad_key(tmp_back),
                save_on_success=True,
                parent_win=key_win
            )
            if success:
                key_win.destroy()
            else:
                # 2-31-h-iii. 검증 실패 시 다시 입력할 수 있도록 버튼 상태를 재검사합니다.
                _check_apply_btn()

        # 2-31-i. [창 크기 자동 조정] 내용물에 맞게 창 크기를 자동으로 결정하고 첫 번째 입력창에 포커스를 줍니다.
        key_win.update_idletasks()
        key_win.geometry("")
        ent_main.focus_set()

    # 2-32. [_validate_and_apply_keys 함수] 메인/백업 인증키를 서울시 API 서버에 실제로 요청하여
    #       유효성을 검증하고, 하나 이상 성공하면 프로그램에 적용하는 핵심 검증 함수입니다.
    #       수동 입력(open_key_dialog)과 자동 로드(toggle_key_lock) 양쪽 경로에서 공통 사용됩니다.
    #       Parameters:
    #         main_input      : 검증할 메인 인증키 문자열
    #         back_input      : 검증할 백업 인증키 문자열 (비어있으면 건너뜀)
    #         on_bad_main     : 메인키 실패 시 호출할 콜백 (예: 입력창 초기화)
    #         on_bad_back     : 백업키 실패 시 호출할 콜백
    #         save_on_success : True이면 검증 성공 후 key.cfg에 저장 (수동 입력 경로)
    #         parent_win      : messagebox의 부모 창 (모달 창 위에 표시하기 위함)
    #       Returns: bool — 하나 이상의 키가 성공하면 True
    def _validate_and_apply_keys(self, main_input, back_input,
                                  on_bad_main=None, on_bad_back=None,
                                  save_on_success=False, parent_win=None):
        from urllib.parse import unquote
        import requests

        self.log("🔑 인증키 유효성 검사 중...")

        # 2-32-a. [검증 대상 준비] 메인키는 필수, 백업키는 있을 때만 검증 목록에 추가합니다.
        keys_to_test = [("메인", main_input, on_bad_main)]
        if back_input:
            keys_to_test.append(("백업", back_input, on_bad_back))

        valid_main, valid_back = "", ""
        any_success = False

        # 2-32-b. [순차 검증] 각 키를 getBusRouteList API에 실제로 요청하여 headerCd=0 응답을 확인합니다.
        for name, key_val, bad_cb in keys_to_test:
            test_url = "http://ws.bus.go.kr/api/rest/busRouteInfo/getBusRouteList"
            try:
                r = requests.get(test_url,
                                 params={'serviceKey': unquote(key_val), 'strSrch': '12'},
                                 timeout=5)
                if "<headerCd>0</headerCd>" in r.text:
                    self.log(f"✅ {name} 인증키: 정상 작동 확인")
                    if name == "메인": valid_main = key_val
                    else:             valid_back = key_val
                    any_success = True
                else:
                    if bad_cb: bad_cb()
                    self.log(f"❌ {name} 인증키: 작동 불가하여 삭제되었습니다.")
            except Exception:
                if bad_cb: bad_cb()
                self.log(f"❌ {name} 통신 에러로 삭제 처리되었습니다.")

        if any_success:
            # 2-32-c. [키 확정] 검증 성공한 키를 프로그램 전체에서 사용할 변수에 저장합니다.
            self.service_key_var.set(valid_main)
            self.backup_key_var.set(valid_back)
            self.final_main_key = valid_main
            self.final_backup_key = valid_back
            self.key_locked = True

            # 2-32-d. [VLD 카운터 업데이트] 인증키 검증에 사용된 API 호출을 VLD 카운터에 반영합니다.
            if valid_main:
                self.api_stats['VLD'] += 1
                self.api_stats_by_key["main"]['VLD'] = \
                    self.api_stats_by_key["main"].get('VLD', 0) + 1
            if valid_back:
                self.api_stats['VLD'] += 1
                self.api_stats_by_key["back"]['VLD'] = \
                    self.api_stats_by_key["back"].get('VLD', 0) + 1
            self._trigger_counter_update()

            # 2-32-e. [인증키 버튼 변경] "인증키 입력" → "인증키 변경"(outline 스타일)으로 바꿉니다.
            _outline = self.get_btn_style("outline"); _outline.pop("state", None)
            self.btn_key_manage.config(text="인증키 변경", **_outline)
            self._update_btn_tracking(self.btn_key_manage, "outline")

            # 2-32-f. [검색 버튼 활성화] 인증키가 확정되면 정류소 검색 버튼을 사용할 수 있게 합니다.
            _s_on = self.get_btn_style("normal"); _s_on.pop("state", None)
            for btn in self.btn_searches:
                btn.config(state="normal", **_s_on)
                self._update_btn_tracking(btn, "normal")

            # 2-32-g. [파일 저장] 수동 입력 경로(save_on_success=True)일 때만 key.cfg에 저장합니다.
            if save_on_success:
                self.save_key_to_file()

            self.log("🔒 인증키가 확정되었습니다. 이제 검색이 가능합니다.")
        else:
            messagebox.showerror("인증 실패", "사용 가능한 인증키가 없습니다.",
                                 **({'parent': parent_win} if parent_win else {}))

        return any_success

    # 2-33. [toggle_key_lock 함수] 프로그램 시작 시 파일에서 불러온 인증키를 서버에 자동 검증합니다.
    #       load_saved_key()에서 root.after()를 통해 호출됩니다.
    #       이미 검증 완료(key_locked=True) 상태이거나 키가 비어있으면 즉시 반환합니다.
    #       save_on_success=False로 호출하여 파일을 다시 저장하지 않습니다.
    def toggle_key_lock(self):
        if self.key_locked:
            return

        main_input = self.service_key_var.get().strip()
        back_input  = self.backup_key_var.get().strip()
        if not main_input:
            return

        self._validate_and_apply_keys(
            main_input, back_input,
            on_bad_main=lambda: self.service_key_var.set(""),
            on_bad_back=lambda: self.backup_key_var.set(""),
            save_on_success=False
        )

    # 2-34. [fetch_api 함수] 서울시 버스 공공 API에 HTTP GET 요청을 보내고 XML 응답을 파싱하여 반환하는
    #       핵심 통신 함수입니다. 프로그램의 모든 API 호출이 이 함수를 통해 이루어집니다.
    #       주 API × 메인키 → 주 API × 백업키 → 폴백 API × 메인키 → 폴백 API × 백업키
    #       순으로 최대 4번 시도하며, 하나라도 성공하면 즉시 XML 루트 요소를 반환합니다.
    #       일일 호출 한도 초과(headerCd=22) 또는 미등록 키 응답 시 자동으로 다음 시도로 넘어갑니다.
    #       운행 차량 없음(NODATA) 응답 시 ('NO_BUS', 첫차시각) 튜플을 반환합니다.
    #       모든 시도가 실패하면 None을 반환합니다.
    def fetch_api(self, url, params, api_type=None):
        from urllib.parse import unquote
        # 2-34-a. [키 준비] 현재 등록된 메인키와 백업키를 리스트로 준비합니다.
        keys = [self.service_key_var.get().strip(), self.backup_key_var.get().strip()]

        # 2-34-b. [API 주소 선택] api_type에 따라 주 API URL(p_url)과 폴백 API URL(f_url)을 결정합니다.
        #         ARR1: 주=getArrInfoByRoute, 폴백=getArrInfoByRouteAll
        #         POS2: 주=getBusPosByRtid, 폴백=getBusPosByRouteSt
        p_url, f_url = url, url
        if api_type == "ARR1":
            p_url = "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRoute"
            f_url = "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRouteAll"
        elif api_type == "POS2":
            p_url = "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRtid"
            f_url = "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRouteSt"

        # 2-34-c. [시도 계획] 주API×메인 → 주API×백업 → 폴백API×메인 → 폴백API×백업 순서로 4번 시도합니다.
        attempts = [
            (p_url, "메인", keys[0]), (p_url, "백업", keys[1]),
            (f_url, "메인", keys[0]), (f_url, "백업", keys[1])
        ]

        for step, (curr_url, name, key) in enumerate(attempts):
            if not key: continue
            try:
                # 2-34-d. [API 호출] URL 인코딩된 인증키를 디코딩하여 serviceKey 파라미터로 전달합니다.
                params['serviceKey'] = unquote(key)
                resp = requests.get(curr_url, params=params, timeout=10)

                # 2-34-e. [일일 한도 초과] 하루 호출 한도를 넘으면 다음 키/URL 조합으로 넘어갑니다.
                if "LIMITED NUMBER OF SERVICE REQUESTS" in resp.text or "<headerCd>22</headerCd>" in resp.text:
                    continue

                # 2-34-f. [미등록 키] 등록되지 않은 인증키 응답이면 다음 시도로 넘어갑니다.
                if "SERVICE KEY IS NOT REGISTERED" in resp.text or "UNREGISTERED_KEY" in resp.text:
                    continue

                if resp.status_code == 200:
                    root = ET.fromstring(resp.text)
                    header_cd = root.findtext(".//headerCd")
                    err_msg = root.findtext('.//headerMsg') or ""

                    # 2-34-g. [성공 판정] headerCd=0(정상 데이터) 또는 NODATA(데이터 없음) 모두 호출 성공으로 처리합니다.
                    is_no_result = ("결과가 없습니다" in err_msg) or ("NODATA" in err_msg)

                    if header_cd == "0" or is_no_result:
                        # 2-34-h. [API 종류 판별 및 카운터 증가] 응답 URL을 분석하여 API 종류를 식별하고
                        #         해당 종류의 합산 카운터와 키별 개별 카운터를 각각 1 증가시킵니다.
                        success_url = curr_url
                        key_type = "VLD"
                        if "getArrInfoByRoute" in success_url and "All" not in success_url: key_type = "ARR1"
                        elif "getArrInfoByRouteAll" in success_url: key_type = "ARR2"
                        elif "getBusPosByRtid" in success_url: key_type = "POS1"
                        elif "getStationByUid" in success_url: key_type = "SINF"
                        elif "getBusPosByRouteSt" in success_url: key_type = "POS2"
                        elif "getStationByName" in success_url: key_type = "SCNM"
                        elif "getRouteByStation" in success_url: key_type = "SCID"
                        elif "getRouteInfo" in success_url: key_type = "RINF"
                        elif "getStaionByRoute" in success_url: key_type = "SLST"
                        elif "getBusPosByVehId" in success_url: key_type = "VID"

                        self.api_stats[key_type] += 1
                        key_slot = "main" if name == "메인" else "back"
                        self.api_stats_by_key[key_slot][key_type] = \
                            self.api_stats_by_key[key_slot].get(key_type, 0) + 1
                        self._trigger_counter_update()

                        if header_cd == "0":
                            # 2-34-i. [정상 데이터] XML 루트 요소를 그대로 반환합니다.
                            return root
                        else:
                            # 2-34-j. [운행 없음 처리] 위치 조회 API에서 운행 차량이 없다는 응답이 오면
                            #         노선의 첫차 시각을 조회하여 ('NO_BUS', 첫차시각) 튜플로 반환합니다.
                            #         호출자(process_station)는 이 튜플을 받아 POS 일시정지를 등록합니다.
                            if any(u in success_url for u in ["getBusPosByRtid", "getBusPosByRouteSt"]):
                                rid_param = params.get('busRouteId', '')
                                rnm_log = self.rid_to_rnm.get(rid_param, rid_param)
                                self.log(f"{rnm_log}번은 운행중인 차량이 없습니다.")
                                f_tm = self._fetch_route_first_time(rid_param)
                                return ('NO_BUS', f_tm)
                            return None

            except Exception:
                # 2-34-k. [예외 처리] 네트워크 타임아웃 등 예외 발생 시 해당 시도를 건너뛰고 다음으로 진행합니다.
                continue
        return None

    # 2-35. [_fetch_route_first_time 함수] 특정 노선의 첫 번째 버스 출발 시각을
    #       노선정보 API(getRouteInfo)의 firstBusTm 필드에서 조회합니다.
    #       POS 일시정지 시간을 계산하는 데 사용됩니다.
    #       Returns: "HH:MM" 형식 문자열 또는 None(조회 실패)
    def _fetch_route_first_time(self, rid):
        if not rid:
            return None
        try:
            root_ri = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo",
                {'busRouteId': rid}
            )
            if root_ri is None or isinstance(root_ri, tuple):
                return None
            f_tm_raw = root_ri.findtext(".//firstBusTm") or ""
            if f_tm_raw:
                return self.format_hhmm(f_tm_raw)
        except Exception:
            pass
        return None

    # 2-36. [open_search_window 함수] 정류소를 검색하고 모니터링할 버스 노선을 선택하는 팝업 창을 엽니다.
    #       상단: 정류소명/ARS번호 검색 입력창 + 검색 결과 Treeview
    #       중단: 선택한 정류소를 지나는 모든 노선 목록 (체크박스 형태로 선택/해제)
    #       하단: '선택한 노선들로 적용' 버튼
    #       노선 선택 확정 시 정류소 ID, 선택된 노선 리스트, 정류소 순번(ord_map)을 저장합니다.
    def open_search_window(self, target_idx):
        # 2-36-a. [인증키 확인] 인증키 없이는 API 호출이 불가능하므로 경고 후 종료합니다.
        if not self.service_key_var.get().strip():
            messagebox.showwarning("알림", "인증키를 먼저 입력해주세요."); return

        T = self.get_theme()

        # 2-36-b. [검색 창 전용 Treeview 스타일] 메인 화면과 독립적인 스타일을 적용합니다.
        sty = ttk.Style()
        if self.is_dark_mode:
            sty.configure("Search.Treeview",
                background=T["bg_tree"], foreground=T["fg_tree"],
                fieldbackground=T["bg_tree"], rowheight=22)
            sty.configure("Search.Treeview.Heading",
                background=T["bg_card"], foreground=T["fg_main"],
                font=(FONT_SUB, SZ_S, "bold"), relief="flat")
            sty.map("Search.Treeview",
                background=[("selected", T["bg_tree_sel"])],
                foreground=[("selected", T["fg_main"])])
            sty.map("Search.Treeview.Heading",
                background=[("active", T["border"])])
        else:
            sty.configure("Search.Treeview",
                background="#ffffff", foreground="#000000",
                fieldbackground="#ffffff", rowheight=22)
            sty.configure("Search.Treeview.Heading",
                background="#dfe6e9", foreground="#2d3436",
                font=(FONT_SUB, SZ_S, "bold"), relief="raised")
            sty.map("Search.Treeview",
                background=[("selected", "#0078d7")],
                foreground=[("selected", "#ffffff")])
            sty.map("Search.Treeview.Heading",
                background=[("active", "#c8d6e5")])

        search_win = tk.Toplevel(self.root)
        search_win.title(f"정류소 {target_idx+1} 검색 및 노선 선택")
        search_win.geometry("1000x850")
        search_win.grab_set()
        search_win.minsize(500, 500)
        search_win.configure(bg=T["bg_panel"])

        # 2-36-c. [검색 입력 영역] 정류소명 또는 ARS 번호를 입력하는 행입니다.
        frame_search = tk.Frame(search_win, pady=10, bg=T["bg_panel"])
        frame_search.pack(fill="x")
        tk.Label(frame_search, text="정류소명/ID:", font=(FONT_MAIN, SZ_S),
                 bg=T["bg_panel"], fg=T["fg_main"]).pack(side="left", padx=(15,5))
        search_ent = tk.Entry(frame_search, width=30,
                              bg=T["bg_entry"], fg=T["fg_main"],
                              insertbackground=T["fg_main"])
        search_ent.pack(side="left", padx=5)
        search_ent.focus_set()

        # 2-36-d. [정류소 검색 결과 표] 검색된 정류소 이름과 ARS-ID를 표시하는 Treeview입니다.
        tree_st_frame = tk.Frame(search_win, bg=T["bg_panel"])
        tree_st_frame.pack(fill="x", padx=15, pady=5)
        st_scroll = ttk.Scrollbar(tree_st_frame, orient="vertical", style="App.Vertical.TScrollbar")
        tree_st = ttk.Treeview(tree_st_frame, columns=("name", "arsid"),
                               show="headings", height=5,
                               style="Search.Treeview",
                               yscrollcommand=st_scroll.set)
        st_scroll.config(command=tree_st.yview)
        tree_st.heading("name",  text="정류소명")
        tree_st.heading("arsid", text="ARS-ID")
        tree_st.column("name",  width=659, stretch=True)
        tree_st.column("arsid", width=141, stretch=True)
        tree_st.bind("<Button-1>", self.prevent_column_resize)
        tree_st.pack(side="left", fill="both", expand=True)
        st_scroll.pack(side="right", fill="y")

        # 2-36-e. [노선 선택 도구 행] "▼ 노선 선택" 라벨과 전체 선택/해제 버튼을 배치합니다.
        btn_action_frame = tk.Frame(search_win, bg=T["bg_panel"])
        btn_action_frame.pack(fill="x", padx=15, pady=5)
        lbl_route_sel = tk.Label(btn_action_frame, text="▼ 노선 선택",
                                  font=(FONT_SUB, SZ_S, "bold"),
                                  fg=T["fg_accent"], bg=T["bg_panel"])
        lbl_route_sel.pack(side="left")

        # 2-36-f. [노선 데이터·정렬 상태 초기화] 현재 검색 세션의 노선 데이터와 정렬 상태를 초기화합니다.
        self.current_route_data = []
        self.sort_state = {"key": None, "reverse": False}
        type_order = {"간선": 0, "지선": 1, "광역": 2, "기타": 3, "경기": 4, "인천": 5}

        def update_route_tree_view():
            render_route_list()

        def set_all_check(status):
            for item in self.current_route_data:
                item["checked"].set(status)
            update_route_tree_view()

        _btn_s = self.get_btn_style("normal", font_size=SZ_XS)
        btn_all_off = AnyButton(btn_action_frame, text="전체 해제",
                                command=lambda: set_all_check(False), **_btn_s)
        btn_all_off.pack(side="right", padx=2)
        btn_all_on = AnyButton(btn_action_frame, text="전체 선택",
                               command=lambda: set_all_check(True), **_btn_s)
        btn_all_on.pack(side="right", padx=2)

        # 2-36-g. [노선 목록 표] 정류소를 지나는 모든 노선을 표시하는 Treeview입니다.
        #         열: 선택여부(V) / 노선번호 / 유형 / 기점↔종점 / 배차간격 / 첫차 / 막차 / 정류장수
        frame_route_list = tk.Frame(search_win, bd=1, relief="sunken",
                                    bg=T["bg_panel"])
        frame_route_list.pack(fill="both", expand=True, padx=15, pady=5)
        route_scroll = ttk.Scrollbar(frame_route_list, orient="vertical", style="App.Vertical.TScrollbar")
        route_cols = ("status", "rnm", "rtype", "path", "term", "f_tm", "l_tm", "st_cnt")
        tree_route = ttk.Treeview(frame_route_list, columns=route_cols,
                                  show="headings",
                                  yscrollcommand=route_scroll.set,
                                  style="Search.Treeview",
                                  selectmode="none")
        route_scroll.config(command=tree_route.yview)

        # 2-36-h. [선택 상태 색상] 체크(V)된 노선은 초록 배경, 해제된 노선은 기본 배경으로 표시합니다.
        if self.is_dark_mode:
            tree_route.tag_configure("selected",   background="#1a472a", foreground="#a8e6b5")
            tree_route.tag_configure("unselected", background=T["bg_tree"],  foreground=T["fg_tree"])
        else:
            tree_route.tag_configure("selected",   background="#d4edda", foreground="#155724")
            tree_route.tag_configure("unselected", background="white",   foreground="black")

        tree_route.heading("status",  text="선택여부")
        tree_route.column("status",   width=42,  anchor="center", stretch=True)
        tree_route.heading("rnm",     text="노선번호", command=lambda: sort_data("rnm"))
        tree_route.column("rnm",      width=117, anchor="center", stretch=True)
        tree_route.heading("rtype",   text="유형",     command=lambda: sort_data("rtype"))
        tree_route.column("rtype",    width=50,  anchor="center", stretch=True)
        tree_route.heading("path",    text="기점↔종점")
        tree_route.column("path",     width=347, anchor="center", stretch=True)
        tree_route.heading("term",    text="배차간격")
        tree_route.column("term",     width=50,  anchor="center", stretch=True)
        tree_route.heading("f_tm",    text="첫차시각")
        tree_route.column("f_tm",     width=67,  anchor="center", stretch=True)
        tree_route.heading("l_tm",    text="막차시각")
        tree_route.column("l_tm",     width=67,  anchor="center", stretch=True)
        tree_route.heading("st_cnt",  text="정류장수")
        tree_route.column("st_cnt",   width=60,  anchor="center", stretch=True)
        tree_route.bind("<Button-1>", self.prevent_column_resize)
        tree_route.pack(side="left", fill="both", expand=True)
        route_scroll.pack(side="right", fill="y")

        # 2-36-i. [검색 실행] 입력한 키워드로 정류소를 검색합니다.
        #         3자리 이상 숫자 → ARS번호(getStationByUid)로 먼저 시도,
        #         결과 없거나 문자열 → 이름 검색(getStationByName)으로 재시도합니다.
        def perform_search(event=None):
            keyword = search_ent.get().strip()
            if not keyword: return
            for i in tree_st.get_children():
                tree_st.delete(i)
            root_res = None
            is_digit_search = keyword.isdigit() and len(keyword) >= 3
            if is_digit_search:
                root_res = self.fetch_api(
                    "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
                    {"arsId": keyword})
            if root_res is None or not root_res.findall(".//itemList"):
                root_res = self.fetch_api(
                    "http://ws.bus.go.kr/api/rest/stationinfo/getStationByName",
                    {"stSrch": keyword})
            insert_count = 0
            seen_stations = set()
            if root_res is not None:
                for item in root_res.findall(".//itemList"):
                    name   = item.findtext("stNm")
                    ars_id = item.findtext("arsId")
                    if ars_id and ars_id != "0":
                        key = (name, str(ars_id))
                        if key not in seen_stations:
                            seen_stations.add(key)
                            tree_st.insert("", "end", values=key)
                            insert_count += 1
            if insert_count == 0:
                tree_st.insert("", "end", values=("검색 결과가 없습니다.", ""))

        search_ent.bind("<Return>", perform_search)
        _btn_search = self.get_btn_style("normal", font_size=SZ_XS)
        AnyButton(frame_search, text="검색", command=perform_search,
                  **_btn_search).pack(side="left", padx=5)

        # 2-36-j. [노선 정렬] 열 제목 클릭 시 해당 열 기준으로 오름차순/내림차순을 토글 정렬합니다.
        def sort_data(key):
            if self.sort_state["key"] == key:
                self.sort_state["reverse"] = not self.sort_state["reverse"]
            else:
                self.sort_state["key"] = key
                self.sort_state["reverse"] = False
            if key == "rtype":
                self.current_route_data.sort(
                    key=lambda x: (type_order.get(x["rtype"], 99), x["rnm"]),
                    reverse=self.sort_state["reverse"])
            else:
                self.current_route_data.sort(
                    key=lambda x: x[key],
                    reverse=self.sort_state["reverse"])
            render_route_list()

        # 2-36-k. [노선 목록 렌더링] current_route_data를 읽어 표에 다시 그립니다.
        def render_route_list():
            for item in tree_route.get_children():
                tree_route.delete(item)
            for item in self.current_route_data:
                status_text = "V" if item["checked"].get() else ""
                tag = "selected" if item["checked"].get() else "unselected"
                vals = [status_text] + item["data_vals"]
                tree_route.insert("", "end", iid=item["rid"], values=vals, tags=(tag,))

        # 2-36-l. [노선 클릭 토글] 표에서 노선 행을 클릭하면 선택(V)/해제가 토글됩니다.
        def on_route_click(event):
            if tree_route.identify_region(event.x, event.y) == "separator":
                return "break"
            item_id = tree_route.identify_row(event.y)
            if not item_id: return
            for item in self.current_route_data:
                if item["rid"] == item_id:
                    new_state = not item["checked"].get()
                    item["checked"].set(new_state)
                    cur_values = list(tree_route.item(item_id, "values"))
                    cur_values[0] = "V" if new_state else ""
                    tree_route.item(item_id, values=cur_values,
                                    tags=("selected" if new_state else "unselected",))
                    break
        tree_route.bind("<Button-1>", on_route_click)

        # 2-36-m. [정류소 선택 이벤트] 정류소 목록에서 하나를 선택하면 그 정류소를 지나는
        #         모든 노선 정보를 API로 조회하여 노선 목록 표에 채웁니다.
        #         이전에 같은 정류소가 설정되어 있었다면 기존 선택 상태를 유지합니다.
        def on_station_select(event):
            selected = tree_st.selection()
            if not selected: return
            ars_id = str(tree_st.item(selected[0])["values"][1]).zfill(5)
            if not ars_id or ars_id == "": return
            # 2-36-m-i. [기존 선택 복원] 동일 정류소를 다시 선택한 경우 이전에 체크했던 노선 ID를 기억합니다.
            saved_checked_routes = set()
            if self.ars_ids[target_idx].get().strip() == ars_id and self.target_st_info[target_idx].get("routes"):
                saved_checked_routes = {r[0] for r in self.target_st_info[target_idx]["routes"]}
            self.current_route_data.clear()
            self._strt_cache.clear()
            self._strt_ord_cache.clear()
            # 2-36-m-ii. [노선 목록 조회] getRouteByStation API로 정류소를 지나는 전체 노선 ID를 가져옵니다.
            root_route = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/stationinfo/getRouteByStation",
                {"arsId": ars_id})
            if root_route is not None:
                type_map = {"1":"공항","2":"마을","3":"간선","4":"지선","5":"순환",
                            "6":"광역","7":"인천","8":"경기","9":"폐지","0":"공용"}
                for it in root_route.findall(".//itemList"):
                    rid, rnm = it.findtext("busRouteId"), it.findtext("busRouteNm")
                    # 2-36-m-iii. [노선 상세 조회] 각 노선의 유형, 기점↔종점, 배차간격, 첫/막차, 운수사명을 가져옵니다.
                    root_info = self.fetch_api(
                        "http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo",
                        {"busRouteId": rid})
                    rtype, path, term, f_tm, l_tm, st_cnt_str = "-","-","-","-","-","-"
                    st_cnt_val = 100
                    if root_info is not None:
                        info = root_info.find(".//itemList")
                        if info is not None:
                            rtype = type_map.get(info.findtext("routeType"), "기타")
                            path  = f"{info.findtext('stStationNm')}↔{info.findtext('edStationNm')}"
                            term  = info.findtext("term") + "분"
                            f_tm  = self.format_hhmm(info.findtext("firstBusTm", "-"))
                            l_tm  = self.format_hhmm(info.findtext("lastBusTm", "-"))
                            self.route_corp_map[rid] = info.findtext("corpNm", "정보없음")
                            # 2-36-m-iv. [전체 정류소 조회 및 캐시] 노선의 전체 정류소 목록을 가져와
                            #            정류소 수를 세고, 해당 정류소의 순번(seq/staOrd)을 캐시에 저장합니다.
                            root_st_list = self.fetch_api(
                                "http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute",
                                {"busRouteId": rid})
                            if root_st_list is not None:
                                st_list_items = root_st_list.findall(".//itemList")
                                st_cnt_val = len(st_list_items)
                                st_cnt_str = f"{st_cnt_val}개"
                                self._strt_cache[rid] = root_st_list
                                for _si in st_list_items:
                                    _item_ars = str(_si.findtext("arsId") or "").zfill(5)
                                    if _item_ars == ars_id:
                                        _ord_val = _si.findtext("seq") or _si.findtext("staOrd") or ""
                                        if _ord_val:
                                            self._strt_ord_cache[rid] = _ord_val
                                        break
                    self.rid_to_rnm[rid] = rnm
                    is_checked = (rid in saved_checked_routes) if saved_checked_routes else True
                    self.current_route_data.append({
                        "rid": rid, "rnm": rnm, "rtype": rtype, "st_cnt": st_cnt_val,
                        "data_vals": [rnm, rtype, path, term, f_tm, l_tm, st_cnt_str],
                        "checked": tk.BooleanVar(value=is_checked)
                    })
            # 2-36-m-v. [초기 정렬] 유형(간선→지선→광역→기타→경기→인천) 순서로 정렬 후 표를 그립니다.
            self.sort_state = {"key": "rtype", "reverse": False}
            self.current_route_data.sort(
                key=lambda x: (type_order.get(x["rtype"], 99), x["rnm"]))
            render_route_list()

        tree_st.bind("<<TreeviewSelect>>", on_station_select)

        # 2-36-n. [기존 정류소 자동 표시] 이미 설정된 정류소가 있으면 자동으로 결과에 표시하고 선택합니다.
        if self.ars_ids[target_idx].get().strip():
            item_id = tree_st.insert("", "end",
                values=(self.target_st_info[target_idx].get("nm", ""),
                        self.ars_ids[target_idx].get()))
            tree_st.selection_set(item_id)

        # 2-36-o. [노선 확정] 선택한 노선들을 target_st_info에 저장하고 검색 창을 닫습니다.
        #         정류소 ID, 이름, 선택된 노선 리스트, ord_map(노선별 정류소 순번)을 저장합니다.
        def confirm_selection():
            chosen = [(item["rid"], item["rnm"], item["st_cnt"])
                      for item in self.current_route_data if item["checked"].get()]
            selected_st = tree_st.selection()
            if not selected_st or not chosen: return
            st_name = tree_st.item(selected_st[0])["values"][0]
            ars_id  = str(tree_st.item(selected_st[0])["values"][1]).zfill(5)
            self.ars_ids[target_idx].set(ars_id)
            self.lbl_st_names[target_idx].config(text=f"[{st_name}] 실시간 현황")
            self.lbl_hist_titles[target_idx].config(text=f"[{st_name}] 도착 기록")
            # 2-36-o-i. [정류소 ID 조회] ARS 번호로 내부 정류소 ID(stId)를 조회합니다.
            root_st_uid = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
                {"arsId": ars_id})
            st_id = root_st_uid.findtext(".//stId") if (
                root_st_uid is not None and not isinstance(root_st_uid, tuple)) else ""
            # 2-36-o-ii. [제거된 노선 정리] 이전 설정에서 선택했으나 이번에 해제된 노선의
            #            일시정지 기록과 차량번호 캐시를 삭제합니다.
            if hasattr(self, "target_st_info") and self.target_st_info[target_idx].get("routes"):
                old_rids = {r[0] for r in self.target_st_info[target_idx]["routes"]}
                new_rids = {item[0] for item in chosen}
                gone = old_rids - new_rids
                for rid in gone:
                    self.pos_suspend_until.pop(rid, None)
                    old_st_id = self.target_st_info[target_idx].get("id", "")
                    self.veh_cache.pop((old_st_id, rid), None)
            # 2-36-o-iii. [ord_map 구성] 각 선택된 노선의 정류소 순번(ord)을 캐시에서 조회하여 맵에 저장합니다.
            #             캐시 누락 시 API를 재호출합니다.
            ord_map = {}
            for rid, rnm, _ in chosen:
                if rid in self._strt_ord_cache:
                    ord_map[rid] = self._strt_ord_cache[rid]
                elif rid in self._strt_cache:
                    root_strt = self._strt_cache[rid]
                    for st_item in root_strt.findall(".//itemList"):
                        item_ars = str(st_item.findtext("arsId") or "").zfill(5)
                        if item_ars == ars_id:
                            ord_val = st_item.findtext("seq") or st_item.findtext("staOrd") or ""
                            if ord_val: ord_map[rid] = ord_val
                            break
                else:
                    self.log(f"⚠ 노선전체정류소 API 캐시 누락: {rnm}번 — 재호출합니다.")
                    root_strt = self.fetch_api(
                        "http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute",
                        {"busRouteId": rid})
                    if root_strt is not None and not isinstance(root_strt, tuple):
                        for st_item in root_strt.findall(".//itemList"):
                            item_ars = str(st_item.findtext("arsId") or "").zfill(5)
                            if item_ars == ars_id:
                                ord_val = st_item.findtext("seq") or st_item.findtext("staOrd") or ""
                                if ord_val: ord_map[rid] = ord_val
                                break
            # 2-36-o-iv. [캐시 초기화 및 저장] 사용이 끝난 SLST 캐시를 비우고 최종 정류소 정보를 저장합니다.
            self._strt_cache.clear()
            self._strt_ord_cache.clear()
            self.target_st_info[target_idx] = {
                "id": st_id, "nm": st_name, "routes": chosen, "ord_map": ord_map}
            self.log(f"정류소 {target_idx+1} 설정 완료: {st_name} "
                     f"(총 {len(ord_map)}개 노선 중 {len(chosen)}개 선택)")
            search_win.destroy()
            self.update_button_states()

        AnyButton(search_win, text="선택한 노선들로 적용",
                  command=confirm_selection, height=2,
                  **self.get_btn_style("normal", font_size=SZ_M)
        ).pack(fill="x", padx=15, pady=10)

    # 2-37. [process_station 함수] 특정 정류소(idx)에 대해 실시간 도착 정보 조회 + 버스 도착 판정을
    #       한 번에 수행하는 핵심 처리 함수입니다. refresh_data()에서 정류소1·2 각각에 대해 호출됩니다.
    #       ① 날짜 변경 감지 및 카운터/일시정지 초기화
    #       ② 각 노선의 도착 예정 정보 조회 (ARR1 → ARR2 폴백 → SINF 폴백)
    #       ③ 실시간 현황 표 갱신
    #       ④ 각 노선의 버스 위치 조회 (POS1 → POS2 폴백, 일시정지/공유 캐시 적용)
    #       ⑤ 도착 판정: lastStnId == 현재 정류소 → 기록 + 엑셀 자동 저장
    def process_station(self, idx):
        # 2-37-a. [날짜 변경 감지] 자정을 넘겼으면 POS 일시정지 테이블과 API 카운터를 초기화합니다.
        today = datetime.now().date()
        if today != self._last_date:
            self._last_date = today
            self.pos_suspend_until.clear()
            self.log("📅 날짜가 바뀌었습니다. 도착 기록 API 일시정지 기록을 초기화합니다.")

            # 2-37-b. [API 카운터 리셋] idx==0(정류소1)일 때만 실행하여 2번 초기화되는 것을 방지합니다.
            if idx == 0:
                for k in self.api_stats:
                    self.api_stats[k] = 0
                for k in self.api_stats_by_key["main"]:
                    self.api_stats_by_key["main"][k] = 0
                for k in self.api_stats_by_key["back"]:
                    self.api_stats_by_key["back"][k] = 0
                self._trigger_counter_update()
                self.log("📊 API 호출 횟수가 00:00 기준으로 초기화되었습니다.")

        info, ars_id = self.target_st_info[idx], self.ars_ids[idx].get().strip()
        st_id, rt_rows = info.get('id', ''), []

        # 2-37-c. [SINF 보조 캐시] ARR1/ARR2가 모두 실패했을 때 getStationByUid(SINF) 응답을
        #         정류소당 1회만 호출하여 캐시하는 변수입니다. None=미호출 상태.
        uid_cache = None
        ord_map = info.get('ord_map', {})

        def get_uid_cache():
            """SINF API를 정류소당 최대 1회만 호출하고 결과를 두 가지 키(rid/노선명)로 조회 가능하게 반환합니다."""
            nonlocal uid_cache
            if uid_cache is not None:
                return uid_cache
            root_uid = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
                {'arsId': ars_id}
            )
            if root_uid is None or isinstance(root_uid, tuple):
                uid_cache = {'by_rid': {}, 'by_nm': {}}
                return uid_cache
            by_rid, by_nm = {}, {}
            for item in root_uid.findall(".//itemList"):
                rid_key = item.findtext("busRouteId") or ""
                nm_key  = item.findtext("busRouteAbrv") or item.findtext("rtNm") or ""
                if rid_key and rid_key not in by_rid:
                    by_rid[rid_key] = item
                if nm_key and nm_key not in by_nm:
                    by_nm[nm_key] = item
            uid_cache = {'by_rid': by_rid, 'by_nm': by_nm}
            # 2-37-c-i. [하루 1회 로그] SINF 보조 호출 안내를 정류소별로 하루 한 번만 출력합니다.
            _today = datetime.now().date()
            if self._sinf_logged_date.get(idx) != _today:
                self._sinf_logged_date[idx] = _today
                self.log(f"ℹ 실시간 현황 API 초과 → 실시간 현황 예비2 API 보조 호출 (정류소 {idx+1}, {len(by_rid)}개 노선 수신) — 이후 동일 알림은 내일까지 표시 생략")
            return uid_cache

        # 2-37-d. [노선별 도착 예정 정보 조회] 각 노선에 대해 ARR1(주) → ARR2(폴백) → SINF(보조) 순으로
        #         도착 예정 정보를 조회하여 실시간 현황 표에 표시할 데이터를 수집합니다.
        for rid, rnm, _ in info.get('routes', []):
            row = None

            # 2-37-d-i. [1순위: ARR1/ARR2] 도착정보 API를 호출합니다.
            ord_val = ord_map.get(rid, '0')
            arr_params = {'stId': st_id, 'busRouteId': rid, 'ord': ord_val}
            root_arr = self.fetch_api("", arr_params, api_type="ARR1")

            if root_arr is not None and not isinstance(root_arr, tuple):
                # 2-37-d-ii. [운행 재개 감지] arrmsg에 실제 도착 정보가 있으면 POS 일시정지를 해제합니다.
                NO_BUS_MSGS = {"운행정보없음", "운행종료", "출발대기", "-", ""}
                arr_items = root_arr.findall(".//itemList")
                has_active_bus = any(
                    (item.findtext("arrmsg1") or "").strip() not in NO_BUS_MSGS
                    or (item.findtext("arrmsg2") or "").strip() not in NO_BUS_MSGS
                    for item in arr_items
                )
                if rid in self.pos_suspend_until:
                    if has_active_bus:
                        del self.pos_suspend_until[rid]
                        self.log(f"🔄 {rnm}번 운행 재개 확인")
                # 2-37-d-iii. [응답 행 매칭] arsId 일치 또는 arsId=0일 때 staOrd로 2차 매칭하여
                #             해당 정류소의 도착 정보 행을 찾습니다.
                for item in root_arr.findall(".//itemList"):
                    res_ars  = str(item.findtext("arsId") or "").zfill(5)
                    res_ord  = str(item.findtext("staOrd") or "")

                    arsid_match = (res_ars == ars_id)
                    arsid_zero  = (not res_ars.strip("0"))
                    ord_match   = (ord_val != '0' and res_ord == ord_val)

                    if arsid_match or (arsid_zero and (ord_match or ord_val == '0')):
                        p1 = item.findtext("plainNo1") or "-"
                        p2 = item.findtext("plainNo2") or "-"
                        # 2-37-d-iv. [차량번호 캐시 저장] 유효한 차량번호를 veh_cache에 저장합니다.
                        cache_key = (st_id, rid)
                        if p1 not in ("-", "", None) or p2 not in ("-", "", None):
                            self.veh_cache[cache_key] = (p1, p2)
                        row = (
                            item.findtext("rtNm") or rnm,
                            p1,
                            item.findtext("arrmsg1") or "-",
                            p2,
                            item.findtext("arrmsg2") or "-"
                        )
                        break
            elif root_arr is None:
                # 2-37-d-v. [2순위: SINF 보조] ARR1/ARR2가 모두 실패(한도 초과)했을 때
                #           SINF 캐시에서 해당 노선 항목을 꺼냅니다.
                uid_data = get_uid_cache()
                uid_item = uid_data['by_rid'].get(rid)
                if uid_item is None:
                    uid_item = uid_data['by_nm'].get(rnm)

                if uid_item is not None:
                    route_display = (
                        uid_item.findtext("busRouteAbrv")
                        or uid_item.findtext("rtNm")
                        or rnm
                    )
                    # 2-37-d-vi. [차량번호 캐시 우선] 이전에 ARR1/ARR2에서 받아둔 번호판이 있으면
                    #            VID API 호출 없이 캐시된 값을 사용합니다.
                    cache_key = (st_id, rid)
                    cached = self.veh_cache.get(cache_key)
                    if cached:
                        p1_display, p2_display = cached
                    else:
                        # 2-37-d-vii. [VID API로 차량번호 조회] 캐시가 없을 때 vehId로 번호판을 조회합니다.
                        veh_id1 = uid_item.findtext("vehId1") or ""
                        veh_id2 = uid_item.findtext("vehId2") or ""

                        def _fetch_plain(veh_id):
                            if not veh_id or veh_id == "-":
                                return "-"
                            root_vid = self.fetch_api(
                                "http://ws.bus.go.kr/api/rest/buspos/getBusPosByVehId",
                                {"vehId": veh_id}
                            )
                            if root_vid is not None and not isinstance(root_vid, tuple):
                                plain = root_vid.findtext(".//plainNo")
                                if plain:
                                    return plain
                            return veh_id

                        p1_display = _fetch_plain(veh_id1)
                        p2_display = _fetch_plain(veh_id2)

                        if (p1_display not in ("-", "", veh_id1) or
                                p2_display not in ("-", "", veh_id2)):
                            self.veh_cache[cache_key] = (p1_display, p2_display)
                    row = (
                        route_display,
                        p1_display,
                        uid_item.findtext("arrmsg1") or "-",
                        p2_display,
                        uid_item.findtext("arrmsg2") or "-"
                    )

            if row:
                rt_rows.append(row)

        # 2-37-e. [실시간 현황 표 갱신] 수집한 도착 예정 데이터를 메인 스레드에서 표에 반영합니다.
        self.root.after(0, lambda d=rt_rows, t=self.trees_rt[idx]: self.refresh_tree(t, d))

        # 2-37-f. [도착 판정 단계] 버스 위치 조회 API로 각 노선의 버스가 이 정류소에 실제로 도착했는지 판정합니다.
        if st_id:
            # 2-37-f-i. [POS 공유 캐시 초기화] 정류소1(idx=0) 처리 시작 시 임시 저장소를 비웁니다.
            #           정류소2 처리 시 같은 노선의 POS 결과를 재사용하기 위함입니다.
            if idx == 0: self.temp_pos_data = {}

            for rid, rnm, st_cnt in info.get('routes', []):
                root_pos = None

                # 2-37-f-ii. [POS 일시정지 확인] 이 노선이 첫차 전까지 일시정지 상태인지 확인합니다.
                now_dt = datetime.now()
                if rid in self.pos_suspend_until:
                    resume_dt = self.pos_suspend_until[rid]
                    if now_dt < resume_dt:
                        continue
                    else:
                        del self.pos_suspend_until[rid]
                        if rid not in self.pos_resume_logged:
                            self.log(f"⏰ {rnm}번 POS 호출 재개 (첫차 5분 전 도달)")
                            self.pos_resume_logged.add(rid)

                # 2-37-f-iii. [POS 공유 캐시 확인] 정류소1에서 이미 조회한 결과가 있으면 재사용합니다.
                if rid in self.temp_pos_data:
                    root_pos = self.temp_pos_data[rid]
                else:
                    # 2-37-f-iv. [POS API 호출] 버스 위치 조회 API(getBusPosByRtid → getBusPosByRouteSt 폴백)를 호출합니다.
                    pos_result = self.fetch_api("", {'busRouteId': rid, 'startOrd': '1', 'endOrd': str(st_cnt)}, api_type="POS2")

                    if isinstance(pos_result, tuple) and pos_result[0] == 'NO_BUS':
                        # 2-37-f-v. [운행 없음 → 일시정지 등록] 첫차 5분 전까지 POS 호출을 건너뛰도록 등록합니다.
                        _, f_tm_str = pos_result
                        if f_tm_str:
                            try:
                                f_hm = datetime.strptime(f_tm_str, "%H:%M")
                                base_dt = now_dt.replace(hour=f_hm.hour, minute=f_hm.minute, second=0, microsecond=0)
                                if base_dt <= now_dt:
                                    base_dt += timedelta(days=1)
                                resume_dt = base_dt - timedelta(minutes=5)
                                if resume_dt > now_dt:
                                    self.pos_suspend_until[rid] = resume_dt
                                    self.pos_resume_logged.discard(rid)
                                    self.log(f"💤 {rnm}번 차량 없음 → 첫차 {f_tm_str} 5분 전({resume_dt.strftime('%H:%M')})까지 POS 정지")
                                else:
                                    self.pos_suspend_until[rid] = now_dt + timedelta(minutes=1)
                            except Exception:
                                self.pos_suspend_until[rid] = now_dt + timedelta(minutes=30)
                                self.pos_resume_logged.discard(rid)
                                self.log(f"💤 {rnm}번 차량 없음 → 첫차 시각 파싱 실패, 30분 후 재시도")
                        else:
                            self.pos_suspend_until[rid] = now_dt + timedelta(minutes=30)
                            self.pos_resume_logged.discard(rid)
                            self.log(f"💤 {rnm}번 차량 없음 → 첫차 정보 없음, 30분 후 재시도")
                        continue

                    root_pos = pos_result
                    # 2-37-f-vi. [POS 공유 캐시 저장] 정류소2 처리 시 재사용할 수 있도록 저장합니다.
                    if root_pos is not None:
                        self.temp_pos_data[rid] = root_pos

                # 2-37-f-vii. [도착 판정] 각 버스의 lastStnId(마지막 경유 정류소)가 현재 정류소 ID와
                #             일치하면 "도착"으로 판정하고 기록합니다.
                if root_pos is not None:
                    for bus in root_pos.findall(".//itemList"):
                        if bus.findtext("lastStnId") == st_id:
                            veh_no = bus.findtext("plainNo")
                            # 2-37-f-viii. [중복 방지] 같은 차량이 5분(300초) 이내에 다시 기록되지 않도록 합니다.
                            if veh_no in self.last_arrival_logs[idx] and (time.time() - self.last_arrival_logs[idx][veh_no] < 300): continue

                            f_time = self.format_datetm(bus.findtext("dataTm"))
                            # 2-37-f-ix. [운수사명 결정] API 응답의 운수사명을 기본으로 하되,
                            #            엑셀에서 공동배차 정보(2개 이상 회사)가 있으면 합쳐서 표시합니다.
                            corp_nm = self.route_corp_map.get(rid, "정보없음")
                            if rnm in self.excel_multi_corp_map and len(self.excel_multi_corp_map[rnm]) >= 2:
                                corp_nm = ", ".join(sorted(list(self.excel_multi_corp_map[rnm])))

                            # 2-37-f-x. [기록 저장 및 자동 저장] 도착 기록을 recorded_data에 추가하고
                            #           즉시 엑셀 자동 저장을 실행합니다. 도착 기록 표에도 행을 추가합니다.
                            log_entry = (f_time, rnm, veh_no, corp_nm, "정류소 도착")
                            self.recorded_data.append((f_time, info['nm'], rnm, veh_no, corp_nm))
                            self.last_arrival_logs[idx][veh_no] = time.time()
                            self.perform_auto_save()
                            self.log(f"★ [{info['nm']}] 도착: {rnm} ({veh_no})")
                            self.root.after(0, lambda r=log_entry, t=self.trees_hist[idx]: (
                                t.insert("", "end", values=r),
                                t.see(t.get_children()[-1])
                            ))

    # 2-38. [start_excel_download_thread 함수] 서울시 버스 운행 노선 엑셀 파일을 백그라운드 스레드에서
    #       자동 다운로드하는 작업을 시작합니다. UI가 멈추지 않도록 daemon 스레드로 실행합니다.
    def start_excel_download_thread(self):
        threading.Thread(target=self.download_and_load_excel, daemon=True).start()

    # 2-39. [download_and_load_excel 함수] Selenium(크롬 자동화)으로 서울 데이터 광장에서
    #       최신 버스 노선 정보 엑셀 파일(.xlsx)을 headless 모드로 다운로드합니다.
    #       다운로드 성공 시 load_excel_routes()로 즉시 로드하고,
    #       실패 시 기존 파일 중 가장 최신 것을 찾아 로드합니다.
    def download_and_load_excel(self):
        target_url = "https://data.seoul.go.kr/dataList/OA-15066/F/1/datasetView.do"
        self.log(f"📥 최신 노선정보 확인 중 (Text Search): {target_url}")

        download_success = False
        downloaded_filename = None
        driver = None

        try:
            # 2-39-a. [크롬 옵션 설정] headless(화면 없음), 봇 감지 우회, 다운로드 경로 등을 설정합니다.
            options = Options()
            options.add_argument("--headless")
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)

            target_download_dir = os.path.abspath(self.current_dir)
            prefs = {
                "download.default_directory": target_download_dir,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True,
                "profile.default_content_settings.popups": 0
            }
            options.add_experimental_option("prefs", prefs)

            # 2-39-b. [크롬 드라이버 자동 설치 및 실행] ChromeDriverManager가 현재 크롬 버전에 맞는 드라이버를 자동 다운로드합니다.
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)

            if not os.path.exists(target_download_dir):
                os.makedirs(target_download_dir, exist_ok=True)

            driver.get(target_url)

            # 2-39-c. [다운로드 대상 탐색] 페이지에서 ".xlsx" 텍스트가 포함된 클릭 가능한 요소를 찾습니다.
            wait = WebDriverWait(driver, 20)
            self.log("   ㄴ페이지 분석 중 (파일명 탐색)...")
            download_target = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//*[contains(text(), '.xlsx')]")
            ))

            found_text = download_target.text.strip()
            self.log(f"   ㄴ다운로드 대상 발견: {found_text}")

            # 2-39-d. [기존 파일 삭제] 동일 이름 파일이 있으면 삭제하여 다운로드 충돌을 방지합니다.
            file_path = os.path.join(self.current_dir, found_text)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    self.log(f"   ℹ 기존 파일 삭제 후 다운로드: {found_text}")
                except Exception as del_err:
                    self.log(f"   ⚠ 기존 파일 삭제 실패 (사용중일 수 있음): {del_err}")

            before_files = set(os.listdir(self.current_dir))

            # 2-39-e. [다운로드 실행] JavaScript로 요소를 스크롤하여 보이게 한 뒤 클릭합니다.
            driver.execute_script("arguments[0].scrollIntoView(true);", download_target)
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", download_target)

            # 2-39-f. [다운로드 완료 대기] 1초마다 폴더를 확인하여 새 파일이 생겼는지 최대 30초간 확인합니다.
            found_file = None
            for _ in range(30):
                time.sleep(1)
                after_files = set(os.listdir(self.current_dir))
                new_files = after_files - before_files
                for f in new_files:
                    if f.endswith(".xlsx") and "서울시버스노선기본정보" in f:
                        found_file = f
                        break
                if found_file:
                    break

            if found_file:
                self.log(f"✅ 다운로드 성공: {found_file}")
                download_success = True
                downloaded_filename = found_file
            else:
                self.log("⚠ 다운로드 시도했으나 파일 생성 안됨 (시간초과)")

        except Exception as e:
            self.log(f"❌ 다운로드 프로세스 오류: {e}")
            self.log("   (Chrome 브라우저와 인터넷 연결을 확인해주세요)")
        finally:
            if driver:
                driver.quit()

        # 2-39-g. [파일 로드] 새로 받은 파일 또는 기존 최신 파일을 메모리에 로드합니다.
        if download_success and downloaded_filename:
            self.load_excel_routes(specific_filename=downloaded_filename)
        else:
            self.log("ℹ 기존 파일 중 최신 파일을 사용합니다.")
            self.load_excel_routes()

    # 2-40. [load_excel_routes 함수] 서울시 버스 운행 노선 엑셀 파일을 pandas로 읽어
    #       노선번호 → 운수사명 집합 매핑(excel_multi_corp_map)을 구성합니다.
    #       공동배차 노선(2개 이상 회사)을 식별하는 데 사용됩니다.
    def load_excel_routes(self, specific_filename=None):
        target_file = specific_filename

        # 2-40-a. [파일 자동 탐색] 지정 파일이 없으면 작업 폴더에서 이름이 비슷한 가장 최신 파일을 찾습니다.
        if not target_file:
            candidates = [f for f in os.listdir(self.current_dir) if "서울시버스노선기본정보" in f and f.endswith(".xlsx")]
            if candidates:
                candidates.sort(reverse=True)
                target_file = candidates[0]
            else:
                target_file = "서울시버스노선기본정보(20260108).xlsx"

        path = os.path.join(self.current_dir, target_file)

        if not os.path.exists(path):
            self.log(f"⚠ 로드할 엑셀 파일이 없습니다: {target_file}")
            return

        try:
            df = pd.read_excel(path)
            r_col = [c for c in df.columns if '노선번호' in str(c)]
            c_col = [c for c in df.columns if '업체명' in str(c)]

            if r_col and c_col:
                r_col_name = r_col[0]
                c_col_name = c_col[0]

                # 2-40-b. [노선-회사 매핑 구성] 각 행을 순회하며 {노선번호: {회사명1, 회사명2, ...}} 구조를 만듭니다.
                temp_map = {}
                for _, row in df.iterrows():
                    r_name = str(row[r_col_name]).strip()
                    c_name = str(row[c_col_name]).strip()
                    if r_name not in temp_map:
                        temp_map[r_name] = set()
                    temp_map[r_name].add(c_name)

                self.excel_multi_corp_map = temp_map
                self.log(f"✅ 엑셀 데이터 로드 완료: {target_file}")
            else:
                self.log(f"⚠ 엑셀 형식 오류: '노선번호' 또는 '업체명' 열 미발견")
        except Exception as e:
            self.log(f"❌ 엑셀 로드 중 오류: {e}")

    # 2-41. [start_monitoring 함수] 자동 기록을 시작합니다.
    #       갱신 주기 검증 → 입력창 잠금 → 검색 버튼 비활성 → 엑셀 파일 생성 →
    #       모니터링 플래그 ON → 백그라운드 스레드(main_loop) 시작 순으로 진행합니다.
    def start_monitoring(self):
        if not self.refresh_interval_var.get().strip():
            messagebox.showwarning("입력 누락", "갱신주기를 입력해야 자동 기록을 시작할 수 있습니다.")
            return

        if not self.service_key_var.get().strip(): return

        # 2-41-a. [입력창·버튼 잠금] 자동 기록 중에는 갱신 주기와 정류소 변경을 막습니다.
        self.entry_refresh_interval.config(state='readonly')
        self.is_monitoring = True
        _s_off2 = self.get_btn_style("disabled")
        _s_off2.pop("state", None)
        for _b in self.btn_searches:
            _b.config(state="disabled", **_s_off2)

        # 2-41-b. [엑셀 파일 생성] 타임스탬프가 포함된 새 엑셀 파일을 생성하고 자동 저장을 허용합니다.
        filename = f"Bus_Arrival_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.auto_save_path = os.path.join(self.current_dir, filename)

        try:
            pd.DataFrame(columns=["데이터시각", "도착정류소명", "노선", "차량번호", "운수사명"]).to_excel(self.auto_save_path, index=False)
            self.can_auto_save = True
            display_name = os.path.basename(self.auto_save_path)
            self.lbl_auto_save_status.config(text=f"[{display_name}] 파일에 자동 기록 중 ......", fg="#27ae60")
        except Exception as e:
            self.can_auto_save = False
            messagebox.showwarning("오류", f"자동 저장 설정 중 오류 발생: {e}")

        self.update_button_states()
        threading.Thread(target=self.main_loop, daemon=True).start()

    # 2-42. [main_loop 함수] 백그라운드 스레드에서 실행되는 자동 갱신 무한 루프입니다.
    #       설정된 갱신 주기(초)마다 refresh_data()를 호출하고,
    #       대기를 0.1초 단위로 분할하여 중지 버튼 클릭에 즉시 반응할 수 있게 합니다.
    def main_loop(self):
        next_call = time.time()
        while self.is_monitoring:
            self.refresh_data()
            try: interval = int(self.refresh_interval_var.get())
            except: interval = 60
            next_call += interval

            # 2-42-a. [분할 대기] 0.1초마다 is_monitoring 플래그를 확인하여 중지 시 즉시 탈출합니다.
            sleep_time = next_call - time.time()
            if sleep_time < 0: next_call = time.time(); sleep_time = 0
            for _ in range(int(sleep_time * 10)):
                if not self.is_monitoring: break
                time.sleep(0.1)

    # 2-43. [manual_refresh 함수] 수동 갱신 버튼 클릭 시 대기 없이 즉시 데이터를 갱신합니다.
    #       중복 클릭 방지를 위해 버튼을 비활성화하고, 갱신 완료 후 다시 활성화합니다.
    def manual_refresh(self):
        if CURRENT_OS == "Darwin" and not self._btn_active.get('manual', True):
            return
        # 2-43-a. [중복 방지] 갱신 완료 전까지 버튼을 비활성화합니다.
        self._btn_active['manual'] = False
        _s_dis = self.get_btn_style("disabled")
        _s_dis.pop("state", None)
        self.btn_manual.config(state="disabled", **_s_dis)
        self.log("🔄 [수동] 데이터 갱신 시작...")

        def run_manual():
            self.refresh_data(manual=True)
            self.log("✅ [수동] 데이터 갱신 완료")
            def _re_enable():
                self._btn_active['manual'] = True
                _s_on = self.get_btn_style("normal")
                _s_on.pop("state", None)
                self.btn_manual.config(state="normal", **_s_on)
            self.root.after(0, _re_enable)

        threading.Thread(target=run_manual, daemon=True).start()

    # 2-44. [refresh_data 함수] 정류소1·2 모두의 데이터를 순차 갱신하는 공용 함수입니다.
    #       자동/수동 갱신 모두 이 함수를 호출합니다.
    #       _refresh_lock으로 동시 갱신을 방지하고, 사이클 종료 후 미저장 기록이 있으면
    #       자동으로 구제 저장을 시도합니다.
    def refresh_data(self, manual=False):
        # 2-44-a. [동시 갱신 방지] 자동 갱신과 수동 갱신이 동시에 실행되면 공유 데이터에
        #         중복 기록이나 누락이 생길 수 있으므로 lock으로 한 스레드만 허용합니다.
        if not self._refresh_lock.acquire(blocking=False):
            if manual:
                self.log("ℹ 자동 갱신 진행 중이므로 수동 갱신을 건너뜁니다.")
            return

        try:
            for i in range(2):
                if self.target_st_info[i].get('routes'):
                    self.process_station(i)

            # 2-44-b. [구제 저장] 갱신 완료 후에도 미저장 데이터가 있으면 재시도합니다.
            #         process_station 내부의 perform_auto_save가 PermissionError 등으로
            #         실패했을 경우를 대비합니다.
            if (self.recorded_data
                    and self.auto_save_path
                    and self.can_auto_save
                    and len(self.recorded_data) > self._saved_record_count):
                missed = len(self.recorded_data) - self._saved_record_count
                self.log(f"⚠ 미저장 기록 {missed}건 감지 → 구제 저장 재시도 중...")
                self.perform_auto_save()
                if len(self.recorded_data) == self._saved_record_count:
                    self.log(f"✅ 구제 저장 성공: {missed}건 복구 완료")
                else:
                    self.log(f"❌ 구제 저장도 실패 — 다음 사이클에 재시도됩니다.")

            if not manual:
                self.log("데이터 갱신 완료")
        finally:
            self._refresh_lock.release()

    # 2-45. [stop_monitoring 함수] 자동 기록을 중지합니다.
    #       종료 확인 팝업 → 플래그 OFF → 마지막 구제 저장 → 입력창·버튼 복원 순으로 진행합니다.
    def stop_monitoring(self):
        if not messagebox.askyesno("중지 확인", "정말 중지하시겠습니까?"):
            return
        self.is_monitoring = False
        # 2-45-a. [검색 버튼 복원] 인증키가 유효하면 정류소 검색 버튼을 다시 활성화합니다.
        if self.key_locked:
            _s_on2 = self.get_btn_style("normal")
            _s_on2.pop("state", None)
            for _b in self.btn_searches:
                _b.config(state="normal", **_s_on2)
        self.log("🛑 자동 기록을 중지합니다.")

        # 2-45-b. [최종 구제 저장] 중지 직전 미저장 기록이 있으면 마지막으로 강제 저장을 시도합니다.
        if (self.recorded_data
                and self.auto_save_path
                and self.can_auto_save
                and len(self.recorded_data) > self._saved_record_count):
            missed = len(self.recorded_data) - self._saved_record_count
            self.log(f"⚠ 중지 시점 미저장 기록 {missed}건 → 최종 구제 저장 시도 중...")
            self.perform_auto_save()
            if len(self.recorded_data) == self._saved_record_count:
                self.log(f"✅ 최종 구제 저장 성공: {missed}건 복구 완료")
            else:
                self.log(f"❌ 최종 구제 저장 실패 — '다른 이름으로 엑셀 저장' 버튼으로 수동 저장해 주세요.")

        self.entry_refresh_interval.config(state='normal')
        self.lbl_auto_save_status.config(text=" ※ 자동 기록 시작 버튼을 작동시키면 도착 기록이 엑셀파일로 자동 저장됩니다.", fg="#e74c3c")
        self.update_button_states()

    # 2-46. [clear_history 함수] 도착 기록 표의 모든 행을 삭제합니다.
    #       실수 방지를 위해 확인 팝업을 먼저 띄우며,
    #       이미 엑셀에 저장된 기록은 영향받지 않습니다.
    def clear_history(self, idx):
        if messagebox.askyesno("삭제", f"정류소 {idx+1}의 모든 기록을 삭제하시겠습니까?\n기록을 삭제하더라도, 이미 저장된 기록은 삭제되지 않습니다."):
            tree = self.trees_hist[idx]
            for item in tree.get_children():
                tree.delete(item)
            self.log(f"정류소 {idx+1} 기록을 초기화했습니다.")

    # 2-47. [format_hhmm 함수] API 응답의 시간 문자열("20260101120000" 또는 "0430" 등)을
    #       "HH:MM" 형식으로 변환합니다. POS 일시정지의 첫차 시각 계산에 사용됩니다.
    def format_hhmm(self, raw_str):
        if not raw_str or len(raw_str) < 4: return raw_str
        if len(raw_str) >= 14: return f"{raw_str[8:10]}:{raw_str[10:12]}"
        return f"{raw_str[:2]}:{raw_str[2:4]}"

    # 2-48. [log 함수] 현재 시각 도장([HH:MM:SS])을 붙여 하단 로그 창에 메시지를 한 줄 추가합니다.
    #       인증키가 로그에 노출되지 않도록 ********로 마스킹하고,
    #       5,000줄을 초과하면 가장 오래된 줄부터 삭제합니다.
    #       백그라운드 스레드에서 호출되어도 root.after(0, ...)로 메인 스레드에 안전하게 위임합니다.
    def log(self, msg):
        now = datetime.now().strftime("%H:%M:%S")
        try:
            key = self.service_key_var.get().strip()
            backup_key = self.backup_key_var.get().strip()
        except Exception:
            key, backup_key = "", ""

        # 2-48-a. [인증키 마스킹] 64자리 인증키가 로그에 그대로 노출되면 보안 위험이 있으므로 가립니다.
        if key and len(key) > 4: msg = msg.replace(key, "********")
        if backup_key and len(backup_key) > 4: msg = msg.replace(backup_key, "********")

        line = f"[{now}] {msg}\n"

        # 2-48-b. [UI 준비 확인] 초기화 중에는 txt_log가 아직 없을 수 있으므로 콘솔에 출력합니다.
        if not hasattr(self, 'txt_log'):
            print(line, end="")
            return

        def _do_insert():
            try:
                self.txt_log.config(state="normal")
                self.txt_log.insert(tk.END, line)
                self.txt_log.see(tk.END)
                # 2-48-c. [줄 수 제한] 5,000줄 초과 시 가장 오래된 줄을 삭제하여 메모리를 절약합니다.
                line_count = int(self.txt_log.index('end-1c').split('.')[0])
                if line_count > 5000:
                    self.txt_log.delete('1.0', '2.0')
                self.txt_log.config(state="disabled")
            except Exception:
                pass

        try:
            self.root.after(0, _do_insert)
        except Exception:
            pass

    # 2-49. [format_datetm 함수] API 응답의 날짜시각 문자열("20260101120000")을
    #       "2026-01-01 12:00:00" 형식으로 변환합니다. 도착 기록의 데이터 시각 표시에 사용됩니다.
    def format_datetm(self, raw_tm):
        try:
            if raw_tm and len(raw_tm) >= 14:
                return datetime.strptime(raw_tm, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
            return raw_tm
        except: return raw_tm

    # 2-50. [_core_excel_save_logic 함수] 수집된 도착 기록(recorded_data)을 영업일(BizDate) 기준으로
    #       시트를 나누어 엑셀 파일에 저장하는 핵심 로직입니다.
    #       자동 저장(perform_auto_save)과 수동 저장(save_to_excel) 양쪽에서 공통 사용됩니다.
    #       save_completed=True이면 완결된 영업일의 데이터를 별도 파일로도 저장합니다.
    #       영업일 기준: 새벽 03:00 이전의 기록은 전날 영업일로 분류합니다.
    def _core_excel_save_logic(self, target_path, save_completed=False):
        try:
            import pandas as pd
            from datetime import datetime, timedelta
            from openpyxl.styles import Font, PatternFill, Alignment

            cols = ["데이터시각", "도착정류소명", "노선", "차량번호", "운수사명"]
            df = pd.DataFrame(self.recorded_data, columns=cols)

            # 2-50-a. [영업일 계산] 새벽 00:00~02:59 사이 도착 기록은 전날 영업일로 분류합니다.
            def get_biz_date(dt_str):
                try:
                    dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
                    if dt.hour < 3: dt -= timedelta(days=1)
                    return dt.strftime("%Y-%m-%d")
                except: return "Unknown"

            df['BizDate'] = df['데이터시각'].apply(get_biz_date)

            # 2-50-b. [완결 영업일 판별] 현재 영업일이 아닌 날짜는 운행이 끝난 "완결된 날"로 간주합니다.
            biz_dates_in_data = set(df['BizDate'].unique())
            now = datetime.now()
            current_biz = (now - timedelta(days=1)).strftime("%Y-%m-%d") if now.hour < 3 else now.strftime("%Y-%m-%d")
            completed_dates = biz_dates_in_data - {current_biz, "Unknown"}

            # 2-50-c. [날짜별 시트 저장] 각 영업일을 별도 시트로 만들고 머리글 서식과 열 너비를 적용합니다.
            with pd.ExcelWriter(target_path, engine='openpyxl') as writer:
                for biz_date, group in df.groupby('BizDate'):
                    sheet_data = group.drop(columns=['BizDate'])
                    sheet_data.to_excel(writer, sheet_name=biz_date, index=False)

                    worksheet = writer.sheets[biz_date]

                    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    header_font = Font(bold=True, size=11)
                    header_alignment = Alignment(horizontal='center', vertical='center')

                    for idx, col in enumerate(sheet_data.columns):
                        cell = worksheet.cell(row=1, column=idx+1)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment

                        max_data_len = sheet_data[col].astype(str).map(len).max()
                        base_width = max(max_data_len, len(str(col)))

                        if col == "데이터시각":
                            adjusted_width = base_width + 1
                        elif col == "도착정류소명":
                            adjusted_width = (base_width + 15) * 1.5
                        elif col == "노선":
                            adjusted_width = base_width + 8
                        elif col == "운수사명":
                            adjusted_width = base_width + 7
                        else:
                            adjusted_width = base_width + 5

                        col_letter = worksheet.cell(row=1, column=idx+1).column_letter
                        worksheet.column_dimensions[col_letter].width = adjusted_width

            # 2-50-d. [완결 파일 자동 저장] save_completed=True이고 새로 완결된 영업일이 있으면
            #         해당 날의 데이터만 "Bus_Arrival_Log_YYYYMMDD_completed.xlsx"로 별도 저장합니다.
            if save_completed:
                new_completed = completed_dates - self._completed_dates_saved
                for biz_date in sorted(new_completed):
                    day_df = df[df['BizDate'] == biz_date].drop(columns=['BizDate'])
                    if day_df.empty:
                        continue
                    safe_date = biz_date.replace("-", "")
                    completed_name = f"Bus_Arrival_Log_{safe_date}_completed.xlsx"
                    completed_path = os.path.join(self.current_dir, completed_name)
                    try:
                        with pd.ExcelWriter(completed_path, engine='openpyxl') as cw:
                            day_df.to_excel(cw, sheet_name=biz_date, index=False)
                            ws = cw.sheets[biz_date]
                            hfill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                            hfont = Font(bold=True, size=11)
                            halign = Alignment(horizontal='center', vertical='center')
                            for ci, col in enumerate(day_df.columns):
                                cell = ws.cell(row=1, column=ci+1)
                                cell.fill = hfill; cell.font = hfont; cell.alignment = halign
                                max_len = day_df[col].astype(str).map(len).max()
                                base_w  = max(max_len, len(str(col)))
                                if col == "도착정류소명":
                                    col_w = (base_w + 15) * 1.5
                                elif col == "데이터시각":
                                    col_w = base_w + 1
                                elif col == "노선":
                                    col_w = base_w + 8
                                elif col == "운수사명":
                                    col_w = base_w + 7
                                else:
                                    col_w = base_w + 5
                                col_letter = ws.cell(row=1, column=ci+1).column_letter
                                ws.column_dimensions[col_letter].width = col_w
                        self._completed_dates_saved.add(biz_date)
                        self.log(f"📁 영업일 완결 파일 저장: {completed_name}")
                    except Exception as ce:
                        self.log(f"⚠ 완결 파일 저장 실패 ({biz_date}): {ce}")

            # 2-50-e. [저장 완료 표시] _saved_record_count를 갱신하여 구제 저장이 불필요하다고 표시합니다.
            self._saved_record_count = len(self.recorded_data)
            return True

        except PermissionError:
            self.log("⚠ 엑셀 파일이 열려 있어 저장을 건너뜁니다. (파일을 닫아주세요)")
            return False
        except Exception as e:
            self.log(f"❌ 엑셀 저장 중 오류 발생: {e}")
            return False

    # 2-51. [save_to_excel 함수] 사용자가 직접 저장 위치와 파일 이름을 지정하여 엑셀 파일을 저장합니다.
    #       파일 탐색기(filedialog)를 열어 경로를 선택하고,
    #       _core_excel_save_logic()을 호출하여 영업일 분류 저장을 수행합니다.
    def save_to_excel(self):
        if not self.recorded_data:
            messagebox.showwarning("알림", "저장할 데이터가 없습니다.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            title="다른 이름으로 저장"
        )

        if path:
            if self._core_excel_save_logic(path):
                messagebox.showinfo("저장 완료", f"영업일 기준으로 분류되어 저장되었습니다.\n경로: {path}")
                self.log(f"💾 수동 엑셀 저장 완료: {path}")

    # 2-52. [perform_auto_save 함수] 도착 기록이 추가될 때마다 자동으로 엑셀 파일에 저장합니다.
    #       _save_lock으로 한 스레드만 파일을 쓰도록 하여 동시 쓰기로 인한 데이터 손실을 방지합니다.
    #       save_completed=True로 호출하여 완결 영업일 파일도 함께 처리합니다.
    def perform_auto_save(self):
        if not self.recorded_data or not self.auto_save_path or not self.can_auto_save:
            return

        # 2-52-a. [비차단 잠금] 다른 스레드가 이미 저장 중이면 기다리지 않고 건너뜁니다.
        #         해당 스레드가 전체 recorded_data 스냅샷을 저장하므로 데이터 손실은 없습니다.
        if not self._save_lock.acquire(blocking=False):
            return
        try:
            self._core_excel_save_logic(self.auto_save_path, save_completed=True)
        finally:
            self._save_lock.release()

    # 2-53. [refresh_tree 함수] Treeview 표의 모든 기존 행을 삭제하고 새 데이터로 다시 채웁니다.
    #       실시간 현황 표 갱신 시 사용됩니다.
    def refresh_tree(self, tree, data):
        for i in tree.get_children(): tree.delete(i)
        for row in data: tree.insert("", "end", values=row)


# O. [프로그램 진입점] 이 파일이 직접 실행될 때(python main.py) 아래 코드가 실행됩니다.
if __name__ == "__main__":
    import traceback

    # 3. [_get_error_log_path 함수] 오류 로그 파일(errorlog.txt)의 저장 경로를 반환합니다.
    #    PyInstaller 빌드(.exe/.app)와 소스 코드(.py) 실행 모두에서 동작하며,
    #    macOS .app 번들의 경우 Contents/MacOS 상위 폴더로 올라가 접근 가능한 위치에 저장합니다.
    def _get_error_log_path():
        if getattr(sys, 'frozen', False):
            base = os.path.dirname(os.path.abspath(sys.executable))
            if sys.platform == "darwin" and base.endswith("Contents/MacOS"):
                base = os.path.dirname(os.path.dirname(os.path.dirname(base)))
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, "errorlog.txt")

    # 4. [_write_error_log 함수] 예외 정보를 errorlog.txt에 타임스탬프와 함께 기록하고,
    #    가능하면 messagebox로도 사용자에게 오류 발생을 알립니다.
    def _write_error_log(exc_type, exc_value, exc_tb):
        try:
            log_path = _get_error_log_path()
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            tb_text = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"\n{'='*60}\n")
                f.write(f"[{timestamp}] 프로그램 오류 발생\n")
                f.write(f"{'='*60}\n")
                f.write(tb_text)
                f.write("\n")
            try:
                import tkinter.messagebox as _mb
                short_msg = f"{exc_type.__name__}: {exc_value}"
                _mb.showerror(
                    "프로그램 오류",
                    f"예기치 않은 오류가 발생했습니다.\n\n"
                    f"{short_msg}\n\n"
                    f"오류 내용이 다음 파일에 저장되었습니다:\n{log_path}"
                )
            except Exception:
                pass
        except Exception:
            pass

    # 5. [_global_exception_handler 함수] sys.excepthook에 등록되어 프로그램 어디서든
    #    처리되지 않은 예외가 발생하면 errorlog.txt에 자동으로 기록합니다.
    def _global_exception_handler(exc_type, exc_value, exc_tb):
        _write_error_log(exc_type, exc_value, exc_tb)
        sys.__excepthook__(exc_type, exc_value, exc_tb)
    sys.excepthook = _global_exception_handler

    # P. [stderr 리다이렉트] PyInstaller --windowed 빌드 시 콘솔이 없어
    #    sys.stderr가 None이 되는 문제를 방지하기 위해 파일로 리다이렉트합니다.
    try:
        if sys.stderr is None or getattr(sys, 'frozen', False):
            sys.stderr = open(_get_error_log_path().replace("errorlog.txt", "stderr.txt"),
                              "a", encoding="utf-8")
    except Exception:
        pass

    # Q. [메인 창 생성 및 실행] tkinter 루트 창을 만들고 SeoulBusArrivalRecorder를 초기화한 뒤
    #    mainloop()로 이벤트 루프에 진입합니다. 예외 발생 시 errorlog.txt에 기록됩니다.
    try:
        root = tk.Tk()
        app = SeoulBusArrivalRecorder(root)
        root.mainloop()
    except Exception as e:
        _write_error_log(type(e), e, e.__traceback__)
