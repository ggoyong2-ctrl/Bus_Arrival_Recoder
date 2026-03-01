import tkinter as tk # 1. [준비물] 컴퓨터 화면에 창을 띄우는 기본 도구상자입니다.
from tkinter import ttk, messagebox, filedialog # 1-1. [준비물] 더 예쁜 버튼, 알림창, 파일 찾기 기능을 가져옵니다.
import threading # 1-2. [준비물] 여러 가지 일을 동시에 처리해주는 '일꾼'을 부릅니다.
import time # 1-3. [준비물] 시간을 재거나 잠시 기다리게 하는 시계를 가져옵니다.
import os # 1-4. [준비물] 컴퓨터의 파일이나 폴더를 다루는 도구를 가져옵니다.
import sys # 1-5. [준비물] 프로그램 설정이나 컴퓨터 정보를 확인하는 도구입니다.
from cryptography.fernet import Fernet # 1-6. [준비물] 글자를 암호로 바꾸어 비밀을 지켜주는 도구입니다.
MASTER_KEY = b'u7_K-5D4fR9zP2mN8xL1qJ6vH3sB0tG9wE8rT7yU4iA=' 
cipher_suite = Fernet(MASTER_KEY) # 1-7. [준비물] 암호를 풀거나 잠그는 열쇠를 만듭니다.
import webbrowser # 1-8. [준비물] 인터넷 사이트를 열어주는 도구입니다.
import platform # 1-9. [준비물] 지금 쓰는 컴퓨터가 어떤 종류인지 알려주는 도구입니다.
from datetime import datetime, timedelta # 1-10. [준비물] 오늘 날짜와 현재 시간을 계산하는 도구입니다.
import xml.etree.ElementTree as ET # 1-11. [준비물] 복잡한 인터넷 문서를 읽기 좋게 정리해주는 도구입니다.
from urllib.parse import unquote # 1-12. [준비물] 인터넷 주소에 들어있는 특수 문자를 글자로 바꿔줍니다.

# 2. [주변 검사] 지금 내 컴퓨터 상태가 어떤지 확인해봅니다.
# 2-1. 윈도우인지 맥인지 확인해서 이름을 기억해둡니다.
CURRENT_OS = platform.system() 
# 2-2. [선택]맥os 앱 실행 권한(Gatekeeper) 격리 속성 제거 터미널 명령어 xattr -d com.apple.quarantine /경로/to/your/app_file(혹은 드래그)

# 3. [예쁜 글씨] 화면에 나올 글자들의 글꼴과 크기를 미리 정합니다.
if CURRENT_OS == "Windows":
    # 3-1. 윈도우 컴퓨터용 글꼴과 크기 설정입니다.
    FONT_MAIN = "맑은 고딕" 
    FONT_SUB = "돋움"       
    FONT_MONO = "Consolas" 
    # 3-1-1. 글자 크기들을 미리 정해둡니다.
    SZ_L = 19  
    SZ_M = 11
    SZ_S = 9
    SZ_XS = 8
    SZ_XXS = 7
else: 
    # 3-2. 맥(Mac) 컴퓨터용 글꼴과 크기 설정입니다.
    FONT_MAIN = "AppleGothic"         
    FONT_SUB = "Apple SD Gothic Neo" 
    FONT_MONO = "Menlo"              
    # 3-2-1. 맥은 글자가 작게 보여서 1.4배 정도 키워줍니다.
    #        버튼 크기 조정은 get_btn_style()의 padx/pady로 별도 제어합니다.
    SZ_L = int(15 * 1.4)
    SZ_M = int(11 * 1.4)
    SZ_S = int(9 * 1.4)
    SZ_XS = int(8 * 1.4)
    SZ_XXS = int(7 * 1.4)

# 4. [창고 검사] 프로그램을 돌릴 때 필요한 특수 도구들이 다 있는지 확인합니다.
try:
    # 4-1. 인터넷 연결과 엑셀 작업에 꼭 필요한 외부 도구들을 불러옵니다.
    import requests # 인터넷 서버에 질문을 보내는 일꾼
    import pandas as pd # 표 형태의 데이터를 엑셀처럼 다루는 도구
    from selenium import webdriver # 인터넷 창을 자동으로 조종하는 로봇
    from selenium.webdriver.chrome.service import Service 
    from selenium.webdriver.chrome.options import Options 
    from selenium.webdriver.common.by import By 
    from selenium.webdriver.support.ui import WebDriverWait 
    from selenium.webdriver.support import expected_conditions as EC 
    from webdriver_manager.chrome import ChromeDriverManager 
except ImportError as e:
    # 4-2. [오류 대처] 필요한 도구가 없으면 안내 창을 띄우고 프로그램을 종료합니다.
    root = tk.Tk() 
    root.withdraw() 
    msg = f"필수 라이브러리가 없습니다.\n아래 명령어를 터미널(CMD)에 실행해 주세요.\n\npip install requests pandas openpyxl cryptography selenium webdriver-manager\n\n에러 내용: {e}"
    messagebox.showerror("실행 오류", msg) 
    exit() 

# 5. [기계 설계도] 서울 버스 도착을 기록하는 프로그램의 전체 기능 모음
class SeoulBusArrivalRecorder:

    # 1그룹 : 초기단계(Initialization)

    # 5-1. [탄생] 프로그램이 처음 켜질 때 실행되는 준비 과정
    def __init__(self, root):
        # 5-1-1. 화면의 기본 정보들을 설정합니다.
        self.root = root 
        self.root.title("서울버스 정류소 듀얼 도착기록 프로그램 v1.3.80") 
        self.root.geometry("1200x800") 
        # 5-1-1-1. 창이 너무 작아지면 화면이 깨지므로 최소 크기를 정합니다.
        self.root.minsize(960, 400) 
        
        # 5-1-2. 디자인 테마를 설정합니다.
        self.style = ttk.Style() 
        try:
            self.style.theme_use('clam') 
        except:
            pass

        # 5-1-3. 프로그램이 돌아가는 위치(폴더)를 정확히 찾아냅니다.
        if getattr(sys, 'frozen', False):
            # 5-1-3-1. 완성된 실행 파일(.exe) 상태일 때 폴더 위치 찾기
            self.resource_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(sys.executable)))
            base_exe_path = os.path.abspath(sys.executable)
            if CURRENT_OS == "Darwin": 
                self.current_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(base_exe_path))))
            else:
                self.current_dir = os.path.dirname(base_exe_path)
        else:
            # 5-1-3-2. 소스 코드(.py) 상태로 실행 중일 때 폴더 위치 찾기
            self.resource_dir = os.path.dirname(os.path.abspath(__file__))
            self.current_dir = self.resource_dir

        # 5-1-4. 필요한 파일들의 위치를 미리 약속해둡니다.
        self.key_file_path = os.path.join(self.current_dir, "key.cfg")
        icon_path_png = os.path.join(self.resource_dir, "icon.png")
        icon_path_ico = os.path.join(self.resource_dir, "icon.ico")
        icon_path_icns = os.path.join(self.resource_dir, "icon.icns")

        # 5-1-5. 윈도우 컴퓨터라면 아이콘이 예쁘게 보이도록 특별 설정을 합니다.
        if CURRENT_OS == "Windows":
            import ctypes
            try:
                myappid = 'Bus_Arrival_Recoder(v1.3.80)' 
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
            except: pass

        # 5-1-6. 프로그램의 아이콘 이미지를 창에 입힙니다.
        try:
            if os.path.exists(icon_path_png):
                img = tk.PhotoImage(file=icon_path_png)
                self.root.iconphoto(True, img) 
            if CURRENT_OS == "Windows" and os.path.exists(icon_path_ico):
                self.root.iconbitmap(icon_path_ico)
        except Exception as e:
            print(f"아이콘 설정 실패: {e}")

        # 5-1-7. 화면 제일 위쪽에 메뉴 바 공간을 만듭니다.
        self.top_container = tk.Frame(self.root)
        self.top_container.pack(side="top", fill="x")

        # 5-1-8. [기억 저장소] 프로그램이 돌아가는 동안 기억해야 할 소중한 정보들
        self.service_key_var = tk.StringVar(value="") # 메인 열쇠(인증키)
        self.backup_key_var = tk.StringVar(value="")  # 비상용 열쇠(백업키)
        self.key_locked = False # 키 입력창 잠금 상태

        # 5-1-8-1. 검색 버튼과 확정된 열쇠들을 관리할 공간입니다.
        self.btn_searches = [] 
        self.final_main_key = ""
        self.final_backup_key = ""

        # 5-1-8-2. 버스 정류소 정보와 감시 상태를 저장합니다.
        self.ars_ids = [tk.StringVar(), tk.StringVar()] # 정류소 번호 2개
        self.refresh_interval_var = tk.StringVar(value="30") # 갱신 시간 (기본 30초)
        self.is_monitoring = False # 지금 감시 중인지 아닌지 (ON/OFF)
        
        # 5-1-8-3. 엑셀 저장 위치와 자동 저장 허용 여부입니다.
        self.auto_save_path = None # 자동 저장할 엑셀 파일 위치
        self.can_auto_save = True # 저장을 해도 되는지 확인하는 스위치
        
        # 5-1-8-4. 문제가 생겼을 때 쓸 비상용 API 사용 상태입니다.
        self.use_fallback_api = False # 비상용 API 1번 사용 여부
        self.use_fallback_pos_api = False # 비상용 API 2번 사용 여부
        
        # 5-1-8-5. 선택된 정류소 정보와 버스 회사 이름 등을 모아둡니다.
        self.target_st_info = [{}, {}] # 선택한 버스 정보 저장소
        self.recorded_data = [] # 기록된 내용을 잠시 담아두는 바구니
        self.last_arrival_logs = [{}, {}] # 중복 기록을 막기 위한 메모장
        self.route_corp_map = {} # 버스 회사 이름 저장소
        self.rid_to_rnm = {} # 버스 ID를 번호로 바꿔주는 사전
        self.excel_multi_corp_map = {} # 엑셀에서 읽어온 회사 정보
        self.temp_pos_data = {} # POS 데이터를 임시로 저장하는 곳
        # 5-1-8-6. 차량번호 캐시: ARR1/ARR2 응답에서 받은 차량번호를 저장해 UID 폴백에 활용합니다.
        #   {(st_id, rid): (plainNo1, plainNo2)}  ← 정류소+노선 조합 키
        self.veh_cache = {}
        # 5-1-8-7. POS1 일시정지 테이블: 차량 없음 확인 시 첫차 시각까지 호출을 건너뜁니다.
        #   {rid: datetime}  — 정지 해제 시각 (이 시각이 되면 다시 호출 허용)
        self.pos_suspend_until = {}
        self.pos_resume_logged = set()  # POS 재개 로그 중복 방지: 한 번만 기록
        # 5-1-8-8. 마지막 날짜 체크: 자정 이후 suspend 초기화에 사용합니다.
        self._last_date = datetime.now().date()
        # 5-1-8-9. [macOS 전용] Aqua 테마에서 state='disabled'가 클릭을 막지 못하는
        #          버그를 우회하기 위해 버튼별 활성화 플래그를 별도 관리합니다.
        #          apply_btn() 호출 시 동기화하고, 각 핸들러 진입 시 guard 체크합니다.
        self._btn_active = {'toggle': False, 'manual': False}
        # 5-1-8-9. SLST(getStaionByRoute) 캐시: 노선 ID → 전체 정류소 목록.
        #   on_station_select 에서 노선당 1회만 호출하고 confirm_selection 에서 재사용합니다.
        #   정류소 선택이 바뀔 때마다 초기화됩니다.
        #   {busRouteId: root_element}
        self._strt_cache = {}
        # 5-1-8-10-1. SLST ord 캐시: on_station_select 에서 미리 구해둔 (ars_id, seq) 를 저장합니다.
        #   {busRouteId: ord_value}  — 정류소가 바뀔 때마다 초기화됩니다.
        self._strt_ord_cache = {}
        # 5-1-8-10. 자동 완결 파일 저장용: 이미 완결 파일을 저장한 영업일을 기억합니다.
        #   {"YYYY-MM-DD"} 형태의 집합으로 중복 저장을 방지합니다.
        self._completed_dates_saved = set()
        # 5-1-8-11. [스레드 안전] 엑셀 파일 쓰기와 데이터 갱신의 동시 접근을 막는 잠금장치입니다.
        #   _save_lock  : perform_auto_save() / _core_excel_save_logic() 를 한 스레드씩만 실행합니다.
        #   _refresh_lock : refresh_data() 를 한 스레드씩만 실행합니다.
        #   (자동 모니터링 스레드 + 수동 갱신 스레드가 동시에 실행될 때 발생하는
        #    Race Condition 및 엑셀 스냅샷 역전으로 인한 데이터 누락을 방지합니다.)
        self._save_lock = threading.Lock()
        self._refresh_lock = threading.Lock()
        # 5-1-8-12. [이중 저장 보호] 마지막으로 엑셀에 성공적으로 저장된 시점의 recorded_data 크기입니다.
        #   refresh_data() 사이클 종료 시 현재 recorded_data 크기와 비교해
        #   미저장 기록이 있으면 자동으로 재저장을 시도합니다 (구제 저장).
        self._saved_record_count = 0

        # 5-1-9. API를 몇 번 호출했는지 세는 통계 숫자판입니다.
        # 5-1-9-1. 합산 통계(메인키 + 백업키 합계) — 메인 화면 상단에 표시됩니다.
        self.api_stats = {
            "ARR1": 0, "ARR2": 0, "SINF": 0, "POS1": 0, "POS2": 0, "SCNM": 0, "SCID": 0, "RINF": 0, "SLST": 0, "VLD": 0
        }
        # 5-1-9-2. 메인/백업 키별 개별 통계 — API 상세 창에서 구분 표시됩니다.
        self.api_stats_by_key = {
            "main": {k: 0 for k in self.api_stats},
            "back": {k: 0 for k in self.api_stats},
        }
        # 5-1-9-3. [어제 스냅샷] 자정 초기화 직전의 통계값을 보관합니다.
        #   통계 창 '어제의 합계' 탭에 표시됩니다.
        self.api_stats_yesterday = {k: 0 for k in self.api_stats}
        self.api_stats_by_key_yesterday = {
            "main": {k: 0 for k in self.api_stats},
            "back": {k: 0 for k in self.api_stats},
        }
        
        # 5-1-10. 인터넷을 통해 정보를 가져올 주소 목록입니다.
        self.api_urls = {
            "ARR1": "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRoute",
            "ARR2": "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRouteAll",
            "POS1": "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRtid",
            "SINF": "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
            "POS2": "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRouteSt",
            "SCNM": "http://ws.bus.go.kr/api/rest/stationinfo/getStationByName",
            "SCID": "http://ws.bus.go.kr/api/rest/stationinfo/getRouteByStation",
            "RINF": "http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo",
            "SLST": "http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute",
            "VLD": "http://ws.bus.go.kr/api/rest/busRouteInfo/getBusRouteList  (인증키 검증)"
        }
        
        self.stats_win = None # 통계 창을 저장할 빈 공간
        # 5-1-12. [다크모드] 현재 테마 상태 (False=라이트, True=다크)
        self.is_dark_mode = False

        # 5-1-11. [최종 시작] 저장된 열쇠를 불러오고 화면을 그린 뒤 정보를 가져옵니다.
        self.load_saved_key() # 저장된 인증키가 있으면 불러옵니다.
        self.setup_ui()
        # 5-1-11-1. [X버튼 연결] 창 닫기(X) 클릭 시 종료 확인 팝업을 표시합니다.
        self.root.protocol("WM_DELETE_WINDOW", self._on_window_close)
        # 5-1-11-2. [자정 초기화 예약] 다음 0시 0분까지 남은 시간을 계산해 예약합니다.
        #   ★ 핵심: _do_midnight_reset() 내부에서 이 함수를 다시 호출해야 매일 반복됩니다.
        #     1회성 after() 만 등록하면 첫날만 작동하는 버그가 발생하므로 반드시 재예약합니다.
        self.root.after(200, self._schedule_midnight_reset)
        self.root.after(500, self.start_excel_download_thread) 

    # 2그룹 : 메인 UI (Main UI)

    # 5-1-A. [종료 확인] X버튼 클릭 시 정말 종료할지 확인하는 함수
    def _on_window_close(self):
        if messagebox.askyesno("종료 확인", "정말 종료하시겠습니까?"):
            self.root.destroy()

    # 5-1-B. [자정 예약] 다음 0시 0분 0초까지 남은 밀리초를 계산해 _do_midnight_reset을 예약합니다.
    #   ★ 이 함수는 _do_midnight_reset() 내부에서도 재호출되어 매일 반복 실행이 보장됩니다.
    def _schedule_midnight_reset(self):
        now = datetime.now()
        from datetime import timedelta
        next_midnight = (now + timedelta(days=1)).replace(
            hour=0, minute=0, second=0, microsecond=0)
        ms_left = int((next_midnight - now).total_seconds() * 1000)
        ms_left = max(ms_left, 1000)  # 자정 직후 즉시 재귀 호출 방지
        self.root.after(ms_left, self._do_midnight_reset)

    # 5-1-C. [자정 초기화] 매일 0시 정각에 실행됩니다.
    #   1) 오늘 통계 → 어제 스냅샷으로 복사
    #   2) 오늘 카운터 0 초기화
    #   3) UI 갱신 및 로그 기록
    #   4) ★ 다음날 자정을 위해 _schedule_midnight_reset() 재호출 (없으면 첫날만 실행됨)
    def _do_midnight_reset(self):
        self.api_stats_yesterday = dict(self.api_stats)
        self.api_stats_by_key_yesterday = {
            slot: dict(vals) for slot, vals in self.api_stats_by_key.items()
        }
        for k in self.api_stats:
            self.api_stats[k] = 0
        for slot in self.api_stats_by_key.values():
            for k in slot:
                slot[k] = 0
        self.update_api_counter_ui()
        self.log("📅 [자정] API 호출 카운터를 초기화했습니다. 어제의 합계는 통계 창에서 확인할 수 있습니다.")
        # ★ 반드시 재예약해야 매일 반복됩니다.
        self._schedule_midnight_reset()


    # 5-1-D. [테마 색상] 라이트/다크 팔레트를 딕셔너리로 반환합니다.
    def get_theme(self):
        if self.is_dark_mode:
            return {
                "bg_root":      "#0f0f0f",
                "bg_panel":     "#212121",
                "bg_card":      "#272727",
                "bg_entry":     "#121212",
                "bg_entry_ro":  "#1a1a1a",
                "fg_main":      "#f1f1f1",
                "fg_sub":       "#aaaaaa",
                "fg_accent":    "#3ea6ff",
                "fg_warn":      "#ff4e45",
                "fg_link":      "#5ebe5e",
                "fg_email":     "#3ea6ff",
                "fg_st1":       "#3ea6ff",
                "fg_st2":       "#ff4e45",
                "fg_counter":   "#aaaaaa",
                "fg_counter_v": "#f1f1f1",
                "bg_log":       "#0f0f0f",
                "fg_log":       "#dfe6e9",
                "bg_tree":      "#212121",
                "fg_tree":      "#f1f1f1",
                "bg_tree_sel":  "#3f3f3f",
                "border":       "#3f3f3f",
                "sash":         "#3f3f3f",
                "btn_normal_bg":      "#272727",
                "btn_normal_fg":      "#f1f1f1",
                "btn_normal_active":  "#3f3f3f",
                "btn_normal_border":  "#3f3f3f",
                "btn_outline_bg":     "#272727",
                "btn_outline_fg":     "#3ea6ff",
                "btn_outline_border": "#3ea6ff",
                "btn_disabled_bg":    "#1a1a1a",
                "btn_disabled_fg":    "#555555",
                "btn_disabled_border":"#2a2a2a",
                "toggle_track_off":  "#3f3f3f",
                "toggle_track_on":   "#3ea6ff",
                "toggle_thumb":      "#f1f1f1",
                "toggle_label":      "#aaaaaa",
                "toggle_bg":         "#272727",
                # 스크롤바
                "sb_bg":      "#3f3f3f",   # 스크롤바 thumb 배경
                "sb_trough":  "#1a1a1a",   # 스크롤바 trough(홈) 배경
                "sb_arrow":   "#aaaaaa",   # 스크롤바 화살표 색
            }
        else:
            return {
                "bg_root":      "#ffffff",
                "bg_panel":     "#ecf0f1",
                "bg_card":      "#f1f2f6",
                "bg_entry":     "#ffffff",
                "bg_entry_ro":  "#f0f0f0",
                "fg_main":      "#2d3436",
                "fg_sub":       "#7f8c8d",
                "fg_accent":    "#0984e3",
                "fg_warn":      "#e74c3c",
                "fg_link":      "#006400",
                "fg_email":     "#2980b9",
                "fg_st1":       "#0984e3",
                "fg_st2":       "#d63031",
                "fg_counter":   "#636e72",
                "fg_counter_v": "#2d3436",
                "bg_log":       "#2d3436",
                "fg_log":       "#dfe6e9",
                "bg_tree":      "#ffffff",
                "fg_tree":      "#000000",
                "bg_tree_sel":  "#0078d7",
                "border":       "#cccccc",
                "sash":         "#cccccc",
                "btn_normal_bg":      "#2ecc71",
                "btn_normal_fg":      "#ffffff",
                "btn_normal_active":  "#27ae60",
                "btn_normal_border":  "#1e8449",
                "btn_outline_bg":     "#eafaf1",
                "btn_outline_fg":     "#27ae60",
                "btn_outline_border": "#27ae60",
                "btn_disabled_bg":    "#e8e8e8",
                "btn_disabled_fg":    "#a0a0a0",
                "btn_disabled_border":"#c0c0c0",
                "toggle_track_off":  "#bbbbbb",
                "toggle_track_on":   "#2ecc71",
                "toggle_thumb":      "#ffffff",
                "toggle_label":      "#7f8c8d",
                "toggle_bg":         "#f1f2f6",
                # 스크롤바
                "sb_bg":      "#c0c0c0",
                "sb_trough":  "#f0f0f0",
                "sb_arrow":   "#606060",
            }

    # 5-1-E. [테마 적용] is_dark_mode에 맞게 전체 위젯 색상을 갱신합니다.
    def apply_theme(self):
        T = self.get_theme()
        self.root.configure(bg=T["bg_root"])
        self.top_container.configure(bg=T["bg_panel"])

        for w in self._tw.get("status_frames", []):
            try: w.configure(bg=T["bg_panel"])
            except: pass
        for key, w in self._tw.get("status_labels", {}).items():
            try:
                if key == "title_text":  w.configure(bg=T["bg_panel"], fg=T["fg_main"])
                elif key == "title_text3":  w.configure(bg=T["bg_panel"], fg=T["fg_main"])
                elif key == "title_blue":w.configure(bg=T["bg_panel"], fg=T["fg_accent"])
                elif key == "maker":     w.configure(bg=T["bg_panel"], fg=T["fg_sub"])
                elif key == "email":     w.configure(bg=T["bg_panel"], fg=T["fg_email"])
                elif key == "link1":     w.configure(bg=T["bg_panel"], fg=T["fg_link"])
                elif key == "link2":     w.configure(bg=T["bg_panel"], fg="#7fb3ff" if self.is_dark_mode else "#000080")
                elif key == "maker_bracket1": w.configure(bg=T["bg_panel"], fg=T["fg_sub"])
                elif key == "maker_bracket2": w.configure(bg=T["bg_panel"], fg=T["fg_sub"])
            except: pass

        for w in self._tw.get("ctrl_frames", []):
            try: w.configure(bg=T["bg_card"])
            except: pass
        try: self.lbl_auto_save_status.configure(bg=T["bg_card"], fg=T["fg_warn"])
        except: pass
        for w in self._tw.get("interval_labels", []):
            try: w.configure(bg=T["bg_card"], fg=T["fg_main"])
            except: pass
        try:
            self.entry_refresh_interval.configure(
                bg=T["bg_entry"], fg=T["fg_main"],
                insertbackground=T["fg_main"],
                readonlybackground=T["bg_entry_ro"])
        except: pass
        for entry in [self.entry_service_key, self.entry_backup_key]:
            try:
                entry.configure(
                    bg=T["bg_entry"], fg=T["fg_main"],
                    insertbackground=T["fg_main"],
                    readonlybackground=T["bg_entry_ro"])
            except: pass
        for w in self._tw.get("key_labels", []):
            try: w.configure(bg=T["bg_card"], fg=T["fg_main"])
            except: pass
        for w in self._tw.get("key_frames", []):
            try: w.configure(bg=T["bg_card"])
            except: pass

        try: self.api_stats_container.configure(bg=T["bg_card"])
        except: pass
        try: self._tw["stats_outer"].configure(bg=T["bg_card"])
        except: pass
        for lbl in self._tw.get("stat_key_labels", []):
            try: lbl.configure(bg=T["bg_card"], fg=T["fg_counter"])
            except: pass
        for key, lbl in self.stat_value_labels.items():
            try: lbl.configure(bg=T["bg_card"], fg=T["fg_counter_v"])
            except: pass

        self._restyle_all_buttons(T)

        for pw in [self.super_paned, self.main_paned]:
            try: pw.configure(bg=T["sash"])
            except: pass

        for w in self._tw.get("tree_frames", []):
            try: w.configure(bg=T["bg_root"])
            except: pass

        for lbl in self.lbl_st_names:
            try: lbl.configure(fg=T["fg_st1"], bg=T["bg_root"])
            except: pass
        for lbl in self.lbl_hist_titles:
            try: lbl.configure(fg=T["fg_st2"], bg=T["bg_root"])
            except: pass
        for w in self._tw.get("ars_entries", []):
            try: w.configure(readonlybackground=T["bg_entry_ro"], fg=T["fg_main"])
            except: pass

        try:
            self.txt_log.configure(bg=T["bg_log"], fg=T["fg_log"])
            self._tw["log_outer"].configure(bg=T["bg_root"])
        except: pass

        self._apply_treeview_style(T)
        self._apply_scrollbar_style(T)
        self._draw_toggle()

    # 5-1-F. [버튼 일괄 재스타일]
    def _restyle_all_buttons(self, T):
        for btn, type_cell in self._tw.get("buttons", []):
            try:
                style = self.get_btn_style(type_cell[0])
                style.pop("state", None)
                btn.configure(**style)
            except: pass

    # 5-1-G. [Treeview 테마]
    def _apply_treeview_style(self, T):
        sty = ttk.Style()
        if self.is_dark_mode:
            sty.configure("Treeview",
                background=T["bg_tree"], foreground=T["fg_tree"],
                fieldbackground=T["bg_tree"], rowheight=22)
            sty.configure("Treeview.Heading",
                background=T["bg_card"], foreground=T["fg_main"], relief="flat")
            sty.map("Treeview",
                background=[("selected", T["bg_tree_sel"])],
                foreground=[("selected", T["fg_main"])])
            sty.map("Treeview.Heading",
                background=[("active", T["border"])])
        else:
            sty.configure("Treeview",
                background="#ffffff", foreground="#000000",
                fieldbackground="#ffffff", rowheight=22)
            sty.configure("Treeview.Heading",
                background="#f0f0f0", foreground="#000000", relief="raised")
            sty.map("Treeview",
                background=[("selected", "#0078d7")],
                foreground=[("selected", "#ffffff")])
            sty.map("Treeview.Heading",
                background=[("active", "#e0e0e0")])

    # 5-1-G2. [스크롤바 테마] 앱 전체 ttk.Scrollbar 색상을 테마에 맞게 적용합니다.
    #   - "App.Vertical.TScrollbar" 스타일로 모든 scrollbar에 일괄 적용합니다.
    #   - macOS Aqua 테마에서는 troughcolor/background 적용이 제한적이나
    #     arrowcolor는 반영되며 시각적 차이를 제공합니다.
    def _apply_scrollbar_style(self, T):
        sty = ttk.Style()
        sty.configure("App.Vertical.TScrollbar",
            background=T["sb_bg"],
            troughcolor=T["sb_trough"],
            arrowcolor=T["sb_arrow"],
            bordercolor=T["sb_trough"],
            darkcolor=T["sb_bg"],
            lightcolor=T["sb_bg"],
            relief="flat",
        )
        sty.map("App.Vertical.TScrollbar",
            background=[
                ("active",   T["border"]),
                ("disabled", T["sb_trough"]),
            ])


    def _build_toggle(self, parent):
        T = self.get_theme()
        # 폰트 크기(SZ_XS)에 맞춘 소형 토글: 트랙 28×14, 썸 반지름 5
        TRACK_W, TRACK_H, THUMB_R = 28, 14, 5
        # ① 문구 먼저 (왼쪽)
        self._toggle_label = tk.Label(
            parent, text="다크", font=(FONT_MAIN, SZ_XS),
            fg=T["toggle_label"], bg=T["bg_panel"],
            cursor="pointinghand" if CURRENT_OS == "Darwin" else "hand2"
        )
        self._toggle_label.pack(side="left", padx=(6, 2))
        self._toggle_label.bind("<Button-1>", self._on_toggle_theme)
        # ② 스위치 나중에 (오른쪽)
        self._toggle_canvas = tk.Canvas(
            parent, width=TRACK_W + 6, height=TRACK_H + 6,
            highlightthickness=0,
            cursor="pointinghand" if CURRENT_OS == "Darwin" else "hand2"
        )
        self._toggle_canvas.pack(side="left", padx=(0, 6))
        self._toggle_canvas.bind("<Button-1>", self._on_toggle_theme)
        self._draw_toggle()

    # 5-1-I. [토글 스위치 그리기] — 소형 (28×14, 썸 반지름 5)
    def _draw_toggle(self):
        if not hasattr(self, "_toggle_canvas"):
            return
        T = self.get_theme()
        c = self._toggle_canvas
        TRACK_W, TRACK_H, THUMB_R = 28, 14, 5
        ox, oy = 3, 3
        c.delete("all")
        parent_bg = T["bg_panel"]
        c.configure(bg=parent_bg)
        track_color = T["toggle_track_on"] if self.is_dark_mode else T["toggle_track_off"]
        r = TRACK_H // 2
        # 둥근 트랙
        c.create_arc(ox, oy, ox+TRACK_H, oy+TRACK_H, start=90, extent=180,
                     fill=track_color, outline=track_color)
        c.create_arc(ox+TRACK_W-TRACK_H, oy, ox+TRACK_W, oy+TRACK_H, start=270, extent=180,
                     fill=track_color, outline=track_color)
        c.create_rectangle(ox+r, oy, ox+TRACK_W-r, oy+TRACK_H,
                           fill=track_color, outline=track_color)
        # 썸(원): 다크=오른쪽, 라이트=왼쪽
        thumb_cx = (ox+TRACK_W-THUMB_R-3) if self.is_dark_mode else (ox+THUMB_R+3)
        thumb_cy = oy + TRACK_H // 2
        c.create_oval(thumb_cx-THUMB_R, thumb_cy-THUMB_R,
                      thumb_cx+THUMB_R, thumb_cy+THUMB_R,
                      fill=T["toggle_thumb"], outline=T["toggle_thumb"])
        # 레이블 업데이트
        try:
            self._toggle_label.configure(
                text="라이트" if self.is_dark_mode else "다크",
                fg=T["toggle_label"], bg=parent_bg)
        except: pass

    # 5-1-J. [토글 클릭]
    def _on_toggle_theme(self, event=None):
        self.is_dark_mode = not self.is_dark_mode
        self.apply_theme()

    # 5-2. [그림 그리기] 화면의 전체적인 레이아웃과 버튼들을 배치하는 함수
    # 5-2. [그림 그리기] 화면의 전체적인 레이아웃과 버튼들을 배치하는 함수
    def setup_ui(self):
        # 5-2-0. 위젯 추적 딕셔너리 초기화 (apply_theme 에서 재색상 적용에 사용합니다)
        self._tw = {
            "status_frames": [], "status_labels": {}, "ctrl_frames": [],
            "interval_labels": [], "key_labels": [], "key_frames": [],
            "stat_key_labels": [], "tree_frames": [], "ars_entries": [],
            "buttons": [],
        }

        # 5-2-1. 맥 컴퓨터라면 글자색을 약간 조절합니다.
        T = self.get_theme()

        # 5-2-2. 화면 제일 상단에 프로그램 안내 문구를 넣습니다.
        frame_status = tk.Frame(self.top_container, pady=2, bg=T["bg_panel"])
        frame_status.pack(fill="x", side="top")
        self._tw["status_frames"].append(frame_status)

        # 5-2-2-1. 왼쪽 영역 안내 글씨들
        frame_left_status = tk.Frame(frame_status, bg=T["bg_panel"])
        frame_left_status.pack(side="left", padx=10, pady=2)
        self._tw["status_frames"].append(frame_left_status)
        self.lbl_info_container = tk.Frame(frame_left_status, bg=T["bg_panel"])
        self.lbl_info_container.pack(anchor="w", pady=(2, 0))
        self._tw["status_frames"].append(self.lbl_info_container)

        lbl_t1 = tk.Label(self.lbl_info_container, text="정류소와 노선을 선택하고 ",
                          font=(FONT_MAIN, SZ_L, "bold"), fg=T["fg_main"], bg=T["bg_panel"])
        lbl_t1.pack(side="left")
        self._tw["status_labels"]["title_text"] = lbl_t1
        lbl_t2 = tk.Label(self.lbl_info_container, text="자동 기록 시작",
                          font=(FONT_MAIN, SZ_L, "bold"), fg=T["fg_accent"], bg=T["bg_panel"])
        lbl_t2.pack(side="left")
        self._tw["status_labels"]["title_blue"] = lbl_t2
        lbl_t3 = tk.Label(self.lbl_info_container, text=" 버튼을 누르세요.",
                          font=(FONT_MAIN, SZ_L, "bold"), fg=T["fg_main"], bg=T["bg_panel"])
        lbl_t3.pack(side="left")
        # title_text 를 리스트로 관리: t3도 같이 처리
        self._tw["status_labels"]["title_text3"] = lbl_t3

        # 5-2-2-2. 오른쪽 영역: 제작자 정보 + 링크 (토글은 박국환 행 우측에 배치)
        frame_right_status = tk.Frame(frame_status, bg=T["bg_panel"])
        frame_right_status.pack(side="right", padx=10, pady=2)
        self._tw["status_frames"].append(frame_right_status)

        # 제작자 정보 행: ① 토글뭉치(문구+스위치)  ②  [만든이 : 박국환 (이메일)]
        frame_right_top = tk.Frame(frame_right_status, bg=T["bg_panel"])
        frame_right_top.pack(side="top", anchor="e")
        self._tw["status_frames"].append(frame_right_top)

        # ── ① 토글 뭉치 (왼쪽) ─────────────────────────────────────────
        self._build_toggle(frame_right_top)

        # ── ② 제작자 문구 (오른쪽) ──────────────────────────────────────
        lbl_m1 = tk.Label(frame_right_top, text=" [ 만든이 : 박 국 환 (",
                           font=(FONT_SUB, SZ_XS), fg=T["fg_sub"], bg=T["bg_panel"])
        lbl_m1.pack(side="left", padx=(5, 0))
        self._tw["status_labels"]["maker_bracket1"] = lbl_m1
        lbl_email = tk.Label(frame_right_top, text="ggoyong2@naver.com",
                             font=(FONT_SUB, SZ_XS, "underline"), fg=T["fg_email"],
                             bg=T["bg_panel"], cursor="hand2")
        lbl_email.pack(side="left")
        lbl_email.bind("<Button-1>", lambda e: self.send_email_link("ggoyong2@naver.com"))
        self._tw["status_labels"]["email"] = lbl_email
        lbl_m2 = tk.Label(frame_right_top, text=") ]",
                           font=(FONT_SUB, SZ_XS), fg=T["fg_sub"], bg=T["bg_panel"])
        lbl_m2.pack(side="left")
        self._tw["status_labels"]["maker_bracket2"] = lbl_m2

        frame_sources = tk.Frame(frame_right_status, bg=T["bg_panel"])
        frame_sources.pack(side="top", anchor="e")
        self._tw["status_frames"].append(frame_sources)
        lbl_link1 = tk.Label(frame_sources,
                             text="공공데이터포털 Open API (https://www.data.go.kr)",
                             font=(FONT_SUB, SZ_XS), fg=T["fg_link"], bg=T["bg_panel"], cursor="hand2")
        lbl_link1.pack(side="top", anchor="e")
        lbl_link1.bind("<Button-1>", lambda e: self.open_link("https://www.data.go.kr"))
        self._tw["status_labels"]["link1"] = lbl_link1
        lbl_link2 = tk.Label(frame_sources,
                             text="서울시 버스운행노선 정보 (https://data.seoul.go.kr)",
                             font=(FONT_SUB, SZ_XS), fg="#000080", bg=T["bg_panel"], cursor="hand2")
        lbl_link2.pack(side="top", anchor="e")
        lbl_link2.bind("<Button-1>", lambda e: self.open_link("https://data.seoul.go.kr/dataList/OA-15066/F/1/datasetView.do"))
        self._tw["status_labels"]["link2"] = lbl_link2

        # 5-2-3. 주요 조종 버튼들이 모여있는 패널을 만듭니다.
        self.frame_ctrl_master = tk.Frame(self.top_container, pady=4, bg=T["bg_card"])
        self.frame_ctrl_master.pack(fill="x", side="top")
        self._tw["ctrl_frames"].append(self.frame_ctrl_master)

        frame_main_content = tk.Frame(self.frame_ctrl_master, bg=T["bg_card"])
        frame_main_content.pack(side="left", fill="x", expand=True)
        self._tw["ctrl_frames"].append(frame_main_content)

        PAD = 5

        # ══════════════════════════════════════════════════════════════════
        # 1행: 좌=자동저장 안내문구  /  우=(갱신주기 + 자동기록시작 + 수동갱신 + 다른이름저장)
        # ══════════════════════════════════════════════════════════════════
        frame_row1 = tk.Frame(frame_main_content, bg=T["bg_card"])
        frame_row1.pack(fill="x", side="top", pady=(2, 1))
        self._tw["ctrl_frames"].append(frame_row1)

        self.lbl_auto_save_status = tk.Label(
            frame_row1,
            text=" ※ 자동 기록 시작 버튼을 작동시키면 도착 기록이 엑셀파일로 자동 저장됩니다.",
            font=(FONT_SUB, SZ_XS, "bold"), fg=T["fg_warn"], bg=T["bg_card"]
        )
        self.lbl_auto_save_status.pack(side="left", padx=(5, 0))

        right_r1 = tk.Frame(frame_row1, bg=T["bg_card"])
        right_r1.pack(side="right", padx=(0, 5))
        self._tw["ctrl_frames"].append(right_r1)

        lbl_ival = tk.Label(right_r1, text="갱신주기(초):", bg=T["bg_card"],
                            font=(FONT_MAIN, SZ_S, "bold"), fg=T["fg_main"])
        lbl_ival.pack(side="left", padx=(PAD * 2, 2))
        self._tw["interval_labels"].append(lbl_ival)

        self.entry_refresh_interval = tk.Entry(
            right_r1, textvariable=self.refresh_interval_var,
            width=4, bg=T["bg_entry"], fg=T["fg_main"],
            insertbackground=T["fg_main"], font=(FONT_MONO, SZ_S),
            readonlybackground=T["bg_entry_ro"]
        )
        self.entry_refresh_interval.pack(side="left", padx=(0, PAD))

        BTN_H = 1
        self.btn_toggle = tk.Button(
            right_r1, text="자동 기록 시작",
            command=self._on_toggle_monitoring,
            width=0 if CURRENT_OS == "Darwin" else 11,
            height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_toggle.pack(side="left", padx=(0, PAD),
                             ipady=0 if CURRENT_OS == "Darwin" else 4)
        self._tw["buttons"].append((self.btn_toggle, ["normal"]))

        self.btn_manual = tk.Button(
            right_r1, text="수동 갱신",
            command=self.manual_refresh,
            width=0 if CURRENT_OS == "Darwin" else 7,
            height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_manual.pack(side="left", padx=(0, PAD),
                             ipady=0 if CURRENT_OS == "Darwin" else 4)
        self._tw["buttons"].append((self.btn_manual, ["normal"]))

        self.btn_save_excel = tk.Button(
            right_r1, text="다른 이름으로 엑셀 저장",
            command=self.save_to_excel, height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_save_excel.pack(side="left", padx=(0, 8),
                                 ipady=0 if CURRENT_OS == "Darwin" else 4)
        self._tw["buttons"].append((self.btn_save_excel, ["normal"]))

        # ══════════════════════════════════════════════════════════════════
        # 2행: 좌=(인증키 입력창 + 버튼)  /  우=(API현황 버튼 + API카운트표)
        # ══════════════════════════════════════════════════════════════════
        frame_row2 = tk.Frame(frame_main_content, bg=T["bg_card"])
        frame_row2.pack(fill="x", side="top", pady=(1, 2))
        self._tw["ctrl_frames"].append(frame_row2)

        left_r2 = tk.Frame(frame_row2, bg=T["bg_card"])
        left_r2.pack(side="left", padx=(5, 0), fill="y")
        self._tw["key_frames"].append(left_r2)

        key_input_area = tk.Frame(left_r2, bg=T["bg_card"])
        key_input_area.pack(side="left", fill="y")
        self._tw["key_frames"].append(key_input_area)

        main_key_row = tk.Frame(key_input_area, bg=T["bg_card"])
        main_key_row.pack(fill="x", pady=1)
        self._tw["key_frames"].append(main_key_row)
        lbl_mk = tk.Label(main_key_row, text="메인인증키 :", font=(FONT_MAIN, SZ_S),
                          bg=T["bg_card"], fg=T["fg_main"])
        lbl_mk.pack(side="left")
        self._tw["key_labels"].append(lbl_mk)
        self.entry_service_key = tk.Entry(
            main_key_row, textvariable=self.service_key_var,
            width=68, show="*", bg=T["bg_entry"], fg=T["fg_main"],
            insertbackground=T["fg_main"], font=(FONT_MONO, SZ_XXS),
            readonlybackground=T["bg_entry_ro"])
        self.entry_service_key.pack(side="left", padx=2)

        backup_key_row = tk.Frame(key_input_area, bg=T["bg_card"])
        backup_key_row.pack(fill="x", pady=1)
        self._tw["key_frames"].append(backup_key_row)
        lbl_bk = tk.Label(backup_key_row, text="백업인증키 :", font=(FONT_MAIN, SZ_S),
                          bg=T["bg_card"], fg=T["fg_main"])
        lbl_bk.pack(side="left")
        self._tw["key_labels"].append(lbl_bk)
        self.entry_backup_key = tk.Entry(
            backup_key_row, textvariable=self.backup_key_var,
            width=68, show="*", bg=T["bg_entry"], fg=T["fg_main"],
            insertbackground=T["fg_main"], font=(FONT_MONO, SZ_XXS),
            readonlybackground=T["bg_entry_ro"])
        self.entry_backup_key.pack(side="left", padx=2)

        self.btn_key_manage = tk.Button(
            left_r2, text="인증키\n입력",
            command=self.toggle_key_lock,
            width=0 if CURRENT_OS == "Darwin" else 7,
            **self.get_btn_style("normal")
        )
        self.btn_key_manage.pack(side="left", padx=(6, 0), fill="y")
        self._tw["buttons"].append((self.btn_key_manage, ["normal"]))

        right_r2 = tk.Frame(frame_row2, bg=T["bg_card"])
        right_r2.pack(side="right", padx=(0, 5), fill="y")
        self._tw["ctrl_frames"].append(right_r2)

        self.btn_api_stats = tk.Button(
            right_r2, text="API\n호출 현황",
            command=self.open_api_stats_window,
            width=0 if CURRENT_OS == "Darwin" else 7,
            **self.get_btn_style("normal")
        )
        self.btn_api_stats.pack(side="left", padx=(0, 4), fill="y")
        self._tw["buttons"].append((self.btn_api_stats, ["normal"]))

        stats_outer = tk.Frame(right_r2, bg=T["bg_card"])
        stats_outer.pack(side="left", fill="both", expand=True)
        stats_outer.grid_rowconfigure(0, weight=1)
        stats_outer.grid_rowconfigure(1, weight=0)
        stats_outer.grid_rowconfigure(2, weight=1)
        stats_outer.grid_columnconfigure(0, weight=1)
        self._tw["stats_outer"] = stats_outer

        self.api_stats_container = tk.Frame(stats_outer, bg=T["bg_card"])
        self.api_stats_container.grid(row=1, column=0, sticky="")
        self.stat_value_labels = {}
        stat_layout = [
            ["ARR1", "ARR2", "SINF", "POS1", "POS2"],
            ["SCNM", "SCID", "RINF", "SLST", "VLD"]
        ]
        for r, row_keys in enumerate(stat_layout):
            for c, key in enumerate(row_keys):
                lbl_key = tk.Label(
                    self.api_stats_container,
                    text=key, font=(FONT_MONO, SZ_XS, "bold"),
                    fg=T["fg_counter"], bg=T["bg_card"], width=5, anchor="w"
                )
                lbl_key.grid(row=r, column=c*2, padx=(6, 0), sticky="w")
                self._tw["stat_key_labels"].append(lbl_key)
                val_lbl = tk.Label(
                    self.api_stats_container,
                    text="0", font=(FONT_MONO, SZ_XS, "bold"),
                    fg=T["fg_counter_v"], bg=T["bg_card"], width=4, anchor="w"
                )
                val_lbl.grid(row=r, column=c*2+1, padx=(0, 4), sticky="w")
                self.stat_value_labels[key] = val_lbl

        # 5-2-6. 초기 화면 상태를 업데이트합니다.
        self.update_api_counter_ui()
        self.update_button_states()

        # 5-2-5-1. 인증키 길이 변화를 감지해 인증키 입력 버튼의 활성화를 제어합니다.
        self.service_key_var.trace_add("write", lambda *_: self._check_key_btn_state())
        self.backup_key_var.trace_add("write",  lambda *_: self._check_key_btn_state())
        self._check_key_btn_state()

        # 5-2-7. 열쇠가 잠겨 있다면 입력칸을 읽기 전용으로 바꿉니다.
        if self.key_locked:
            self.entry_service_key.config(state='readonly')
            self.entry_backup_key.config(state='readonly')
            _ok = self.get_btn_style("outline")
            _ok.pop("state", None)
            self.btn_key_manage.config(text="인증키\n변경", **_ok)
            # 버튼 타입 추적 업데이트
            for item in self._tw["buttons"]:
                if item[0] is self.btn_key_manage:
                    item[1][0] = "outline"

        # 5-2-8. 화면 중간과 아래쪽의 위아래 크기를 조절할 수 있는 공간을 만듭니다.
        self.super_paned = tk.PanedWindow(self.root, orient=tk.VERTICAL,
                                          sashrelief=tk.RAISED, sashwidth=6)
        self.super_paned.pack(fill="both", expand=True)
        self.main_paned = tk.PanedWindow(self.super_paned, orient=tk.VERTICAL,
                                         sashrelief=tk.RAISED, sashwidth=6)
        self.setup_trees()
        self.super_paned.add(self.main_paned, stretch="always", minsize=220)

        # 5-2-9. 화면 제일 아래쪽에 작업 일지(로그)를 적는 창을 만듭니다.
        log_outer_frame = tk.Frame(self.super_paned, padx=10, pady=5, bg=T["bg_root"])
        self._tw["log_outer"] = log_outer_frame
        log_scroll_frame = tk.Frame(log_outer_frame, bg=T["bg_root"])
        log_scroll_frame.pack(fill="both", expand=True)
        self.txt_log = tk.Text(log_scroll_frame, height=10, bg=T["bg_log"], fg=T["fg_log"],
                               font=(FONT_MONO, SZ_S), state="disabled")
        log_scrollbar = ttk.Scrollbar(log_scroll_frame, orient="vertical", command=self.txt_log.yview, style="App.Vertical.TScrollbar")
        self.txt_log.configure(yscrollcommand=log_scrollbar.set)
        self.txt_log.pack(side="left", fill="both", expand=True)
        log_scrollbar.pack(side="right", fill="y")
        self.super_paned.add(log_outer_frame, minsize=60, height=180)

        # 5-2-10. [초기 스타일 일괄 적용] 생성 직후 스크롤바·Treeview 스타일을 즉시 반영합니다.
        self._apply_scrollbar_style(self.get_theme())

    # 5-3. [버스 정보판 만들기] 실시간 버스 위치와 도착 기록을 보여주는 표 생성 함수
    def setup_trees(self):
        T = self.get_theme()
        # 5-3-1. 화면 상단에 2개의 실시간 현황 표를 배치합니다.
        frame_rt_container = tk.Frame(self.main_paned, bg=T["bg_root"])
        self._tw["tree_frames"].append(frame_rt_container)
        rt_inner = tk.Frame(frame_rt_container, bg=T["bg_root"])
        rt_inner.pack(fill="both", expand=True)
        self._tw["tree_frames"].append(rt_inner)
        rt_inner.grid_columnconfigure(0, weight=1); rt_inner.grid_columnconfigure(1, weight=1)
        rt_inner.grid_rowconfigure(0, weight=1)

        self.btn_searches = []
        self.trees_rt = []; self.lbl_st_names = []

        for i in range(2):
            f = tk.Frame(rt_inner, bg=T["bg_root"])
            f.grid(row=0, column=i, sticky="nsew", padx=2)
            self._tw["tree_frames"].append(f)
            header = tk.Frame(f, bg=T["bg_root"])
            header.pack(fill="x", pady=2)
            self._tw["tree_frames"].append(header)
            inner_header = tk.Frame(header, bg=T["bg_root"])
            inner_header.pack(anchor="center")
            self._tw["tree_frames"].append(inner_header)

            lbl = tk.Label(inner_header, text=f"[정류소 {i+1}] 실시간 현황",
                           fg=T["fg_st1"], bg=T["bg_root"],
                           font=(FONT_SUB, SZ_M, "bold"))
            lbl.pack(side="left")
            self.lbl_st_names.append(lbl)

            ars_entry = tk.Entry(inner_header, textvariable=self.ars_ids[i],
                                 width=10, state="readonly",
                                 readonlybackground=T["bg_entry_ro"],
                                 fg=T["fg_main"], font=(FONT_MONO, SZ_M, "bold"))
            ars_entry.pack(side="left", padx=5)
            self._tw["ars_entries"].append(ars_entry)

            _s_search = self.get_btn_style("disabled")
            _s_search.pop("state", None)
            btn_search = tk.Button(inner_header, text="검색",
                                   command=lambda idx=i: self.open_search_window(idx),
                                   width=0 if CURRENT_OS == "Darwin" else 5,
                                   state="disabled", **_s_search)
            btn_search.pack(side="left", padx=(4, 0),
                            ipady=0 if CURRENT_OS == "Darwin" else 1)
            self.btn_searches.append(btn_search)
            self._tw["buttons"].append((btn_search, ["disabled"]))

            tree_frame = tk.Frame(f, bg=T["bg_root"])
            tree_frame.pack(fill="both", expand=True)
            self._tw["tree_frames"].append(tree_frame)
            cols = ("route", "bus1_no", "bus1_msg", "bus2_no", "bus2_msg")
            tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
            for col in cols: tree.heading(col, text=self.get_col_name(col))
            tree.column("route", width=69); tree.column("bus1_no", width=86)
            tree.column("bus1_msg", width=130); tree.column("bus2_no", width=86)
            tree.column("bus2_msg", width=129)
            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview, style="App.Vertical.TScrollbar")
            tree.configure(yscrollcommand=vsb.set)
            tree.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")
            self.trees_rt.append(tree)
        self.main_paned.add(frame_rt_container, minsize=100)

        # 5-3-2. 화면 하단에 2개의 도착 기록 표를 배치합니다.
        frame_hist_container = tk.Frame(self.main_paned, bg=T["bg_root"])
        self._tw["tree_frames"].append(frame_hist_container)
        hist_inner = tk.Frame(frame_hist_container, bg=T["bg_root"])
        hist_inner.pack(fill="both", expand=True)
        self._tw["tree_frames"].append(hist_inner)
        hist_inner.grid_columnconfigure(0, weight=1); hist_inner.grid_columnconfigure(1, weight=1)
        hist_inner.grid_rowconfigure(0, weight=1)

        self.trees_hist = []; self.lbl_hist_titles = []
        for i in range(2):
            f = tk.Frame(hist_inner, bg=T["bg_root"])
            f.grid(row=0, column=i, sticky="nsew", padx=2)
            self._tw["tree_frames"].append(f)
            header = tk.Frame(f, bg=T["bg_root"])
            header.pack(fill="x", pady=2)
            self._tw["tree_frames"].append(header)
            inner_header = tk.Frame(header, bg=T["bg_root"])
            inner_header.pack(anchor="center")
            self._tw["tree_frames"].append(inner_header)

            lbl = tk.Label(inner_header, text=f"[정류소 {i+1}] 도착 기록",
                           fg=T["fg_st2"], bg=T["bg_root"],
                           font=(FONT_SUB, SZ_M, "bold"))
            lbl.pack(side="left")
            self.lbl_hist_titles.append(lbl)

            _s_del = self.get_btn_style("normal", font_size=SZ_XXS)
            btn_del = tk.Button(inner_header, text="기록 삭제",
                                command=lambda idx=i: self.clear_history(idx),
                                **_s_del)
            btn_del.pack(side="left", padx=10,
                         ipady=0 if CURRENT_OS == "Darwin" else 1)
            self._tw["buttons"].append((btn_del, ["normal"]))

            tree_frame = tk.Frame(f, bg=T["bg_root"])
            tree_frame.pack(fill="both", expand=True)
            self._tw["tree_frames"].append(tree_frame)
            cols = ("data_time", "route", "veh_no", "corp", "status")
            tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
            for col in cols: tree.heading(col, text=self.get_col_name(col))
            tree.column("data_time", width=135); tree.column("route", width=63)
            tree.column("veh_no", width=94); tree.column("corp", width=135)
            tree.column("status", width=73)
            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview, style="App.Vertical.TScrollbar")
            tree.configure(yscrollcommand=vsb.set)
            tree.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")
            self.trees_hist.append(tree)
        self.main_paned.add(frame_hist_container, minsize=100)

        # 초기 Treeview 스타일 적용
        self._apply_treeview_style(T)

    # 5-4. [버튼 옷입히기] 현재 테마와 버튼 타입에 따라 스타일 딕셔너리를 반환합니다.
    def get_btn_style(self, btn_type="normal", font_size=None):
        # 5-4-0-1. 하위 호환: 구버전 색상 문자열 매핑
        _color_map = {
            "#28a745": "normal", "#007bff": "normal", "#8e44ad": "normal",
            "#f1c40f": "normal", "#2ecc71": "normal",
            "#d63031": "outline", "#7f8c8d": "outline",
        }
        if btn_type.startswith("#"):
            btn_type = _color_map.get(btn_type, "normal")

        target_sz = font_size if font_size else SZ_S
        T = self.get_theme()
        CURSOR_HAND = "pointinghand" if CURRENT_OS == "Darwin" else "hand2"

        if CURRENT_OS == "Darwin":
            # ──────────────────────────────────────────────────────────────────
            # macOS Aqua 테마: bg/activebackground 는 시스템이 덮어씁니다.
            # 유효 속성: fg, activeforeground, font, highlightbackground,
            #            highlightthickness, relief, bd, padx, pady, overrelief, cursor
            #
            # [변경 4-6] 요구사항:
            #   - 버튼 크기: 내부 문구에 딱 맞게 타이트 (padx=2, pady=0)
            #   - 테두리 제거: highlightthickness=0, bd=0, relief="flat"
            #   - 라이트 fg: #0f0f0f (YouTube 라이트 검정)
            #   - 다크   fg: #f1f1f1 (YouTube 다크 흰색)
            # ──────────────────────────────────────────────────────────────────
            _fg      = "#0f0f0f" if not self.is_dark_mode else "#f1f1f1"
            _fg_dim  = "#888888" if not self.is_dark_mode else "#555555"

            if btn_type == "disabled":
                return {
                    "fg": _fg_dim,
                    "font": (FONT_MAIN, target_sz, "normal"),
                    "cursor": "arrow",
                    "highlightthickness": 0,
                    "relief": "flat", "bd": 0,
                    "padx": 2, "pady": 0,
                    "state": "normal",
                }
            elif btn_type == "outline":
                return {
                    "fg": _fg,
                    "activeforeground": _fg,
                    "font": (FONT_MAIN, target_sz, "bold"),
                    "cursor": CURSOR_HAND,
                    "highlightthickness": 0,
                    "relief": "flat", "bd": 0,
                    "padx": 2, "pady": 0,
                    "overrelief": "flat",
                }
            else:  # normal
                return {
                    "fg": _fg,
                    "activeforeground": _fg,
                    "font": (FONT_MAIN, target_sz, "bold"),
                    "cursor": CURSOR_HAND,
                    "highlightthickness": 0,
                    "relief": "flat", "bd": 0,
                    "padx": 2, "pady": 0,
                    "overrelief": "flat",
                }
        else:
            # Windows / Linux
            if btn_type == "disabled":
                return {
                    "bg": T["btn_disabled_bg"],
                    "fg": T["btn_disabled_fg"],
                    "font": (FONT_MAIN, target_sz, "normal"),
                    "relief": "groove", "bd": 2,
                    "highlightbackground": T["btn_disabled_border"],
                    "highlightthickness": 1,
                    "cursor": "arrow",
                    "state": "disabled",
                    "disabledforeground": T["btn_disabled_fg"],
                }
            elif btn_type == "outline":
                return {
                    "bg": T["btn_outline_bg"],
                    "fg": T["btn_outline_fg"],
                    "activebackground": T["btn_outline_bg"],
                    "activeforeground": T["btn_outline_fg"],
                    "font": (FONT_MAIN, target_sz, "bold"),
                    "relief": "ridge", "bd": 2,
                    "highlightbackground": T["btn_outline_border"],
                    "highlightthickness": 1,
                    "cursor": CURSOR_HAND,
                    "overrelief": "solid",
                }
            else:  # normal
                return {
                    "bg": T["btn_normal_bg"],
                    "fg": T["btn_normal_fg"],
                    "activebackground": T["btn_normal_active"],
                    "activeforeground": T["btn_normal_fg"],
                    "font": (FONT_MAIN, target_sz, "bold"),
                    "relief": "raised", "bd": 2,
                    "highlightbackground": T["btn_normal_border"],
                    "highlightthickness": 1,
                    "cursor": CURSOR_HAND,
                    "overrelief": "solid",
                }

    def get_col_name(self, code):
        mapping = {"route": "노선", "bus1_no": "1번차량", "bus1_msg": "도착정보", "bus2_no": "2번차량", "bus2_msg": "도착정보", "data_time": "데이터 시각", "veh_no": "차량번호", "corp": "운수사명", "status": "상태"}
        return mapping.get(code, code)
    
    # 5-6. [링크 열기] 인터넷 주소를 웹 브라우저로 열어주는 함수
    def open_link(self, url):
        webbrowser.open_new(url)
        
    # 5-7. [메일 보내기] 클릭하면 바로 이메일을 보낼 수 있게 창을 띄우는 함수
    def send_email_link(self, email):
        webbrowser.open_new(f"mailto:{email}")

    # 5-8. [칸 고정] 표의 칸 너비를 실수로 바꾸지 못하게 고정하는 함수
    def prevent_column_resize(self, event):
        if event.widget.identify_region(event.x, event.y) == "separator":
            return "break" 

    # 5-9. [통계 창] API를 몇 번 호출했는지 상세히 보여주는 새 창 열기 함수
    def open_api_stats_window(self):
        # 5-9-1. 창이 이미 떠 있다면 그 창을 맨 앞으로 가져옵니다.
        if self.stats_win is not None and self.stats_win.winfo_exists():
            self.stats_win.lift()
            self.stats_win.focus_force()
            return

        T = self.get_theme()

        # 5-9-2. 새로운 창을 만듭니다.
        self.stats_win = tk.Toplevel(self.root)
        self.stats_win.title("API 호출 상세 현황")
        self.stats_win.geometry("735x460")
        self.stats_win.minsize(735, 420)
        self.stats_win.transient(self.root)
        self.stats_win.protocol("WM_DELETE_WINDOW", self.on_stats_win_close)
        self.stats_win.configure(bg=T["bg_panel"])

        # 제목 레이블
        lbl_title = tk.Label(self.stats_win,
                 text="API 호출 상세 통계  (메인키 / 백업키 / 합계 순)",
                 font=(FONT_MAIN, SZ_M, "bold"),
                 bg=T["bg_panel"], fg=T["fg_main"], pady=8)
        lbl_title.pack(side="top", fill="x")

        # 5-9-3. Notebook 스타일 (다크/라이트)
        sty = ttk.Style()
        if self.is_dark_mode:
            sty.configure("StatsNB.TNotebook",
                          background=T["bg_panel"], borderwidth=0)
            sty.configure("StatsNB.TNotebook.Tab",
                          background=T["bg_card"], foreground=T["fg_sub"],
                          padding=[8, 3])
            sty.map("StatsNB.TNotebook.Tab",
                    background=[("selected", T["bg_tree_sel"])],
                    foreground=[("selected", T["fg_main"])])
        else:
            sty.configure("StatsNB.TNotebook",
                          background="#ecf0f1", borderwidth=0)
            sty.configure("StatsNB.TNotebook.Tab",
                          background="#dfe6e9", foreground="#2d3436",
                          padding=[8, 3])
            sty.map("StatsNB.TNotebook.Tab",
                    background=[("selected", "#ffffff")],
                    foreground=[("selected", "#2d3436")])

        notebook = ttk.Notebook(self.stats_win, style="StatsNB.TNotebook")
        notebook.pack(fill="both", expand=True, padx=6, pady=(0, 6))

        # 5-9-3-1. Treeview + Scrollbar 헬퍼
        def make_tree(parent):
            parent.configure(bg=T["bg_panel"])
            cols = ("abbr", "url", "main", "back", "total")
            t = ttk.Treeview(parent, columns=cols, show="headings",
                             style="Stats.Treeview")
            t.heading("abbr",  text="약칭")
            t.heading("url",   text="API 실제 호출 주소")
            t.heading("main",  text="메인키")
            t.heading("back",  text="백업키")
            t.heading("total", text="합계")
            t.column("abbr",  width=65,  anchor="center", stretch=False)
            t.column("url",   width=395, anchor="w",      stretch=True)
            t.column("main",  width=60,  anchor="center", stretch=False)
            t.column("back",  width=60,  anchor="center", stretch=False)
            t.column("total", width=65,  anchor="center", stretch=False)
            t.bind('<Button-1>', self.prevent_column_resize)
            sb = ttk.Scrollbar(parent, orient="vertical", command=t.yview, style="App.Vertical.TScrollbar")
            t.configure(yscrollcommand=sb.set)
            t.pack(side="left", fill="both", expand=True, padx=(4, 0), pady=4)
            sb.pack(side="right", fill="y", pady=4, padx=(0, 4))
            return t

        # Stats.Treeview 스타일 (다크/라이트)
        if self.is_dark_mode:
            sty.configure("Stats.Treeview",
                background=T["bg_tree"], foreground=T["fg_tree"],
                fieldbackground=T["bg_tree"], rowheight=22)
            sty.configure("Stats.Treeview.Heading",
                background=T["bg_card"], foreground=T["fg_main"], relief="flat")
            sty.map("Stats.Treeview",
                background=[("selected", T["bg_tree_sel"])],
                foreground=[("selected", T["fg_main"])])
            sty.map("Stats.Treeview.Heading",
                background=[("active", T["border"])])
        else:
            sty.configure("Stats.Treeview",
                background="#ffffff", foreground="#000000",
                fieldbackground="#ffffff", rowheight=22)
            sty.configure("Stats.Treeview.Heading",
                background="#f0f0f0", foreground="#000000", relief="raised")
            sty.map("Stats.Treeview",
                background=[("selected", "#0078d7")],
                foreground=[("selected", "#ffffff")])
            sty.map("Stats.Treeview.Heading",
                background=[("active", "#e0e0e0")])

        # 5-9-3-2. 탭 1: 오늘의 합계
        tab_today = tk.Frame(notebook, bg=T["bg_panel"])
        notebook.add(tab_today, text="  오늘의 합계  ")
        tree_today = make_tree(tab_today)

        # 5-9-3-3. 탭 2: 어제의 합계
        tab_yest = tk.Frame(notebook, bg=T["bg_panel"])
        notebook.add(tab_yest, text="  어제의 합계  ")
        tree_yest = make_tree(tab_yest)

        # 5-9-4. 트리 채우기 헬퍼
        def fill_tree(t, stats_dict, by_key_dict, label_row):
            for item in t.get_children():
                t.delete(item)
            total_all = total_main = total_back = 0
            for key, count in stats_dict.items():
                url   = self.api_urls.get(key, "-")
                m_cnt = by_key_dict["main"].get(key, 0)
                b_cnt = by_key_dict["back"].get(key, 0)
                t.insert("", "end", values=(key, url, f"{m_cnt}", f"{b_cnt}", f"{count}"))
                total_all  += count
                total_main += m_cnt
                total_back += b_cnt
            t.insert("", "end", values=(" ", " ", " ", " ", " "))
            t.insert("", "end", values=(label_row, " - ",
                                        f"{total_main}", f"{total_back}", f"{total_all}"))

        # 5-9-5. 1초마다 양쪽 탭을 모두 갱신합니다.
        def update_stats_loop():
            if not self.stats_win or not self.stats_win.winfo_exists():
                return
            fill_tree(tree_today, self.api_stats,
                      self.api_stats_by_key, "오늘의 합계")
            fill_tree(tree_yest, self.api_stats_yesterday,
                      self.api_stats_by_key_yesterday, "어제의 합계")
            self.stats_win.after(1000, update_stats_loop)

        update_stats_loop()

    # 5-10. [수첩 닫기] 통계 창을 닫을 때 메모리에서 지워주는 함수
    def on_stats_win_close(self):
        if self.stats_win:
            self.stats_win.destroy()
            self.stats_win = None

    # 5-11. [스위치 조절] 상황(감시 중인지 등)에 따라 버튼들을 켜거나 끄는 함수
    def update_button_states(self):
        # 5-11-1. 정류소 정보가 있는지 먼저 확인합니다.
        has_info = (bool(self.target_st_info[0].get('routes')) or
                    bool(self.target_st_info[1].get('routes')))

        # 5-11-2. 버튼의 작동 여부를 설정하는 도우미 함수
        def apply_btn(btn, btn_type, is_active, cmd, text=None):
            # 5-11-2-1. 비활성 상태라면 disabled 스타일로 덮어씁니다.
            effective_type = btn_type if is_active else "disabled"
            style = self.get_btn_style(effective_type)
            if text:
                style["text"] = text
            style.pop("state", None)
            # 5-11-2-3. _tw["buttons"] 추적 셀 갱신 (apply_theme 재스타일용)
            for item in self._tw.get("buttons", []):
                if item[0] is btn:
                    item[1][0] = effective_type
                    break
            if CURRENT_OS == "Darwin":
                if btn is self.btn_toggle:
                    self._btn_active['toggle'] = is_active
                elif btn is self.btn_manual:
                    self._btn_active['manual'] = is_active
                btn.config(state=("normal" if is_active else "disabled"),
                           command=(cmd if is_active else lambda: None), **style)
            else:
                btn.config(state=("normal" if is_active else "disabled"),
                           command=(cmd if is_active else lambda: None), **style)

        # 5-11-3. 감시 중일 때 → 통합 버튼을 "중지"(outline), 수동 갱신 활성화
        if self.is_monitoring:
            apply_btn(self.btn_toggle, "outline", True,
                      self.stop_monitoring, text="중지")
            apply_btn(self.btn_manual, "normal", True, self.manual_refresh)
        # 5-11-4. 감시 중이 아닐 때 → 통합 버튼을 "자동 기록 시작"
        else:
            apply_btn(self.btn_toggle, "normal", has_info,
                      self.start_monitoring, text="자동 기록 시작")
            apply_btn(self.btn_manual, "normal", has_info, self.manual_refresh)

    # 5-11-5. [통합 토글] 시작/중지 버튼 클릭 시 현재 상태에 따라 분기합니다.
    def _on_toggle_monitoring(self):
        # macOS guard: Aqua 테마의 state='disabled' 클릭 통과 버그 차단
        if CURRENT_OS == "Darwin" and not self._btn_active.get('toggle', True):
            return
        if self.is_monitoring:
            self.stop_monitoring()
        else:
            self.start_monitoring()

    # 5-11-6. [인증키 길이 감시] 인증키 입력 버튼 활성화 조건을 검사합니다.
    #   활성화 조건 (두 조건 모두 충족해야 함):
    #     ① 메인키: 정확히 64글자
    #     ② 백업키: 비어있거나(0자) 정확히 64글자  ← 그 외 길이면 비활성
    #   - key_locked=True(잠금 상태)이면 "인증키 변경" 버튼이므로 항상 활성화합니다.
    def _check_key_btn_state(self):
        if self.key_locked:
            # 잠금 상태에서는 "인증키 변경" 버튼이 항상 눌릴 수 있어야 합니다.
            return
        main_len = len(self.service_key_var.get().strip())
        back_len = len(self.backup_key_var.get().strip())
        # 메인키 64자 AND (백업키 0자 OR 백업키 64자)
        is_valid = (main_len == 64) and (back_len == 0 or back_len == 64)
        if is_valid:
            _s = self.get_btn_style("normal")
            _s.pop("state", None)
            self.btn_key_manage.config(state="normal", **_s)
        else:
            _s = self.get_btn_style("disabled")
            _s.pop("state", None)
            self.btn_key_manage.config(state="disabled", **_s)


    # 5-12. [숫자판 업데이트] 메인 화면의 API 호출 횟수를 실시간으로 갱신하는 함수
    def update_api_counter_ui(self):
        # 5-12-1. 저장된 숫자판들을 돌면서 최신 숫자로 글자를 바꿉니다.
        for key, val_lbl in self.stat_value_labels.items():
            if key in self.api_stats:
                val_lbl.config(text=str(self.api_stats[key]))

    # 4그룹 : 보안 및 인증 관리 (Security & Auth)

    # 5-13. [열쇠 불러오기] 프로그램이 켜질 때 저장해둔 비밀 열쇠(인증키)를 꺼내서 확인하는 단계
    def load_saved_key(self):
        # 5-13-1. 열쇠가 저장된 파일이 있는지 확인합니다.
        if os.path.exists(self.key_file_path):
            try:
                # 5-13-2. 파일을 열어서 암호화된 내용을 읽어옵니다.
                with open(self.key_file_path, "rb") as f:
                    encrypted_data = f.read()
                    if encrypted_data:
                        # 5-13-2-1. [비밀 풀기] 암호화된 내용을 다시 글자로 바꿉니다.
                        decrypted_data = cipher_suite.decrypt(encrypted_data).decode('utf-8')
                        main_k, back_k = "", ""
                        # 5-13-2-2. 메인 열쇠와 백업 열쇠가 같이 있다면 나누어서 저장합니다.
                        if "|" in decrypted_data:
                            main_k, back_k = decrypted_data.split("|", 1)
                        else:
                            main_k = decrypted_data
                        
                        # 5-13-2-3. 화면에 있는 입력창에 불러온 열쇠들을 넣어줍니다.
                        self.service_key_var.set(main_k)
                        self.backup_key_var.set(back_k)
                        
                        # 5-13-2-4. [자동 검사] 불러온 열쇠가 진짜인지 바로 확인을 시작합니다.
                        if main_k.strip() or back_k.strip():
                            # 5-13-2-4-1. 로그창이 준비될 때까지 아주 잠시(0.1초) 기다렸다가 메시지를 띄웁니다.
                            self.root.after(100, lambda: self.log("🚀 저장된 인증키를 불러와 유효성을 검사합니다..."))
                            self.root.after(150, self.toggle_key_lock)

            except Exception as e:
                # 5-13-3. 만약 열쇠를 읽다가 실수를 하면 무엇이 틀렸는지 알려줍니다.
                print(f"인증키 복호화 실패: {e}")

    # 5-14. [열쇠 저장하기] 입력한 소중한 열쇠(인증키)를 잊어버리지 않게 파일로 잠궈두는 함수
    def save_key_to_file(self):
        # 5-14-1. 입력창에 적힌 메인 열쇠와 백업 열쇠를 가져옵니다.
        main_k = self.service_key_var.get().strip()
        back_k = self.backup_key_var.get().strip()
        
        # 5-14-2. 두 개의 열쇠를 하나로 묶어줍니다.
        combined_data = f"{main_k}|{back_k}"
        
        try:
            # 5-14-3. [비밀 잠금] 남들이 못 보게 암호로 바꾸어 파일에 기록합니다.
            encrypted_data = cipher_suite.encrypt(combined_data.encode('utf-8'))
            with open(self.key_file_path, "wb") as f:
                f.write(encrypted_data)
            self.log("🔒 메인/백업 인증키가 암호화되어 안전하게 저장되었습니다.")
        except Exception as e:
            # 5-14-4. 저장하다가 문제가 생기면 로그창에 알립니다.
            self.log(f"⚠ 인증키 암호화 저장 실패: {e}")

    # 5-15. [열쇠 확인 및 잠금] 입력한 열쇠가 진짜인지 나라 서버에 물어보고 창을 잠그는 함수
    def toggle_key_lock(self):
        import xml.etree.ElementTree as ET
        from urllib.parse import unquote
        import requests

        # 5-15-1. 아직 열쇠가 확인되지 않아 열려있는 상태라면 검사를 시작합니다.
        if not self.key_locked:
            # 5-15-1-1. 입력창에 적힌 글자들을 가져옵니다.
            main_input = self.service_key_var.get().strip()
            back_input = self.backup_key_var.get().strip()

            # 5-15-1-2. 메인 열쇠가 없으면 경고창을 띄우고 멈춥니다.
            if not main_input:
                messagebox.showwarning("알림", "메인 인증키를 입력해주세요.")
                return

            self.log("🔑 인증키 유효성 검사 및 적용 중...")
            
            # 5-15-1-3. 검사할 열쇠들을 목록으로 만듭니다.
            keys_to_test = [("메인", main_input, self.service_key_var)]
            if back_input:
                keys_to_test.append(("백업", back_input, self.backup_key_var))

            valid_main, valid_back = "", ""
            any_success = False

            # 5-15-1-4. [진짜인지 확인] 목록에 있는 열쇠들을 하나씩 API 서버에 물어봅니다.
            for name, key_val, var_obj in keys_to_test:
                test_url = "http://ws.bus.go.kr/api/rest/busRouteInfo/getBusRouteList"
                try:
                    # 5-15-1-4-1. 서버에 아주 짧은 질문을 던져봅니다.
                    r = requests.get(test_url, params={'serviceKey': unquote(key_val), 'strSrch': '12'}, timeout=5)
                    # 5-15-1-4-2. 서버가 'OK(0)'라고 대답하면 진짜 열쇠입니다.
                    if "<headerCd>0</headerCd>" in r.text:
                        self.log(f"✅ {name} 인증키: 정상 작동 확인")
                        if name == "메인": valid_main = key_val
                        else: valid_back = key_val
                        any_success = True
                    else:
                        # 5-15-1-4-3. 가짜거나 망가진 열쇠면 입력창에서 지워버립니다.
                        var_obj.set("") 
                        self.log(f"❌ {name} 인증키: 작동 불가하여 삭제되었습니다.")
                except Exception as e:
                    # 5-15-1-4-4. 인터넷 연결 문제 등으로 확인이 안 되면 삭제 처리합니다.
                    var_obj.set("")
                    self.log(f"❌ {name} 통신 에러로 삭제 처리되었습니다.")

            # 5-15-1-5. [성공 시 처리] 하나라도 진짜 열쇠가 있다면 프로그램을 활성화합니다.
            if any_success:
                # 5-15-1-5-1. 성공한 열쇠들을 기억하고 잠금 상태로 바꿉니다.
                self.final_main_key = valid_main
                self.final_backup_key = valid_back
                self.key_locked = True
                
                # 5-15-1-5-2. 실수로 지우지 못하게 입력창을 회색(읽기전용)으로 잠급니다.
                self.entry_service_key.config(state='readonly', readonlybackground=self.get_theme()["bg_entry_ro"])
                self.entry_backup_key.config(state='readonly', readonlybackground=self.get_theme()["bg_entry_ro"])
                # 잠금 완료 → 버튼을 outline(테두리) 스타일로 변경
                _outline = self.get_btn_style("outline")
                _outline.pop("state", None)
                self.btn_key_manage.config(text="인증키\n변경", **_outline)
                for item in self._tw.get("buttons", []):
                    if item[0] is self.btn_key_manage: item[1][0] = "outline"; break
                
                # 5-15-1-5-3. 이제 버스를 찾을 수 있도록 검색 버튼들을 켭니다.
                _s_on = self.get_btn_style("normal")
                _s_on.pop("state", None)
                for btn in self.btn_searches:
                    btn.config(state="normal", **_s_on)
                    for item in self._tw.get("buttons", []):
                        if item[0] is btn: item[1][0] = "normal"; break
                
                # 5-15-1-5-4. VLD 통계 숫자를 올리고 열쇠를 파일에 저장합니다.
                #             성공한 각 키(메인/백업)별로 개별 카운터도 올립니다.
                if valid_main:
                    self.api_stats['VLD'] += 1
                    self.api_stats_by_key["main"]['VLD'] = self.api_stats_by_key["main"].get('VLD', 0) + 1
                if valid_back:
                    self.api_stats['VLD'] += 1
                    self.api_stats_by_key["back"]['VLD'] = self.api_stats_by_key["back"].get('VLD', 0) + 1
                self.root.after(0, self.update_api_counter_ui)
                self.save_key_to_file()
                self.log("🔒 정상 인증키가 확정되었습니다. 이제 검색이 가능합니다.")
            else:
                # 5-15-1-6. 진짜 열쇠가 하나도 없으면 슬픈 알림창을 띄웁니다.
                messagebox.showerror("인증 실패", "사용 가능한 인증키가 없습니다.")

        # 5-15-2. 이미 잠겨있는 상태에서 버튼을 누르면 잠금을 풀고 수정 모드로 바꿉니다.
        else:
            self.key_locked = False
            # 5-15-2-1. 입력창을 다시 하얀색으로 바꿔서 글자를 쓸 수 있게 합니다.
            _bg_e = self.get_theme()["bg_entry"]
            _fg_e = self.get_theme()["fg_main"]
            self.entry_service_key.config(state='normal', background=_bg_e, fg=_fg_e, insertbackground=_fg_e)
            self.entry_backup_key.config(state='normal', background=_bg_e, fg=_fg_e, insertbackground=_fg_e)
            # 잠금 해제 → 버튼을 normal(녹색) 스타일로 복원
            _normal = self.get_btn_style("normal")
            _normal.pop("state", None)
            self.btn_key_manage.config(text="인증키\n입력", **_normal)
            for item in self._tw.get("buttons", []):
                if item[0] is self.btn_key_manage: item[1][0] = "normal"; break
            
            # 5-15-2-2. 열쇠를 고치는 동안에는 검색 버튼을 못 누르게 끕니다.
            _s_off = self.get_btn_style("disabled")
            _s_off.pop("state", None)
            for btn in self.btn_searches:
                btn.config(state="disabled", **_s_off)
                for item in self._tw.get("buttons", []):
                    if item[0] is btn: item[1][0] = "disabled"; break
            
            # 5-15-2-3. 잠금 해제 즉시 64자 검증을 다시 적용합니다.
            self._check_key_btn_state()
            self.log("🔓 인증키 수정 모드입니다. (검증 전까지 검색 버튼 잠금)")

    # 5그룹: 데이터 수집 및 통신 (Fetching & API)

    # 5-16. [인터넷 대화하기] 나라 서버에 버스 정보를 물어보고 답변을 받아오는 핵심 함수
    def fetch_api(self, url, params, api_type=None):
        from urllib.parse import unquote
        # 5-16-1. 우리가 가진 메인 열쇠와 백업 열쇠를 준비합니다.
        keys = [self.service_key_var.get().strip(), self.backup_key_var.get().strip()]
        
        # 5-16-2. 어떤 정보를 물어볼지에 따라 인터넷 주소를 미리 정합니다. API 타입별 주소 설정 (ARR1->SINF / POS1->POS2 순서)
        p_url, f_url = url, url
        if api_type == "ARR1":
            p_url = "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRoute"
            f_url = "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRouteAll"
        elif api_type == "POS2":
            p_url = "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRtid"
            f_url = "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRouteSt"

        # 5-16-3. 메인 주소와 비상 주소를 섞어서 총 4번의 시도 계획을 세웁니다.
        attempts = [
            (p_url, "메인", keys[0]), (p_url, "백업", keys[1]), 
            (f_url, "메인", keys[0]), (f_url, "백업", keys[1])
        ]

        # 5-16-4. 계획에 따라 서버에 질문을 던집니다.
        for step, (curr_url, name, key) in enumerate(attempts):
            if not key: continue
            try:
                # 5-16-4-1. 열쇠를 꽂고 서버에 연결합니다.
                params['serviceKey'] = unquote(key)
                resp = requests.get(curr_url, params=params, timeout=10)
                
                # 5-16-4-2. 일일호출량이 초과되어 서버가 거절하면 다음 시도로 넘어갑니다.
                if "LIMITED NUMBER OF SERVICE REQUESTS" in resp.text or "<headerCd>22</headerCd>" in resp.text:
                    continue 

                # 5-16-4-3. 열쇠가 잘못되었다고 하면 다음 열쇠로 시도합니다.
                if "SERVICE KEY IS NOT REGISTERED" in resp.text or "UNREGISTERED_KEY" in resp.text:
                    continue

                # 5-16-4-4. 서버가 대답을 잘 해주었다면 내용을 꼼꼼히 읽어봅니다.
                if resp.status_code == 200:
                    root = ET.fromstring(resp.text)
                    header_cd = root.findtext(".//headerCd")
                    err_msg = root.findtext('.//headerMsg') or ""
                    
                    # 5-16-4-4-1. 정보를 찾았거나, 정보가 없다는 확실한 대답을 들었을 때만 횟수를 셉니다.
                    is_no_result = ("결과가 없습니다" in err_msg) or ("NODATA" in err_msg)
                    
                    if header_cd == "0" or is_no_result:
                        # 5-16-4-4-1-1. 어떤 질문을 했는지 종류별로 통계 숫자를 올립니다.
                        success_url = curr_url
                        key_type = "VLD"  # 기본값: 알 수 없는 호출
                        if "getArrInfoByRoute" in success_url and "All" not in success_url: key_type = "ARR1"
                        elif "getArrInfoByRouteAll" in success_url: key_type = "ARR2"
                        elif "getBusPosByRtid" in success_url: key_type = "POS1"
                        elif "getStationByUid" in success_url: key_type = "SINF"
                        elif "getBusPosByRouteSt" in success_url: key_type = "POS2"
                        elif "getStationByName" in success_url: key_type = "SCNM"
                        elif "getRouteByStation" in success_url: key_type = "SCID"
                        elif "getRouteInfo" in success_url: key_type = "RINF"
                        elif "getStaionByRoute" in success_url: key_type = "SLST"
                        
                        # 5-16-4-4-1-1-1. 합산 카운터를 올립니다.
                        self.api_stats[key_type] += 1
                        # 5-16-4-4-1-1-2. 어떤 키(메인/백업)가 성공했는지 개별 카운터도 올립니다.
                        key_slot = "main" if name == "메인" else "back"
                        self.api_stats_by_key[key_slot][key_type] = \
                            self.api_stats_by_key[key_slot].get(key_type, 0) + 1
                        self.root.after(0, self.update_api_counter_ui)

                        # 5-16-4-4-1-2. 진짜 데이터가 들어있으면 내용을 전달하고, 없으면 이유를 적습니다.
                        if header_cd == "0":
                            return root
                        else:
                            # 5-16-4-4-1-2-1. 버스 위치를 물어봤는데 버스가 한 대도 없으면 로그를 남기고
                            #   첫차 시간(f_tm)을 함께 반환합니다. — 호출자는 (None, f_tm) 형태를 처리해야 합니다.
                            if any(u in success_url for u in ["getBusPosByRtid", "getBusPosByRouteSt"]):
                                rid_param = params.get('busRouteId', '')
                                rnm_log = self.rid_to_rnm.get(rid_param, rid_param)
                                self.log(f"{rnm_log}번은 운행중인 차량이 없습니다.")
                                # 첫차 시간을 노선정보 API(getRouteInfo)에서 가져옵니다.
                                f_tm = self._fetch_route_first_time(rid_param)
                                return ('NO_BUS', f_tm)
                            return None
                        
            except Exception:
                # 5-16-4-5. 중간에 실수가 생기면 다음 방법으로 다시 시도합니다.
                continue
        return None
    
    # 5-16b. [첫차 시각 조회] 해당 노선의 첫 버스 시각(f_tm)을 getRouteInfo API로 조회합니다.
    def _fetch_route_first_time(self, rid):
        """getRouteInfo에서 firstBusTm 필드를 가져옵니다 (HHmm 또는 HHmmss 형식).
        실패 또는 없는 오전 시리즈 노선은 None 반환."""
        if not rid:
            return None
        try:
            root_ri = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo",
                {'busRouteId': rid}
            )
            if root_ri is None or isinstance(root_ri, tuple):
                return None
            f_tm_raw = root_ri.findtext(".//firstBusTm") or ""
            if f_tm_raw:
                return self.format_hhmm(f_tm_raw)  # HH:MM 형식
        except Exception:
            pass
        return None

    # 5-17. [검색 창 열기] 버스 정류소를 찾고 기록할 버스를 고르는 팝업 창 함수
    def open_search_window(self, target_idx):
        # 5-17-1. 열쇠가 없으면 검색을 할 수 없습니다.
        if not self.service_key_var.get().strip():
            messagebox.showwarning("알림", "인증키를 먼저 입력해주세요."); return

        T = self.get_theme()

        # ── Search 전용 ttk.Style ─────────────────────────────────────────
        sty = ttk.Style()
        if self.is_dark_mode:
            sty.configure("Search.Treeview",
                background=T["bg_tree"], foreground=T["fg_tree"],
                fieldbackground=T["bg_tree"], rowheight=22)
            sty.configure("Search.Treeview.Heading",
                background=T["bg_card"], foreground=T["fg_main"],
                font=(FONT_SUB, SZ_S, "bold"), relief="flat")
            sty.map("Search.Treeview",
                background=[("selected", T["bg_tree_sel"])],
                foreground=[("selected", T["fg_main"])])
            sty.map("Search.Treeview.Heading",
                background=[("active", T["border"])])
        else:
            sty.configure("Search.Treeview",
                background="#ffffff", foreground="#000000",
                fieldbackground="#ffffff", rowheight=22)
            sty.configure("Search.Treeview.Heading",
                background="#dfe6e9", foreground="#2d3436",
                font=(FONT_SUB, SZ_S, "bold"), relief="raised")
            sty.map("Search.Treeview",
                background=[("selected", "#0078d7")],
                foreground=[("selected", "#ffffff")])
            sty.map("Search.Treeview.Heading",
                background=[("active", "#c8d6e5")])

        # 5-17-2. 검색용 새 창
        search_win = tk.Toplevel(self.root)
        search_win.title(f"정류소 {target_idx+1} 검색 및 노선 선택")
        search_win.geometry("1000x850")
        search_win.grab_set()
        search_win.minsize(500, 500)
        search_win.configure(bg=T["bg_panel"])

        # 5-17-3. 검색 입력 행
        frame_search = tk.Frame(search_win, pady=10, bg=T["bg_panel"])
        frame_search.pack(fill="x")
        tk.Label(frame_search, text="정류소명/ID:", font=(FONT_MAIN, SZ_S),
                 bg=T["bg_panel"], fg=T["fg_main"]).pack(side="left", padx=(15,5))
        search_ent = tk.Entry(frame_search, width=30,
                              bg=T["bg_entry"], fg=T["fg_main"],
                              insertbackground=T["fg_main"])
        search_ent.pack(side="left", padx=5)
        search_ent.focus_set()

        # 5-17-4. 정류소 목록 표
        tree_st_frame = tk.Frame(search_win, bg=T["bg_panel"])
        tree_st_frame.pack(fill="x", padx=15, pady=5)
        st_scroll = ttk.Scrollbar(tree_st_frame, orient="vertical", style="App.Vertical.TScrollbar")
        tree_st = ttk.Treeview(tree_st_frame, columns=("name", "arsid"),
                               show="headings", height=5,
                               style="Search.Treeview",
                               yscrollcommand=st_scroll.set)
        st_scroll.config(command=tree_st.yview)
        tree_st.heading("name",  text="정류소명")
        tree_st.heading("arsid", text="ARS-ID")
        tree_st.column("name",  width=659, stretch=True)
        tree_st.column("arsid", width=141, stretch=True)
        tree_st.bind("<Button-1>", self.prevent_column_resize)
        tree_st.pack(side="left", fill="both", expand=True)
        st_scroll.pack(side="right", fill="y")

        # 5-17-5. 노선 선택 버튼 행
        btn_action_frame = tk.Frame(search_win, bg=T["bg_panel"])
        btn_action_frame.pack(fill="x", padx=15, pady=5)
        lbl_route_sel = tk.Label(btn_action_frame, text="▼ 노선 선택",
                                  font=(FONT_SUB, SZ_S, "bold"),
                                  fg=T["fg_accent"], bg=T["bg_panel"])
        lbl_route_sel.pack(side="left")

        self.current_route_data = []
        self.sort_state = {"key": None, "reverse": False}
        type_order = {"간선": 0, "지선": 1, "광역": 2, "기타": 3, "경기": 4, "인천": 5}

        def update_route_tree_view():
            render_route_list()

        def set_all_check(status):
            for item in self.current_route_data:
                item["checked"].set(status)
            update_route_tree_view()

        _btn_s = self.get_btn_style("normal", font_size=SZ_XS)
        btn_all_off = tk.Button(btn_action_frame, text="전체 해제",
                                command=lambda: set_all_check(False), **_btn_s)
        btn_all_off.pack(side="right", padx=2)
        btn_all_on = tk.Button(btn_action_frame, text="전체 선택",
                               command=lambda: set_all_check(True), **_btn_s)
        btn_all_on.pack(side="right", padx=2)

        # 5-17-8. 노선 목록 표
        frame_route_list = tk.Frame(search_win, bd=1, relief="sunken",
                                    bg=T["bg_panel"])
        frame_route_list.pack(fill="both", expand=True, padx=15, pady=5)
        route_scroll = ttk.Scrollbar(frame_route_list, orient="vertical", style="App.Vertical.TScrollbar")
        route_cols = ("status", "rnm", "rtype", "path", "term", "f_tm", "l_tm", "st_cnt")
        tree_route = ttk.Treeview(frame_route_list, columns=route_cols,
                                  show="headings",
                                  yscrollcommand=route_scroll.set,
                                  style="Search.Treeview",
                                  selectmode="none")
        route_scroll.config(command=tree_route.yview)

        # 5-17-8-1. 선택/미선택 태그 색상 (테마 반응)
        if self.is_dark_mode:
            tree_route.tag_configure("selected",   background="#1a472a", foreground="#a8e6b5")
            tree_route.tag_configure("unselected", background=T["bg_tree"],  foreground=T["fg_tree"])
        else:
            tree_route.tag_configure("selected",   background="#d4edda", foreground="#155724")
            tree_route.tag_configure("unselected", background="white",   foreground="black")

        # 5-17-8-2. 열 제목
        tree_route.heading("status",  text="선택여부")
        tree_route.column("status",   width=42,  anchor="center", stretch=True)
        tree_route.heading("rnm",     text="노선번호", command=lambda: sort_data("rnm"))
        tree_route.column("rnm",      width=117, anchor="center", stretch=True)
        tree_route.heading("rtype",   text="유형",     command=lambda: sort_data("rtype"))
        tree_route.column("rtype",    width=50,  anchor="center", stretch=True)
        tree_route.heading("path",    text="기점↔종점")
        tree_route.column("path",     width=347, anchor="center", stretch=True)
        tree_route.heading("term",    text="배차간격")
        tree_route.column("term",     width=50,  anchor="center", stretch=True)
        tree_route.heading("f_tm",    text="첫차시각")
        tree_route.column("f_tm",     width=67,  anchor="center", stretch=True)
        tree_route.heading("l_tm",    text="막차시각")
        tree_route.column("l_tm",     width=67,  anchor="center", stretch=True)
        tree_route.heading("st_cnt",  text="정류장수")
        tree_route.column("st_cnt",   width=60,  anchor="center", stretch=True)
        tree_route.bind("<Button-1>", self.prevent_column_resize)
        tree_route.pack(side="left", fill="both", expand=True)
        route_scroll.pack(side="right", fill="y")

        # 5-17-9. 검색 실행 함수
        def perform_search(event=None):
            keyword = search_ent.get().strip()
            if not keyword: return
            for i in tree_st.get_children():
                tree_st.delete(i)
            root_res = None
            is_digit_search = keyword.isdigit() and len(keyword) >= 3
            if is_digit_search:
                root_res = self.fetch_api(
                    "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
                    {"arsId": keyword})
            if root_res is None or not root_res.findall(".//itemList"):
                root_res = self.fetch_api(
                    "http://ws.bus.go.kr/api/rest/stationinfo/getStationByName",
                    {"stSrch": keyword})
            insert_count = 0
            seen_stations = set()
            if root_res is not None:
                for item in root_res.findall(".//itemList"):
                    name   = item.findtext("stNm")
                    ars_id = item.findtext("arsId")
                    if ars_id and ars_id != "0":
                        key = (name, str(ars_id))
                        if key not in seen_stations:
                            seen_stations.add(key)
                            tree_st.insert("", "end", values=key)
                            insert_count += 1
            if insert_count == 0:
                tree_st.insert("", "end", values=("검색 결과가 없습니다.", ""))

        search_ent.bind("<Return>", perform_search)
        _btn_search = self.get_btn_style("normal", font_size=SZ_XS)
        tk.Button(frame_search, text="검색", command=perform_search,
                  **_btn_search).pack(side="left", padx=5)

        # 5-17-10. 정렬 함수
        def sort_data(key):
            if self.sort_state["key"] == key:
                self.sort_state["reverse"] = not self.sort_state["reverse"]
            else:
                self.sort_state["key"] = key
                self.sort_state["reverse"] = False
            if key == "rtype":
                self.current_route_data.sort(
                    key=lambda x: (type_order.get(x["rtype"], 99), x["rnm"]),
                    reverse=self.sort_state["reverse"])
            else:
                self.current_route_data.sort(
                    key=lambda x: x[key],
                    reverse=self.sort_state["reverse"])
            render_route_list()

        # 5-17-11. 노선 목록 렌더링
        def render_route_list():
            for item in tree_route.get_children():
                tree_route.delete(item)
            for item in self.current_route_data:
                status_text = "V" if item["checked"].get() else ""
                tag = "selected" if item["checked"].get() else "unselected"
                vals = [status_text] + item["data_vals"]
                tree_route.insert("", "end", iid=item["rid"], values=vals, tags=(tag,))

        # 5-17-12. 노선 클릭 체크토글
        def on_route_click(event):
            if tree_route.identify_region(event.x, event.y) == "separator":
                return "break"
            item_id = tree_route.identify_row(event.y)
            if not item_id: return
            for item in self.current_route_data:
                if item["rid"] == item_id:
                    new_state = not item["checked"].get()
                    item["checked"].set(new_state)
                    cur_values = list(tree_route.item(item_id, "values"))
                    cur_values[0] = "V" if new_state else ""
                    tree_route.item(item_id, values=cur_values,
                                    tags=("selected" if new_state else "unselected",))
                    break
        tree_route.bind("<Button-1>", on_route_click)

        # 5-17-13. 정류소 선택 시 노선 목록 불러오기
        def on_station_select(event):
            selected = tree_st.selection()
            if not selected: return
            ars_id = str(tree_st.item(selected[0])["values"][1]).zfill(5)
            if not ars_id or ars_id == "": return
            saved_checked_routes = set()
            if self.ars_ids[target_idx].get().strip() == ars_id and self.target_st_info[target_idx].get("routes"):
                saved_checked_routes = {r[0] for r in self.target_st_info[target_idx]["routes"]}
            self.current_route_data.clear()
            self._strt_cache.clear()
            self._strt_ord_cache.clear()
            root_route = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/stationinfo/getRouteByStation",
                {"arsId": ars_id})
            if root_route is not None:
                type_map = {"1":"공항","2":"마을","3":"간선","4":"지선","5":"순환",
                            "6":"광역","7":"인천","8":"경기","9":"폐지","0":"공용"}
                for it in root_route.findall(".//itemList"):
                    rid, rnm = it.findtext("busRouteId"), it.findtext("busRouteNm")
                    root_info = self.fetch_api(
                        "http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo",
                        {"busRouteId": rid})
                    rtype, path, term, f_tm, l_tm, st_cnt_str = "-","-","-","-","-","-"
                    st_cnt_val = 100
                    if root_info is not None:
                        info = root_info.find(".//itemList")
                        if info is not None:
                            rtype = type_map.get(info.findtext("routeType"), "기타")
                            path  = f"{info.findtext('stStationNm')}↔{info.findtext('edStationNm')}"
                            term  = info.findtext("term") + "분"
                            f_tm  = self.format_hhmm(info.findtext("firstBusTm", "-"))
                            l_tm  = self.format_hhmm(info.findtext("lastBusTm", "-"))
                            self.route_corp_map[rid] = info.findtext("corpNm", "정보없음")
                            root_st_list = self.fetch_api(
                                "http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute",
                                {"busRouteId": rid})
                            if root_st_list is not None:
                                st_list_items = root_st_list.findall(".//itemList")
                                st_cnt_val = len(st_list_items)
                                st_cnt_str = f"{st_cnt_val}개"
                                self._strt_cache[rid] = root_st_list
                                for _si in st_list_items:
                                    _item_ars = str(_si.findtext("arsId") or "").zfill(5)
                                    if _item_ars == ars_id:
                                        _ord_val = _si.findtext("seq") or _si.findtext("staOrd") or ""
                                        if _ord_val:
                                            self._strt_ord_cache[rid] = _ord_val
                                        break
                    self.rid_to_rnm[rid] = rnm
                    is_checked = (rid in saved_checked_routes) if saved_checked_routes else True
                    self.current_route_data.append({
                        "rid": rid, "rnm": rnm, "rtype": rtype, "st_cnt": st_cnt_val,
                        "data_vals": [rnm, rtype, path, term, f_tm, l_tm, st_cnt_str],
                        "checked": tk.BooleanVar(value=is_checked)
                    })
            self.sort_state = {"key": "rtype", "reverse": False}
            self.current_route_data.sort(
                key=lambda x: (type_order.get(x["rtype"], 99), x["rnm"]))
            render_route_list()

        tree_st.bind("<<TreeviewSelect>>", on_station_select)

        # 5-17-14. 이미 입력된 정류소 자동 표시
        if self.ars_ids[target_idx].get().strip():
            item_id = tree_st.insert("", "end",
                values=(self.target_st_info[target_idx].get("nm", ""),
                        self.ars_ids[target_idx].get()))
            tree_st.selection_set(item_id)

        # 5-17-15. 노선 확정 함수
        def confirm_selection():
            chosen = [(item["rid"], item["rnm"], item["st_cnt"])
                      for item in self.current_route_data if item["checked"].get()]
            selected_st = tree_st.selection()
            if not selected_st or not chosen: return
            st_name = tree_st.item(selected_st[0])["values"][0]
            ars_id  = str(tree_st.item(selected_st[0])["values"][1]).zfill(5)
            self.ars_ids[target_idx].set(ars_id)
            self.lbl_st_names[target_idx].config(text=f"[{st_name}] 실시간 현황")
            self.lbl_hist_titles[target_idx].config(text=f"[{st_name}] 도착 기록")
            root_st_uid = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
                {"arsId": ars_id})
            st_id = root_st_uid.findtext(".//stId") if (
                root_st_uid is not None and not isinstance(root_st_uid, tuple)) else ""
            if hasattr(self, "target_st_info") and self.target_st_info[target_idx].get("routes"):
                old_rids = {r[0] for r in self.target_st_info[target_idx]["routes"]}
                new_rids = {item[0] for item in chosen}
                gone = old_rids - new_rids
                for rid in gone:
                    self.pos_suspend_until.pop(rid, None)
                    old_st_id = self.target_st_info[target_idx].get("id", "")
                    self.veh_cache.pop((old_st_id, rid), None)
            ord_map = {}
            for rid, rnm, _ in chosen:
                if rid in self._strt_ord_cache:
                    ord_map[rid] = self._strt_ord_cache[rid]
                elif rid in self._strt_cache:
                    root_strt = self._strt_cache[rid]
                    for st_item in root_strt.findall(".//itemList"):
                        item_ars = str(st_item.findtext("arsId") or "").zfill(5)
                        if item_ars == ars_id:
                            ord_val = st_item.findtext("seq") or st_item.findtext("staOrd") or ""
                            if ord_val: ord_map[rid] = ord_val
                            break
                else:
                    self.log(f"⚠ SLST 캐시 누락: {rnm}번 — 재호출합니다.")
                    root_strt = self.fetch_api(
                        "http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute",
                        {"busRouteId": rid})
                    if root_strt is not None and not isinstance(root_strt, tuple):
                        for st_item in root_strt.findall(".//itemList"):
                            item_ars = str(st_item.findtext("arsId") or "").zfill(5)
                            if item_ars == ars_id:
                                ord_val = st_item.findtext("seq") or st_item.findtext("staOrd") or ""
                                if ord_val: ord_map[rid] = ord_val
                                break
            self._strt_cache.clear()
            self._strt_ord_cache.clear()
            self.target_st_info[target_idx] = {
                "id": st_id, "nm": st_name, "routes": chosen, "ord_map": ord_map}
            self.log(f"정류소 {target_idx+1} 설정 완료: {st_name} "
                     f"(총 {len(ord_map)}개 노선 중 {len(chosen)}개 선택)")
            search_win.destroy()
            self.update_button_states()

        # 5-17-16. 최종 확인 버튼
        tk.Button(search_win, text="선택한 노선들로 적용",
                  command=confirm_selection, height=2,
                  **self.get_btn_style("normal", font_size=SZ_M)
        ).pack(fill="x", padx=15, pady=10)


    # 5-18. [정보 가져오기] 고른 정류소의 버스 도착 시간과 위치를 실제로 확인하는 함수
    def process_station(self, idx):
        # 5-18-0. [날짜 전환 체크] 자정을 넘으면 POS 정지 테이블을 초기화합니다.
        today = datetime.now().date()
        if today != self._last_date:
            self._last_date = today
            self.pos_suspend_until.clear()
            # veh_cache는 (st_id, rid) 기반으로 날짜와 무관하게 유효하므로 자정에 초기화하지 않습니다.
            self.log("📅 날짜가 바뀌었습니다. POS1 일시정지 기록을 초기화합니다.")

            # 5-18-0-1. [자정 API 카운트 초기화] 00:00 날짜 전환 시 모든 API 호출 횟수를 0으로 리셋합니다.
            #   메인창 상단의 합산 카운트(api_stats)와
            #   API 상세 통계 창의 키별 카운트(api_stats_by_key)를 동시에 초기화합니다.
            #   idx == 0 일 때만 실행해 정류소 1·2 처리 중 두 번 초기화되는 것을 막습니다.
            if idx == 0:
                for k in self.api_stats:
                    self.api_stats[k] = 0
                for k in self.api_stats_by_key["main"]:
                    self.api_stats_by_key["main"][k] = 0
                for k in self.api_stats_by_key["back"]:
                    self.api_stats_by_key["back"][k] = 0
                self.root.after(0, self.update_api_counter_ui)
                self.log("📊 API 호출 횟수가 00:00 기준으로 초기화되었습니다.")

        # 5-18-1. 어떤 정류소를 확인해야 하는지 정보를 가져옵니다.
        info, ars_id = self.target_st_info[idx], self.ars_ids[idx].get().strip()
        st_id, rt_rows = info.get('id', ''), []

        # 5-18-2. [실시간 시간] 각 버스가 몇 분 뒤에 오는지 서버에 물어봅니다.
        # ── ARR1(주) 초과 대비: SINF(보조)는 정류소당 1회만 호출하는 캐시 ──────────
        uid_cache = None  # None=미호출 / {}=호출했으나 결과없음 / {rid:item}=정상수신
        ord_map = info.get('ord_map', {})  # {busRouteId: 정류소순번} — confirm_selection에서 미리 저장

        def get_uid_cache():
            """SINF API를 정류소당 최대 1회만 호출해 두 가지 키로 조회 가능한 딕셔너리를 반환합니다.
            기본 키: busRouteId (숫자형 ID)
            보조 키: rtNm 또는 busRouteAbrv (노선 번호 문자열) ← busRouteId가 없거나 0인 경우 대비
            반환값: {'by_rid': {busRouteId: item}, 'by_nm': {rtNm: item}}
            """
            nonlocal uid_cache
            if uid_cache is not None:
                return uid_cache
            root_uid = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
                {'arsId': ars_id}
            )
            if root_uid is None or isinstance(root_uid, tuple):
                uid_cache = {'by_rid': {}, 'by_nm': {}}
                return uid_cache
            by_rid, by_nm = {}, {}
            for item in root_uid.findall(".//itemList"):
                rid_key = item.findtext("busRouteId") or ""
                nm_key  = item.findtext("busRouteAbrv") or item.findtext("rtNm") or ""
                # rid 키 등록: 비어있거나 "0"이 아닌 경우만
                if rid_key and rid_key not in by_rid:
                    by_rid[rid_key] = item
                # 노선명 키 등록: 항상 보조 수단으로 저장
                if nm_key and nm_key not in by_nm:
                    by_nm[nm_key] = item
            uid_cache = {'by_rid': by_rid, 'by_nm': by_nm}
            self.log(f"ℹ ARR1 초과 → SINF 보조 호출 완료 (정류소 {idx+1}, {len(by_rid)}개 노선 수신)")
            return uid_cache

        # 노선마다 ARR1 우선 시도, 실패 시 SINF 캐시에서 해당 노선 항목만 추출
        for rid, rnm, _ in info.get('routes', []):
            row = None

            # 5-18-2-1. [주 API] getArrInfoByRoute — stId + busRouteId + ord 로 해당 정류소 1개 행만 조회
            ord_val = ord_map.get(rid, '0')
            arr_params = {'stId': st_id, 'busRouteId': rid, 'ord': ord_val}
            root_arr = self.fetch_api("", arr_params, api_type="ARR1")

            if root_arr is not None and not isinstance(root_arr, tuple):
                # 5-18-2-1-0. [노선 재출현] ARR1 응답에 실제 운행 버스가 있을 때만 POS 정지 해제
                #   판단 기준: arrmsg1 또는 arrmsg2 가 운행 없음 메시지 이외의 값(예: "3분 후")인 경우만 운행 중으로 판정합니다.
                #   ※ plainNo(차량번호)는 서울시 API가 운행 종료 후에도 지우지 않고 남겨두는
                #     경우가 있어 신뢰도가 낮습니다. arrmsg가 실시간 운행 상태를 더 정확히 반영하므로
                #     arrmsg 기준만으로 단일화하여 유령 차량번호로 인한 오판정을 방지합니다.
                NO_BUS_MSGS = {"운행정보없음", "운행종료", "출발대기", "-", ""}
                arr_items = root_arr.findall(".//itemList")
                has_active_bus = any(
                    (item.findtext("arrmsg1") or "").strip() not in NO_BUS_MSGS
                    or (item.findtext("arrmsg2") or "").strip() not in NO_BUS_MSGS
                    for item in arr_items
                )
                if rid in self.pos_suspend_until:
                    if has_active_bus:
                        del self.pos_suspend_until[rid]
                        self.log(f"🔄 {rnm}번 운행 재개 확인")
                    # else: 아직 운행 정보 없음 → 정지 유지 (불필요한 로그 없음)
                for item in root_arr.findall(".//itemList"):
                    res_ars  = str(item.findtext("arsId") or "").zfill(5)
                    res_ord  = str(item.findtext("staOrd") or "")

                    # ① 주 API(getArrInfoByRoute) 응답: 항상 해당 정류소 1행 → arsId 무조건 수용
                    # ② ARR2 폴백 응답: arsId 일치 행 우선, arsId=0인 경우 staOrd로 2차 검증
                    #    - arsId가 정확히 일치하면 OK
                    #    - arsId=0("00000")이고 ord_val이 일치하면 인천/경기 버스 해당 정류소 행 OK
                    #    - arsId=0이고 ord를 모를 때(ord_val='0')는 첫 번째 0행 수용(기존 동작 유지)
                    arsid_match = (res_ars == ars_id)
                    arsid_zero  = (not res_ars.strip("0"))   # "00000" 등 0으로만 구성
                    ord_match   = (ord_val != '0' and res_ord == ord_val)

                    if arsid_match or (arsid_zero and (ord_match or ord_val == '0')):
                        p1 = item.findtext("plainNo1") or "-"
                        p2 = item.findtext("plainNo2") or "-"
                        # 5-18-2-1-1. [차량번호 캐시] ARR1/ARR2에서 얻은 실제 번호판을 저장합니다.
                        cache_key = (st_id, rid)
                        if p1 not in ("-", "", None) or p2 not in ("-", "", None):
                            self.veh_cache[cache_key] = (p1, p2)
                        row = (
                            item.findtext("rtNm") or rnm,
                            p1,
                            item.findtext("arrmsg1") or "-",
                            p2,
                            item.findtext("arrmsg2") or "-"
                        )
                        break
            elif root_arr is None:
                # 5-18-2-2. [보조] ARR 전부 실패(횟수 초과) → SINF 캐시에서 해당 노선 항목 추출
                uid_data = get_uid_cache()
                # busRouteId(숫자형) 우선 조회, 없으면 노선명(문자열)으로 재시도 (DeprecationWarning 수정: 명시적 None 비교)
                uid_item = uid_data['by_rid'].get(rid)
                if uid_item is None:
                    uid_item = uid_data['by_nm'].get(rnm)

                if uid_item is not None:
                    route_display = (
                        uid_item.findtext("busRouteAbrv")
                        or uid_item.findtext("rtNm")
                        or rnm
                    )
                    # 5-18-2-2-1. [차량번호 캐시 우선 활용] ARR에서 받은 번호판이 있으면 vehId 대신 사용합니다.
                    cache_key = (st_id, rid)
                    cached = self.veh_cache.get(cache_key)
                    if cached:
                        p1_display, p2_display = cached
                    else:
                        # 캐시 없음: 9자리 vehId를 표시합니다 (개선 여지가 있습니다).
                        p1_display = uid_item.findtext("vehId1") or "-"
                        p2_display = uid_item.findtext("vehId2") or "-"
                    row = (
                        route_display,
                        p1_display,
                        uid_item.findtext("arrmsg1") or "-",
                        p2_display,
                        uid_item.findtext("arrmsg2") or "-"
                    )

            if row:
                rt_rows.append(row)

        # 5-18-2-3. 알아낸 시간 정보를 실시간 표에 보여줍니다.
        self.root.after(0, lambda d=rt_rows, t=self.trees_rt[idx]: self.refresh_tree(t, d))

        # 5-18-3. [스마트 도착 여부] 중복 노선은 API 호출을 아끼고 정보를 공유합니다.
        if st_id:
            # 5-18-3-1. 첫 번째 정류소(idx=0)를 시작할 때는 임시 저장소를 비웁니다.
            if idx == 0: self.temp_pos_data = {}

            for rid, rnm, st_cnt in info.get('routes', []):
                root_pos = None

                # 5-18-3-1-1. [POS 정지 확인] 첫차 시각 이전이면 호출 건너뜁니다.
                now_dt = datetime.now()
                if rid in self.pos_suspend_until:
                    resume_dt = self.pos_suspend_until[rid]
                    if now_dt < resume_dt:
                        # 아직 재개 시각 전 → POS 호출 생략
                        continue
                    else:
                        # 재개 시각 도달 → 정지 해제 (최초 1회만 로그 출력)
                        del self.pos_suspend_until[rid]
                        if rid not in self.pos_resume_logged:
                            self.log(f"⏰ {rnm}번 POS 호출 재개 (첫차 5분 전 도달)")
                            self.pos_resume_logged.add(rid)

                # 5-18-3-2. [스마트 판단] 이미 다른 정류소에서 이 노선의 위치 정보를 가져왔는지 확인합니다.
                if rid in self.temp_pos_data:
                    # 5-18-3-2-1. 이미 있다면 서버에 묻지 않고 저장된 정보를 꺼내 씁니다.
                    root_pos = self.temp_pos_data[rid]
                else:
                    # 5-18-3-2-2. 저장된 정보가 없다면 서버에 직접 좌표를 물어봅니다.
                    pos_result = self.fetch_api("", {'busRouteId': rid, 'startOrd': '1', 'endOrd': str(st_cnt)}, api_type="POS2")

                    if isinstance(pos_result, tuple) and pos_result[0] == 'NO_BUS':
                        # 5-18-3-2-2-1. [운행 없음] 첫차 시각을 조회하여 5분 전 재개 시각을 등록합니다.
                        _, f_tm_str = pos_result
                        if f_tm_str:
                            try:
                                f_hm = datetime.strptime(f_tm_str, "%H:%M")
                                # 내일 날짜로 첫차 시각 조합 (첫차가 다음 날인 경우 대비)
                                base_dt = now_dt.replace(hour=f_hm.hour, minute=f_hm.minute, second=0, microsecond=0)
                                if base_dt <= now_dt:
                                    base_dt += timedelta(days=1)
                                resume_dt = base_dt - timedelta(minutes=5)
                                if resume_dt > now_dt:
                                    # 아직 재개 시각이 오지 않음 → 재개 시각까지 정지
                                    self.pos_suspend_until[rid] = resume_dt
                                    self.pos_resume_logged.discard(rid)
                                    self.log(f"💤 {rnm}번 차량 없음 → 첫차 {f_tm_str} 5분 전({resume_dt.strftime('%H:%M')})까지 POS 정지")
                                else:
                                    # 이미 5분 전 창 안에 있지만 아직 차량 없음 → 1분 후 재시도 (로그 없음)
                                    self.pos_suspend_until[rid] = now_dt + timedelta(minutes=1)
                                    # pos_resume_logged는 유지 (중복 로그 방지)
                            except Exception:
                                # 파싱 실패시 30분 후 재시도
                                self.pos_suspend_until[rid] = now_dt + timedelta(minutes=30)
                                self.pos_resume_logged.discard(rid)
                                self.log(f"💤 {rnm}번 차량 없음 → 첫차 시각 파싱 실패, 30분 후 재시도")
                        else:
                            # 첫차 시각 없음: 30분 후 재시도
                            self.pos_suspend_until[rid] = now_dt + timedelta(minutes=30)
                            self.pos_resume_logged.discard(rid)
                            self.log(f"💤 {rnm}번 차량 없음 → 첫차 정보 없음, 30분 후 재시도")
                        continue  # POS 처리 건너뜀

                    root_pos = pos_result
                    # 5-18-3-2-3. 가져온 정보는 다음 정류소를 위해 임시 저장소에 보관합니다.
                    if root_pos is not None:
                        self.temp_pos_data[rid] = root_pos

                # 5-18-3-3. 가져온(또는 공유받은) 위치 정보로 도착 여부를 판정합니다.
                if root_pos is not None:
                    for bus in root_pos.findall(".//itemList"):
                        # 5-18-3-3-1. 버스의 현재 위치가 지금 처리 중인 정류소(st_id)인지 확인합니다.
                        if bus.findtext("lastStnId") == st_id:
                            veh_no = bus.findtext("plainNo")
                            # 5-18-3-3-2. 방금 기록한 버스는 5분 동안 다시 기록하지 않습니다.
                            if veh_no in self.last_arrival_logs[idx] and (time.time() - self.last_arrival_logs[idx][veh_no] < 300): continue

                            f_time = self.format_datetm(bus.findtext("dataTm"))
                            # 5-18-3-3-3. 버스 회사 이름을 찾아보고 공동배차 정보를 확인합니다.
                            corp_nm = self.route_corp_map.get(rid, "정보없음")
                            if rnm in self.excel_multi_corp_map and len(self.excel_multi_corp_map[rnm]) >= 2:
                                corp_nm = ", ".join(sorted(list(self.excel_multi_corp_map[rnm])))

                            # 5-18-3-3-4. 도착 기록판에 한 줄 적고 엑셀에도 자동으로 저장합니다.
                            log_entry = (f_time, rnm, veh_no, corp_nm, "정류소 도착")
                            self.recorded_data.append((f_time, info['nm'], rnm, veh_no, corp_nm))
                            self.last_arrival_logs[idx][veh_no] = time.time()
                            self.perform_auto_save()
                            self.log(f"★ [{info['nm']}] 도착: {rnm} ({veh_no})")
                            # 5-18-3-3-4-1. "end" 로 변경하여 오래된 기록이 위, 새 기록이 아래에 누적됩니다 (오름차순)
                            # 5-18-3-3-4-2. 삽입 후 see()로 스크롤을 새 항목 위치로 자동 이동합니다.
                            self.root.after(0, lambda r=log_entry, t=self.trees_hist[idx]: (
                                t.insert("", "end", values=r),
                                t.see(t.get_children()[-1])
                            ))
       
    # 5-19. [공동배차확인용 엑셀 파일 부르기] 버스 회사 정보를 담은 엑셀을 뒷단에서 몰래 받아오기 시작하는 함수
    def start_excel_download_thread(self):
        # 5-19-1. 프로그램이 멈추지 않게 스레드(별도의 일꾼)를 써서 다운로드합니다.
        threading.Thread(target=self.download_and_load_excel, daemon=True).start()

    # 5-20. [공동배차확인용 엑셀 파일 다운로드] 서울 데이터 광장에서 최신 버스 정보 엑셀을 직접 받아오는 함수
    def download_and_load_excel(self):
        target_url = "https://data.seoul.go.kr/dataList/OA-15066/F/1/datasetView.do" 
        self.log(f"📥 최신 노선정보 확인 중 (Text Search): {target_url}")
        
        download_success = False
        downloaded_filename = None
        driver = None

        try:
            # 5-20-1. 인터넷 로봇(크롬 드라이버) 설정을 합니다.
            options = Options()
            options.add_argument("--headless") # 5-20-1-1. 화면에는 안 보이게 숨깁니다.
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--window-size=1920,1080")
            
            # 5-20-1-2. 로봇이 아니라 진짜 사람인 것처럼 꾸밉니다.
            options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            options.add_argument("--disable-blink-features=AutomationControlled") 
            options.add_experimental_option("excludeSwitches", ["enable-automation"]) 
            options.add_experimental_option('useAutomationExtension', False)
            
            # 5-20-1-3. 파일을 저장할 우리 컴퓨터의 폴더를 정합니다.
            target_download_dir = os.path.abspath(self.current_dir)
            prefs = {
                "download.default_directory": target_download_dir,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True,
                "profile.default_content_settings.popups": 0
            }
            options.add_experimental_option("prefs", prefs)

            # 5-20-2. 크롬 로봇을 실행해서 사이트로 보냅니다.
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            
            if not os.path.exists(target_download_dir):
                os.makedirs(target_download_dir, exist_ok=True)

            driver.get(target_url) 
            
            # 5-20-3. 사이트가 다 열릴 때까지 최대 20초 동안 기다려줍니다.
            wait = WebDriverWait(driver, 20)
            self.log("   ㄴ페이지 분석 중 (파일명 탐색)...")
            
            # 5-20-4. '.xlsx'라고 써진 엑셀 파일 다운로드 단추를 찾습니다.
            download_target = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//*[contains(text(), '.xlsx')]")
            ))
            
            found_text = download_target.text.strip()
            self.log(f"   ㄴ다운로드 대상 발견: {found_text}")
            
            # 5-20-5. 이미 같은 파일이 있다면 지우고 새로 받기로 합니다.
            file_path = os.path.join(self.current_dir, found_text)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    self.log(f"   ℹ 기존 파일 삭제 후 다운로드: {found_text}")
                except Exception as del_err:
                    self.log(f"   ⚠ 기존 파일 삭제 실패 (사용중일 수 있음): {del_err}")

            before_files = set(os.listdir(self.current_dir)) 
            
            # 5-20-6. 마우스로 다운로드 버튼을 클릭합니다.
            driver.execute_script("arguments[0].scrollIntoView(true);", download_target)
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", download_target)
            
            # 5-20-7. 파일이 다 받아질 때까지 1초마다 폴더를 확인합니다.
            found_file = None
            for _ in range(30):
                time.sleep(1)
                after_files = set(os.listdir(self.current_dir))
                new_files = after_files - before_files 
                
                for f in new_files:
                    if f.endswith(".xlsx") and "서울시버스노선기본정보" in f:
                        found_file = f
                        break
                if found_file:
                    break
            
            # 5-20-8. 성공하면 알려주고 실패하면 이유를 적습니다.
            if found_file:
                self.log(f"✅ 다운로드 성공: {found_file}")
                download_success = True
                downloaded_filename = found_file
            else:
                self.log("⚠ 다운로드 시도했으나 파일 생성 안됨 (시간초과)")

        except Exception as e:
            self.log(f"❌ 다운로드 프로세스 오류: {e}")
            self.log("   (Chrome 브라우저와 인터넷 연결을 확인해주세요)")
        finally:
            if driver:
                driver.quit() # 5-20-9. 로봇을 퇴근시킵니다(인터넷 창 닫기).
        
        # 5-20-10. 새로 받은 파일을 열고, 없으면 옛날 파일이라도 찾아봅니다.
        if download_success and downloaded_filename:
            self.load_excel_routes(specific_filename=downloaded_filename)
        else:
            self.log("ℹ 기존 파일 중 최신 파일을 사용합니다.")
            self.load_excel_routes()

    # 5-21. [공동배차확인용 엑셀 읽기] 받아온 엑셀 파일을 열어서 버스 노선과 회사 이름을 정리하는 함수
    def load_excel_routes(self, specific_filename=None):
        target_file = specific_filename
        
        # 5-21-1. 어떤 파일인지 모르면 이름이 비슷한 것 중에 가장 최근 것을 고릅니다.
        if not target_file:
            candidates = [f for f in os.listdir(self.current_dir) if "서울시버스노선기본정보" in f and f.endswith(".xlsx")]
            if candidates:
                candidates.sort(reverse=True) 
                target_file = candidates[0]
            else:
                target_file = "서울시버스노선기본정보(20260108).xlsx" 

        path = os.path.join(self.current_dir, target_file)
        
        # 5-21-2. 파일이 진짜 있는지 확인합니다.
        if not os.path.exists(path):
            self.log(f"⚠ 로드할 엑셀 파일이 없습니다: {target_file}")
            return

        try:
            # 5-21-3. 엑셀을 표 모양으로 읽어서 '노선번호'와 '업체명'을 기억합니다.
            df = pd.read_excel(path)
            r_col = [c for c in df.columns if '노선번호' in str(c)]
            c_col = [c for c in df.columns if '업체명' in str(c)]
            
            if r_col and c_col:
                r_col_name = r_col[0]
                c_col_name = c_col[0]
                
                temp_map = {}
                # 5-21-3-1. 표를 한 줄씩 읽으면서 버스 번호에 회사를 짝지어줍니다.
                for _, row in df.iterrows():
                    r_name = str(row[r_col_name]).strip()
                    c_name = str(row[c_col_name]).strip()
                    if r_name not in temp_map:
                        temp_map[r_name] = set()
                    temp_map[r_name].add(c_name) 
                
                self.excel_multi_corp_map = temp_map 
                self.log(f"✅ 엑셀 데이터 로드 완료: {target_file}")
            else:
                self.log(f"⚠ 엑셀 형식 오류: '노선번호' 또는 '업체명' 열 미발견")
        except Exception as e:
            self.log(f"❌ 엑셀 로드 중 오류: {e}")

    # 6그룹 : 실시간 모니터링 로직 (Monitoring)        

    # 5-22. [기록 엔진 켜기] 감시와 기록을 실제로 시작하는 함수
    def start_monitoring(self):
        # 5-22-1. 몇 초마다 확인할지 안 적었다면 시작할 수 없습니다.
        if not self.refresh_interval_var.get().strip():
            messagebox.showwarning("입력 누락", "갱신주기를 입력해야 자동 기록을 시작할 수 있습니다.")
            return

        if not self.service_key_var.get().strip(): return 
        
        # 5-22-2. 기록 중에는 주기 입력창을 못 건드리게 막습니다.
        self.entry_refresh_interval.config(state='readonly')
        
        self.save_key_to_file()
        self.is_monitoring = True
        # 자동 기록 중에는 정류소 검색 버튼을 비활성화합니다.
        _s_off2 = self.get_btn_style("disabled")
        _s_off2.pop("state", None)
        for _b in self.btn_searches:
            _b.config(state="disabled", **_s_off2)
        
        # 5-22-3. 오늘 기록을 저장할 새 엑셀 파일을 하나 만듭니다.
        filename = f"Bus_Arrival_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.auto_save_path = os.path.join(self.current_dir, filename)
        
        try:
            pd.DataFrame(columns=["데이터시각", "도착정류소명", "노선", "차량번호", "운수사명"]).to_excel(self.auto_save_path, index=False)
            self.can_auto_save = True 
            display_name = os.path.basename(self.auto_save_path)
            self.lbl_auto_save_status.config(text=f"[{display_name}] 파일에 자동 기록 중 ......", fg="#27ae60")
        except Exception as e:
            self.can_auto_save = False
            messagebox.showwarning("오류", f"자동 저장 설정 중 오류 발생: {e}")

        # 5-22-4. 일꾼(스레드)을 불러서 무한 루프(main_loop)를 돌립니다.
        self.update_button_states()
        threading.Thread(target=self.main_loop, daemon=True).start()

    # 5-23. [무한 반복] 정해진 시간마다 계속해서 정보를 새로고침하는 함수
    def main_loop(self):
        next_call = time.time()
        # 5-23-1. 우리가 멈추라고 할 때까지 계속 반복합니다.
        while self.is_monitoring:
            self.refresh_data() # 5-23-1-1. 서버에 물어보고 데이터를 갱신합니다.
            try: interval = int(self.refresh_interval_var.get())
            except: interval = 60
            next_call += interval
            
            # 5-23-2. 다음 갱신 때까지 기다립니다. (중간에 멈춤을 누를 수 있게 0.1초씩 나눠서 쉽니다.)
            sleep_time = next_call - time.time()
            if sleep_time < 0: next_call = time.time(); sleep_time = 0
            for _ in range(int(sleep_time * 10)):
                if not self.is_monitoring: break
                time.sleep(0.1)

    # 5-24. [지금 바로] 기다리지 않고 당장 정보를 새로 가져오는 함수
    def manual_refresh(self):
        # macOS guard: Aqua 테마의 state='disabled' 클릭 통과 버그 차단
        if CURRENT_OS == "Darwin" and not self._btn_active.get('manual', True):
            return
        # 갱신 중 중복 클릭 방지: 플래그와 state 동시 비활성화
        self._btn_active['manual'] = False
        _s_dis = self.get_btn_style("disabled")
        _s_dis.pop("state", None)
        self.btn_manual.config(state="disabled", **_s_dis)
        self.log("🔄 [수동] 데이터 갱신 시작...")

        # 5-24-1. 별도의 일꾼을 써서 지금 바로 정보를 갱신합니다.
        def run_manual():
            self.refresh_data(manual=True)
            self.log("✅ [수동] 데이터 갱신 완료")
            def _re_enable():
                self._btn_active['manual'] = True
                _s_on = self.get_btn_style("normal")
                _s_on.pop("state", None)
                self.btn_manual.config(state="normal", **_s_on)
            self.root.after(0, _re_enable)

        threading.Thread(target=run_manual, daemon=True).start()

    # 5-25. [모두 새로고침] 모든 정류소의 정보를 한꺼번에 갱신하는 공용 함수
    def refresh_data(self, manual=False):
        # 5-25-0. [노선 변동 체크] ARR1 응답에서 기존에 없던 노선이 나타나거나 사라진 경우를 감지합니다.
        #   pos_suspend_until 에 등록된 노선이 ARR1 응답에서 새로 보이면 정지를 해제합니다.
        #   (실제 체크는 process_station 내부에서 수행하며, 여기서는 플래그만 기록합니다.)

        # 5-25-0-1. [스레드 안전] _refresh_lock 으로 자동/수동 갱신이 동시 진입하지 못하게 막습니다.
        #   수동 갱신이 자동 갱신과 겹치면 last_arrival_logs / veh_cache 등의 공유 상태를
        #   두 스레드가 동시에 읽고 쓸 수 있어 중복 기록 또는 누락이 발생할 수 있습니다.
        if not self._refresh_lock.acquire(blocking=False):
            # 이미 다른 스레드가 갱신 중 → 이번 호출은 건너뜁니다.
            if manual:
                self.log("ℹ 자동 갱신 진행 중이므로 수동 갱신을 건너뜁니다.")
            return

        try:
            # 5-25-1. 정류소 1번과 2번을 차례대로 확인합니다.
            for i in range(2):
                if self.target_st_info[i].get('routes'): 
                    self.process_station(i) 
            
            # 5-25-2. [이중 저장 보호 — 구제 저장] 사이클이 끝난 뒤에도 미저장 기록이 있으면 재시도합니다.
            #   process_station() 내부에서 perform_auto_save()를 즉시 호출하지만,
            #   PermissionError나 OS 쓰기 오류로 저장이 실패했다면 recorded_data의 크기가
            #   _saved_record_count보다 크게 됩니다. 이를 감지해 구제 저장을 실행합니다.
            #   최악의 경우에도 누락 데이터는 다음 사이클 종료까지만 지연됩니다.
            if (self.recorded_data
                    and self.auto_save_path
                    and self.can_auto_save
                    and len(self.recorded_data) > self._saved_record_count):
                missed = len(self.recorded_data) - self._saved_record_count
                self.log(f"⚠ 미저장 기록 {missed}건 감지 → 구제 저장 재시도 중...")
                self.perform_auto_save()
                if len(self.recorded_data) == self._saved_record_count:
                    self.log(f"✅ 구제 저장 성공: {missed}건 복구 완료")
                else:
                    self.log(f"❌ 구제 저장도 실패 — 다음 사이클에 재시도됩니다.")

            # 5-25-3. 자동으로 갱신될 때만 로그를 남깁니다. (너무 많이 남으면 지저분하니까요.)
            if not manual:
                self.log("데이터 갱신 완료")
        finally:
            self._refresh_lock.release()

    # 5-26. [기록 멈춤] 하던 일을 멈추고 쉬는 단계로 돌아가는 함수
    def stop_monitoring(self):
        # 5-26-0-0. [중지 확인] 실수 클릭 방지를 위해 한 번 더 확인합니다.
        if not messagebox.askyesno("중지 확인", "정말 중지하시겠습니까?"):
            return
        self.is_monitoring = False
        # 자동 기록 중지 후 키가 잠긴 상태면 검색 버튼을 다시 활성화합니다.
        if self.key_locked:
            _s_on2 = self.get_btn_style("normal")
            _s_on2.pop("state", None)
            for _b in self.btn_searches:
                _b.config(state="normal", **_s_on2)
        self.log("🛑 자동 기록을 중지합니다.")

        # 5-26-0. [최종 구제 저장] 모니터링 중지 시점에도 미저장 기록이 있으면 마지막으로 저장합니다.
        #   마지막 사이클의 즉시 저장이 실패한 채로 중지하면 데이터가 영구 손실되므로
        #   중지 직전에 한 번 더 강제 저장을 시도합니다.
        if (self.recorded_data
                and self.auto_save_path
                and self.can_auto_save
                and len(self.recorded_data) > self._saved_record_count):
            missed = len(self.recorded_data) - self._saved_record_count
            self.log(f"⚠ 중지 시점 미저장 기록 {missed}건 → 최종 구제 저장 시도 중...")
            self.perform_auto_save()
            if len(self.recorded_data) == self._saved_record_count:
                self.log(f"✅ 최종 구제 저장 성공: {missed}건 복구 완료")
            else:
                self.log(f"❌ 최종 구제 저장 실패 — '다른 이름으로 엑셀 저장' 버튼으로 수동 저장해 주세요.")
        
        # 5-26-1. 이제 주기를 다시 고칠 수 있게 입력창을 엽니다.
        self.entry_refresh_interval.config(state='normal')
        self.lbl_auto_save_status.config(text=" ※ 자동 기록 시작 버튼을 작동시키면 도착 기록이 엑셀파일로 자동 저장됩니다.", fg="#e74c3c")
        self.update_button_states()

    # 7그룹 : 데이터 분석 및 기록, 출력 (Analysis, Log, Export)

    # 5-27. [기록 지우기] 표에 쌓인 도착 기록들을 싹 청소하는 함수
    def clear_history(self, idx):
        # 5-27-1. 혹시 실수로 누른 걸 수도 있으니 한 번 더 물어봅니다.
        if messagebox.askyesno("삭제", f"정류소 {idx+1}의 모든 기록을 삭제하시겠습니까?\n기록을 삭제하더라도, 이미 저장된 기록은 삭제되지 않습니다."): 
            tree = self.trees_hist[idx]
            # 5-27-2. 표의 모든 줄을 하나씩 지웁니다.
            for item in tree.get_children():
                tree.delete(item)
            self.log(f"정류소 {idx+1} 기록을 초기화했습니다.")

    # 5-28. [시간 예쁘게] 숫자만 있는 시간을 사람이 읽기 좋게(12:34) 바꿔주는 함수
    def format_hhmm(self, raw_str):
        if not raw_str or len(raw_str) < 4: return raw_str
        # 5-28-1. 긴 시간 데이터면 중간에 시간만 쏙 뽑아냅니다.
        if len(raw_str) >= 14: return f"{raw_str[8:10]}:{raw_str[10:12]}"
        return f"{raw_str[:2]}:{raw_str[2:4]}" 

    # 5-29. [일지 적기] 화면 하단 로그창에 메시지를 남기는 함수 (최대 5,000줄까지만 기억해요)
    def log(self, msg):
        # 5-29-0. [맥OS 스레드 안전] 백그라운드 스레드에서 호출될 수 있으므로
        #         tk.Text 조작은 반드시 메인 스레드(root.after)를 통해 실행합니다.
        #         txt_log가 아직 생성되지 않은 경우(초기화 중)는 콘솔 출력으로 대체합니다.
        now = datetime.now().strftime("%H:%M:%S")
        try:
            key = self.service_key_var.get().strip()
            backup_key = self.backup_key_var.get().strip()
        except Exception:
            key, backup_key = "", ""

        # 5-29-1. 비밀번호가 로그에 보이면 안 되니 별표(****)로 가려줍니다.
        if key and len(key) > 4: msg = msg.replace(key, "********")
        if backup_key and len(backup_key) > 4: msg = msg.replace(backup_key, "********")
        
        line = f"[{now}] {msg}\n"

        # 5-29-2. txt_log가 준비됐는지 확인하고 메인 스레드로 UI를 업데이트합니다.
        if not hasattr(self, 'txt_log'):
            # 5-29-2-1. UI 초기화 전에는 콘솔에 출력합니다(맥 터미널에서도 확인 가능).
            print(line, end="")
            return

        def _do_insert():
            try:
                # 5-29-2-2. 읽기전용(disabled) 위젯에 쓰려면 잠시 NORMAL로 열었다가 다시 닫습니다.
                self.txt_log.config(state="normal")
                self.txt_log.insert(tk.END, line)
                self.txt_log.see(tk.END)
                # 5-29-3. 로그 5,000줄이 넘어가면 가장 오래된(제일 위) 줄부터 지웁니다.
                line_count = int(self.txt_log.index('end-1c').split('.')[0])
                if line_count > 5000:
                    self.txt_log.delete('1.0', '2.0')
                self.txt_log.config(state="disabled")
            except Exception:
                pass  # 5-29-3-1. 창이 닫히는 도중 예외가 발생해도 무시합니다.

        # 5-29-4. 메인 스레드에서 실행되도록 예약합니다.
        try:
            self.root.after(0, _do_insert)
        except Exception:
            pass

    # 5-30. [날짜 예쁘게] 숫자 덩어리 날짜를 달력처럼(2026-01-01 12:00:00) 바꾸는 함수
    def format_datetm(self, raw_tm):
        try:
            if raw_tm and len(raw_tm) >= 14: 
                return datetime.strptime(raw_tm, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
            return raw_tm
        except: return raw_tm

    # 5-31. [엑셀 만들기] 기록한 데이터들을 멋진 엑셀 파일로 완성하는 핵심 함수
    def _core_excel_save_logic(self, target_path, save_completed=False):
        # 5-31-0. save_completed=True 일 때:
        #   영업일(BizDate)이 처음 '완결'되는 순간(새 날짜 시트가 열릴 때)
        #   해당 날짜의 데이터만 Bus_Arrival_Log_YYYYMMDD_completed.xlsx 로 별도 저장합니다.
        try:
            import pandas as pd
            from datetime import datetime, timedelta
            from openpyxl.styles import Font, PatternFill, Alignment

            # 5-31-1. 우리가 모은 기록들을 엑셀 표 모양으로 정리합니다.
            cols = ["데이터시각", "도착정류소명", "노선", "차량번호", "운수사명"]
            df = pd.DataFrame(self.recorded_data, columns=cols)
            
            # 5-31-2. [영업일 계산] 새벽 3시 전에 들어온 버스는 '어제' 버스로 쳐줍니다.
            def get_biz_date(dt_str):
                try:
                    dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
                    if dt.hour < 3: dt -= timedelta(days=1) 
                    return dt.strftime("%Y-%m-%d")
                except: return "Unknown"
            
            df['BizDate'] = df['데이터시각'].apply(get_biz_date)

            # 5-31-2-1. 현재 존재하는 영업일 목록을 구합니다.
            biz_dates_in_data = set(df['BizDate'].unique())
            # 5-31-2-2. 오늘(현재 진행 중인) 영업일을 계산합니다.
            now = datetime.now()
            current_biz = (now - timedelta(days=1)).strftime("%Y-%m-%d") if now.hour < 3 else now.strftime("%Y-%m-%d")
            # 5-31-2-3. 완결된 날짜: 데이터에 있으나 현재 영업일이 아닌 과거 날짜입니다.
            completed_dates = biz_dates_in_data - {current_biz, "Unknown"}
            
            # 5-31-3. 엑셀을 저장하면서 날짜별로 시트(종이)를 나누어 담습니다.
            with pd.ExcelWriter(target_path, engine='openpyxl') as writer:
                for biz_date, group in df.groupby('BizDate'):
                    sheet_data = group.drop(columns=['BizDate'])
                    sheet_data.to_excel(writer, sheet_name=biz_date, index=False)
                    
                    worksheet = writer.sheets[biz_date]
                    
                    # 5-31-3-1. 첫 번째 줄(제목줄)을 하늘색으로 칠하고 글씨를 굵게 합니다.
                    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    header_font = Font(bold=True, size=11)
                    header_alignment = Alignment(horizontal='center', vertical='center')

                    # 5-31-3-2. 각 칸의 너비를 글자 길이에 맞춰서 보기 좋게 조절합니다.
                    for idx, col in enumerate(sheet_data.columns):
                        cell = worksheet.cell(row=1, column=idx+1)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment

                        max_data_len = sheet_data[col].astype(str).map(len).max()
                        base_width = max(max_data_len, len(str(col)))
                        
                        # 5-31-3-2-1. 정류소 이름처럼 긴 글자는 더 넓게, 시각은 좁게 개별 조절합니다.
                        if col == "데이터시각":
                            adjusted_width = base_width + 1 
                        elif col == "도착정류소명":
                            # 기존 너비(base_width + 15)의 1.5배로 확장합니다.
                            adjusted_width = (base_width + 15) * 1.5
                        elif col == "노선":
                            adjusted_width = base_width + 8 
                        elif col == "운수사명":
                            adjusted_width = base_width + 7 
                        else:
                            adjusted_width = base_width + 5 
                        
                        col_letter = worksheet.cell(row=1, column=idx+1).column_letter
                        worksheet.column_dimensions[col_letter].width = adjusted_width

            # 5-31-4. [완결 파일 자동 저장]
            #   save_completed=True 이면서, 아직 저장하지 않은 완결 영업일이 있을 때만 저장합니다.
            if save_completed:
                new_completed = completed_dates - self._completed_dates_saved
                for biz_date in sorted(new_completed):
                    # 5-31-4-1. 해당 영업일 데이터만 추출합니다.
                    day_df = df[df['BizDate'] == biz_date].drop(columns=['BizDate'])
                    if day_df.empty:
                        continue
                    # 5-31-4-2. 파일명: Bus_Arrival_Log_YYYYMMDD_completed.xlsx
                    safe_date = biz_date.replace("-", "")
                    completed_name = f"Bus_Arrival_Log_{safe_date}_completed.xlsx"
                    completed_path = os.path.join(self.current_dir, completed_name)
                    try:
                        with pd.ExcelWriter(completed_path, engine='openpyxl') as cw:
                            day_df.to_excel(cw, sheet_name=biz_date, index=False)
                            ws = cw.sheets[biz_date]
                            hfill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                            hfont = Font(bold=True, size=11)
                            halign = Alignment(horizontal='center', vertical='center')
                            for ci, col in enumerate(day_df.columns):
                                cell = ws.cell(row=1, column=ci+1)
                                cell.fill = hfill; cell.font = hfont; cell.alignment = halign
                                max_len = day_df[col].astype(str).map(len).max()
                                base_w  = max(max_len, len(str(col)))
                                if col == "도착정류소명":
                                    col_w = (base_w + 15) * 1.5
                                elif col == "데이터시각":
                                    col_w = base_w + 1
                                elif col == "노선":
                                    col_w = base_w + 8
                                elif col == "운수사명":
                                    col_w = base_w + 7
                                else:
                                    col_w = base_w + 5
                                col_letter = ws.cell(row=1, column=ci+1).column_letter
                                ws.column_dimensions[col_letter].width = col_w
                        # 5-31-4-3. 저장 성공 시 중복 저장 방지 집합에 추가합니다.
                        self._completed_dates_saved.add(biz_date)
                        self.log(f"📁 영업일 완결 파일 저장: {completed_name}")
                    except Exception as ce:
                        self.log(f"⚠ 완결 파일 저장 실패 ({biz_date}): {ce}")

            # 5-31-5. [저장 성공 확인] 저장 성공 시 _saved_record_count를 갱신합니다.
            #   이 값은 refresh_data() 사이클 종료 시 누락 여부 판단에 사용됩니다.
            self._saved_record_count = len(self.recorded_data)
            return True

        except PermissionError:
            self.log("⚠ 엑셀 파일이 열려 있어 저장을 건너뜁니다. (파일을 닫아주세요)")
            return False
        except Exception as e:
            self.log(f"❌ 엑셀 저장 중 오류 발생: {e}")
            return False
        
    # 5-32. [수동 저장] 우리가 직접 버튼을 눌러서 원하는 이름으로 엑셀을 저장하는 함수
    def save_to_excel(self):
        # 5-32-1. 기록이 하나도 없으면 저장할 수 없습니다.
        if not self.recorded_data:
            messagebox.showwarning("알림", "저장할 데이터가 없습니다.")
            return 
        
        # 5-32-2. 어디에 어떤 이름으로 저장할지 물어보는 창을 띄웁니다.
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel 파일", "*.xlsx")],
            title="다른 이름으로 저장"
        ) 
        
        if path: 
            # 5-32-3. 진짜로 파일을 만듭니다.
            if self._core_excel_save_logic(path):
                messagebox.showinfo("저장 완료", f"영업일 기준으로 분류되어 저장되었습니다.\n경로: {path}")
                self.log(f"💾 수동 엑셀 저장 완료: {path}")

    # 5-33. [자동 저장] 프로그램이 정해진 장소에 알아서 척척 기록을 백업하는 함수
    def perform_auto_save(self):
        # 5-33-1. 저장할 게 있는지, 자동 저장이 켜져 있는지 확인합니다.
        if not self.recorded_data or not self.auto_save_path or not self.can_auto_save: 
            return 
        
        # 5-33-2. [스레드 안전] _save_lock 으로 한 스레드씩만 파일을 씁니다.
        #   자동 모니터링 스레드와 수동 갱신 스레드가 동시에 파일을 덮어써
        #   더 이른 스냅샷이 나중에 기록되면 데이터가 사라지는 버그를 방지합니다.
        #   lock 취득 전에 재진입 여부를 확인해 불필요한 중복 저장을 줄입니다.
        if not self._save_lock.acquire(blocking=False):
            # 다른 스레드가 이미 저장 중 → 대기하지 않고 건너뜁니다.
            # 해당 스레드가 완료될 때 recorded_data 의 전체 스냅샷을 포함하므로 데이터 손실은 없습니다.
            return
        try:
            # 5-33-3. 소리 없이 조용히 파일에 기록합니다.
            #         save_completed=True 로 전달해 완결된 영업일 파일도 함께 저장합니다.
            self._core_excel_save_logic(self.auto_save_path, save_completed=True)
        finally:
            self._save_lock.release()

    # 5-34. [표 새로고침] 표의 내용을 다 지우고 새로운 버스 정보로 채우는 함수
    def refresh_tree(self, tree, data):
        # 5-34-1. 옛날 정보는 싹 지웁니다.
        for i in tree.get_children(): tree.delete(i)
        # 5-34-2. 새로 가져온 정보를 한 줄씩 넣습니다.
        for row in data: tree.insert("", "end", values=row)

# 6. [진짜 실행] 여기가 프로그램의 시작 버튼입니다!
if __name__ == "__main__":
    import traceback

    # 6-0. [오류 기록기] 프로그램이 예기치 않게 종료될 때 오류 내용을 errorlog.txt 에 저장합니다.
    #   - --windowed (pyinstaller) 로 빌드된 실행파일에서도 동작합니다.
    #   - 실행 파일(또는 .py) 이 있는 폴더에 errorlog.txt 를 생성합니다.
    def _get_error_log_path():
        """실행 파일 또는 소스 파일이 위치한 폴더에 errorlog.txt 경로를 반환합니다."""
        if getattr(sys, 'frozen', False):
            # PyInstaller --windowed 빌드: sys.executable 위치
            base = os.path.dirname(os.path.abspath(sys.executable))
            # macOS .app 번들일 경우 Contents/MacOS 위로 3단계 올라갑니다.
            if sys.platform == "darwin" and base.endswith("Contents/MacOS"):
                base = os.path.dirname(os.path.dirname(os.path.dirname(base)))
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, "errorlog.txt")

    def _write_error_log(exc_type, exc_value, exc_tb):
        """예외 정보를 errorlog.txt 에 기록하고, 가능하면 messagebox 로도 알립니다."""
        try:
            log_path = _get_error_log_path()
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            tb_text = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"\n{'='*60}\n")
                f.write(f"[{timestamp}] 프로그램 오류 발생\n")
                f.write(f"{'='*60}\n")
                f.write(tb_text)
                f.write("\n")
            # GUI 메시지박스로도 알립니다 (창이 살아있을 때)
            try:
                import tkinter.messagebox as _mb
                short_msg = f"{exc_type.__name__}: {exc_value}"
                _mb.showerror(
                    "프로그램 오류",
                    f"예기치 않은 오류가 발생했습니다.\n\n"
                    f"{short_msg}\n\n"
                    f"오류 내용이 다음 파일에 저장되었습니다:\n{log_path}"
                )
            except Exception:
                pass
        except Exception:
            pass  # 로그 기록 자체가 실패해도 프로그램이 이상하게 종료되지 않습니다.

    # 6-0-1. 처리되지 않은 예외를 가로채서 errorlog.txt 에 기록합니다.
    def _global_exception_handler(exc_type, exc_value, exc_tb):
        _write_error_log(exc_type, exc_value, exc_tb)
        sys.__excepthook__(exc_type, exc_value, exc_tb)  # 원래 동작도 유지
    sys.excepthook = _global_exception_handler

    # 6-0-2. stderr 를 파일로 리다이렉트 (--windowed 빌드 시 stderr 가 없어 오류 발생 방지)
    try:
        if sys.stderr is None or getattr(sys, 'frozen', False):
            sys.stderr = open(_get_error_log_path().replace("errorlog.txt", "stderr.txt"),
                              "a", encoding="utf-8")
    except Exception:
        pass

    # 6-1. 윈도우 창을 하나 새로 만듭니다.
    try:
        root = tk.Tk()
        # 6-2. 우리가 만든 '버스 기록기' 설계도대로 프로그램을 조립합니다.
        app = SeoulBusArrivalRecorder(root)
        # 6-3. 창이 꺼지지 않고 계속 우리를 기다리게 만듭니다.
        root.mainloop()
    except Exception as e:
        _write_error_log(type(e), e, e.__traceback__)