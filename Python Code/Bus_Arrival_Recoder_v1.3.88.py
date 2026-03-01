import tkinter as tk # 1. [준비물] 컴퓨터 화면에 창을 띄우는 기본 도구상자입니다.
from tkinter import ttk, messagebox, filedialog # 1-1. [준비물] 더 예쁜 버튼, 알림창, 파일 찾기 기능을 가져옵니다.
import threading # 1-2. [준비물] 여러 가지 일을 동시에 처리해주는 '일꾼'을 부릅니다.
import time # 1-3. [준비물] 시간을 재거나 잠시 기다리게 하는 시계를 가져옵니다.
import os # 1-4. [준비물] 컴퓨터의 파일이나 폴더를 다루는 도구를 가져옵니다.
import sys # 1-5. [준비물] 프로그램 설정이나 컴퓨터 정보를 확인하는 도구입니다.
from cryptography.fernet import Fernet # 1-6. [준비물] 글자를 암호로 바꾸어 비밀을 지켜주는 도구입니다.
MASTER_KEY = b'u7_K-5D4fR9zP2mN8xL1qJ6vH3sB0tG9wE8rT7yU4iA=' 
cipher_suite = Fernet(MASTER_KEY) # 1-7. [준비물] 암호를 풀거나 잠그는 열쇠를 만듭니다.
import webbrowser # 1-8. [준비물] 인터넷 사이트를 열어주는 도구입니다.
import platform # 1-9. [준비물] 지금 쓰는 컴퓨터가 어떤 종류인지 알려주는 도구입니다.
from datetime import datetime, timedelta # 1-10. [준비물] 오늘 날짜와 현재 시간을 계산하는 도구입니다.
import xml.etree.ElementTree as ET # 1-11. [준비물] 복잡한 인터넷 문서를 읽기 좋게 정리해주는 도구입니다.
from urllib.parse import unquote # 1-12. [준비물] 인터넷 주소에 들어있는 특수 문자를 글자로 바꿔줍니다.

# A. [주변 검사] 지금 내 컴퓨터 상태가 어떤지 확인해봅니다.
# B. 윈도우인지 맥인지 확인해서 이름을 기억해둡니다.
CURRENT_OS = platform.system() 
# C. [선택]맥os 앱 실행 권한(Gatekeeper) 격리 속성 제거 터미널 명령어 xattr -d com.apple.quarantine /경로/to/your/app_file(혹은 드래그)

# D. [예쁜 글씨] 화면에 나올 글자들의 글꼴과 크기를 미리 정합니다.
if CURRENT_OS == "Windows":
    # 1. 윈도우 컴퓨터용 글꼴과 크기 설정입니다.
    FONT_MAIN = "맑은 고딕" 
    FONT_SUB = "돋움"       
    FONT_MONO = "Consolas" 
    # 2. 글자 크기들을 미리 정해둡니다.
    SZ_L = 19  
    SZ_M = 11
    SZ_S = 9
    SZ_XS = 8
    SZ_XXS = 7
else: 
    # 3. 맥(Mac) 컴퓨터용 글꼴과 크기 설정입니다.
    FONT_MAIN = "AppleGothic"         
    FONT_SUB = "Apple SD Gothic Neo" 
    FONT_MONO = "Menlo"              
    # 4. 맥은 글자가 작게 보여서 1.4배 정도 키워줍니다.
    #        버튼 크기 조정은 get_btn_style()의 padx/pady로 별도 제어합니다.
    SZ_L = int(15 * 1.4)
    SZ_M = int(11 * 1.4)
    SZ_S = int(9 * 1.4)
    SZ_XS = int(8 * 1.4)
    SZ_XXS = int(7 * 1.4)

# E. [창고 검사] 프로그램을 돌릴 때 필요한 특수 도구들이 다 있는지 확인합니다.
try:
    # 1. 인터넷 연결과 엑셀 작업에 꼭 필요한 외부 도구들을 불러옵니다.
    import requests # 인터넷 서버에 질문을 보내는 일꾼
    import pandas as pd # 표 형태의 데이터를 엑셀처럼 다루는 도구
    from selenium import webdriver # 인터넷 창을 자동으로 조종하는 로봇
    from selenium.webdriver.chrome.service import Service 
    from selenium.webdriver.chrome.options import Options 
    from selenium.webdriver.common.by import By 
    from selenium.webdriver.support.ui import WebDriverWait 
    from selenium.webdriver.support import expected_conditions as EC 
    from webdriver_manager.chrome import ChromeDriverManager 
except ImportError as e:
    # 2. [오류 대처] 필요한 도구가 없으면 안내 창을 띄우고 프로그램을 종료합니다.
    root = tk.Tk() 
    root.withdraw() 
    msg = f"필수 라이브러리가 없습니다.\n아래 명령어를 터미널(CMD)에 실행해 주세요.\n\npip install requests pandas openpyxl cryptography selenium webdriver-manager\n\n에러 내용: {e}"
    messagebox.showerror("실행 오류", msg) 
    exit() 


# ══════════════════════════════════════════════════════════════════════════════
# F. [macOS 전용 커스텀 버튼 위젯] MacButton
# ──────────────────────────────────────────────────────────────────────────────
# G. [MacButton 클래스 소개] 맥OS에서 tkinter 기본 버튼은 배경색을 바꿀 수 없습니다.
# H. 이 클래스는 Frame(배경판) + Label(글씨)을 합쳐 배경색을 자유롭게 바꿀 수 있는 커스텀 버튼입니다.
# I. (텍스트)  조합으로 완전한 배경색 제어를 구현하며, tk.Button 과 동일한
# J. configure, config, cget, pack, state 등 tk.Button과 똑같은 방식으로 사용할 수 있습니다.
# K. 윈도우와 리눅스는 기본 tk.Button을 쓰고, 맥에서만 이 클래스를 씁니다.
# ══════════════════════════════════════════════════════════════════════════════
class MacButton(tk.Frame):
    """tk.Button drop-in replacement with full bg/fg control on macOS."""

    def __init__(self, parent, text="", command=None,
                 bg="#272727", fg="#f1f1f1",
                 activebackground=None, activeforeground=None,
                 disabledforeground="#555555",
                 font=None, cursor="pointinghand",
                 state="normal",
                 padx=4, pady=2,
                 width=0, height=0,
                 relief="flat", bd=0,
                 highlightthickness=0, highlightbackground=None,
                 overrelief=None,   # 호환성용 (무시)
                 **_ignored):
        _frm = dict(bg=bg, relief=relief, bd=bd,
                    highlightthickness=highlightthickness)
        if highlightbackground:
            _frm["highlightbackground"] = highlightbackground
        super().__init__(parent, **_frm)

        self._norm_bg  = bg
        self._norm_fg  = fg
        self._act_bg   = activebackground  or bg
        self._act_fg   = activeforeground  or fg
        self._dis_fg   = disabledforeground
        self._state    = state
        self._command  = command
        self._hand_cur = cursor

        _en  = (state == "normal")
        _cur = cursor if _en else "arrow"
        _fg  = fg     if _en else disabledforeground

        lbl_kw = dict(bg=bg, fg=_fg, padx=padx, pady=pady, cursor=_cur)
        if font:   lbl_kw["font"]   = font
        if width:  lbl_kw["width"]  = width
        if height: lbl_kw["height"] = height
        self._lbl = tk.Label(self, text=text, **lbl_kw)
        self._lbl.pack(fill="both", expand=True)

        for w in (self, self._lbl):
            w.bind("<Enter>",           self._on_enter,   "+")
            w.bind("<Leave>",           self._on_leave,   "+")
            w.bind("<Button-1>",        self._on_press,   "+")
            w.bind("<ButtonRelease-1>", self._on_release, "+")

    # ── 마우스 이벤트 ────────────────────────────────────────────────────────
    def _on_enter(self, _e):
        if self._state == "normal":
            self._lbl.configure(bg=self._act_bg, fg=self._act_fg)
            tk.Frame.configure(self, bg=self._act_bg)

    def _on_leave(self, _e):
        _fg = self._norm_fg if self._state == "normal" else self._dis_fg
        self._lbl.configure(bg=self._norm_bg, fg=_fg)
        tk.Frame.configure(self, bg=self._norm_bg)

    def _on_press(self, _e):
        self._pressed = True

    def _on_release(self, _e):
        if getattr(self, "_pressed", False) and self._state == "normal":
            self._pressed = False
            if self._command:
                self._command()

    # ── configure / config ────────────────────────────────────────────────────
    def configure(self, **kw):
        if "command" in kw:
            self._command = kw.pop("command")

        if "state" in kw:
            self._state = kw.pop("state")
            _en  = (self._state == "normal")
            _cur = self._hand_cur if _en else "arrow"
            _fg  = self._norm_fg  if _en else self._dis_fg
            self._lbl.configure(fg=_fg, bg=self._norm_bg, cursor=_cur)
            tk.Frame.configure(self, bg=self._norm_bg, cursor=_cur)

        if "text" in kw:
            self._lbl.configure(text=kw.pop("text"))

        # 2-1. bg
        _bg = kw.pop("bg", kw.pop("background", None))
        if _bg is not None:
            self._norm_bg = _bg
            self._act_bg  = _bg
            _fg = self._norm_fg if self._state == "normal" else self._dis_fg
            self._lbl.configure(bg=_bg, fg=_fg)
            tk.Frame.configure(self, bg=_bg)

        # 2-2. fg
        _fg = kw.pop("fg", kw.pop("foreground", None))
        if _fg is not None:
            self._norm_fg = _fg
            if self._state == "normal":
                self._lbl.configure(fg=_fg)

        if "activebackground" in kw: self._act_bg = kw.pop("activebackground")
        if "activeforeground" in kw: self._act_fg = kw.pop("activeforeground")
        if "disabledforeground" in kw: self._dis_fg = kw.pop("disabledforeground")

        if "font"   in kw: self._lbl.configure(font=kw.pop("font"))
        if "padx"   in kw: self._lbl.configure(padx=kw.pop("padx"))
        if "pady"   in kw: self._lbl.configure(pady=kw.pop("pady"))
        if "width"  in kw: self._lbl.configure(width=kw.pop("width"))
        if "height" in kw: self._lbl.configure(height=kw.pop("height"))
        if "cursor" in kw:
            _c = kw.pop("cursor")
            self._hand_cur = _c
            if self._state == "normal":
                self._lbl.configure(cursor=_c)
                tk.Frame.configure(self, cursor=_c)

        # 2-3. [호환 처리] tk.Button에서만 쓰는 속성(highlightthickness 등)을 Frame 수준에서 처리합니다.
        for _k in ["highlightthickness", "highlightbackground", "relief", "bd"]:
            if _k in kw:
                try: tk.Frame.configure(self, **{_k: kw.pop(_k)})
                except: kw.pop(_k, None)
        for _k in ["overrelief"]:
            kw.pop(_k, None)
        if kw:
            try: tk.Frame.configure(self, **kw)
            except: pass

    config = configure   # config 별칭

    def lock_size(self):
        """렌더링 후 Frame 크기를 자연 크기로 고정 — 텍스트 변경 시에도 크기 유지."""
        self.update_idletasks()
        w = self.winfo_reqwidth()
        h = self.winfo_reqheight()
        tk.Frame.configure(self, width=w, height=h)
        self.pack_propagate(False)
        self._size_locked = True

    def cget(self, key):
        if key == "state":  return self._state
        if key == "text":   return self._lbl.cget("text")
        if key in ("bg", "background"):  return self._norm_bg
        if key in ("fg", "foreground"):  return self._norm_fg
        try: return tk.Frame.cget(self, key)
        except: return ""


# L. macOS(맥): MacButton 사용 / 윈도우·리눅스: 기본 tk.Button 사용
AnyButton = MacButton if CURRENT_OS == "Darwin" else tk.Button


# M. 공공누리 저작권 마크 이미지를 base64 문자열로 코드 안에 내장했습니다. 별도 이미지 파일 불필요.
_GONGGONG_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAJUAAAA2CAYAAADOKtsPAAAvwUlEQVR42u2deZiU1ZX/P+9S+9LV+0bTTUODyCIgm6LgvhDQcWSMEqKiiCFG1F80Ehd0HkdlXKIxo1nGJGKMRB7HjaCRoIIhbIKAtIKAiNA0Te9dXXu9y++PqnupbtEmoDP/5OWpp5vuqne599xzvud7vve0Ylu2/edly3jmmWfYvHkzhmGgqiq2bcuv/zyO/1BVFdM0SSaTFBQUoKoqDoeDVCpFJBJBUZSv/bxpmrhcLkzTBMCyLGzbxul0kkqlcKraCd2frany3OJeLctC0zQ0TZO/+6rDsiy8Xi/9+vXjnnvu4TvTpqFs/XCLfdddd/Hee+8xZMgQRo4cicvlIp1O43Q6sSzrn5ZxAkfvCVu1ahXhcJhx48ZRW1tLOp0+pnNYloWu6/I8qqpmDPIE58fIGpC4T6fTCUA6nZbX/LojGo2yc+dOPvvsM+rq6nj22Wfhpnk/tD0ut33mpDPsFX95204nU7ZtWnYiFrdty/7n60RfpiW///yzvfZ555xrV5ZX2L/972dtM230+XnLMHueK+d8tmXblmWd2Euc37R6XMtIpY/p+SzDtN9ftdq+4LzzbRXFvuq7V9r6ihUr8Pl8XHXVVUyZMgXd4QDAkf1q88/wdyKHoihg26Ao2LZNV1cXLS0tGe+laX2Or6KqPTyW8FC28FCaeoL3d+QbRVGwTJN0Oo2qqmjHMP+KqnLm5Ml897vfpaGhgXfffRc9HA4zYMAAxo4di9PlIpVM4nQ6UTXtyKD88zgB0JKdOdtGURTcbjculwvtWMc3O6e2ZaHlhCLTNNEdjhNf8jlGLwzdlb03bBtF7eP+rMxnTz31VFwuF4FAAF3TNJxOp/RMTpcrE1NTqcyD6+o/DeNEbMowM8agKFiWRTqdJpVKYRhGZk76MItYdwR/IICiqpiGIY1QdzgyxnDivlQarWma2LaNI4urTNNE6SMRUBWFZCKR8XKWhWEY6OJBE4lED4vVdR1FVf8Z/E40+9N1bMtCUVVUVcXlcuF2uzPGlUzicLu+9vP+QABsm3Xr1rFv3z4Mw8CyLFwuF7FYDGcfQLqvw8p6SwHKRdY3ePBgThk1qs/5t7P3Yts2FRUVNDU1HfFUAvXnxkrbsrC1f4a/E/MDGSwlRjGRSJBOp9E0DafL1aenUlB46aWXuP/++zlw4ACBQIBYLIbH48mmb+aJZafZGzMMA7fbLamOSy65hIceeoiBQ+q+3iizYTgajdLR0YGqqpngZhkGmqJkwKRloQjwpyoYKDgtQIUYBiom7pQNmk5KU3GgEo+n8XocWd7CQFXBtszsQGqk0klUlwcTMLFxoqHaoNpgWIZMWwV9oeaAU9u2JW8i/i9CQI/UGkgmk7iy4ds0zcxnjsHVplMpHE4npmEcwS22nbmWpsprplIpnE5nxsXL9yE9fCqZlPDBMs0MLrXtDCDPSf0F1MC2UfvAVIlolOd+9zvCnZ2UFBWhKApBvx8jGwpVMgA7ZaQxTZPCwkL2NzTg8Xhwu90UlRSzf/9+LMvC7/VlqAPLwuFwZAB/dnwty0JRFLxuN0UFBbz3zjv85c03uXnIrX16YkAauW3b9Ok7nVGDmE8hjYIvZqBrbqIu2GR1sH7vLvaGEzRHO4gqJkkrRcBSGBEs5fSyAYwtq8ENBF06Whr0eBSCHpJYpLBw2Rq6rktjyAWtYiIVRUHTNFJZjCe+V1X1SxxKrrcV3IumqH0CVU3TsATpBxnj0jSZeQmjdjqdJBIJ3G43QOY+UGTG7HS5pDGJROdEj1Qq1YPrEmMkxkZBwTRNfD4fhmGwd+9eRo0Zw8yZM5kwYQIOl5O2tjZ+/etf89e3VxAKhXBkxyYej+N0OuW5BNlt2zamaUrc948efQdkn43TNPDiBqebnakE/7VtHa/s2YYj4GN/woR0GlRQUVEti782HeKXez+hpKiYef1GMnXASQxRAIcHYglUr4aiOLA1hXg8ztq1a/H5fGiahmEYktUXhhMOhwkGgwwbNgyv19vDoITXaGpqYufOnTidzgzYdDiIxWK4HM4+GW9d16mrqyOYl0fz4cPs3r0bTdNIp9PUDRlMWVkZsVgMr9eLw+GQVQan0wmWTbiri40bN5KXl0cqlWL48OHkhUIY6XSf5GFfRzKZlJOb68Fzs8B0Oo0Vs3G5XNTW1nLnnXdy8cUXZxZFdp0WFxejorBy5Urys/fpdrt7RAdVVSVjL877rRhVCgunrYMKf2r+nCfq/87mcCumSwcjDqoDPA6wLSxsLM0JKZVUysTsDvPQxnfZ0LCXG0eO56xQEUrSgSMch6CDOCZv/nk5CxYskN5HDJyu68RiMTRNw+12o6oqyWSS2tpaZs2axdSpUykoKEDXdSzD5JWX/4f/+I//kPjQNE0cDkefqy0cDjNy5Eh+/vOfc3IgwFNPPcUf/vCHTOYVi/H/bv8xd955p5wA6QGFJ1IU1qxZw5w5c1BVFY/Hw8KFC5k1a9Y3kqGJSRYeShi0KKMZponX6yWeTJBMJjnnnHOYOnUqKEomg3dkIsHQoUO56qqrWLlypQT7vasluSW5EynP9ckXOOM2UV3n4f2bmLf1z2yMH8Z0uXCpLjBsBqkqE/0hpgSKOMOdzxjNS4WughYnkmikNaDyUuun/HTb2yxt24vtcoDuh4SJQmZ1dXZ29kgWhIF5vV4Z7hKJBKlUit27d3PHHXdw9dVXs2PHDsmtpNNpWltbZW0tmUySTCbl4H3Vy+Px0NnZmZk4VZWG7PV68fl8PPfcc7z99tuoqko6nZaTmTvwmqZlvKLLRSqVkpOVSia/EfI010OJ+xbXFqUbXdexbZva2lrIpvmOrNf2eDyk02lOPvlkSkpKsCxLjrUId73PK6DGtxP+3G5eaPyEJ3eso9NMQjAEze1Uaz5mnDSRM0tLqMkroUzzolkKh6w4q8MNLGv6hPcb99CdMLGCLjZ0H+TRrX+jclw+Z3jywVJwYaMCmqLg93ppbW3Ftm0ikQjhzk5ZgxKEoWmahEIhVK+XDzdt4rZbbuGPf/wjJeXluJ1OQsEgDk2js70dwzBIJ5PE+5jYVCpFSUnJkQHR9R5pe2trK08//TQjR46ksrLyCH8jJtvOTIBt2+i6TiKRyPI7aga0n6CnysVRwgB6G10ikUDVM9AhEon0wJfi/Q6Hg+7ubqLRqHwGkXT0Pq+45rdmVO8raX6z/m80K2l0Xx580cbUgScxf9xkznTl4zTMDLdl2aAqBDQPtUV1XOarYb/nILM3L2dXKk7C62RzZzu//3QbFePOotZWUIyUBIRdXV0AzJ07l9NOO01OlJiscDjMmjVrePvtt+nq6qKgoID6+nqWLVvG7NmzSSQSckILCwu58sorGTlyJM4sqP66Yq3H46Guri6T+WYzIkVRMAyD/Px81q9fz7PPPsuPf/xj/H5/j2xUeA9haJqmZTyVbffMEk+EQD3KhAtvKK4rMOT69evpaG8nv6BALpJoNIrP5+Nvf/sbhw8fprK8XI6XuD8B1nuUl47z6POJ/7BtHfuSMSjwYsQTzKoZxdPjLySoQNjswKnlZdJqwwJFxTbA0iDgcVBbU8N/1HyPn760mE8cGrbbwfLdnzC+vJqZ/QYQMDODEggE5KoZNWoU5190UWaFWxaoKmY6jeZ0cskllzBo0CCefPJJuru7cbvdvPbaa1w/dy5er1e6cdu2GTduXOY8fc8YlmXJbE2oAAQfpCrgcrlYsmQJkydP5uyzz0bTtAyeMQx0TZeGKCZYYK1vAlOJhZWbDYvrCQ+UTqdJp9P4/X4aGhro7OwkGAxmDF9R8Pl8JJNJ6uvrCQQCpNNpvF4vqVTqKw3paF7xmL1rWFUgbRNUdWxMUhqYQALY1drO6sZGkh4PxCIM9Pu4dPRIggpgglvLA1MFVEyXTkKzSDoNLDWO00oSUpNMt0P85NRz0IMqKN20uFK8d2g3sRTYLheOLHFmWRbJZLKnfifLnWkOB7ZlYQHnnHce1QMGYFgWpm0TjcexLYuUYWAButNJMp0mkUphmSZ2X/8UUHVNZkmJRAKPy0U8GkUFkvEE+Xkh2lpaWfTQw7QcbkZFwUxnDCpX3mLleDrBNn9TXqo3pyQ8lUWGS4tGo7g8Hn71m99Q0a8y80zZul06ncblcnH//fdz1VVX0draKuU0YhHmJgDiOn1pqb7SqJy6A1tV6E7GM24UG9u0SQA7Is18lm4hqnfhC/m4uvYUpvrLwTCxXCniJOmyoqQVC82yccct3Gkdt+FBMVykOzMA85IhI5iiFBPEj+rS2LBvB+FkFCWWoQScTueXwC+WdaSEniUeVU3D7/fL+pmqqhmeJes1bNuWbLXD4UDVtB6r+6teuQBVkIIAXq+XvLw8wuEwgUCALVu28Itf/CJj6LougbjgsLTs9QzDwEineygMvq1DkKDBYJA5c+YwZsyYTIktJ0QKsrWktJQZM2Zw+umn09HRcdyYqe/SlJXJnpJ2lmCzFHRVIQ580HEAf3cUJWUTiCQI6Y6M0sKhYSsqCpCn+tAVlUOqyRrCPHVwO/O2ruTaXau5PbmLD5ItpIApFbW4uiMYzhj7QknWJZuwVbsHW527ysm6biPLlbiy2Gjnzp20tbXh8/kkl2JmeS2hqDRNk66uLuKxmMQOX/eSzH3OylRVFb/fzzXXXENJSQmJRAKfz8fSpUtZ9sYbEq+YWV4tlUpJjysIxf+VgrVtE4/H6d+/P5dffjker7cHQSreIziniRMnctZZZ0mv+m0cumJYKLqG7cgATNVUQIUWLcbW5i8I52toLg9tXTFe2f8J/bwFXOSswOvS8Tp1PkxG2HrgAH85UM/7rfs4TBLcHlQU2GWwp9ti/rSruHzwqXzcuJc3EvuIxbr5pHUfkaKBmYnJckqhUCjzoKaZMapsNV5M9tatW3nhhRdobGykoKCA1tZWRo8eLd21w+GQ6fPzzz/PihUr+nThXr+PGTNmcPbZZ6NoGfIvHo9LL3D22Wfj9/tZtGgRuq5z+PBhnnnmGU499VQqstmgw+HA7XbLgqyUtvwvSLEFfzdu3LhMFiv4q5wkwchWCMT9jB07loqKCsLh8DeSSHwZqGcBpq7r2MqR0rWuKgyq7A+fdqMkwcLF+tZmDm9ayX87/KQwafVCc8qgq7uTeKSbEo+Xy6qHcXH/oZziLqQwBVvjDVQoboaacHv1eH5QcyE79u/hZH8xPiOT2YnYHo1GefXVV9m+fbs0FKfbzd69e4lEIjQ2NrJv3z5M0ySVSlFcXHyEZATi8bjUgW/ZsoVUKtXnAMSTCYYOHcrkyZPRNV0apzAuXdf53ve+x+rVq3n77bcpLi7mo48+4re//S333nuvDM2CgjAMg0QiIXmvb9uwbNvG7/czYcKEHuEsl9sSpLJtZsb5lFNOobq6mk2bNn07RqWqKnbaQLFsUNRMPmhBpeLh/Orh1Ect1rXuI04S0+FgZ7qNna5uSBgQtkDX0JJwRkF/rhw8mulVJ9HfUiAJYJJfdjJegINd/PuV16IEPJx77rn0+/53UYOKXFF+v59YLMbGjRtZt26dJBkVTSMajcowGQgE0HUdp9PJzJkzOeWUU3rU+oSbF0y83Yen0hx6j7qjkICIMJxKpQjl53PTTTexdetWWd5YvHgxEydO5PwLLkBVVSKRCHl5ebLU9L91mKZJaWkplZWVmWdQFNLp9JGidXYshAFquk5RaSl+v/9buyddURTcio6aMrGBlGLhNCFgqaQ27uKx6lF8VFLLex2fsyfRxWexNsJpC5/tpJ/uolp1c9boUVxYO4xaFJxGRvdl+jIZpA+bFAppO8YXkUM07Wli18cfYRS4+X83/FDG9nQ6LTkrgYtUVSWd3U0iPJFt29TV1XHhhRdy1913Z9x7Oi1VlbquE4/HGThwYMYA+wDL3mCAyspKubI1TSOZTGLbNj6fT2KuKWedxaxZs3j66afJy8ujo6ODxx9/nNGjRxMIBCSl4XA4pBbtf+OwLAu/3y+L3F9VYkmn0zg0vQd98G0BdT2VTuNxudAtEflU0FXUqM3nK9ez4JV53HPLT7hp9EjMqiEknTp2LI3L4SRQXkqJ7sAFBAElLVRfYKmgKRouA9At3nxnBXvaGgmVFdC+v5WSsmKwjpQg4vE4Xq+X6upqCgsLZXFTzRqJ1+ulsLCQmpoaLrvsMqpraqQiIHc7kQiBCxYs4OKpU/sOPzlyWTNtSA8nito+nw8BEW6//XY2bNjAmjVrKC0tZffu3SxZsoTq6mpZGxSlHCl/+ZaNS3BjXq8XRVWxsky58FYiEgjOCtsG25bk8rdiVBoG3VYCy+NCt1VU0wQdEo4klaMG0PLLJm76yXxKSkqYM2cOP73rLjmZlmmi5nAbiuPIAOq2jQMFy0rjsh387k8vUe4txug2KKodyIiaIaBnVpqoXXWGw9x77bX864zLexRNc6UZIm3PpK4atmlJSkFkZAJo99CG9UEuqqqKlk3Fhac0DAPTtkBVsEwbr9/HnT9dwK7rr6e9owO/38+LL7zAuHHj8Hk82NnVbyuZhaVoKicqnVU0jUQqha0oON1uuru7cblcqIqCadsomsbBQ4dYt2EDJWVlRyS9uk5SScgyl4AFLpeLXbt2cbilBSXL/otFKcYil9w97vCXm1paloVKZrWOGDGCU089lfr6eqLRKEuXLuXSSy/l5GHD6OrszMg7cmQZuRV1uSHV4eC/f/Urdu/ejWVZhMNhzrvgfIYPH56ZNNOUVX9d1/F6vT0EX19lCDJtzuG3tByvZWUlvH3tW8zdQKmr2pdkNwLsC4wyatQofvCDH/DYY4+RTqc5ePAgnZ2dJJNJPD4f4XBYjoncSXOCeipRcQiHw3LyE4kEWpa3SyQSLFq0iDfeeEMWtqXx5YyBnS0daZrGrl270HUdh8NBIpHAMIweBfxM2dd9AkA9697FQAjvMXjwYMaPH8/HH3+M0+mkqamJu+66i9/85jeUlJYSjUTwBfw9iLjcCrhpmry74q88+eSTxONxYokE+fn5XHbZZUcmPgcgi3JDD6P5Ci7lCElqS3wgiEtVVeXAHEv4ERmoMB4hxhNZce4EB4NBrr76atasWcP69evxe31Sa6U7HPh8PqnpMgyjB445nkMYk7g3kWH6fD6pDQ+HwyiKwrvvvivLL7IGmYMTvV6vJIelHDmLyeT9ZhdPrtriuDyVmFAATdex7CMDfN111/H222/T3t6Oy+Vi06ZNXH311Tz66KOcfPLJPdy7rulkN4txYP9+tm7dyv0LFxKJRHC5XMSTSS699FLOO++8HqFNeCmtl7v9OnJO/i6LiQzDIBqNomkaJSUlR37fh6MwLVPiMVV3yPKEGI/cVS52bBcXFzN79mw+++wzwu0d5Ofn093dTVtbGwUFBfKzDocjU2g/QZGex+PhtttuY9KkSfJ+vF4vzc3NAIRCIZLJZA9PnSt9DgaDcpGZpkk0GiUYDEonkEqlCIVCtLW1EQgEaG9vZ9GiRaSPs0yjC6/Um9cQ3NGAAQO4//77ufXWW4lEIoRCIbZu3crMmTO5+OKLufzyywkGg+Tl5aGqKl1dXezYsYM33niD5cuXkxcI4PF42LdvH5dedhm33HKLzOZyNei5pZLc+lNvCe3XeRuPx4PH48GyLFavXs3hw4f7xAWKlvFqo0aN4qwpZ8nwIjLJo7HTANOmTWP79u089bMnULJFWzNHa/VNsdXifoYNG8apY8dKPX0sGmXLli18/PHHTJ06lQkTJuB0uXjxj3+kqKiIs846iw0bNlBbW8u2bduorKyktbWVQYMGMfKUU2htaSGRSBAKhdiwYQPjxo1j4MCBsgj+7LPP0t3dfXxGJTyG1HdnRfyKomRcvw3TL7mEzs5OHnroIZLJpJRT/OEPf+CF55+nrq6Oqqoq4vE4+/bto7m5GdM0cTudxGIx0uk0F198MbfddhvVNTU9cIuo+wnDyjXw3qHraBhL4cimB/H7trY2XnrpJeLxOI4+0uaUlSnp3HTTTZw15SzptUWYyb2m0KcLLzBnzhy2fLCJFStW4HK5cLlcdHV19QinKsoJG5XYgQPIPXmvvPIKixcvJpFI8PHHH/Pggw+SSqV45plnqKio4NChQzQ0NJCXl8cjjzzC+eefz7Zt2zj//POZM2cOy5YtIxwOc/a55/Dq668xfOQISktLJaWTMtI9MuPjCn+5K1EhB7xnPcbMmTOpq6vj3nvvZdeuXaRSqQw3Y5rs3buXjz/+WComhRpRZB3Tp0/n3//93ykpL8/sOMnudTMMg1gshmEYElQK8H40DqW3t+hdWM0V11mWhc/n65OnCrpdsigrykGmaWZkL1kvKdJzAVxFWCktLWXu3Lns37+fxsZGNIejBw6zc/ZRnqhhCayTTqVobGzkr3/9Kw0NDQwaNIhPPvmEl19+mWnTplFRUYHX6+XnP/85kyZNwu/3Ew6H8Xq9MiExDIMDBw5gWRbBYBCHwyHHWyxsUSQ/bqAuJqR3NmXbNqqiyi1GEydO5MUXX+Stt97i97//PZ9//jlWdr+YAI2iwJmfn89JJ53E7bffzmmnnYaSreo7sx1lHA4HqqpSWVnJSSedREFBAeFIhMrKyh4g82hAOze0JOMJXG43gwYNYty4cVIvJMjLdF/KT8ukrq6O2tpaVE3j5JNP5rTTTiORSFBRUUFJSYks26hZubFYOLZtc8HFF7Nr1y5ef/11nG43TqeTioqKHlnXicqJhfJBeKqSkpKMsjWdZs+ePSiKQl5eHkVFRXJRXHTRRXg8HpLJJIFAAFVVCQaDFBYW0tDQwObNmwmHwwysG4Rt25K8Ffp+kTQdd/gTvEQut5MrBlMUBUXLGF9xaQlXX3sN0y+9hKamJt544w06Ojpob2/Htm1KSkoYOnQoo0ePZuDAgT3SUuGhxKpTVZXR48by57+8Jf8vJquHYuFrDpcnQzpOv/QSvjN9Wg9sdqyMtFiZpmUy6+rvc9X3ZspV7fP5ehi31+v9knHfdMt8fvCjm3osBKllt+wvKSvl/R2DF0ulUui6noUimc95vF4GDx7MqlWrJEAvKSnhnXfe4YorriAQCLB9+3aam5vZuXMneXl5RKNRPB4PjY2NbNu2jblz51JZWcnSpUvBsknE4ng9XrBsUokkik3f29tyniGXS9SPZaUc7fv8/Hzcbjd33nlnD+4ol0w81pQ59/O5nztWsCvCVO8mbb3P91VG1ZuryyUCj9WbHG1PXm9O7XjUlMLbhMNhKQXSNI1hw4bx/e9/n1AoRDqdZvz48RQWFlKY9VaVlZVs376dCRMmUFdXR1FRkTTAoqIi3B4P2DbV1dW89dZbRxaLouDJcoXHUpDnKCFe7/szPQclV1Lh8Xi+ZES9PdyxGMQ3cfQ2jH9EOvJNnO8feQ5pXMeAufLy8qiurmbp0qUMGDCAsrIykskkkyZNYsqUKTidTrkbKRKJyI0PBQUFnHHGGcRiMSoqKujq6sLn89HZ2Uk0GqW5uZlktsPP2LFjCYfD7N27l1AoxEcffcSuXbuYNGnScakm9GMdgFxPILdc55RQektfTzSlzj3Xsbzvq7zrsXiGr9ql8k08xwmDdF1n7ty53HPPPcyePTvD98XjUqQoFrTITAUe6r1bOx6PSyFjPB6XtI4I8bFYTJ7P7XYzfvx4LrnkkuPUUx3jBH9VKBENsnrzM717IByLJ/xHjOlYDO1Yw/D/heEc6zVj0Shnn3MOzz33HJ999hm6ruPz+WhtbZWGJUotoqIgdlfnlqH8fj+dnZ0EAoGMDj8bVkV2a9s2RUVFHD58GI/Hw4QJE/Bm8eQ3blRf0o4fhXTMxVK9mfJva7D7ev+xht9v6sj1drkeLnd3zZe83zHcn8fjwbYsBg4cSN3gwXLrl6pmsvLcpiJWto4qtVSiBCaahYB8v5HrDLJYzTAMTh07Vr7vWMJzb6CeaRryD0yeKGGIDCuX3xKerHfHlv8rL9EbHB/Lq7eB/COgWjma0RzHe770GVXtocIQDUVUTSPS3f2lhetwOmUrSLExQ1GUjJH0wn6qKB7bNrrDgdvjkVzd8XamPiajyuUqcrMs8TWXi+m9P/9YeJreN567d+9Y9URHe//R9sr19TpRPHXUz3wDi8a2LJwuF/FYTF7HSKfxBwJyw4ads6UqnUrJHdKmYaBkZT1Wdue0LdSt2fOKNkjC0MRn/pEtWrnPros+3Tt27JBxVpB9ototdDi5+CkWi+Hz+eiORmUIFCDQ7XZLklBMutvtpqOjg7y8PNm5TwBLcb3c4raotuvZ6+WWdnJrgkJhkduMQ7j/WCwm700oF4QeSywAXddlPwe/3y/BqmgCZmW5ItG5rqioiJaWlh4teBRFkWK53Gvpuo6W/Vk6naatrY1kPE7A56OpsZEd9fXYikI0GiUUCkkuSaT+6XRaPr+maXR3d8vOMrKXgtOZKUdlFaddXV2EQiH5M9u2SaVSUiIjWjf5/X6ppM0lm8U4SrXK10QUuY/A6WTXzp3Eo1HyAgGU6sp+tqIoVFVVSYGbkG6IGpZY8aK+JrZ+J5NJtCwzLupTBQUFdHV1ycERLQlFxiFuxOVyEYlE8Pv9pFIpWW+zbZtYLIbb7c6UCrL3ZNs2gUBAttYR50wmk4RCoR4DJq4pJlNVVcnyi00JLpdLqhqExkg09ujRSzyrKxKT1ZEV54n7VbJGIQCwYL+DwSDR7IZUsRhSqZTsTlxVVZUhVnMWqcfjIZFISL2+w+FAyRq4UBJ0d3fj8/lkcb8zHMbj8UjZSiKR+NJ4a5pGOByW5RjxPF6vF6fTSXt7Oz6fT/Zb8Pv9EsS7s7bQOxrkLiAhpzlw4ADpdBplaN1gW+iwhfBLy242EHyUYMXFdqBcojOVFXfleoxIJCK1OyK9FWmw0OmIc4pMRExGrldKpVLY2d4Ign8RLtnr9RKLxeQAxONxaeymacpsSKTO0WhUZkLiZ0LyYpomeXl5dOdgFLGrJpXl5dJZICsMW9yf6BojVBKifCL0SkpO8iI8nWh8H41GsYBgMChrjclkEp/PRzQazRTbs+MudFXied1uN5FIBHe2X1cymZTXFdEmFxvJ58mOv+gGE41GZclM7A8U13A6nZg58Odombno4eD1emXNVRlQ1d/OTT2FNQuLz8VUYmvU+PHjGTlyJPF4HCtHYLd27VoaGxu5/vrrpcfJrSW9+uqr9OvXj+HDh7Nu3ToikYjUTwsPNGjQIDmJ0WgUXVW59NJL2bRpEx9++KEMXyLkiNBw0kknyYkRCoPGxkba29spLy+nvLxcGnsymSSRSHDo0CEMw6CiokL2Ni8tLaWpqYna2lo6Ojro6OqivLxcGkskEqGwsBDTNGltbcUwDOrq6qTRBINBWltbaWlpyXjELCTw+/14vV46OjrweDyoqkppaSnxZJLdu3fj8XgIBoNUVlbK8WpsbCQZj8uQXF1djWVZHDhwAEVROOecc1j57ruyk96YMWPIy8vDMAx27NjBwYMHZRirqKjAzPay2rlzJ8OHDycQCPDRRx9RVVVFIpHAmVWViL9IsXHjRlKJhBzjXMiRi1lzsa0sKOfySbliL3EygSmEhxg1ahRTp07NrMzsiUaPGcMt8+fz6quvcvrppxMMBikoKCAejxMKhQiFQixdupThw4ezcOFC5s+fz8qVK2VtzTAMioqKuPXWW2VP7lgsxn333sv06dPp6uqSRiUeQlEUurq6mDRpEj/84Q8JhUL4/X7a29upq6tj1apVXHPNNcybN48LL7xQYrFUKsXYsWM599xzaWlpYcqUKQwaNIi8vDw2b97Mli1bmDFjBu+88w6KpnH++edz2mmn0dDQwJ///GeCwSBFRUU8/PDD3HDDDfzkJz9h1apVFBcXE4vFqKmp4Qc/+AHbtm3LNBfJdvs977zz6N+/P59++inbt29n5MiRpAyDJ554gpqaGm677TZqa2tl85ENGzbw9C9+QSwWY9asWVx66aWk02nuv/9+tm3bxsKFC3nnvfeoqqrigQceIBwO09jYSF5eHvPnz+f3v/89ixcvZsyYMdxzzz2MGDGCLVu2cM899zBx4kTOPPNM7rvvPqZMmUJRURH9+vWjvr6e+vp6HnroIa688kqam5qOSpv0DoUioqmqmgHqArDmAmZFUeQGAvFmEUYee+wxfvazn3HLLbewYuVKdF3nueeeIxwO03Cwge9///sUFhYyc+ZMnn/+ea666irmzZsna1jBYJCamhpGjx4t43o4HGbdunX86le/oqGhgalTp7JgwQKpWxfuXdyb6HZSUFDAmjVr+Pvf/05tbS2lpaVs3LiRhQsXyvc++eSTPPjgg1xxxRU0NjayatUqPvjgA7xeL93d3SxbtoxFixZhGAZbtmyhpqZGaov+tno1XV1d/Nu//Rsvvvgizz77LD/60Y8YM2YMc+fOpaWlhY0bN/LII49QWlqKoijMnj2bjo4OGXKSWW8UCAQ4/fTTWb58OWvXrmXo0KE0HDxIa2srN910EwUFBUyYMIFINMKUyVN4/vnn2bBuHZs3b+bKK6/k2muvZcaMGdx9993MmDFDzk9ZWRlVVVWsWLGCZcuWUV5ezqmnnio9aHl5OStXrmT58uWEQiH69+8vQ2NXVxeRSATDMBg+fDhFRUU0NTVRXVPD5Zdfzu9/+1sJKYTH7M0ACBsS0ETPFaPldvsV1ieKq+FwmHQ6zciRIykuLsa2bRYuXMjIUaPYsmULHR0dlJeXc8W/XUE4HGbChAncddddhEIhfD4f9fX11NXV4fP5yAuFeOCBB6RIPx6Ps2TJEjZu3MisWbPo378/5eXlMgQKT5a7yyN3AdTW1qJpGldccQWDBg0C4NChQ6TTaUaMGIGqqhQWFnLDDTfQ0NBAd3c3W7ZsQVEUPB4PH3/yMV1dXeTn51NUVMSkSZMoLi6WevUZM2awY8cOLrvsMj755BMJlDs6Opg4cSL5+fk888wzcst7Y2Mj06ZN4y9/+QuHDh4EoKWlhba2Nvx+P9FolIMHD2LbNrNnz+ajjz7i2Wef5Y477mDJkiUSTL///vts27aN/v37s3//frZt28ZJJ53EaaedJjePaprGypUrefzxx3nggQdIJpOMGDGCPXv28OCDD2KaJocOHeK6664jEAjwxRdfEA6H6d+/P06nk7q6OmbNmsXatWtZuXIlnZ2dVFVV0dLcLHeDH42zE/9PZ/uair/qZdt2xlMJkJzLPeVao2j+1dzcTF1dHU8//TRdXV043G4uueQSLrzwQgn0amtr8Xq9mYl3Ornhhhvktqcf/ehHvPDCC/x0wQIKCgrkTfj9fnbt2kV+fj5nn302Bw8e5PXXX+fTTz+lpaVFatiFhxIrQwzqjBkzJBju6uri9NNPx+PxEIvFOPfcc+Vmz+3bt6NmMVpTUxOTJ09m06ZNTJ82ndGjR1NTU8NLL73E4sWLufbaa/F6vdx88834fD6uu+46xo4dK3HLihUrWLt2LWVlZUQiEWpra6murmbjxo2Ew2G540WI9txuN5MnT6a2ro4zzzyT9evXY1kWa9eupaCggO7ubpYuXcrgwYPliv/ggw9wuVz4/X7pBcvLy2ltbZU4tF+/flLt6fF4mDFjhjTKl19+mf/6r/9i9OjRdHR08Mgjj3D99dczfvx4IpEI8Xic7u5umWlWVlYyYsQICgoKaGlp4Z133pEJjoBEvVsaiaRGLHhVVTN6KuGResdI8dXlcsnMY/Xq1Vx00UVMnz6dyZMnY2W9yEsvvcT69esJh8OMHj2af/3Xf6WmpgaXy0V9fT1LliwhEonw8MMPEwgEMqrLYJB4NEoikeCKGTPobG+nraWFro4OPt6+neLiYq677joKCwtlfaq3KjQSifD6669z2WWXMWXKFAoKCrAsi3fffZc333yTPXv20L9/f2688UaGDRuG3++nq6uLlStX8uKLLzJkyBCeeOIJ3nrrLVavXs19993Hc889J8PXU08+yd13383DDz4oB3Ls2LE8//zzNDc1UV9fz9y5c2U4LisrY8yYMdx4440cOnQId1YlOn/+fKZNm8YF557LT37yExYtWsTevXvZt3cvjQ0NXH755YwcOVJ6qDPOOIOKsjLq6+t58803cblc3HvvvYwePZq9e/eyf/9+nE4nzU1NPLJoEZMnT+a0006jrKyMVCrF2rVrWbt2LevXr+eiCy4Ay6K9tRWfx8PGjRtpaGhg1qxZ7N+3j5899hjXXXcdhw4dIhmP09XRIfm0ZDzeo+4rDKl3i2yRsVuWdWx6qpaWFsrLy0kmk7S2tjJ//nyuvvpqfv3rX7P7s884//zz+c///E8uv/xy9uzZw3333Uc0GmXJkiUoisLMmTN56KGHOPvss7n22mtpbW1l6NChvPLKK9y1YAF///vfJQ0wb948Bg8ezI9//GN0Xc/wQlnCNJdYzOXKhg0bxjXXXMNdd93FqlWrKCoq4qmnnsLlcjFnzhzuvvtuioqKmDdvHvv27eP000/nT3/6E+vXrycYDPLaa6/xy1/+EkVRuPfee6moqKCgoEC2JHr11VdlZ71UKsXcuXNlD1Kfz0d1dTU7d+6U3FBzc7MkUi3LoqqqilAoxB133MG+ffu4+eab+Zd/+RdGjx5Nd3c3TU1NPProozz++ONomsabb77JsGHDWL58OatXr8bj8TB37lxuvPFGNm7cyPPPP080GqW4uFjyWj/84Q9ZvHgxTzzxBIFAgIcffhin08mKFSt49NFHuf/++1m8eDFbt25lyZIlTJ8+XXJriUSCDz74QN7voEGDmDBhwrHpqY6noGyaJsFgENM06ezsxOfzUVRURCKRoLGxkdbWVvbv34+aFX8J6UQikWDnzp2SligrK5PVdZfLJfuhi+ZlqVQKh8PBggUL+Pzzz+Vf8zz55JP505/+RF5enuRTxO/E91988QXbt2/nO9/5DkOGDCEYDNLW1sb777+PpmmsWrWKmTNncu2119Lc3MyoUaNYvnw5e/fu5cCBA6xatUoWxzdv3szevXsZMGAAhmGQl5fHJ598InFmMplkzZo1eDweCgoK8Pv9tLW1sXv3bindFfhTGN6hQ4dYsGCB3KFjGAYvv/yyNAixG7qjo0NSK3v37pV0jG3bNDc3c+utt8oNpP369ePDDz+UVMbatWupqqriggsuIBQK0dDQwBdffIHL5aK9vZ2bb75ZbtgQbH19fT35+fnU19ezdetWfD4fTU1NjBs3jiFDhuB2u+Wew3+oXDWwuuZri2wul4ukqAtlu5uMGDGC2bNnM3ToUKzsz//nf/6H1157jc7OTqZMmcK0adMYOnQopmnS1tbG7373O95//32p3amsrGTevHkszmrdRdwWxuLI2USwcOFCli1bxnvvvSd3MAsSz+l00tXVRWFhIRMmTKCkpIR0Os3GjRvZs2ePlNL269ePsWPHUl5ezs6dO9m+fTsdHR2yP1YuVtB1ndLSUhKJBM3NzZL7EX/9qn///qRSKZqamigqKaG6ulpu0BQM++eff05HRweOLA4RZSnDMORfuKisrKSjowPDMOju7qaqqgqn00ljYyNlZWW0t7fLaoKoWhQXF0vMNmTIEPbs2UM4HKasrIyJEydSlP1TIx988AEfffQRHo9HZvYA3d3dqKpKIBCgqqqKTZs2UVVVlWkSF4/LXqEDBgzg008/7dmJ+ShUwlH1bH0ZlWidk7vDtbm5mYKCAvr374+iaUQiEZqbm6XHcTqd+P1+QqEQAO3t7bS0tEhJq/B+kUgEn8dDV1cXeXl5EiOJnRwyRdV12apH1BaFa04kEgSDQVkmEu0eI5EIRUVFRCIRyYK3t7dLplvXdYLBoOy1LrKXQCBAPB6ns7NTUiki48zPz5c93ePxOIWFhUTjcVmWyt0YAWRKUFnyUFAzovwh+liJJiCCc8stWwm9lEhUxCLr7u7G6/XKslAwGCQejxOJRGRNNpVKUVhYKK8liG1FUeT4iD7ykUgEj8cjgXhnZyeubD/W4zEqrSAUuv/rjEow7GKCNU0jFArh8Xj44osviESjxGIxWaTM3cbe0dFBLBajra2NAQMG0NjYSDAYBDJdXhRFwZ0lOUXdqrCwkO7u7i/VHcW1o9GoHCi3200gEODQoUPk5+fLARXbp9rb22XHPZ/PJ0Ouw+GguLiYtrY22X3Ptm1ZC0wkEpSWlkojE2y605n5Oy9Cn9/d3Y2m65RnW0iL3SgCHpimiStbeBYgX5wzFosRCoXIy8ujpaVFpuXCI1VWVsqanCClRUnE7/fL6kAgEKChoUH2m6qoqCCRbS8gDDh3K7t4hry8PFkvFTuvhQfVNI28vDza29ulIR6toPxVKo7/D9MIwVcT9HmyAAAAAElFTkSuQmCC"
)
# N. [기계 설계도] 서울 버스 도착을 기록하는 프로그램의 전체 기능 모음
class SeoulBusArrivalRecorder:

    # 1. [그룹 1] 초기단계 — 프로그램 시작 시 한 번만 실행되는 준비 작업 모음

    # 2. [탄생] 프로그램이 처음 켜질 때 실행되는 준비 과정
    def __init__(self, root):
        # 2-1. 화면의 기본 정보들을 설정합니다.
        self.root = root 
        self.root.title("서울버스 정류소 듀얼 도착기록 프로그램 v1.3.88") 
        self.root.geometry("1200x800") 
        # 2-2. 창이 너무 작아지면 화면이 깨지므로 최소 크기를 정합니다.
        self.root.minsize(960, 400) 
        
        # 2-3. 디자인 테마를 설정합니다.
        self.style = ttk.Style() 
        try:
            self.style.theme_use('clam') 
        except:
            pass

        # 2-4. 프로그램이 돌아가는 위치(폴더)를 정확히 찾아냅니다.
        if getattr(sys, 'frozen', False):
            # 2-4-1. 완성된 실행 파일(.exe) 상태일 때 폴더 위치 찾기
            self.resource_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(sys.executable)))
            base_exe_path = os.path.abspath(sys.executable)
            if CURRENT_OS == "Darwin": 
                self.current_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(base_exe_path))))
            else:
                self.current_dir = os.path.dirname(base_exe_path)
        else:
            # 2-4-2. 소스 코드(.py) 상태로 실행 중일 때 폴더 위치 찾기
            self.resource_dir = os.path.dirname(os.path.abspath(__file__))
            self.current_dir = self.resource_dir

        # 2-5. 필요한 파일들의 위치를 미리 약속해둡니다.
        self.key_file_path = os.path.join(self.current_dir, "key.cfg")
        icon_path_png = os.path.join(self.resource_dir, "icon.png")
        icon_path_ico = os.path.join(self.resource_dir, "icon.ico")
        icon_path_icns = os.path.join(self.resource_dir, "icon.icns")

        # 2-6. 윈도우 컴퓨터라면 아이콘이 예쁘게 보이도록 특별 설정을 합니다.
        if CURRENT_OS == "Windows":
            import ctypes
            try:
                myappid = 'Bus_Arrival_Recoder(v1.3.88)' 
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
            except: pass

        # 2-7. 프로그램의 아이콘 이미지를 창에 입힙니다.
        try:
            if os.path.exists(icon_path_png):
                img = tk.PhotoImage(file=icon_path_png)
                self.root.iconphoto(True, img) 
            if CURRENT_OS == "Windows" and os.path.exists(icon_path_ico):
                self.root.iconbitmap(icon_path_ico)
        except Exception as e:
            print(f"아이콘 설정 실패: {e}")

        # 2-8. 화면 제일 위쪽에 메뉴 바 공간을 만듭니다.
        self.top_container = tk.Frame(self.root)
        self.top_container.pack(side="top", fill="x")

        # 2-9. [기억 저장소] 프로그램이 돌아가는 동안 기억해야 할 소중한 정보들
        self.service_key_var = tk.StringVar(value="") # 메인 열쇠(인증키)
        self.backup_key_var = tk.StringVar(value="")  # 비상용 열쇠(백업키)
        self.key_locked = False # 키 입력창 잠금 상태

        # 2-10. 검색 버튼과 확정된 열쇠들을 관리할 공간입니다.
        self.btn_searches = [] 
        self.final_main_key = ""
        self.final_backup_key = ""

        # 2-11. 버스 정류소 정보와 감시 상태를 저장합니다.
        self.ars_ids = [tk.StringVar(), tk.StringVar()] # 정류소 번호 2개
        self.refresh_interval_var = tk.StringVar(value="20") # 갱신 시간 (기본 20초)
        self.is_monitoring = False # 지금 감시 중인지 아닌지 (ON/OFF)
        
        # 2-12. 엑셀 저장 위치와 자동 저장 허용 여부입니다.
        self.auto_save_path = None # 자동 저장할 엑셀 파일 위치
        self.can_auto_save = True # 저장을 해도 되는지 확인하는 스위치
        
        # 2-13. 문제가 생겼을 때 쓸 비상용 API 사용 상태입니다.
        self.use_fallback_api = False # 비상용 API 1번 사용 여부
        self.use_fallback_pos_api = False # 비상용 API 2번 사용 여부
        
        # 2-14. 선택된 정류소 정보와 버스 회사 이름 등을 모아둡니다.
        self.target_st_info = [{}, {}] # 선택한 버스 정보 저장소
        self.recorded_data = [] # 기록된 내용을 잠시 담아두는 바구니
        self.last_arrival_logs = [{}, {}] # 중복 기록을 막기 위한 메모장
        self.route_corp_map = {} # 버스 회사 이름 저장소
        self.rid_to_rnm = {} # 버스 ID를 번호로 바꿔주는 사전
        self.excel_multi_corp_map = {} # 엑셀에서 읽어온 회사 정보
        self.temp_pos_data = {} # POS 데이터를 임시로 저장하는 곳
        # 2-15. 차량번호 캐시: ARR1/ARR2 응답에서 받은 차량번호를 저장해 UID 폴백에 활용합니다.
        #   {(st_id, rid): (plainNo1, plainNo2)}  ← 정류소+노선 조합 키
        self.veh_cache = {}
        # 2-16. POS1 일시정지 테이블: 차량 없음 확인 시 첫차 시각까지 호출을 건너뜁니다.
        #   {rid: datetime}  — 정지 해제 시각 (이 시각이 되면 다시 호출 허용)
        self.pos_suspend_until = {}
        self.pos_resume_logged = set()  # POS 재개 로그 중복 방지: 한 번만 기록
        # 2-17. 마지막 날짜 체크: 자정 이후 suspend 초기화에 사용합니다.
        self._last_date = datetime.now().date()
        # 2-18. SINF 보조 호출 로그: 정류소 인덱스별로 오늘 날짜를 기억해 하루 1회만 출력합니다.
        #   {0: date, 1: date}  — 자정마다 초기화됩니다.
        self._sinf_logged_date = {}
        # 2-19. [macOS 전용 버튼 잠금 방지] 맥OS에서 state=disabled가 클릭을 막지 못하는 버그를 우회합니다.
        #          버그를 우회하기 위해 버튼별 활성화 플래그를 별도 관리합니다.
        #          apply_btn() 호출 시 동기화하고, 각 핸들러 진입 시 guard 체크합니다.
        self._btn_active = {'toggle': False, 'manual': False}
        # 2-20. [노선전체정류소 캐시] 정류소 선택창에서 노선별 전체 정류소 목록을 한 번만 받아 저장해둡니다.
        #   on_station_select 에서 노선당 1회만 호출하고 confirm_selection 에서 재사용합니다.
        #   정류소 선택이 바뀔 때마다 초기화됩니다.
        #   {busRouteId: root_element}
        self._strt_cache = {}
        # 2-21. SLST ord 캐시: on_station_select 에서 미리 구해둔 (ars_id, seq) 를 저장합니다.
        #   {busRouteId: ord_value}  — 정류소가 바뀔 때마다 초기화됩니다.
        self._strt_ord_cache = {}
        # 2-22. 자동 완결 파일 저장용: 이미 완결 파일을 저장한 영업일을 기억합니다.
        #   {"YYYY-MM-DD"} 형태의 집합으로 중복 저장을 방지합니다.
        self._completed_dates_saved = set()
        # 2-23. [스레드 안전] 엑셀 파일 쓰기와 데이터 갱신의 동시 접근을 막는 잠금장치입니다.
        #   _save_lock  : perform_auto_save() / _core_excel_save_logic() 를 한 스레드씩만 실행합니다.
        #   _refresh_lock : refresh_data() 를 한 스레드씩만 실행합니다.
        #   (자동 모니터링 스레드 + 수동 갱신 스레드가 동시에 실행될 때 발생하는
        #    Race Condition 및 엑셀 스냅샷 역전으로 인한 데이터 누락을 방지합니다.)
        self._save_lock = threading.Lock()
        self._refresh_lock = threading.Lock()
        # 2-24. [이중 저장 보호] 마지막으로 엑셀에 성공적으로 저장된 시점의 recorded_data 크기입니다.
        #   refresh_data() 사이클 종료 시 현재 recorded_data 크기와 비교해
        #   미저장 기록이 있으면 자동으로 재저장을 시도합니다 (구제 저장).
        self._saved_record_count = 0

        # 2-25. API를 몇 번 호출했는지 세는 통계 숫자판입니다.
        # 2-26. 합산 통계(메인키 + 백업키 합계) — 메인 화면 상단에 표시됩니다.
        self.api_stats = {
            "ARR1": 0, "ARR2": 0, "SINF": 0, "POS1": 0, "POS2": 0, "SCNM": 0, "SCID": 0, "RINF": 0, "SLST": 0, "VID": 0, "VLD": 0
        }
        # 2-27. 메인/백업 키별 개별 통계 — API 상세 창에서 구분 표시됩니다.
        self.api_stats_by_key = {
            "main": {k: 0 for k in self.api_stats},
            "back": {k: 0 for k in self.api_stats},
        }
        # 2-28. [어제 스냅샷] 자정 초기화 직전의 통계값을 보관합니다.
        #   통계 창 '어제의 합계' 탭에 표시됩니다.
        self.api_stats_yesterday = {k: 0 for k in self.api_stats}
        self.api_stats_by_key_yesterday = {
            "main": {k: 0 for k in self.api_stats},
            "back": {k: 0 for k in self.api_stats},
        }
        
        # 2-29. [API 주소 모음] 서울시 공공 버스 API 11종의 인터넷 주소를 미리 모아둡니다.
        self.api_urls = {
            "ARR1": "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRoute",
            "ARR2": "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRouteAll",
            "POS1": "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRtid",
            "SINF": "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
            "POS2": "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRouteSt",
            "SCNM": "http://ws.bus.go.kr/api/rest/stationinfo/getStationByName",
            "SCID": "http://ws.bus.go.kr/api/rest/stationinfo/getRouteByStation",
            "RINF": "http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo",
            "SLST": "http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute",
            "VID": "http://ws.bus.go.kr/api/rest/buspos/getBusPosByVehId",
            "VLD": "http://ws.bus.go.kr/api/rest/busRouteInfo/getBusRouteList"
        }
        
        self.stats_win = None # 통계 창을 저장할 빈 공간
        # 2-30. [다크 모드 상태] False=라이트(밝은 화면), True=다크(어두운 화면). 초기값은 라이트 모드.
        self.is_dark_mode = False

        # 2-31. [초기화 마무리] 저장된 인증키 불러오기 → 화면 그리기 → 종료 이벤트 연결 → 자정 초기화 예약 → 엑셀 다운로드 예약.
        self.load_saved_key() # 저장된 인증키가 있으면 불러옵니다.
        self.setup_ui()
        # 2-32. [X버튼 연결] 창 닫기(X) 클릭 시 종료 확인 팝업을 표시합니다.
        self.root.protocol("WM_DELETE_WINDOW", self._on_window_close)
        # 2-33. [자정 초기화 예약] 다음 0시 0분까지 남은 시간을 계산해 예약합니다.
        #   ★ 핵심: _do_midnight_reset() 내부에서 이 함수를 다시 호출해야 매일 반복됩니다.
        #     1회성 after() 만 등록하면 첫날만 작동하는 버그가 발생하므로 반드시 재예약합니다.
        self.root.after(200, self._schedule_midnight_reset)
        self.root.after(500, self.start_excel_download_thread) 

    # 3. [그룹 2] 메인 UI — 화면 레이아웃, 버튼 배치, 표 구성 관련 함수 모음

    # 4. [종료 확인] X버튼 클릭 시 정말 종료할지 확인하는 함수
    def _on_window_close(self):
        if messagebox.askyesno("종료 확인", "정말 종료하시겠습니까?"):
            self.root.destroy()

    # 5. [자정 예약] 다음 0시 0분 0초까지 남은 밀리초를 계산해 _do_midnight_reset을 예약합니다.
    #   ★ 이 함수는 _do_midnight_reset() 내부에서도 재호출되어 매일 반복 실행이 보장됩니다.
    def _schedule_midnight_reset(self):
        now = datetime.now()
        from datetime import timedelta
        next_midnight = (now + timedelta(days=1)).replace(
            hour=0, minute=0, second=0, microsecond=0)
        ms_left = int((next_midnight - now).total_seconds() * 1000)
        ms_left = max(ms_left, 1000)  # 자정 직후 즉시 재귀 호출 방지
        self.root.after(ms_left, self._do_midnight_reset)

    # 6. [자정 초기화] 매일 0시 정각에 실행됩니다.
    #   1) 오늘 통계 → 어제 스냅샷으로 복사
    #   2) 오늘 카운터 0 초기화
    #   3) UI 갱신 및 로그 기록
    #   4) ★ 다음날 자정을 위해 _schedule_midnight_reset() 재호출 (없으면 첫날만 실행됨)
    def _do_midnight_reset(self):
        self.api_stats_yesterday = dict(self.api_stats)
        self.api_stats_by_key_yesterday = {
            slot: dict(vals) for slot, vals in self.api_stats_by_key.items()
        }
        for k in self.api_stats:
            self.api_stats[k] = 0
        for slot in self.api_stats_by_key.values():
            for k in slot:
                slot[k] = 0
        self._sinf_logged_date.clear()   # SINF 보조 로그 날짜 초기화 (하루 1회 재설정)
        self.update_api_counter_ui()
        self.log("📅 [자정] API 호출 카운터를 초기화했습니다. 어제의 합계는 통계 창에서 확인할 수 있습니다.")
        # 6-1. ★ 반드시 재예약해야 매일 반복됩니다.
        self._schedule_midnight_reset()


    # 7. [테마 색상] 라이트/다크 팔레트를 딕셔너리로 반환합니다.
    def get_theme(self):
        if self.is_dark_mode:
            return {
                "bg_root":      "#0f0f0f",
                "bg_panel":     "#212121",
                "bg_card":      "#272727",
                "bg_entry":     "#121212",
                "bg_entry_ro":  "#1a1a1a",
                "fg_main":      "#f1f1f1",
                "fg_sub":       "#aaaaaa",
                "fg_accent":    "#3ea6ff",
                "fg_warn":      "#ff4e45",
                "fg_link":      "#5ebe5e",
                "fg_email":     "#3ea6ff",
                "fg_st1":       "#3ea6ff",
                "fg_st2":       "#ff4e45",
                "fg_counter":   "#aaaaaa",
                "fg_counter_v": "#f1f1f1",
                "bg_log":       "#0f0f0f",
                "fg_log":       "#dfe6e9",
                "bg_tree":      "#212121",
                "fg_tree":      "#f1f1f1",
                "bg_tree_sel":  "#3f3f3f",
                "border":       "#3f3f3f",
                "sash":         "#3f3f3f",
                "btn_normal_bg":      "#272727",
                "btn_normal_fg":      "#f1f1f1",
                "btn_normal_active":  "#3f3f3f",
                "btn_normal_border":  "#3f3f3f",
                "btn_outline_bg":     "#272727",
                "btn_outline_fg":     "#3ea6ff",
                "btn_outline_border": "#3ea6ff",
                "btn_disabled_bg":    "#1a1a1a",
                "btn_disabled_fg":    "#555555",
                "btn_disabled_border":"#2a2a2a",
                "toggle_track_off":  "#3f3f3f",
                "toggle_track_on":   "#3ea6ff",
                "toggle_thumb":      "#f1f1f1",
                "toggle_label":      "#aaaaaa",
                "toggle_bg":         "#272727",
                # 7-1-2-1. 스크롤바
                "sb_bg":      "#3f3f3f",   # 스크롤바 thumb 배경
                "sb_trough":  "#1a1a1a",   # 스크롤바 trough(홈) 배경
                "sb_arrow":   "#aaaaaa",   # 스크롤바 화살표 색
                # 7-1-2-2. macOS 입력창 테두리
                "entry_hl":       "#48484a",   # 비활성 테두리
                "entry_hl_focus": "#0a84ff",   # 포커스 테두리
            }
        else:
            return {
                "bg_root":      "#ffffff",
                "bg_panel":     "#ecf0f1",
                "bg_card":      "#f1f2f6",
                "bg_entry":     "#ffffff",
                "bg_entry_ro":  "#f0f0f0",
                "fg_main":      "#2d3436",
                "fg_sub":       "#7f8c8d",
                "fg_accent":    "#0984e3",
                "fg_warn":      "#e74c3c",
                "fg_link":      "#006400",
                "fg_email":     "#2980b9",
                "fg_st1":       "#0984e3",
                "fg_st2":       "#d63031",
                "fg_counter":   "#636e72",
                "fg_counter_v": "#2d3436",
                "bg_log":       "#2d3436",
                "fg_log":       "#dfe6e9",
                "bg_tree":      "#ffffff",
                "fg_tree":      "#000000",
                "bg_tree_sel":  "#0078d7",
                "border":       "#cccccc",
                "sash":         "#cccccc",
                "btn_normal_bg":      "#2ecc71",
                "btn_normal_fg":      "#ffffff",
                "btn_normal_active":  "#27ae60",
                "btn_normal_border":  "#1e8449",
                "btn_outline_bg":     "#eafaf1",
                "btn_outline_fg":     "#27ae60",
                "btn_outline_border": "#27ae60",
                "btn_disabled_bg":    "#e8e8e8",
                "btn_disabled_fg":    "#a0a0a0",
                "btn_disabled_border":"#c0c0c0",
                "toggle_track_off":  "#bbbbbb",
                "toggle_track_on":   "#2ecc71",
                "toggle_thumb":      "#ffffff",
                "toggle_label":      "#7f8c8d",
                "toggle_bg":         "#f1f2f6",
                # 7-1-2-3. 스크롤바
                "sb_bg":      "#c0c0c0",
                "sb_trough":  "#f0f0f0",
                "sb_arrow":   "#606060",
                # 7-1-2-4. macOS 입력창 테두리
                "entry_hl":       "#c7c7c7",   # 비활성 테두리
                "entry_hl_focus": "#007aff",   # 포커스 테두리
            }

    # 8. [테마 적용] is_dark_mode에 맞게 전체 위젯 색상을 갱신합니다.
    def apply_theme(self):
        T = self.get_theme()
        self.root.configure(bg=T["bg_root"])
        self.top_container.configure(bg=T["bg_panel"])

        for w in self._tw.get("status_frames", []):
            try: w.configure(bg=T["bg_panel"])
            except: pass
        for key, w in self._tw.get("status_labels", {}).items():
            try:
                if key == "title_text":  w.configure(bg=T["bg_panel"], fg=T["fg_main"])
                elif key == "title_text3":  w.configure(bg=T["bg_panel"], fg=T["fg_main"])
                elif key == "title_blue":w.configure(bg=T["bg_panel"], fg=T["fg_accent"])
                elif key == "maker":     w.configure(bg=T["bg_panel"], fg=T["fg_sub"])
                elif key == "email":     w.configure(bg=T["bg_panel"], fg=T["fg_email"])
                elif key == "link1":     w.configure(bg=T["bg_panel"], fg=T["fg_link"])
                elif key == "link2":     w.configure(bg=T["bg_panel"], fg="#7fb3ff" if self.is_dark_mode else "#000080")
                elif key == "maker_bracket1": w.configure(bg=T["bg_panel"], fg=T["fg_sub"])
                elif key == "maker_bracket2": w.configure(bg=T["bg_panel"], fg=T["fg_sub"])
            except: pass

        for w in self._tw.get("ctrl_frames", []):
            try: w.configure(bg=T["bg_card"])
            except: pass
        try: self.lbl_auto_save_status.configure(bg=T["bg_panel"], fg=T["fg_warn"])
        except: pass
        for w in self._tw.get("interval_labels", []):
            try: w.configure(bg=T["bg_card"], fg=T["fg_main"])
            except: pass
        try:
            self.entry_refresh_interval.configure(
                bg=T["bg_entry"], fg=T["fg_main"],
                insertbackground=T["fg_main"],
                readonlybackground=T["bg_entry_ro"])
        except: pass
        for w in self._tw.get("key_labels", []):
            try: w.configure(bg=T["bg_card"], fg=T["fg_main"])
            except: pass
        for w in self._tw.get("key_frames", []):
            try: w.configure(bg=T["bg_card"])
            except: pass

        try: self.api_stats_container.configure(bg=T["bg_card"])
        except: pass
        try: self._tw["stats_outer"].configure(bg=T["bg_card"])
        except: pass
        for lbl in self._tw.get("stat_key_labels", []):
            try: lbl.configure(bg=T["bg_card"], fg=T["fg_counter"])
            except: pass
        for key, lbl in self.stat_value_labels.items():
            try: lbl.configure(bg=T["bg_card"], fg=T["fg_counter_v"])
            except: pass

        self._restyle_all_buttons(T)

        for pw in [self.super_paned, self.main_paned]:
            try: pw.configure(bg=T["sash"])
            except: pass

        for w in self._tw.get("tree_frames", []):
            try: w.configure(bg=T["bg_root"])
            except: pass

        for lbl in self.lbl_st_names:
            try: lbl.configure(fg=T["fg_st1"], bg=T["bg_root"])
            except: pass
        for lbl in self.lbl_hist_titles:
            try: lbl.configure(fg=T["fg_st2"], bg=T["bg_root"])
            except: pass
        for w in self._tw.get("ars_entries", []):
            try: w.configure(readonlybackground=T["bg_entry_ro"], fg=T["fg_main"])
            except: pass

        try:
            self.txt_log.configure(bg=T["bg_log"], fg=T["fg_log"])
            self._tw["log_outer"].configure(bg=T["bg_root"])
        except: pass

        self._apply_treeview_style(T)
        self._apply_scrollbar_style(T)
        if CURRENT_OS == "Darwin":
            self._mac_style_all_entries(T)
        self._draw_toggle()

    # 9. [버튼 일괄 재스타일]
    def _restyle_all_buttons(self, T):
        for btn, type_cell in self._tw.get("buttons", []):
            try:
                style = self.get_btn_style(type_cell[0])
                style.pop("state", None)
                btn.configure(**style)
            except: pass

    # 10. [Treeview 테마]
    def _apply_treeview_style(self, T):
        sty = ttk.Style()
        if self.is_dark_mode:
            sty.configure("Treeview",
                background=T["bg_tree"], foreground=T["fg_tree"],
                fieldbackground=T["bg_tree"], rowheight=22)
            sty.configure("Treeview.Heading",
                background=T["bg_card"], foreground=T["fg_main"], relief="flat")
            sty.map("Treeview",
                background=[("selected", T["bg_tree_sel"])],
                foreground=[("selected", T["fg_main"])])
            sty.map("Treeview.Heading",
                background=[("active", T["border"])])
        else:
            sty.configure("Treeview",
                background="#ffffff", foreground="#000000",
                fieldbackground="#ffffff", rowheight=22)
            sty.configure("Treeview.Heading",
                background="#f0f0f0", foreground="#000000", relief="raised")
            sty.map("Treeview",
                background=[("selected", "#0078d7")],
                foreground=[("selected", "#ffffff")])
            sty.map("Treeview.Heading",
                background=[("active", "#e0e0e0")])

    # 11. 5-1-G2. [스크롤바 테마] 앱 전체 ttk.Scrollbar 색상을 테마에 맞게 적용합니다.
    #   - "App.Vertical.TScrollbar" 스타일로 모든 scrollbar에 일괄 적용합니다.
    #   - macOS Aqua 테마에서는 troughcolor/background 적용이 제한적이나
    #     arrowcolor는 반영되며 시각적 차이를 제공합니다.
    def _apply_scrollbar_style(self, T):
        sty = ttk.Style()
        sty.configure("App.Vertical.TScrollbar",
            background=T["sb_bg"],
            troughcolor=T["sb_trough"],
            arrowcolor=T["sb_arrow"],
            bordercolor=T["sb_trough"],
            darkcolor=T["sb_bg"],
            lightcolor=T["sb_bg"],
            relief="flat",
        )
        sty.map("App.Vertical.TScrollbar",
            background=[
                ("active",   T["border"]),
                ("disabled", T["sb_trough"]),
            ])


    # 12. 5-1-G3. [macOS 입력창 스타일] 모든 tk.Entry 에 macOS 기본 스타일 적용
    #   - 라이트: 흰 배경, 연회색 테두리(#c7c7c7), 포커스 시 애플 블루(#007aff)
    #   - 다크  : 어두운 배경(#1c1c1e), 중간회색 테두리(#48484a), 포커스 시 iOS 블루(#0a84ff)
    #   - relief="flat" + highlightthickness=1 로 macOS TextField 느낌 구현
    def _mac_style_entry(self, entry, T):
        try:
            entry.configure(
                relief="flat",
                highlightthickness=1,
                highlightbackground=T.get("entry_hl", "#c7c7c7"),
                highlightcolor=T.get("entry_hl_focus", "#007aff"),
            )
        except Exception:
            pass

    def _mac_style_all_entries(self, T):
        """현재 존재하는 모든 Entry 에 macOS 스타일 일괄 적용."""
        targets = []
        try: targets.append(self.entry_refresh_interval)
        except: pass
        for e in self._tw.get("ars_entries", []):
            targets.append(e)
        for e in targets:
            self._mac_style_entry(e, T)

    # 13. [토글 스위치 생성] — 이모티콘 내장 토글
    def _build_toggle(self, parent):
        T = self.get_theme()
        # 13-1. [토글 크기 설정] 트랙 50×24px. 원래 42×20 비율(2.1:1)을 유지하면서 버튼 높이에 맞게 확대한 크기입니다.
        # 13-2.   42/20=2.1 → 50/24=2.08 (거의 동일 비율).
        TRACK_W, TRACK_H = 50, 24
        self._toggle_canvas = tk.Canvas(
            parent, width=TRACK_W + 6, height=TRACK_H + 6,
            highlightthickness=0,
            cursor="pointinghand" if CURRENT_OS == "Darwin" else "hand2"
        )
        self._toggle_canvas.pack(side="left", padx=(6, 6))
        self._toggle_canvas.bind("<Button-1>", self._on_toggle_theme)
        self._draw_toggle()

    # 14. [토글 스위치 그리기] — 이모티콘 내장형 (☀ / 🌙)
    def _draw_toggle(self):
        if not hasattr(self, "_toggle_canvas"):
            return
        T = self.get_theme()
        c = self._toggle_canvas
        TRACK_W, TRACK_H, THUMB_R = 50, 24, 10   # 비율 유지 확대: 42→50, 20→24, 8→10
        ox, oy = 3, 3
        c.delete("all")
        c.configure(bg=T["bg_card"])   # 컨트롤바(bg_card) 영역으로 이동했으므로
        track_color = T["toggle_track_on"] if self.is_dark_mode else T["toggle_track_off"]
        r = TRACK_H // 2   # = 12

        # ── 둥근 트랙 ───────────────────────────────────────────────────────
        c.create_arc(ox, oy, ox+TRACK_H, oy+TRACK_H,
                     start=90, extent=180, fill=track_color, outline=track_color)
        c.create_arc(ox+TRACK_W-TRACK_H, oy, ox+TRACK_W, oy+TRACK_H,
                     start=270, extent=180, fill=track_color, outline=track_color)
        c.create_rectangle(ox+r, oy, ox+TRACK_W-r, oy+TRACK_H,
                           fill=track_color, outline=track_color)

        # ── 썸(원): 다크=오른쪽, 라이트=왼쪽 ──────────────────────────────
        thumb_cx = (ox+TRACK_W-THUMB_R-3) if self.is_dark_mode else (ox+THUMB_R+3)
        thumb_cy = oy + r
        c.create_oval(thumb_cx-THUMB_R, thumb_cy-THUMB_R,
                      thumb_cx+THUMB_R, thumb_cy+THUMB_R,
                      fill=T["toggle_thumb"], outline=T["toggle_thumb"])

        # ── 이모티콘 : 썸 반대쪽 빈 공간 ──────────────────────────────────
        if self.is_dark_mode:
            emoji_x = ox + r
            emoji   = "🌙"
        else:
            emoji_x = ox + TRACK_W - r
            emoji   = "☀"
        c.create_text(emoji_x, thumb_cy, text=emoji,
                      font=(FONT_MAIN, 8), fill="#ffffff", anchor="center")

        # ── 구 레이블 호환 처리 (이미 없지만 방어) ─────────────────────────
        try:
            if hasattr(self, "_toggle_label"):
                self._toggle_label.configure(text="", bg=T["bg_card"])
        except: pass

    # 15. [토글 클릭]
    def _on_toggle_theme(self, event=None):
        self.is_dark_mode = not self.is_dark_mode
        self.apply_theme()

    # 16. [그림 그리기] 화면의 전체적인 레이아웃과 버튼들을 배치하는 함수
    def setup_ui(self):
        # 16-1. [위젯 추적 저장소] apply_theme()에서 색상을 다시 적용할 때 필요한 모든 위젯을 카테고리별로 기억합니다.
        self._tw = {
            "status_frames": [], "status_labels": {}, "ctrl_frames": [],
            "interval_labels": [], "key_labels": [], "key_frames": [],
            "stat_key_labels": [], "tree_frames": [], "ars_entries": [],
            "buttons": [],
        }

        # 16-2. [테마 색상 로드] 현재 라이트/다크 모드에 맞는 색상 팔레트를 가져옵니다.
        T = self.get_theme()

        # 16-3. [상단 상태 바] 화면 맨 위에 프로그램 사용법 안내 문구와 제작자 정보를 표시하는 영역을 만듭니다.
        frame_status = tk.Frame(self.top_container, pady=2, bg=T["bg_panel"])
        frame_status.pack(fill="x", side="top")
        self._tw["status_frames"].append(frame_status)

        # 16-4. [좌측 안내 영역] "정류소와 노선을 선택하고 자동 기록 시작 버튼을 누르세요." 문구 배치.
        frame_left_status = tk.Frame(frame_status, bg=T["bg_panel"])
        frame_left_status.pack(side="left", padx=10, pady=2)
        self._tw["status_frames"].append(frame_left_status)
        self.lbl_info_container = tk.Frame(frame_left_status, bg=T["bg_panel"])
        self.lbl_info_container.pack(anchor="w", pady=(2, 0))
        self._tw["status_frames"].append(self.lbl_info_container)

        lbl_t1 = tk.Label(self.lbl_info_container, text="정류소와 노선을 선택하고 ",
                          font=(FONT_MAIN, SZ_L, "bold"), fg=T["fg_main"], bg=T["bg_panel"])
        lbl_t1.pack(side="left")
        self._tw["status_labels"]["title_text"] = lbl_t1
        lbl_t2 = tk.Label(self.lbl_info_container, text="자동 기록 시작",
                          font=(FONT_MAIN, SZ_L, "bold"), fg=T["fg_accent"], bg=T["bg_panel"])
        lbl_t2.pack(side="left")
        self._tw["status_labels"]["title_blue"] = lbl_t2
        lbl_t3 = tk.Label(self.lbl_info_container, text=" 버튼을 누르세요.",
                          font=(FONT_MAIN, SZ_L, "bold"), fg=T["fg_main"], bg=T["bg_panel"])
        lbl_t3.pack(side="left")
        # 16-5. [테마 추적 등록] 이 라벨도 apply_theme()에서 색상이 바뀌도록 추적 목록에 등록합니다.
        self._tw["status_labels"]["title_text3"] = lbl_t3

        # ── 안내문구 두 번째 줄: ※ 자동 기록 시작 버튼 안내 ─────────────
        self.lbl_auto_save_status = tk.Label(
            frame_left_status,
            text=" ※ 자동 기록 시작 버튼을 작동시키면 도착 기록이 엑셀파일로 자동 저장됩니다.",
            font=(FONT_SUB, SZ_XS, "bold"), fg=T["fg_warn"], bg=T["bg_panel"]
        )
        self.lbl_auto_save_status.pack(anchor="w", pady=(0, 2))
        frame_right_status = tk.Frame(frame_status, bg=T["bg_panel"])
        frame_right_status.pack(side="right", padx=10, pady=2)
        self._tw["status_frames"].append(frame_right_status)

        # ── 수평 컨테이너: [3줄 정보] + [공공누리 이미지] ───────────────────
        frame_right_horiz = tk.Frame(frame_right_status, bg=T["bg_panel"])
        frame_right_horiz.pack(side="top", anchor="e")
        self._tw["status_frames"].append(frame_right_horiz)

        # ── 좌측: 3줄 정보 (이메일·링크) ────────────────────────────────────
        frame_info_block = tk.Frame(frame_right_horiz, bg=T["bg_panel"])
        frame_info_block.pack(side="left", anchor="e")
        self._tw["status_frames"].append(frame_info_block)

        # 16-6. [제작자 행] "[ 만든이 : 박 국 환 (이메일) ]" 형태의 라벨을 배치합니다.
        frame_right_top = tk.Frame(frame_info_block, bg=T["bg_panel"])
        frame_right_top.pack(side="top", anchor="e")
        self._tw["status_frames"].append(frame_right_top)

        # ── 제작자 문구 ──────────────────────────────────────────────────────
        lbl_m1 = tk.Label(frame_right_top, text=" [ 만든이 : 박 국 환 (",
                           font=(FONT_SUB, SZ_XS), fg=T["fg_sub"], bg=T["bg_panel"])
        lbl_m1.pack(side="left", padx=(5, 0))
        self._tw["status_labels"]["maker_bracket1"] = lbl_m1
        lbl_email = tk.Label(frame_right_top, text="ggoyong2@naver.com",
                             font=(FONT_SUB, SZ_XS, "underline"), fg=T["fg_email"],
                             bg=T["bg_panel"], cursor="hand2")
        lbl_email.pack(side="left")
        lbl_email.bind("<Button-1>", lambda e: self.send_email_link("ggoyong2@naver.com"))
        self._tw["status_labels"]["email"] = lbl_email
        lbl_m2 = tk.Label(frame_right_top, text=") ]",
                           font=(FONT_SUB, SZ_XS), fg=T["fg_sub"], bg=T["bg_panel"])
        lbl_m2.pack(side="left")
        self._tw["status_labels"]["maker_bracket2"] = lbl_m2

        frame_sources = tk.Frame(frame_info_block, bg=T["bg_panel"])
        frame_sources.pack(side="top", anchor="e")
        self._tw["status_frames"].append(frame_sources)
        lbl_link1 = tk.Label(frame_sources,
                             text="공공데이터포털 Open API (https://www.data.go.kr)",
                             font=(FONT_SUB, SZ_XS), fg=T["fg_link"], bg=T["bg_panel"], cursor="hand2")
        lbl_link1.pack(side="top", anchor="e")
        lbl_link1.bind("<Button-1>", lambda e: self.open_link("https://www.data.go.kr"))
        self._tw["status_labels"]["link1"] = lbl_link1
        lbl_link2 = tk.Label(frame_sources,
                             text="서울시 버스운행노선 정보 (https://data.seoul.go.kr)",
                             font=(FONT_SUB, SZ_XS), fg="#000080", bg=T["bg_panel"], cursor="hand2")
        lbl_link2.pack(side="top", anchor="e")
        lbl_link2.bind("<Button-1>", lambda e: self.open_link("https://data.seoul.go.kr/dataList/OA-15066/F/1/datasetView.do"))
        self._tw["status_labels"]["link2"] = lbl_link2

        # ── 우측: 공공누리 이미지 (원본 크기 그대로) ────────────────────────
        try:
            import base64 as _b64s, io as _ios
            _img_data_s = _b64s.b64decode(_GONGGONG_B64)
            try:
                from PIL import Image as _PILs, ImageTk as _PILTks
                _pil_s = _PILs.open(_ios.BytesIO(_img_data_s))
                # 16-6-2-1. 원본 크기 그대로 표시 (149×54)
                self._gonggong_photo_status = _PILTks.PhotoImage(_pil_s)
            except ImportError:
                self._gonggong_photo_status = tk.PhotoImage(data=_GONGGONG_B64)
            lbl_gonggong_s = tk.Label(
                frame_right_horiz, image=self._gonggong_photo_status,
                bg=T["bg_panel"], bd=0, relief="flat"
            )
            lbl_gonggong_s.pack(side="left", padx=(10, 0), anchor="center")
            self._tw["status_frames"].append(lbl_gonggong_s)
        except Exception:
            pass  # 이미지 로드 실패 시 무시

        # 16-7. [컨트롤 바] 인증키 입력, 갱신주기, 자동기록시작, 수동갱신, 엑셀저장, API현황 버튼이 모인 패널.
        self.frame_ctrl_master = tk.Frame(self.top_container, pady=4, bg=T["bg_card"])
        self.frame_ctrl_master.pack(fill="x", side="top")
        self._tw["ctrl_frames"].append(self.frame_ctrl_master)

        frame_main_content = tk.Frame(self.frame_ctrl_master, bg=T["bg_card"])
        frame_main_content.pack(side="left", fill="x", expand=True)
        self._tw["ctrl_frames"].append(frame_main_content)

        PAD = 5
        # 16-8. [버튼 공통 패딩] 윈도우: 상하 4px, 좌우 6px / 맥OS: 0px (맥은 자체 여백 사용).
        BTN_IPADY = 0 if CURRENT_OS == "Darwin" else 4   # 상하 내부 여백 (높이 기준)
        BTN_IPADX = 0 if CURRENT_OS == "Darwin" else 6   # 좌우 내부 여백 (문구 기준)

        # ══════════════════════════════════════════════════════════════════
        # 16-9. [1행 레이아웃]
                #   좌측: [인증키 입력/변경] 버튼
                #   중앙: 갱신주기 입력창 + 자동기록시작 + 수동갱신 + 다른이름으로 엑셀 저장
                #   우측: API 호출 현황 버튼 + 실시간현황/도착기록/기타 카운터 + 라이트/다크 토글
        #   좌=[인증키 버튼] + [갱신주기 + 자동기록시작 + 수동갱신 + 다른이름저장]
        #   우=[API 호출 현황 버튼] + [실시간현황/도착기록/기타 카운터] + [공공누리 이미지]
        # ══════════════════════════════════════════════════════════════════
        frame_row2 = tk.Frame(frame_main_content, bg=T["bg_card"])
        frame_row2.pack(fill="x", side="top", pady=(2, 1))
        self._tw["ctrl_frames"].append(frame_row2)

        # ── 좌측: [인증키] 버튼 ─────────────────────────────────────────
        left_r2 = tk.Frame(frame_row2, bg=T["bg_card"])
        left_r2.pack(side="left", padx=(5, 0), fill="y")
        self._tw["key_frames"].append(left_r2)

        BTN_H = 1
        self.btn_key_manage = AnyButton(
            left_r2, text="인증키 입력",
            command=self.open_key_dialog,
            width=0 if CURRENT_OS == "Darwin" else 10,
            height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_key_manage.pack(side="left", padx=(0, 6), fill="y",
                                 ipadx=BTN_IPADX, ipady=BTN_IPADY)
        self._tw["buttons"].append((self.btn_key_manage, ["normal"]))

        # ── 중앙: 갱신주기 + 자동기록시작 + 수동갱신 + 다른이름저장 ──────
        mid_r2 = tk.Frame(frame_row2, bg=T["bg_card"])
        mid_r2.pack(side="left", fill="y")
        self._tw["ctrl_frames"].append(mid_r2)

        lbl_ival = tk.Label(mid_r2, text="갱신주기(초):", bg=T["bg_card"],
                            font=(FONT_MAIN, SZ_S, "bold"), fg=T["fg_main"])
        lbl_ival.pack(side="left", padx=(0, 2))
        self._tw["interval_labels"].append(lbl_ival)

        self.entry_refresh_interval = tk.Entry(
            mid_r2, textvariable=self.refresh_interval_var,
            width=4, bg=T["bg_entry"], fg=T["fg_main"],
            insertbackground=T["fg_main"], font=(FONT_MONO, SZ_S),
            readonlybackground=T["bg_entry_ro"]
        )
        self.entry_refresh_interval.pack(side="left", padx=(0, PAD))
        if CURRENT_OS == "Darwin":
            self._mac_style_entry(self.entry_refresh_interval, T)

        self.btn_toggle = AnyButton(
            mid_r2, text="자동 기록 시작",
            command=self._on_toggle_monitoring,
            width=0 if CURRENT_OS == "Darwin" else 11,
            height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_toggle.pack(side="left", padx=(0, PAD),
                             ipadx=BTN_IPADX, ipady=BTN_IPADY)
        self._tw["buttons"].append((self.btn_toggle, ["normal"]))

        self.btn_manual = AnyButton(
            mid_r2, text="수동 갱신",
            command=self.manual_refresh,
            width=0 if CURRENT_OS == "Darwin" else 7,
            height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_manual.pack(side="left", padx=(0, PAD),
                             ipadx=BTN_IPADX, ipady=BTN_IPADY)
        self._tw["buttons"].append((self.btn_manual, ["normal"]))

        self.btn_save_excel = AnyButton(
            mid_r2, text="다른 이름으로 엑셀 저장",
            command=self.save_to_excel, height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_save_excel.pack(side="left", padx=(0, 8),
                                 ipadx=BTN_IPADX, ipady=BTN_IPADY)
        self._tw["buttons"].append((self.btn_save_excel, ["normal"]))

        # ── 우측: API 호출 현황 버튼 + 카운터 + 토글 ──────────────────────
        right_r2 = tk.Frame(frame_row2, bg=T["bg_card"])
        right_r2.pack(side="right", padx=(0, 5), fill="y")
        self._tw["ctrl_frames"].append(right_r2)

        self.btn_api_stats = AnyButton(
            right_r2, text="API 호출 현황",
            command=self.open_api_stats_window,
            width=0 if CURRENT_OS == "Darwin" else 11,
            height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_api_stats.pack(side="left", padx=(0, 4), fill="y",
                                ipadx=BTN_IPADX, ipady=BTN_IPADY)
        self._tw["buttons"].append((self.btn_api_stats, ["normal"]))

        stats_outer = tk.Frame(right_r2, bg=T["bg_card"])
        stats_outer.pack(side="left", fill="y")
        stats_outer.grid_rowconfigure(0, weight=1)
        stats_outer.grid_rowconfigure(1, weight=0)
        stats_outer.grid_rowconfigure(2, weight=1)
        stats_outer.grid_columnconfigure(0, weight=1)
        self._tw["stats_outer"] = stats_outer

        self.api_stats_container = tk.Frame(stats_outer, bg=T["bg_card"])
        self.api_stats_container.grid(row=1, column=0, sticky="")
        self.stat_value_labels = {}
        # 16-10. [카운터 항목] 실시간 현황(도착 조회API 합계) / 도착 기록(위치 조회API 합계) / 기타(나머지) 3개 숫자판.
        stat_layout_3 = [
            ("실시간 현황", "REALTIME"),
            ("도착 기록",   "ARRIVAL"),
            ("기타",        "OTHER"),
        ]
        for c, (label, key) in enumerate(stat_layout_3):
            lbl_key = tk.Label(
                self.api_stats_container,
                text=label, font=(FONT_MONO, SZ_S, "bold"),
                fg=T["fg_counter"], bg=T["bg_card"], anchor="w"
            )
            lbl_key.grid(row=0, column=c*2, padx=(4, 1), pady=2, sticky="w")
            self._tw["stat_key_labels"].append(lbl_key)
            val_lbl = tk.Label(
                self.api_stats_container,
                text="0", font=(FONT_MONO, SZ_S, "bold"),
                fg=T["fg_counter_v"], bg=T["bg_card"], width=5, anchor="w"
            )
            val_lbl.grid(row=0, column=c*2+1, padx=(0, 4), pady=2, sticky="w")
            self.stat_value_labels[key] = val_lbl

        # ── 변경 7: 토글 스위치를 카운터 우측에 배치 ────────────────────
        toggle_wrapper = tk.Frame(right_r2, bg=T["bg_card"])
        toggle_wrapper.pack(side="left", fill="y", padx=(4, 0))
        self._tw["ctrl_frames"].append(toggle_wrapper)
        # 16-11. [세로 중앙 정렬] 토글 스위치가 버튼 높이에 맞게 세로 가운데에 오도록 프레임을 추가합니다.
        toggle_inner = tk.Frame(toggle_wrapper, bg=T["bg_card"])
        toggle_inner.pack(anchor="center", expand=True, fill="y")
        self._tw["ctrl_frames"].append(toggle_inner)
        self._build_toggle(toggle_inner)

        # 16-12. [초기 상태 반영] 화면이 처음 열릴 때 카운터와 버튼 활성 상태를 즉시 업데이트합니다.
        self.update_api_counter_ui()
        self.update_button_states()

        # 16-13. [인증키 로드 후 버튼 표시] 저장된 인증키가 이미 검증된 상태라면 버튼을 "인증키 변경" 스타일로 바꿉니다.
        if self.key_locked:
            _ok = self.get_btn_style("outline")
            _ok.pop("state", None)
            self.btn_key_manage.config(text="인증키 변경", **_ok)
            self._update_btn_tracking(self.btn_key_manage, "outline")

        # 16-14. [분할 패널] 위(실시간 현황 표)와 아래(도착 기록 표 + 로그)의 세로 크기를 드래그로 조절할 수 있는 분할창.
        self.super_paned = tk.PanedWindow(self.root, orient=tk.VERTICAL,
                                          sashrelief=tk.RAISED, sashwidth=6)
        self.super_paned.pack(fill="both", expand=True)
        self.main_paned = tk.PanedWindow(self.super_paned, orient=tk.VERTICAL,
                                         sashrelief=tk.RAISED, sashwidth=6)
        self.setup_trees()
        self.super_paned.add(self.main_paned, stretch="always", minsize=220)

        # 16-15. [로그 창] 화면 맨 아래에 작업 내역을 실시간으로 기록하는 텍스트 창을 만듭니다.
        log_outer_frame = tk.Frame(self.super_paned, padx=10, pady=5, bg=T["bg_root"])
        self._tw["log_outer"] = log_outer_frame
        log_scroll_frame = tk.Frame(log_outer_frame, bg=T["bg_root"])
        log_scroll_frame.pack(fill="both", expand=True)
        self.txt_log = tk.Text(log_scroll_frame, height=10, bg=T["bg_log"], fg=T["fg_log"],
                               font=(FONT_MONO, SZ_S), state="disabled")
        log_scrollbar = ttk.Scrollbar(log_scroll_frame, orient="vertical", command=self.txt_log.yview, style="App.Vertical.TScrollbar")
        self.txt_log.configure(yscrollcommand=log_scrollbar.set)
        self.txt_log.pack(side="left", fill="both", expand=True)
        log_scrollbar.pack(side="right", fill="y")
        self.super_paned.add(log_outer_frame, minsize=60, height=180)

        # 16-16. [초기 스크롤바·표 스타일] 화면이 처음 그려지자마자 스크롤바와 표의 색상을 현재 테마에 맞게 즉시 적용합니다.
        self._apply_scrollbar_style(self.get_theme())

        # 16-17. [맥OS 버튼 크기 고정] 맥OS에서만: 자동기록시작 버튼이 "중지"로 바뀔 때 크기가 줄어들지 않도록 초기 크기를 잠급니다.
        #          "중지"로 전환 시에도 동일한 크기를 유지합니다.
        if CURRENT_OS == "Darwin":
            self.root.after_idle(self.btn_toggle.lock_size)

    # 17. [버스 정보판 만들기] 실시간 버스 위치와 도착 기록을 보여주는 표 생성 함수
    def setup_trees(self):
        T = self.get_theme()
        # 17-1. [실시간 현황 표 2개] 정류소1, 정류소2 각각의 버스 도착 예정 시간을 보여주는 표를 화면 상단에 만듭니다.
        frame_rt_container = tk.Frame(self.main_paned, bg=T["bg_root"])
        self._tw["tree_frames"].append(frame_rt_container)
        rt_inner = tk.Frame(frame_rt_container, bg=T["bg_root"])
        rt_inner.pack(fill="both", expand=True)
        self._tw["tree_frames"].append(rt_inner)
        rt_inner.grid_columnconfigure(0, weight=1); rt_inner.grid_columnconfigure(1, weight=1)
        rt_inner.grid_rowconfigure(0, weight=1)

        self.btn_searches = []
        self.trees_rt = []; self.lbl_st_names = []

        for i in range(2):
            f = tk.Frame(rt_inner, bg=T["bg_root"])
            f.grid(row=0, column=i, sticky="nsew", padx=2)
            self._tw["tree_frames"].append(f)
            header = tk.Frame(f, bg=T["bg_root"])
            header.pack(fill="x", pady=2)
            self._tw["tree_frames"].append(header)
            inner_header = tk.Frame(header, bg=T["bg_root"])
            inner_header.pack(anchor="center")
            self._tw["tree_frames"].append(inner_header)

            lbl = tk.Label(inner_header, text=f"[정류소 {i+1}] 실시간 현황",
                           fg=T["fg_st1"], bg=T["bg_root"],
                           font=(FONT_SUB, SZ_M, "bold"))
            lbl.pack(side="left")
            self.lbl_st_names.append(lbl)

            ars_entry = tk.Entry(inner_header, textvariable=self.ars_ids[i],
                                 width=10, state="readonly",
                                 readonlybackground=T["bg_entry_ro"],
                                 fg=T["fg_main"], font=(FONT_MONO, SZ_M, "bold"))
            ars_entry.pack(side="left", padx=5)
            if CURRENT_OS == "Darwin":
                self._mac_style_entry(ars_entry, T)
            self._tw["ars_entries"].append(ars_entry)

            _s_search = self.get_btn_style("disabled")
            _s_search.pop("state", None)
            btn_search = AnyButton(inner_header, text="검색",
                                   command=lambda idx=i: self.open_search_window(idx),
                                   width=0 if CURRENT_OS == "Darwin" else 5,
                                   state="disabled", **_s_search)
            btn_search.pack(side="left", padx=(4, 0),
                            ipady=0 if CURRENT_OS == "Darwin" else 1)
            self.btn_searches.append(btn_search)
            self._tw["buttons"].append((btn_search, ["disabled"]))

            tree_frame = tk.Frame(f, bg=T["bg_root"])
            tree_frame.pack(fill="both", expand=True)
            self._tw["tree_frames"].append(tree_frame)
            cols = ("route", "bus1_no", "bus1_msg", "bus2_no", "bus2_msg")
            tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
            for col in cols: tree.heading(col, text=self.get_col_name(col))
            tree.column("route", width=69); tree.column("bus1_no", width=86)
            tree.column("bus1_msg", width=130); tree.column("bus2_no", width=86)
            tree.column("bus2_msg", width=129)
            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview, style="App.Vertical.TScrollbar")
            tree.configure(yscrollcommand=vsb.set)
            tree.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")
            self.trees_rt.append(tree)
        self.main_paned.add(frame_rt_container, minsize=100)

        # 17-2. [도착 기록 표 2개] 버스가 실제로 도착한 기록을 누적하여 보여주는 표를 화면 하단에 만듭니다.
        frame_hist_container = tk.Frame(self.main_paned, bg=T["bg_root"])
        self._tw["tree_frames"].append(frame_hist_container)
        hist_inner = tk.Frame(frame_hist_container, bg=T["bg_root"])
        hist_inner.pack(fill="both", expand=True)
        self._tw["tree_frames"].append(hist_inner)
        hist_inner.grid_columnconfigure(0, weight=1); hist_inner.grid_columnconfigure(1, weight=1)
        hist_inner.grid_rowconfigure(0, weight=1)

        self.trees_hist = []; self.lbl_hist_titles = []
        for i in range(2):
            f = tk.Frame(hist_inner, bg=T["bg_root"])
            f.grid(row=0, column=i, sticky="nsew", padx=2)
            self._tw["tree_frames"].append(f)
            header = tk.Frame(f, bg=T["bg_root"])
            header.pack(fill="x", pady=2)
            self._tw["tree_frames"].append(header)
            inner_header = tk.Frame(header, bg=T["bg_root"])
            inner_header.pack(anchor="center")
            self._tw["tree_frames"].append(inner_header)

            lbl = tk.Label(inner_header, text=f"[정류소 {i+1}] 도착 기록",
                           fg=T["fg_st2"], bg=T["bg_root"],
                           font=(FONT_SUB, SZ_M, "bold"))
            lbl.pack(side="left")
            self.lbl_hist_titles.append(lbl)

            _s_del = self.get_btn_style("normal", font_size=SZ_XXS)
            btn_del = AnyButton(inner_header, text="기록 삭제",
                                command=lambda idx=i: self.clear_history(idx),
                                **_s_del)
            btn_del.pack(side="left", padx=10,
                         ipady=0 if CURRENT_OS == "Darwin" else 1)
            self._tw["buttons"].append((btn_del, ["normal"]))

            tree_frame = tk.Frame(f, bg=T["bg_root"])
            tree_frame.pack(fill="both", expand=True)
            self._tw["tree_frames"].append(tree_frame)
            cols = ("data_time", "route", "veh_no", "corp", "status")
            tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
            for col in cols: tree.heading(col, text=self.get_col_name(col))
            tree.column("data_time", width=135); tree.column("route", width=63)
            tree.column("veh_no", width=94); tree.column("corp", width=135)
            tree.column("status", width=73)
            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview, style="App.Vertical.TScrollbar")
            tree.configure(yscrollcommand=vsb.set)
            tree.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")
            self.trees_hist.append(tree)
        self.main_paned.add(frame_hist_container, minsize=100)

        # 17-3. [표 초기 스타일] 방금 만든 표들에 현재 테마 색상을 즉시 적용합니다.
        self._apply_treeview_style(T)

    # 18. [버튼 옷입히기] 현재 테마와 버튼 타입에 따라 스타일 딕셔너리를 반환합니다.
    def get_btn_style(self, btn_type="normal", font_size=None):
        # 18-1. [구버전 색상 코드 변환] 예전에 "#28a745" 같은 색상 코드로 버튼 타입을 지정하던 방식을 "normal", "outline" 등으로 자동 변환합니다.
        _color_map = {
            "#28a745": "normal", "#007bff": "normal", "#8e44ad": "normal",
            "#f1c40f": "normal", "#2ecc71": "normal",
            "#d63031": "outline", "#7f8c8d": "outline",
        }
        if btn_type.startswith("#"):
            btn_type = _color_map.get(btn_type, "normal")

        target_sz = font_size if font_size else SZ_S
        T = self.get_theme()
        CURSOR_HAND = "pointinghand" if CURRENT_OS == "Darwin" else "hand2"

        if CURRENT_OS == "Darwin":
            # ──────────────────────────────────────────────────────────────────
            # 18-1-1. [맥OS 버튼 스타일] MacButton은 배경색을 직접 제어할 수 있으므로 정밀한 색상 값을 사용합니다.
            # 18-1-2.   얇은 테두리(1px)를 적용하여 맥OS 스타일의 버튼처럼 보이게 합니다.
            # ──────────────────────────────────────────────────────────────────
            if not self.is_dark_mode:
                _bdr_n = "#c7c7c7"   # 라이트 normal 테두리
                _bdr_o = "#065fd4"   # 라이트 outline 테두리 (파랑)
                _bdr_d = "#d0d0d0"   # 라이트 disabled 테두리
                if btn_type == "disabled":
                    return {
                        "bg": "#e5e5e5", "fg": "#888888",
                        "activebackground": "#e5e5e5", "activeforeground": "#888888",
                        "disabledforeground": "#888888",
                        "font": (FONT_MAIN, target_sz, "normal"), "cursor": "arrow",
                        "relief": "solid", "bd": 0,
                        "highlightthickness": 1, "highlightbackground": _bdr_d,
                        "padx": 2, "pady": 0, "state": "normal",
                    }
                elif btn_type == "outline":
                    return {
                        "bg": "#f2f2f2", "fg": "#065fd4",
                        "activebackground": "#e0e0e0", "activeforeground": "#065fd4",
                        "font": (FONT_MAIN, target_sz, "bold"), "cursor": CURSOR_HAND,
                        "relief": "solid", "bd": 0,
                        "highlightthickness": 1, "highlightbackground": _bdr_o,
                        "padx": 2, "pady": 0,
                    }
                else:
                    return {
                        "bg": "#f2f2f2", "fg": "#0f0f0f",
                        "activebackground": "#e0e0e0", "activeforeground": "#0f0f0f",
                        "font": (FONT_MAIN, target_sz, "bold"), "cursor": CURSOR_HAND,
                        "relief": "solid", "bd": 0,
                        "highlightthickness": 1, "highlightbackground": _bdr_n,
                        "padx": 2, "pady": 0,
                    }
            else:
                _bdr_n = "#555555"   # 다크 normal 테두리
                _bdr_o = "#3ea6ff"   # 다크 outline 테두리 (밝은 파랑)
                _bdr_d = "#333333"   # 다크 disabled 테두리
                if btn_type == "disabled":
                    return {
                        "bg": "#1a1a1a", "fg": "#555555",
                        "activebackground": "#1a1a1a", "activeforeground": "#555555",
                        "disabledforeground": "#555555",
                        "font": (FONT_MAIN, target_sz, "normal"), "cursor": "arrow",
                        "relief": "solid", "bd": 0,
                        "highlightthickness": 1, "highlightbackground": _bdr_d,
                        "padx": 2, "pady": 0, "state": "normal",
                    }
                elif btn_type == "outline":
                    return {
                        "bg": "#272727", "fg": "#3ea6ff",
                        "activebackground": "#3f3f3f", "activeforeground": "#3ea6ff",
                        "font": (FONT_MAIN, target_sz, "bold"), "cursor": CURSOR_HAND,
                        "relief": "solid", "bd": 0,
                        "highlightthickness": 1, "highlightbackground": _bdr_o,
                        "padx": 2, "pady": 0,
                    }
                else:
                    return {
                        "bg": "#272727", "fg": "#f1f1f1",
                        "activebackground": "#3f3f3f", "activeforeground": "#f1f1f1",
                        "font": (FONT_MAIN, target_sz, "bold"), "cursor": CURSOR_HAND,
                        "relief": "solid", "bd": 0,
                        "highlightthickness": 1, "highlightbackground": _bdr_n,
                        "padx": 2, "pady": 0,
                    }
        else:
            # 18-1-3. [윈도우/리눅스 버튼 스타일] 기본 tk.Button의 relief, bd 설정으로 윈도우 스타일 버튼을 만듭니다.
            if btn_type == "disabled":
                return {
                    "bg": T["btn_disabled_bg"],
                    "fg": T["btn_disabled_fg"],
                    "font": (FONT_MAIN, target_sz, "normal"),
                    "relief": "groove", "bd": 2,
                    "highlightbackground": T["btn_disabled_border"],
                    "highlightthickness": 1,
                    "cursor": "arrow",
                    "state": "disabled",
                    "disabledforeground": T["btn_disabled_fg"],
                }
            elif btn_type == "outline":
                return {
                    "bg": T["btn_outline_bg"],
                    "fg": T["btn_outline_fg"],
                    "activebackground": T["btn_outline_bg"],
                    "activeforeground": T["btn_outline_fg"],
                    "font": (FONT_MAIN, target_sz, "bold"),
                    "relief": "ridge", "bd": 2,
                    "highlightbackground": T["btn_outline_border"],
                    "highlightthickness": 1,
                    "cursor": CURSOR_HAND,
                    "overrelief": "solid",
                }
            else:  # normal
                return {
                    "bg": T["btn_normal_bg"],
                    "fg": T["btn_normal_fg"],
                    "activebackground": T["btn_normal_active"],
                    "activeforeground": T["btn_normal_fg"],
                    "font": (FONT_MAIN, target_sz, "bold"),
                    "relief": "raised", "bd": 2,
                    "highlightbackground": T["btn_normal_border"],
                    "highlightthickness": 1,
                    "cursor": CURSOR_HAND,
                    "overrelief": "solid",
                }

    def get_col_name(self, code):
        mapping = {"route": "노선", "bus1_no": "1번차량", "bus1_msg": "도착정보", "bus2_no": "2번차량", "bus2_msg": "도착정보", "data_time": "데이터 시각", "veh_no": "차량번호", "corp": "운수사명", "status": "상태"}
        return mapping.get(code, code)
    
    # 19. [링크 열기] 인터넷 주소를 웹 브라우저로 열어주는 함수
    def open_link(self, url):
        webbrowser.open_new(url)
        
    # 20. [메일 보내기] 클릭하면 바로 이메일을 보낼 수 있게 창을 띄우는 함수
    def send_email_link(self, email):
        webbrowser.open_new(f"mailto:{email}")

    # 21. [칸 고정] 표의 칸 너비를 실수로 바꾸지 못하게 고정하는 함수
    def prevent_column_resize(self, event):
        if event.widget.identify_region(event.x, event.y) == "separator":
            return "break" 

    # 22. [API 호출 통계 창] API 종류별로 오늘과 어제의 호출 횟수를 메인키/백업키 구분하여 보여주는 별도 창을 엽니다.
    def open_api_stats_window(self):
        # 22-1. [창 중복 방지] 통계 창이 이미 열려있으면 새 창을 만들지 않고 기존 창을 맨 앞으로 가져옵니다.
        if self.stats_win is not None and self.stats_win.winfo_exists():
            self.stats_win.lift()
            self.stats_win.focus_force()
            return

        T = self.get_theme()

        # 22-2. [통계 창 생성] 새 Toplevel 창을 만들고 제목, 크기, 배경색을 설정합니다.
        self.stats_win = tk.Toplevel(self.root)
        self.stats_win.title("API 호출 상세 현황")
        self.stats_win.geometry("785x460")
        self.stats_win.minsize(735, 420)
        self.stats_win.transient(self.root)
        self.stats_win.protocol("WM_DELETE_WINDOW", self.on_stats_win_close)
        self.stats_win.configure(bg=T["bg_panel"])

        # 22-3. [제목 레이블] "API 호출 상세 통계" 제목을 창 상단에 표시합니다.
        lbl_title = tk.Label(self.stats_win,
                 text="API 호출 상세 통계",
                 font=(FONT_MAIN, SZ_M, "bold"),
                 bg=T["bg_panel"], fg=T["fg_main"], pady=8)
        lbl_title.pack(side="top", fill="x")

        # 22-4. [탭 스타일 적용] 오늘/어제 탭 스타일을 현재 테마(라이트/다크)에 맞게 설정합니다.
        sty = ttk.Style()
        if self.is_dark_mode:
            sty.configure("StatsNB.TNotebook",
                          background=T["bg_panel"], borderwidth=0)
            sty.configure("StatsNB.TNotebook.Tab",
                          background=T["bg_card"], foreground=T["fg_sub"],
                          padding=[8, 3])
            sty.map("StatsNB.TNotebook.Tab",
                    background=[("selected", T["bg_tree_sel"])],
                    foreground=[("selected", T["fg_main"])])
        else:
            sty.configure("StatsNB.TNotebook",
                          background="#ecf0f1", borderwidth=0)
            sty.configure("StatsNB.TNotebook.Tab",
                          background="#dfe6e9", foreground="#2d3436",
                          padding=[8, 3])
            sty.map("StatsNB.TNotebook.Tab",
                    background=[("selected", "#ffffff")],
                    foreground=[("selected", "#2d3436")])

        notebook = ttk.Notebook(self.stats_win, style="StatsNB.TNotebook")
        notebook.pack(fill="both", expand=True, padx=6, pady=(0, 6))

        # 22-5. [표+스크롤바 생성 헬퍼] 통계 창 내부의 표(Treeview)와 스크롤바를 만드는 내부 함수입니다.
        def make_tree(parent):
            parent.configure(bg=T["bg_panel"])
            cols = ("abbr", "url", "main", "back", "total")
            t = ttk.Treeview(parent, columns=cols, show="headings",
                             style="Stats.Treeview")
            t.heading("abbr",  text="사용 API")
            t.heading("url",   text="API 실제 호출 주소")
            t.heading("main",  text="메인키")
            t.heading("back",  text="백업키")
            t.heading("total", text="합계")
            t.column("abbr",  width=149, anchor="w",      stretch=False)
            t.column("url",   width=395, anchor="w",      stretch=True)
            t.column("main",  width=60,  anchor="center", stretch=False)
            t.column("back",  width=60,  anchor="center", stretch=False)
            t.column("total", width=65,  anchor="center", stretch=False)
            t.bind('<Button-1>', self.prevent_column_resize)
            sb = ttk.Scrollbar(parent, orient="vertical", command=t.yview, style="App.Vertical.TScrollbar")
            t.configure(yscrollcommand=sb.set)
            t.pack(side="left", fill="both", expand=True, padx=(4, 0), pady=4)
            sb.pack(side="right", fill="y", pady=4, padx=(0, 4))
            return t

        # 22-6. 통계 창 전용 표 색상 스타일을 현재 테마에 맞게 적용합니다.
        if self.is_dark_mode:
            sty.configure("Stats.Treeview",
                background=T["bg_tree"], foreground=T["fg_tree"],
                fieldbackground=T["bg_tree"], rowheight=22)
            sty.configure("Stats.Treeview.Heading",
                background=T["bg_card"], foreground=T["fg_main"], relief="flat")
            sty.map("Stats.Treeview",
                background=[("selected", T["bg_tree_sel"])],
                foreground=[("selected", T["fg_main"])])
            sty.map("Stats.Treeview.Heading",
                background=[("active", T["border"])])
        else:
            sty.configure("Stats.Treeview",
                background="#ffffff", foreground="#000000",
                fieldbackground="#ffffff", rowheight=22)
            sty.configure("Stats.Treeview.Heading",
                background="#f0f0f0", foreground="#000000", relief="raised")
            sty.map("Stats.Treeview",
                background=[("selected", "#0078d7")],
                foreground=[("selected", "#ffffff")])
            sty.map("Stats.Treeview.Heading",
                background=[("active", "#e0e0e0")])

        # 22-7. [오늘 탭] 오늘(현재 세션) API 호출 통계를 보여주는 첫 번째 탭.
        tab_today = tk.Frame(notebook, bg=T["bg_panel"])
        notebook.add(tab_today, text="  오늘의 합계  ")
        tree_today = make_tree(tab_today)

        # 22-8. [어제 탭] 자정에 초기화되기 직전에 저장된 어제의 통계를 보여주는 두 번째 탭.
        tab_yest = tk.Frame(notebook, bg=T["bg_panel"])
        notebook.add(tab_yest, text="  어제의 합계  ")
        tree_yest = make_tree(tab_yest)

        # 22-9. [표 데이터 채우기 헬퍼] API 통계 데이터를 표의 각 행에 채워넣는 내부 함수입니다.
        _api_name_map = {
            "ARR1": "실시간 현황 API",
            "ARR2": "실시간 현황 예비 API",
            "SINF": "실시간 현황 예비2 API",
            "POS1": "도착 기록 API",
            "POS2": "도착 기록 예비 API",
            "SCNM": "정류소명 검색 API",
            "SCID": "정류소번호 검색 API",
            "RINF": "노선기본정보 API",
            "SLST": "노선전체정류소 API",
            "VID":  "차량번호확인 API",
            "VLD":  "인증키검증용 API",
        }
        def fill_tree(t, stats_dict, by_key_dict, label_row):
            for item in t.get_children():
                t.delete(item)
            total_all = total_main = total_back = 0
            for key, count in stats_dict.items():
                url   = self.api_urls.get(key, "-")
                m_cnt = by_key_dict["main"].get(key, 0)
                b_cnt = by_key_dict["back"].get(key, 0)
                api_label = _api_name_map.get(key, key)
                t.insert("", "end", values=(api_label, url, f"{m_cnt}", f"{b_cnt}", f"{count}"))
                total_all  += count
                total_main += m_cnt
                total_back += b_cnt
            t.insert("", "end", values=(" ", " ", " ", " ", " "))
            t.insert("", "end", values=(label_row, " - ",
                                        f"{total_main}", f"{total_back}", f"{total_all}"))

        # 22-10. [실시간 갱신 루프] 1초마다 오늘·어제 탭 양쪽을 모두 최신 데이터로 갱신합니다.
        def update_stats_loop():
            if not self.stats_win or not self.stats_win.winfo_exists():
                return
            fill_tree(tree_today, self.api_stats,
                      self.api_stats_by_key, "합계")
            fill_tree(tree_yest, self.api_stats_yesterday,
                      self.api_stats_by_key_yesterday, "합계")
            self.stats_win.after(1000, update_stats_loop)

        update_stats_loop()

    # 23. [수첩 닫기] 통계 창을 닫을 때 메모리에서 지워주는 함수
    def on_stats_win_close(self):
        if self.stats_win:
            self.stats_win.destroy()
            self.stats_win = None

    # 24. [버튼 상태 관리] 현재 상황(인증키 있음/없음, 기록 중/아님)에 따라 각 버튼을 활성 또는 비활성 상태로 업데이트합니다.
    def update_button_states(self):
        # 24-1. [기본 조건 확인] 정류소1이나 정류소2 중 하나라도 설정되어 있는지 확인합니다.
        has_info = (bool(self.target_st_info[0].get('routes')) or
                    bool(self.target_st_info[1].get('routes')))

        # 24-2. [버튼 설정 헬퍼] 버튼 1개의 타입, 활성 여부, 연결 함수, 텍스트를 한 번에 설정하는 내부 함수.
        def apply_btn(btn, btn_type, is_active, cmd, text=None):
            # 24-2-1. [비활성 처리] is_active=False이면 버튼 타입을 강제로 "disabled"로 바꿔 회색 스타일을 적용합니다.
            effective_type = btn_type if is_active else "disabled"
            style = self.get_btn_style(effective_type)
            if text:
                style["text"] = text
            style.pop("state", None)
            # 24-2-2. _tw["buttons"] 추적 셀 갱신 (apply_theme 재스타일용)
            self._update_btn_tracking(btn, effective_type)
            if CURRENT_OS == "Darwin":
                if btn is self.btn_toggle:
                    self._btn_active['toggle'] = is_active
                elif btn is self.btn_manual:
                    self._btn_active['manual'] = is_active
                btn.config(state=("normal" if is_active else "disabled"),
                           command=(cmd if is_active else lambda: None), **style)
            else:
                btn.config(state=("normal" if is_active else "disabled"),
                           command=(cmd if is_active else lambda: None), **style)

        # 24-3. 감시 중일 때 → 통합 버튼을 "중지"(outline), 수동 갱신 활성화
        if self.is_monitoring:
            apply_btn(self.btn_toggle, "outline", True,
                      self.stop_monitoring, text="중지")
            apply_btn(self.btn_manual, "normal", True, self.manual_refresh)
        # 24-4. 감시 중이 아닐 때 → 통합 버튼을 "자동 기록 시작"
        else:
            apply_btn(self.btn_toggle, "normal", has_info,
                      self.start_monitoring, text="자동 기록 시작")
            apply_btn(self.btn_manual, "normal", has_info, self.manual_refresh)

    # 25. [통합 토글] 시작/중지 버튼 클릭 시 현재 상태에 따라 분기합니다.
    def _on_toggle_monitoring(self):
        # 25-1. [맥OS 중복 클릭 방지] 맥OS에서 disabled 상태인 버튼이 클릭되는 버그를 직접 플래그로 차단합니다.
        if CURRENT_OS == "Darwin" and not self._btn_active.get('toggle', True):
            return
        if self.is_monitoring:
            self.stop_monitoring()
        else:
            self.start_monitoring()

    # 26. [카운터 갱신 헬퍼] 어떤 스레드에서 호출해도 안전하게 메인 화면의 API 호출 숫자판을 갱신합니다.
    #   - fetch_api(백그라운드 스레드)·_validate_and_apply_keys(메인 스레드) 양쪽에서 호출됩니다.
    #   - root.after(0, ...)은 tkinter 이벤트 루프에 갱신을 위임하므로 스레드 안전합니다.
    def _trigger_counter_update(self):
        self.root.after(0, self.update_api_counter_ui)

    # 27. [버튼 타입 추적 헬퍼] 테마 변경 시 버튼 색상을 다시 칠할 수 있도록, 버튼의 현재 타입(normal/disabled/outline)을 저장해둡니다.
    #   apply_theme() 재스타일 시 이 타입 셀을 참조하므로 버튼 스타일 변경 후 반드시 호출해야 합니다.
    #   btn      : 추적 대상 버튼 위젯
    #   type_str : 새로 지정할 타입 문자열 ("normal" / "disabled" / "outline" 등)
    def _update_btn_tracking(self, btn, type_str):
        for item in self._tw.get("buttons", []):
            if item[0] is btn:
                item[1][0] = type_str
                break

    # 28. [카운터 숫자 갱신] 화면 상단 카운터의 실시간 현황/도착 기록/기타 숫자를 최신 집계값으로 다시 표시합니다.
    def update_api_counter_ui(self):
        s = self.api_stats
        realtime = s.get("ARR1",0) + s.get("ARR2",0) + s.get("SINF",0)
        arrival  = s.get("POS1",0) + s.get("POS2",0)
        other    = s.get("SCNM",0) + s.get("SCID",0) + s.get("RINF",0)                  + s.get("SLST",0) + s.get("VID",0)  + s.get("VLD",0)
        for key, val_lbl in self.stat_value_labels.items():
            if key == "REALTIME": val_lbl.config(text=str(realtime))
            elif key == "ARRIVAL": val_lbl.config(text=str(arrival))
            elif key == "OTHER":   val_lbl.config(text=str(other))

    # 29. 4그룹 : 보안 및 인증 관리 (Security & Auth)

    # 30. [인증키 자동 로드] 프로그램 시작 시, 이전에 저장해둔 인증키 파일(key.cfg)이 있으면 자동으로 불러옵니다.
    def load_saved_key(self):
        # 30-1. 열쇠가 저장된 파일이 있는지 확인합니다.
        if os.path.exists(self.key_file_path):
            try:
                # 30-1-2-1. 파일을 열어서 암호화된 내용을 읽어옵니다.
                with open(self.key_file_path, "rb") as f:
                    encrypted_data = f.read()
                    if encrypted_data:
                        # 30-1-2-1-0-1. [비밀 풀기] 암호화된 내용을 다시 글자로 바꿉니다.
                        decrypted_data = cipher_suite.decrypt(encrypted_data).decode('utf-8')
                        main_k, back_k = "", ""
                        # 30-1-2-1-0-2. 메인 열쇠와 백업 열쇠가 같이 있다면 나누어서 저장합니다.
                        if "|" in decrypted_data:
                            main_k, back_k = decrypted_data.split("|", 1)
                        else:
                            main_k = decrypted_data
                        
                        # 30-1-2-1-0-3. 화면에 있는 입력창에 불러온 열쇠들을 넣어줍니다.
                        self.service_key_var.set(main_k)
                        self.backup_key_var.set(back_k)
                        
                        # 30-1-2-1-0-4. [자동 검사] 불러온 열쇠가 진짜인지 바로 확인을 시작합니다.
                        if main_k.strip() or back_k.strip():
                            # 30-1-2-1-0-4-1. 로그창이 준비될 때까지 아주 잠시(0.1초) 기다렸다가 메시지를 띄웁니다.
                            self.root.after(100, lambda: self.log("🚀 저장된 인증키를 불러왔습니다. 유효성 검사를 시작합니다..."))
                            self.root.after(150, self.toggle_key_lock)

            except Exception as e:
                # 30-1-2-2. 만약 열쇠를 읽다가 실수를 하면 무엇이 틀렸는지 알려줍니다.
                print(f"인증키 복호화 실패: {e}")

    # 31. [열쇠 저장하기] 입력한 소중한 열쇠(인증키)를 잊어버리지 않게 파일로 잠궈두는 함수
    def save_key_to_file(self):
        # 31-1. 입력창에 적힌 메인 열쇠와 백업 열쇠를 가져옵니다.
        main_k = self.service_key_var.get().strip()
        back_k = self.backup_key_var.get().strip()
        
        # 31-2. 두 개의 열쇠를 하나로 묶어줍니다.
        combined_data = f"{main_k}|{back_k}"
        
        try:
            # 31-2-1. [비밀 잠금] 남들이 못 보게 암호로 바꾸어 파일에 기록합니다.
            encrypted_data = cipher_suite.encrypt(combined_data.encode('utf-8'))
            with open(self.key_file_path, "wb") as f:
                f.write(encrypted_data)
            self.log("🔒 메인/백업 인증키가 암호화되어 안전하게 저장되었습니다.")
        except Exception as e:
            # 31-2-2. 저장하다가 문제가 생기면 로그창에 알립니다.
            self.log(f"⚠ 인증키 암호화 저장 실패: {e}")

    # 32. [인증키 입력 창] 인증키를 새로 입력하거나 변경하는 팝업 창을 띄웁니다. 키를 입력하면 서버에 즉시 유효성 검사를 합니다.
    def open_key_dialog(self):
        """인증키 입력 전용 모달 창을 띄웁니다 (정류소 검색창과 동일한 grab_set 방식)."""
        T = self.get_theme()

        key_win = tk.Toplevel(self.root)
        key_win.title("인증키 입력")
        key_win.grab_set()                     # 메인창 조작 차단 (모달)
        key_win.resizable(False, False)
        key_win.configure(bg=T["bg_card"])

        # ── 창 내용 ────────────────────────────────────────────────────
        pad_f = tk.Frame(key_win, bg=T["bg_card"], padx=18, pady=16)
        pad_f.pack(fill="both", expand=True)

        # 32-1. 현재 키 값 (임시 변수 — 확정 전까지 service_key_var 에 쓰지 않음)
        tmp_main = tk.StringVar(value=self.service_key_var.get())
        tmp_back = tk.StringVar(value=self.backup_key_var.get())

        def _make_row(parent, label_text, var):
            row = tk.Frame(parent, bg=T["bg_card"])
            row.pack(fill="x", pady=4)
            tk.Label(row, text=label_text, font=(FONT_MAIN, SZ_S),
                     bg=T["bg_card"], fg=T["fg_main"], width=10, anchor="e").pack(side="left")
            ent = tk.Entry(row, textvariable=var, show="*",
                           width=68, bg=T["bg_entry"], fg=T["fg_main"],
                           insertbackground=T["fg_main"], font=(FONT_MONO, SZ_XXS),
                           readonlybackground=T["bg_entry_ro"])
            ent.pack(side="left", padx=(6, 0))
            if CURRENT_OS == "Darwin":
                self._mac_style_entry(ent, T)
            return ent

        ent_main = _make_row(pad_f, "메인 인증키 :", tmp_main)
        ent_back = _make_row(pad_f, "백업 인증키 :", tmp_back)

        # 32-2. 안내 문구
        lbl_hint = tk.Label(pad_f,
            text="  ※ 메인 인증키는 64자리 필수, 백업 인증키는 선택(0 또는 64자리)입니다.",
            font=(FONT_SUB, SZ_XS), fg=T["fg_warn"], bg=T["bg_card"])
        lbl_hint.pack(anchor="w", pady=(2, 8))

        # 32-3. 버튼 영역
        btn_row = tk.Frame(pad_f, bg=T["bg_card"])
        btn_row.pack(fill="x")

        _dis_style = self.get_btn_style("disabled")
        _dis_style.pop("state", None)   # ★ state 중복 방지
        btn_apply = AnyButton(btn_row, text="인증키 입력",
                              command=lambda: _do_apply(),
                              state="disabled",
                              **_dis_style)
        btn_apply.pack(side="left", padx=(0, 8),
                       ipady=0 if CURRENT_OS == "Darwin" else 4)

        AnyButton(btn_row, text="취소",
                  command=key_win.destroy,
                  **self.get_btn_style("normal")).pack(
            side="left", ipady=0 if CURRENT_OS == "Darwin" else 4)

        # ── 64자 검증 → 입력 버튼 활성/비활성 ────────────────────────
        def _check_apply_btn(*_):
            ml = len(tmp_main.get().strip())
            bl = len(tmp_back.get().strip())
            ok = (ml == 64) and (bl == 0 or bl == 64)
            if ok:
                _s = self.get_btn_style("normal"); _s.pop("state", None)
                btn_apply.config(state="normal", **_s)
            else:
                _s = self.get_btn_style("disabled"); _s.pop("state", None)
                btn_apply.config(state="disabled", **_s)

        tmp_main.trace_add("write", _check_apply_btn)
        tmp_back.trace_add("write", _check_apply_btn)
        _check_apply_btn()

        # ── 인증키 입력 버튼 로직 ──────────────────────────────────────
        def _do_apply():
            main_input = tmp_main.get().strip()
            back_input = tmp_back.get().strip()
            if not main_input:
                messagebox.showwarning("알림", "메인 인증키를 입력해주세요.", parent=key_win)
                return
            _ds2 = self.get_btn_style("disabled"); _ds2.pop("state", None)
            btn_apply.config(state="disabled", **_ds2)
            key_win.update()

            def _on_bad_key(var_obj):
                var_obj.set("")

            success = self._validate_and_apply_keys(
                main_input, back_input,
                on_bad_main=lambda: _on_bad_key(tmp_main),
                on_bad_back=lambda: _on_bad_key(tmp_back),
                save_on_success=True,
                parent_win=key_win
            )
            if success:
                key_win.destroy()
            else:
                _check_apply_btn()  # 버튼 상태 복원

        # 32-4. 창 크기를 내용에 맞게 조정
        key_win.update_idletasks()
        key_win.geometry("")   # 자동 크기
        ent_main.focus_set()

    # 33. [인증키 공통 검증 핵심 함수] 수동 입력(창에서 입력)과 자동 로드(파일에서 불러오기) 양쪽 경로에서 공통으로 사용하는 인증키 서버 검증 함수.
    def _validate_and_apply_keys(self, main_input, back_input,
                                  on_bad_main=None, on_bad_back=None,
                                  save_on_success=False, parent_win=None):
        """
        메인/백업 인증키를 서버에 검증하고 성공 시 프로그램을 활성화합니다.

        Parameters
        ----------
        main_input      : str  — 검증할 메인 인증키
        back_input      : str  — 검증할 백업 인증키 (비어있으면 건너뜀)
        on_bad_main     : callable | None  — 메인키 실패 시 호출 (예: 입력창 초기화)
        on_bad_back     : callable | None  — 백업키 실패 시 호출
        save_on_success : bool  — 검증 성공 시 key.cfg 에 저장할지 여부
                          수동 입력 경로 → True / 자동 로드 경로 → False
        parent_win      : tk.Toplevel | None  — messagebox 의 부모 창

        Returns
        -------
        bool — 하나 이상의 키가 검증 성공하면 True
        """
        from urllib.parse import unquote
        import requests

        self.log("🔑 인증키 유효성 검사 중...")

        keys_to_test = [("메인", main_input, on_bad_main)]
        if back_input:
            keys_to_test.append(("백업", back_input, on_bad_back))

        valid_main, valid_back = "", ""
        any_success = False

        for name, key_val, bad_cb in keys_to_test:
            test_url = "http://ws.bus.go.kr/api/rest/busRouteInfo/getBusRouteList"
            try:
                r = requests.get(test_url,
                                 params={'serviceKey': unquote(key_val), 'strSrch': '12'},
                                 timeout=5)
                if "<headerCd>0</headerCd>" in r.text:
                    self.log(f"✅ {name} 인증키: 정상 작동 확인")
                    if name == "메인": valid_main = key_val
                    else:             valid_back = key_val
                    any_success = True
                else:
                    if bad_cb: bad_cb()
                    self.log(f"❌ {name} 인증키: 작동 불가하여 삭제되었습니다.")
            except Exception:
                if bad_cb: bad_cb()
                self.log(f"❌ {name} 통신 에러로 삭제 처리되었습니다.")

        if any_success:
            # ── 확정: service_key_var / backup_key_var 에 반영 ────────
            self.service_key_var.set(valid_main)
            self.backup_key_var.set(valid_back)
            self.final_main_key = valid_main
            self.final_backup_key = valid_back
            self.key_locked = True

            # 33-4-1. VLD 카운터 업데이트
            if valid_main:
                self.api_stats['VLD'] += 1
                self.api_stats_by_key["main"]['VLD'] = \
                    self.api_stats_by_key["main"].get('VLD', 0) + 1
            if valid_back:
                self.api_stats['VLD'] += 1
                self.api_stats_by_key["back"]['VLD'] = \
                    self.api_stats_by_key["back"].get('VLD', 0) + 1
            self._trigger_counter_update()

            # 33-4-2. 인증키 버튼 → outline("인증키 변경") 스타일
            _outline = self.get_btn_style("outline"); _outline.pop("state", None)
            self.btn_key_manage.config(text="인증키 변경", **_outline)
            self._update_btn_tracking(self.btn_key_manage, "outline")

            # 33-4-3. 검색 버튼 활성화
            _s_on = self.get_btn_style("normal"); _s_on.pop("state", None)
            for btn in self.btn_searches:
                btn.config(state="normal", **_s_on)
                self._update_btn_tracking(btn, "normal")

            if save_on_success:
                self.save_key_to_file()

            self.log("🔒 인증키가 확정되었습니다. 이제 검색이 가능합니다.")
        else:
            messagebox.showerror("인증 실패", "사용 가능한 인증키가 없습니다.",
                                 **({'parent': parent_win} if parent_win else {}))

        return any_success

    # 34. [자동 로드 전용 검증] 프로그램 시작 시 파일에서 불러온 인증키를 서버에 확인하는 함수. 파일 재저장은 하지 않습니다.
    def toggle_key_lock(self):
        """저장된 키 자동 로드 시 호출되는 무음 유효성 검사."""
        if self.key_locked:
            return  # 이미 잠긴 상태 — 무시

        main_input = self.service_key_var.get().strip()
        back_input  = self.backup_key_var.get().strip()
        if not main_input:
            return

        # 34-1. service_key_var / backup_key_var 에 직접 set 하는 콜백
        self._validate_and_apply_keys(
            main_input, back_input,
            on_bad_main=lambda: self.service_key_var.set(""),
            on_bad_back=lambda: self.backup_key_var.set(""),
            save_on_success=False   # ★ 자동 로드 경로에서는 파일 저장 안 함
        )

    # 35. 5그룹: 데이터 수집 및 통신 (Fetching & API)

    # 36. [API 호출 핵심 함수] 서울시 버스 서버에 인터넷으로 질문을 보내고 응답을 받아오는 가장 중요한 함수입니다.
    def fetch_api(self, url, params, api_type=None):
        from urllib.parse import unquote
        # 36-1. [키 준비] 현재 등록된 메인 인증키와 백업 인증키를 순서대로 준비합니다.
        keys = [self.service_key_var.get().strip(), self.backup_key_var.get().strip()]
        
        # 36-2. [API 주소 선택] 요청 종류에 따라 주 API 주소와 폴백 API 주소를 결정합니다.
                #   ARR1(도착 조회) → 폴백: ARR2(전체 도착 조회) / POS2(구간 위치) → 주 API: POS1(노선 위치)
        p_url, f_url = url, url
        if api_type == "ARR1":
            p_url = "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRoute"
            f_url = "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRouteAll"
        elif api_type == "POS2":
            p_url = "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRtid"
            f_url = "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRouteSt"

        # 36-3. [시도 계획 수립] 주 API × 메인키, 주 API × 백업키, 폴백 API × 메인키, 폴백 API × 백업키 순으로 총 4번 시도합니다.
        attempts = [
            (p_url, "메인", keys[0]), (p_url, "백업", keys[1]), 
            (f_url, "메인", keys[0]), (f_url, "백업", keys[1])
        ]

        # 36-4. [시도 루프] 4가지 시도를 순서대로 실행합니다. 하나라도 성공하면 즉시 결과를 반환합니다.
        for step, (curr_url, name, key) in enumerate(attempts):
            if not key: continue
            try:
                # 36-4-3-1. [API 호출] 인증키를 파라미터에 포함하여 서버에 요청을 보냅니다.
                params['serviceKey'] = unquote(key)
                resp = requests.get(curr_url, params=params, timeout=10)
                
                # 36-4-3-2. [일일 초과 처리] 하루 호출 한도를 초과하면 서버가 거절합니다. 다음 키/URL 조합으로 넘어갑니다.
                if "LIMITED NUMBER OF SERVICE REQUESTS" in resp.text or "<headerCd>22</headerCd>" in resp.text:
                    continue 

                # 36-4-3-3. [잘못된 키 처리] 등록되지 않은 인증키라는 응답이 오면 다음 시도로 넘어갑니다.
                if "SERVICE KEY IS NOT REGISTERED" in resp.text or "UNREGISTERED_KEY" in resp.text:
                    continue

                # 36-4-3-4. [응답 분석] HTTP 200 응답이 왔을 때 XML 내용을 파싱하여 성공 여부를 판정합니다.
                if resp.status_code == 200:
                    root = ET.fromstring(resp.text)
                    header_cd = root.findtext(".//headerCd")
                    err_msg = root.findtext('.//headerMsg') or ""
                    
                    # 36-4-3-4-1. [성공 판정] 데이터가 있거나(headerCd=0), 데이터가 없다는 정상 응답(NODATA)이면 호출 성공으로 처리합니다.
                    is_no_result = ("결과가 없습니다" in err_msg) or ("NODATA" in err_msg)
                    
                    if header_cd == "0" or is_no_result:
                        # 36-4-3-4-1-1. [API 종류 판별] 응답받은 URL을 확인해 어떤 API를 호출했는지 판별하고, 해당 종류의 호출 카운터를 1 올립니다.
                        success_url = curr_url
                        key_type = "VLD"  # 기본값: 알 수 없는 호출
                        if "getArrInfoByRoute" in success_url and "All" not in success_url: key_type = "ARR1"
                        elif "getArrInfoByRouteAll" in success_url: key_type = "ARR2"
                        elif "getBusPosByRtid" in success_url: key_type = "POS1"
                        elif "getStationByUid" in success_url: key_type = "SINF"
                        elif "getBusPosByRouteSt" in success_url: key_type = "POS2"
                        elif "getStationByName" in success_url: key_type = "SCNM"
                        elif "getRouteByStation" in success_url: key_type = "SCID"
                        elif "getRouteInfo" in success_url: key_type = "RINF"
                        elif "getStaionByRoute" in success_url: key_type = "SLST"
                        elif "getBusPosByVehId" in success_url: key_type = "VID"
                        
                        # 36-4-3-4-1-2. [전체 합계 카운터 +1] 메인키+백업키 통합 카운터에 1을 더합니다.
                        self.api_stats[key_type] += 1
                        # 36-4-3-4-1-3. [개별 키 카운터 +1] 메인키 또는 백업키 중 어떤 키가 성공했는지 개별 카운터에도 1을 더합니다.
                        key_slot = "main" if name == "메인" else "back"
                        self.api_stats_by_key[key_slot][key_type] = \
                            self.api_stats_by_key[key_slot].get(key_type, 0) + 1
                        self._trigger_counter_update()

                        # 36-4-3-4-1-4. [결과 반환] 정상 데이터(headerCd=0)면 XML 루트 요소를 반환합니다. 데이터 없음이면 None 또는 특수 튜플을 반환합니다.
                        if header_cd == "0":
                            return root
                        else:
                            # 36-4-3-4-1-4-1. [운행 없음 처리] 위치 조회API에서 운행 차량 없음 응답 시, 첫차 시각을 함께 반환합니다. 호출자는 ('NO_BUS', 첫차시각) 형태를 처리해야 합니다.
                            #   첫차 시간(f_tm)을 함께 반환합니다. — 호출자는 (None, f_tm) 형태를 처리해야 합니다.
                            if any(u in success_url for u in ["getBusPosByRtid", "getBusPosByRouteSt"]):
                                rid_param = params.get('busRouteId', '')
                                rnm_log = self.rid_to_rnm.get(rid_param, rid_param)
                                self.log(f"{rnm_log}번은 운행중인 차량이 없습니다.")
                                # 36-4-3-4-1-4-1-1. 첫차 시간을 노선정보 API(getRouteInfo)에서 가져옵니다.
                                f_tm = self._fetch_route_first_time(rid_param)
                                return ('NO_BUS', f_tm)
                            return None
                        
            except Exception:
                # 36-4-3-5. [예외 처리] 네트워크 오류 등 예외가 발생하면 해당 시도를 건너뛰고 다음 시도로 진행합니다.
                continue
        return None
    
    # 37. [첫차 시각 조회] 특정 노선의 첫 번째 버스 출발 시각을 노선정보 API(노선상세정보 조회)에서 가져옵니다.
    def _fetch_route_first_time(self, rid):
        """getRouteInfo에서 firstBusTm 필드를 가져옵니다 (HHmm 또는 HHmmss 형식).
        실패 또는 없는 오전 시리즈 노선은 None 반환."""
        if not rid:
            return None
        try:
            root_ri = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo",
                {'busRouteId': rid}
            )
            if root_ri is None or isinstance(root_ri, tuple):
                return None
            f_tm_raw = root_ri.findtext(".//firstBusTm") or ""
            if f_tm_raw:
                return self.format_hhmm(f_tm_raw)  # HH:MM 형식
        except Exception:
            pass
        return None

    # 38. [버스 정류소 검색 창] 버스 정류소를 검색하고, 해당 정류소에서 모니터링할 버스 노선을 선택하는 팝업 창을 엽니다.
    def open_search_window(self, target_idx):
        # 38-1. [인증키 확인] 인증키가 없으면 버스 정보를 조회할 수 없으므로 경고 메시지를 보여주고 종료합니다.
        if not self.service_key_var.get().strip():
            messagebox.showwarning("알림", "인증키를 먼저 입력해주세요."); return

        T = self.get_theme()

        # ── Search 전용 ttk.Style ─────────────────────────────────────────
        sty = ttk.Style()
        if self.is_dark_mode:
            sty.configure("Search.Treeview",
                background=T["bg_tree"], foreground=T["fg_tree"],
                fieldbackground=T["bg_tree"], rowheight=22)
            sty.configure("Search.Treeview.Heading",
                background=T["bg_card"], foreground=T["fg_main"],
                font=(FONT_SUB, SZ_S, "bold"), relief="flat")
            sty.map("Search.Treeview",
                background=[("selected", T["bg_tree_sel"])],
                foreground=[("selected", T["fg_main"])])
            sty.map("Search.Treeview.Heading",
                background=[("active", T["border"])])
        else:
            sty.configure("Search.Treeview",
                background="#ffffff", foreground="#000000",
                fieldbackground="#ffffff", rowheight=22)
            sty.configure("Search.Treeview.Heading",
                background="#dfe6e9", foreground="#2d3436",
                font=(FONT_SUB, SZ_S, "bold"), relief="raised")
            sty.map("Search.Treeview",
                background=[("selected", "#0078d7")],
                foreground=[("selected", "#ffffff")])
            sty.map("Search.Treeview.Heading",
                background=[("active", "#c8d6e5")])

        # 38-2. 검색용 새 창
        search_win = tk.Toplevel(self.root)
        search_win.title(f"정류소 {target_idx+1} 검색 및 노선 선택")
        search_win.geometry("1000x850")
        search_win.grab_set()
        search_win.minsize(500, 500)
        search_win.configure(bg=T["bg_panel"])

        # 38-3. 검색 입력 행
        frame_search = tk.Frame(search_win, pady=10, bg=T["bg_panel"])
        frame_search.pack(fill="x")
        tk.Label(frame_search, text="정류소명/ID:", font=(FONT_MAIN, SZ_S),
                 bg=T["bg_panel"], fg=T["fg_main"]).pack(side="left", padx=(15,5))
        search_ent = tk.Entry(frame_search, width=30,
                              bg=T["bg_entry"], fg=T["fg_main"],
                              insertbackground=T["fg_main"])
        search_ent.pack(side="left", padx=5)
        search_ent.focus_set()

        # 38-4. 정류소 목록 표
        tree_st_frame = tk.Frame(search_win, bg=T["bg_panel"])
        tree_st_frame.pack(fill="x", padx=15, pady=5)
        st_scroll = ttk.Scrollbar(tree_st_frame, orient="vertical", style="App.Vertical.TScrollbar")
        tree_st = ttk.Treeview(tree_st_frame, columns=("name", "arsid"),
                               show="headings", height=5,
                               style="Search.Treeview",
                               yscrollcommand=st_scroll.set)
        st_scroll.config(command=tree_st.yview)
        tree_st.heading("name",  text="정류소명")
        tree_st.heading("arsid", text="ARS-ID")
        tree_st.column("name",  width=659, stretch=True)
        tree_st.column("arsid", width=141, stretch=True)
        tree_st.bind("<Button-1>", self.prevent_column_resize)
        tree_st.pack(side="left", fill="both", expand=True)
        st_scroll.pack(side="right", fill="y")

        # 38-5. 노선 선택 버튼 행
        btn_action_frame = tk.Frame(search_win, bg=T["bg_panel"])
        btn_action_frame.pack(fill="x", padx=15, pady=5)
        lbl_route_sel = tk.Label(btn_action_frame, text="▼ 노선 선택",
                                  font=(FONT_SUB, SZ_S, "bold"),
                                  fg=T["fg_accent"], bg=T["bg_panel"])
        lbl_route_sel.pack(side="left")

        self.current_route_data = []
        self.sort_state = {"key": None, "reverse": False}
        type_order = {"간선": 0, "지선": 1, "광역": 2, "기타": 3, "경기": 4, "인천": 5}

        def update_route_tree_view():
            render_route_list()

        def set_all_check(status):
            for item in self.current_route_data:
                item["checked"].set(status)
            update_route_tree_view()

        _btn_s = self.get_btn_style("normal", font_size=SZ_XS)
        btn_all_off = AnyButton(btn_action_frame, text="전체 해제",
                                command=lambda: set_all_check(False), **_btn_s)
        btn_all_off.pack(side="right", padx=2)
        btn_all_on = AnyButton(btn_action_frame, text="전체 선택",
                               command=lambda: set_all_check(True), **_btn_s)
        btn_all_on.pack(side="right", padx=2)

        # 38-6. [노선 목록 표] 선택한 정류소에 서는 모든 버스 노선을 목록으로 보여주는 표를 만듭니다.
        frame_route_list = tk.Frame(search_win, bd=1, relief="sunken",
                                    bg=T["bg_panel"])
        frame_route_list.pack(fill="both", expand=True, padx=15, pady=5)
        route_scroll = ttk.Scrollbar(frame_route_list, orient="vertical", style="App.Vertical.TScrollbar")
        route_cols = ("status", "rnm", "rtype", "path", "term", "f_tm", "l_tm", "st_cnt")
        tree_route = ttk.Treeview(frame_route_list, columns=route_cols,
                                  show="headings",
                                  yscrollcommand=route_scroll.set,
                                  style="Search.Treeview",
                                  selectmode="none")
        route_scroll.config(command=tree_route.yview)

        # 38-7. [선택 상태 색상] 체크된(V) 노선은 초록색, 체크 해제된 노선은 기본 색상으로 표시합니다.
        if self.is_dark_mode:
            tree_route.tag_configure("selected",   background="#1a472a", foreground="#a8e6b5")
            tree_route.tag_configure("unselected", background=T["bg_tree"],  foreground=T["fg_tree"])
        else:
            tree_route.tag_configure("selected",   background="#d4edda", foreground="#155724")
            tree_route.tag_configure("unselected", background="white",   foreground="black")

        # 38-8. [표 열 설정] 선택여부, 노선번호, 유형, 기점↔종점, 배차간격, 첫차, 막차, 정류장수 열을 설정합니다.
        tree_route.heading("status",  text="선택여부")
        tree_route.column("status",   width=42,  anchor="center", stretch=True)
        tree_route.heading("rnm",     text="노선번호", command=lambda: sort_data("rnm"))
        tree_route.column("rnm",      width=117, anchor="center", stretch=True)
        tree_route.heading("rtype",   text="유형",     command=lambda: sort_data("rtype"))
        tree_route.column("rtype",    width=50,  anchor="center", stretch=True)
        tree_route.heading("path",    text="기점↔종점")
        tree_route.column("path",     width=347, anchor="center", stretch=True)
        tree_route.heading("term",    text="배차간격")
        tree_route.column("term",     width=50,  anchor="center", stretch=True)
        tree_route.heading("f_tm",    text="첫차시각")
        tree_route.column("f_tm",     width=67,  anchor="center", stretch=True)
        tree_route.heading("l_tm",    text="막차시각")
        tree_route.column("l_tm",     width=67,  anchor="center", stretch=True)
        tree_route.heading("st_cnt",  text="정류장수")
        tree_route.column("st_cnt",   width=60,  anchor="center", stretch=True)
        tree_route.bind("<Button-1>", self.prevent_column_resize)
        tree_route.pack(side="left", fill="both", expand=True)
        route_scroll.pack(side="right", fill="y")

        # 38-9. [정류소 검색 실행] 입력한 키워드(이름 또는 ARS번호)로 정류소를 검색합니다. 숫자 3자리 이상이면 ARS번호로 먼저 시도합니다.
        def perform_search(event=None):
            keyword = search_ent.get().strip()
            if not keyword: return
            for i in tree_st.get_children():
                tree_st.delete(i)
            root_res = None
            is_digit_search = keyword.isdigit() and len(keyword) >= 3
            if is_digit_search:
                root_res = self.fetch_api(
                    "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
                    {"arsId": keyword})
            if root_res is None or not root_res.findall(".//itemList"):
                root_res = self.fetch_api(
                    "http://ws.bus.go.kr/api/rest/stationinfo/getStationByName",
                    {"stSrch": keyword})
            insert_count = 0
            seen_stations = set()
            if root_res is not None:
                for item in root_res.findall(".//itemList"):
                    name   = item.findtext("stNm")
                    ars_id = item.findtext("arsId")
                    if ars_id and ars_id != "0":
                        key = (name, str(ars_id))
                        if key not in seen_stations:
                            seen_stations.add(key)
                            tree_st.insert("", "end", values=key)
                            insert_count += 1
            if insert_count == 0:
                tree_st.insert("", "end", values=("검색 결과가 없습니다.", ""))

        search_ent.bind("<Return>", perform_search)
        _btn_search = self.get_btn_style("normal", font_size=SZ_XS)
        AnyButton(frame_search, text="검색", command=perform_search,
                  **_btn_search).pack(side="left", padx=5)

        # 38-10. [노선 목록 정렬] 열 제목 클릭 시 해당 열 기준으로 오름차순/내림차순을 토글하여 정렬합니다.
        def sort_data(key):
            if self.sort_state["key"] == key:
                self.sort_state["reverse"] = not self.sort_state["reverse"]
            else:
                self.sort_state["key"] = key
                self.sort_state["reverse"] = False
            if key == "rtype":
                self.current_route_data.sort(
                    key=lambda x: (type_order.get(x["rtype"], 99), x["rnm"]),
                    reverse=self.sort_state["reverse"])
            else:
                self.current_route_data.sort(
                    key=lambda x: x[key],
                    reverse=self.sort_state["reverse"])
            render_route_list()

        # 38-11. [목록 화면 갱신] 현재 노선 데이터를 표에 다시 그립니다. 선택 상태(V)와 색상도 함께 반영합니다.
        def render_route_list():
            for item in tree_route.get_children():
                tree_route.delete(item)
            for item in self.current_route_data:
                status_text = "V" if item["checked"].get() else ""
                tag = "selected" if item["checked"].get() else "unselected"
                vals = [status_text] + item["data_vals"]
                tree_route.insert("", "end", iid=item["rid"], values=vals, tags=(tag,))

        # 38-12. [노선 선택 토글] 표에서 노선을 클릭하면 선택(V)/해제가 토글됩니다.
        def on_route_click(event):
            if tree_route.identify_region(event.x, event.y) == "separator":
                return "break"
            item_id = tree_route.identify_row(event.y)
            if not item_id: return
            for item in self.current_route_data:
                if item["rid"] == item_id:
                    new_state = not item["checked"].get()
                    item["checked"].set(new_state)
                    cur_values = list(tree_route.item(item_id, "values"))
                    cur_values[0] = "V" if new_state else ""
                    tree_route.item(item_id, values=cur_values,
                                    tags=("selected" if new_state else "unselected",))
                    break
        tree_route.bind("<Button-1>", on_route_click)

        # 38-13. [정류소 선택 이벤트] 정류소 목록에서 하나를 선택하면 그 정류소를 지나는 모든 노선을 자동으로 불러옵니다.
        def on_station_select(event):
            selected = tree_st.selection()
            if not selected: return
            ars_id = str(tree_st.item(selected[0])["values"][1]).zfill(5)
            if not ars_id or ars_id == "": return
            saved_checked_routes = set()
            if self.ars_ids[target_idx].get().strip() == ars_id and self.target_st_info[target_idx].get("routes"):
                saved_checked_routes = {r[0] for r in self.target_st_info[target_idx]["routes"]}
            self.current_route_data.clear()
            self._strt_cache.clear()
            self._strt_ord_cache.clear()
            root_route = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/stationinfo/getRouteByStation",
                {"arsId": ars_id})
            if root_route is not None:
                type_map = {"1":"공항","2":"마을","3":"간선","4":"지선","5":"순환",
                            "6":"광역","7":"인천","8":"경기","9":"폐지","0":"공용"}
                for it in root_route.findall(".//itemList"):
                    rid, rnm = it.findtext("busRouteId"), it.findtext("busRouteNm")
                    root_info = self.fetch_api(
                        "http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo",
                        {"busRouteId": rid})
                    rtype, path, term, f_tm, l_tm, st_cnt_str = "-","-","-","-","-","-"
                    st_cnt_val = 100
                    if root_info is not None:
                        info = root_info.find(".//itemList")
                        if info is not None:
                            rtype = type_map.get(info.findtext("routeType"), "기타")
                            path  = f"{info.findtext('stStationNm')}↔{info.findtext('edStationNm')}"
                            term  = info.findtext("term") + "분"
                            f_tm  = self.format_hhmm(info.findtext("firstBusTm", "-"))
                            l_tm  = self.format_hhmm(info.findtext("lastBusTm", "-"))
                            self.route_corp_map[rid] = info.findtext("corpNm", "정보없음")
                            root_st_list = self.fetch_api(
                                "http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute",
                                {"busRouteId": rid})
                            if root_st_list is not None:
                                st_list_items = root_st_list.findall(".//itemList")
                                st_cnt_val = len(st_list_items)
                                st_cnt_str = f"{st_cnt_val}개"
                                self._strt_cache[rid] = root_st_list
                                for _si in st_list_items:
                                    _item_ars = str(_si.findtext("arsId") or "").zfill(5)
                                    if _item_ars == ars_id:
                                        _ord_val = _si.findtext("seq") or _si.findtext("staOrd") or ""
                                        if _ord_val:
                                            self._strt_ord_cache[rid] = _ord_val
                                        break
                    self.rid_to_rnm[rid] = rnm
                    is_checked = (rid in saved_checked_routes) if saved_checked_routes else True
                    self.current_route_data.append({
                        "rid": rid, "rnm": rnm, "rtype": rtype, "st_cnt": st_cnt_val,
                        "data_vals": [rnm, rtype, path, term, f_tm, l_tm, st_cnt_str],
                        "checked": tk.BooleanVar(value=is_checked)
                    })
            self.sort_state = {"key": "rtype", "reverse": False}
            self.current_route_data.sort(
                key=lambda x: (type_order.get(x["rtype"], 99), x["rnm"]))
            render_route_list()

        tree_st.bind("<<TreeviewSelect>>", on_station_select)

        # 38-14. [기존 정류소 자동 표시] 이미 정류소가 설정된 상태라면 기존 정류소를 자동으로 표시하고 선택 상태로 만듭니다.
        if self.ars_ids[target_idx].get().strip():
            item_id = tree_st.insert("", "end",
                values=(self.target_st_info[target_idx].get("nm", ""),
                        self.ars_ids[target_idx].get()))
            tree_st.selection_set(item_id)

        # 38-15. [노선 확정] 선택한 노선들로 모니터링 대상을 확정합니다. 정류소 정보, 노선 ID, 정류소 순번(ord)을 저장합니다.
        def confirm_selection():
            chosen = [(item["rid"], item["rnm"], item["st_cnt"])
                      for item in self.current_route_data if item["checked"].get()]
            selected_st = tree_st.selection()
            if not selected_st or not chosen: return
            st_name = tree_st.item(selected_st[0])["values"][0]
            ars_id  = str(tree_st.item(selected_st[0])["values"][1]).zfill(5)
            self.ars_ids[target_idx].set(ars_id)
            self.lbl_st_names[target_idx].config(text=f"[{st_name}] 실시간 현황")
            self.lbl_hist_titles[target_idx].config(text=f"[{st_name}] 도착 기록")
            root_st_uid = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
                {"arsId": ars_id})
            st_id = root_st_uid.findtext(".//stId") if (
                root_st_uid is not None and not isinstance(root_st_uid, tuple)) else ""
            if hasattr(self, "target_st_info") and self.target_st_info[target_idx].get("routes"):
                old_rids = {r[0] for r in self.target_st_info[target_idx]["routes"]}
                new_rids = {item[0] for item in chosen}
                gone = old_rids - new_rids
                for rid in gone:
                    self.pos_suspend_until.pop(rid, None)
                    old_st_id = self.target_st_info[target_idx].get("id", "")
                    self.veh_cache.pop((old_st_id, rid), None)
            ord_map = {}
            for rid, rnm, _ in chosen:
                if rid in self._strt_ord_cache:
                    ord_map[rid] = self._strt_ord_cache[rid]
                elif rid in self._strt_cache:
                    root_strt = self._strt_cache[rid]
                    for st_item in root_strt.findall(".//itemList"):
                        item_ars = str(st_item.findtext("arsId") or "").zfill(5)
                        if item_ars == ars_id:
                            ord_val = st_item.findtext("seq") or st_item.findtext("staOrd") or ""
                            if ord_val: ord_map[rid] = ord_val
                            break
                else:
                    self.log(f"⚠ 노선전체정류소 API 캐시 누락: {rnm}번 — 재호출합니다.")
                    root_strt = self.fetch_api(
                        "http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute",
                        {"busRouteId": rid})
                    if root_strt is not None and not isinstance(root_strt, tuple):
                        for st_item in root_strt.findall(".//itemList"):
                            item_ars = str(st_item.findtext("arsId") or "").zfill(5)
                            if item_ars == ars_id:
                                ord_val = st_item.findtext("seq") or st_item.findtext("staOrd") or ""
                                if ord_val: ord_map[rid] = ord_val
                                break
            self._strt_cache.clear()
            self._strt_ord_cache.clear()
            self.target_st_info[target_idx] = {
                "id": st_id, "nm": st_name, "routes": chosen, "ord_map": ord_map}
            self.log(f"정류소 {target_idx+1} 설정 완료: {st_name} "
                     f"(총 {len(ord_map)}개 노선 중 {len(chosen)}개 선택)")
            search_win.destroy()
            self.update_button_states()

        # 38-16. [적용 버튼] "선택한 노선들로 적용" 버튼을 만들어 confirm_selection() 함수와 연결합니다.
        AnyButton(search_win, text="선택한 노선들로 적용",
                  command=confirm_selection, height=2,
                  **self.get_btn_style("normal", font_size=SZ_M)
        ).pack(fill="x", padx=15, pady=10)


    # 39. [정보 가져오기] 고른 정류소의 버스 도착 시간과 위치를 실제로 확인하는 함수
    def process_station(self, idx):
        # 39-1. [자정 날짜 전환] 날짜가 바뀌면 도착 기록 API 일시정지 목록을 초기화합니다. 어제의 버스 운행 정지 기록을 새날에 이어쓰지 않도록 합니다.
        today = datetime.now().date()
        if today != self._last_date:
            self._last_date = today
            self.pos_suspend_until.clear()
            # 39-1-1.   차량번호 캐시(veh_cache)는 날짜와 무관하게 유효하므로 날짜가 바뀌어도 초기화하지 않습니다.
            self.log("📅 날짜가 바뀌었습니다. 도착 기록 API 일시정지 기록을 초기화합니다.")

            # 39-1-2. [자정 API 카운터 리셋] 00시 기준으로 API 호출 횟수를 0으로 초기화합니다. idx==0일 때만 실행하여 정류소1·2 처리 중 두 번 리셋되는 것을 막습니다.
            #   메인창 상단의 합산 카운트(api_stats)와
            #   API 상세 통계 창의 키별 카운트(api_stats_by_key)를 동시에 초기화합니다.
            #   idx == 0 일 때만 실행해 정류소 1·2 처리 중 두 번 초기화되는 것을 막습니다.
            if idx == 0:
                for k in self.api_stats:
                    self.api_stats[k] = 0
                for k in self.api_stats_by_key["main"]:
                    self.api_stats_by_key["main"][k] = 0
                for k in self.api_stats_by_key["back"]:
                    self.api_stats_by_key["back"][k] = 0
                self._trigger_counter_update()
                self.log("📊 API 호출 횟수가 00:00 기준으로 초기화되었습니다.")

        # 39-2. [처리 대상 확인] 현재 처리할 정류소(idx번)의 설정 정보(노선목록, 정류소ID 등)와 ARS 번호를 가져옵니다.
        info, ars_id = self.target_st_info[idx], self.ars_ids[idx].get().strip()
        st_id, rt_rows = info.get('id', ''), []

        # 39-3. [도착 예정 시간 조회] 등록된 각 버스 노선의 도착 예정 시간 정보를 서버에서 가져옵니다.
        # ── ARR1(주) 초과 대비: SINF(보조)는 정류소당 1회만 호출하는 캐시 ──────────
        uid_cache = None  # None=미호출 / {}=호출했으나 결과없음 / {rid:item}=정상수신
        ord_map = info.get('ord_map', {})  # {busRouteId: 정류소순번} — confirm_selection에서 미리 저장

        def get_uid_cache():
            """SINF API를 정류소당 최대 1회만 호출해 두 가지 키로 조회 가능한 딕셔너리를 반환합니다.
            기본 키: busRouteId (숫자형 ID)
            보조 키: rtNm 또는 busRouteAbrv (노선 번호 문자열) ← busRouteId가 없거나 0인 경우 대비
            반환값: {'by_rid': {busRouteId: item}, 'by_nm': {rtNm: item}}
            """
            nonlocal uid_cache
            if uid_cache is not None:
                return uid_cache
            root_uid = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
                {'arsId': ars_id}
            )
            if root_uid is None or isinstance(root_uid, tuple):
                uid_cache = {'by_rid': {}, 'by_nm': {}}
                return uid_cache
            by_rid, by_nm = {}, {}
            for item in root_uid.findall(".//itemList"):
                rid_key = item.findtext("busRouteId") or ""
                nm_key  = item.findtext("busRouteAbrv") or item.findtext("rtNm") or ""
                # 39-3-2-1. rid 키 등록: 비어있거나 "0"이 아닌 경우만
                if rid_key and rid_key not in by_rid:
                    by_rid[rid_key] = item
                # 39-3-2-2. 노선명 키 등록: 항상 보조 수단으로 저장
                if nm_key and nm_key not in by_nm:
                    by_nm[nm_key] = item
            uid_cache = {'by_rid': by_rid, 'by_nm': by_nm}
            # 39-3-1. 하루 1회만 로그 출력 (정류소 인덱스별로 날짜 기억)
            _today = datetime.now().date()
            if self._sinf_logged_date.get(idx) != _today:
                self._sinf_logged_date[idx] = _today
                self.log(f"ℹ 실시간 현황 API 초과 → 실시간 현황 예비2 API 보조 호출 (정류소 {idx+1}, {len(by_rid)}개 노선 수신) — 이후 동일 알림은 내일까지 표시 생략")
            return uid_cache

        # 39-4. 노선마다 ARR1 우선 시도, 실패 시 SINF 캐시에서 해당 노선 항목만 추출
        for rid, rnm, _ in info.get('routes', []):
            row = None

            # 39-4-1. [1순위 도착 조회] 도착정보 API(getArrInfoByRoute)를 호출합니다. 정류소 ID + 노선 ID + 정류소 순번으로 해당 정류소 1개 행만 가져옵니다.
            ord_val = ord_map.get(rid, '0')
            arr_params = {'stId': st_id, 'busRouteId': rid, 'ord': ord_val}
            root_arr = self.fetch_api("", arr_params, api_type="ARR1")

            if root_arr is not None and not isinstance(root_arr, tuple):
                # 39-4-1-1. [운행 재개 감지] ARR1 응답에 실제로 운행 중인 버스 정보가 있으면 일시정지 상태를 해제합니다.
                                #   판단 기준: arrmsg(도착 메시지)가 "운행종료" 등 비운행 메시지가 아닌 경우.
                #   판단 기준: arrmsg1 또는 arrmsg2 가 운행 없음 메시지 이외의 값(예: "3분 후")인 경우만 운행 중으로 판정합니다.
                #   ※ plainNo(차량번호)는 서울시 API가 운행 종료 후에도 지우지 않고 남겨두는
                #     경우가 있어 신뢰도가 낮습니다. arrmsg가 실시간 운행 상태를 더 정확히 반영하므로
                #     arrmsg 기준만으로 단일화하여 유령 차량번호로 인한 오판정을 방지합니다.
                NO_BUS_MSGS = {"운행정보없음", "운행종료", "출발대기", "-", ""}
                arr_items = root_arr.findall(".//itemList")
                has_active_bus = any(
                    (item.findtext("arrmsg1") or "").strip() not in NO_BUS_MSGS
                    or (item.findtext("arrmsg2") or "").strip() not in NO_BUS_MSGS
                    for item in arr_items
                )
                if rid in self.pos_suspend_until:
                    if has_active_bus:
                        del self.pos_suspend_until[rid]
                        self.log(f"🔄 {rnm}번 운행 재개 확인")
                    # 39-4-1-1-1. else: 아직 운행 정보 없음 → 정지 유지 (불필요한 로그 없음)
                for item in root_arr.findall(".//itemList"):
                    res_ars  = str(item.findtext("arsId") or "").zfill(5)
                    res_ord  = str(item.findtext("staOrd") or "")

                    # 39-4-1-1-2. ① 주 API(getArrInfoByRoute) 응답: 항상 해당 정류소 1행 → arsId 무조건 수용
                    # 39-4-1-1-3. ② ARR2 폴백 응답: arsId 일치 행 우선, arsId=0인 경우 staOrd로 2차 검증
                    #    - arsId가 정확히 일치하면 OK
                    #    - arsId=0("00000")이고 ord_val이 일치하면 인천/경기 버스 해당 정류소 행 OK
                    #    - arsId=0이고 ord를 모를 때(ord_val='0')는 첫 번째 0행 수용(기존 동작 유지)
                    arsid_match = (res_ars == ars_id)
                    arsid_zero  = (not res_ars.strip("0"))   # "00000" 등 0으로만 구성
                    ord_match   = (ord_val != '0' and res_ord == ord_val)

                    if arsid_match or (arsid_zero and (ord_match or ord_val == '0')):
                        p1 = item.findtext("plainNo1") or "-"
                        p2 = item.findtext("plainNo2") or "-"
                        # 39-4-1-1-3-1. [차량번호 캐시] ARR1/ARR2에서 얻은 실제 번호판을 저장합니다.
                        cache_key = (st_id, rid)
                        if p1 not in ("-", "", None) or p2 not in ("-", "", None):
                            self.veh_cache[cache_key] = (p1, p2)
                        row = (
                            item.findtext("rtNm") or rnm,
                            p1,
                            item.findtext("arrmsg1") or "-",
                            p2,
                            item.findtext("arrmsg2") or "-"
                        )
                        break
            elif root_arr is None:
                # 39-4-1-2. [2순위 보조 조회] 도착 조회API가 모두 실패(한도 초과)했을 때, 실시간 현황 예비2 API(getStationByUid) 캐시에서 해당 노선 정보를 꺼냅니다.
                uid_data = get_uid_cache()
                # 39-4-1-3.   노선ID(숫자)로 먼저 찾고, 없으면 노선명(문자열)으로 다시 찾습니다.
                uid_item = uid_data['by_rid'].get(rid)
                if uid_item is None:
                    uid_item = uid_data['by_nm'].get(rnm)

                if uid_item is not None:
                    route_display = (
                        uid_item.findtext("busRouteAbrv")
                        or uid_item.findtext("rtNm")
                        or rnm
                    )
                    # 39-4-1-3-1. [캐시 차량번호 우선] 이전에 도착 조회API에서 받아뒀던 차량번호가 있으면, 차량번호 조회API 호출 없이 그 번호를 바로 사용합니다.
                    cache_key = (st_id, rid)
                    cached = self.veh_cache.get(cache_key)
                    if cached:
                        p1_display, p2_display = cached
                    else:
                        # 39-4-1-3-1-1. [차량번호 API 조회] 캐시에 번호판이 없을 때, 차량 ID로 차량번호 조회API(getBusPosByVehId)를 호출하여 실제 번호판을 찾아옵니다.
                        veh_id1 = uid_item.findtext("vehId1") or ""
                        veh_id2 = uid_item.findtext("vehId2") or ""

                        def _fetch_plain(veh_id):
                            """vehId 한 개를 VID API에 물어보고 plainNo 반환. 실패 시 원본 반환."""
                            if not veh_id or veh_id == "-":
                                return "-"
                            root_vid = self.fetch_api(
                                "http://ws.bus.go.kr/api/rest/buspos/getBusPosByVehId",
                                {"vehId": veh_id}
                            )
                            if root_vid is not None and not isinstance(root_vid, tuple):
                                plain = root_vid.findtext(".//plainNo")
                                if plain:
                                    return plain
                            return veh_id  # API 실패 시 9자리 vehId 그대로 표시

                        p1_display = _fetch_plain(veh_id1)
                        p2_display = _fetch_plain(veh_id2)

                        # 39-4-1-3-1-2. [번호판 캐시 저장] 조회 성공한 번호판을 veh_cache에 저장해 다음 번 같은 버스가 올 때는 API 호출 없이 바로 쓸 수 있도록 합니다.
                        if (p1_display not in ("-", "", veh_id1) or
                                p2_display not in ("-", "", veh_id2)):
                            self.veh_cache[cache_key] = (p1_display, p2_display)
                    row = (
                        route_display,
                        p1_display,
                        uid_item.findtext("arrmsg1") or "-",
                        p2_display,
                        uid_item.findtext("arrmsg2") or "-"
                    )

            if row:
                rt_rows.append(row)

        # 39-5. [실시간 표 업데이트] 방금 조회한 도착 예정 시간 정보를 화면 상단의 실시간 현황 표에 표시합니다.
        self.root.after(0, lambda d=rt_rows, t=self.trees_rt[idx]: self.refresh_tree(t, d))

        # 39-6. [도착 여부 판정] 버스가 실제로 이 정류소에 도착했는지 판정합니다. 정류소1·2가 같은 노선을 처리할 때 API를 중복 호출하지 않고 결과를 공유합니다.
        if st_id:
            # 39-6-1. [임시 저장소 초기화] 정류소1(idx=0) 처리 시작 시 버스 위치 임시 저장소를 비웁니다. 이 저장소가 정류소1·2 간 정보 공유에 사용됩니다.
            if idx == 0: self.temp_pos_data = {}

            for rid, rnm, st_cnt in info.get('routes', []):
                root_pos = None

                # 39-6-1-1. [위치 조회 일시정지 확인] 이 노선이 현재 일시정지 중인지 확인합니다. 첫차 시각 이전이면 API 호출을 건너뜁니다.
                now_dt = datetime.now()
                if rid in self.pos_suspend_until:
                    resume_dt = self.pos_suspend_until[rid]
                    if now_dt < resume_dt:
                        # 39-6-1-1-1-1. 재개 예정 시각이 아직 안 됐으므로 이번에는 위치 조회를 건너뜁니다.
                        continue
                    else:
                        # 39-6-1-1-1-2. 재개 예정 시각이 지났으므로 일시정지를 해제하고 위치 조회를 다시 시작합니다.
                        del self.pos_suspend_until[rid]
                        if rid not in self.pos_resume_logged:
                            self.log(f"⏰ {rnm}번 POS 호출 재개 (첫차 5분 전 도달)")
                            self.pos_resume_logged.add(rid)

                # 39-6-1-2. [공유 캐시 확인] 정류소1에서 이미 이 노선의 버스 위치를 조회했다면 정류소2는 같은 결과를 재사용합니다.
                if rid in self.temp_pos_data:
                    # 39-6-1-2-1. [캐시 재사용] 이미 조회된 버스 위치 데이터가 있으므로 API를 다시 호출하지 않고 저장된 값을 꺼냅니다.
                    root_pos = self.temp_pos_data[rid]
                else:
                    # 39-6-1-2-2. [새 조회 실행] 이 노선의 버스 위치 정보가 아직 없으므로 위치 조회API(getBusPosByRtid)를 직접 호출합니다.
                    pos_result = self.fetch_api("", {'busRouteId': rid, 'startOrd': '1', 'endOrd': str(st_cnt)}, api_type="POS2")

                    if isinstance(pos_result, tuple) and pos_result[0] == 'NO_BUS':
                        # 39-6-1-2-2-1. [운행 없음 → 일시정지 등록] 이 시간대에 운행 차량이 없습니다. 첫차 5분 전까지 위치 조회를 일시정지하여 불필요한 API 호출을 줄입니다.
                        _, f_tm_str = pos_result
                        if f_tm_str:
                            try:
                                f_hm = datetime.strptime(f_tm_str, "%H:%M")
                                # 39-6-1-2-2-1-1-1. 첫차가 내일인 경우(막차 이후)를 대비하여 내일 날짜로 첫차 시각을 조합합니다.
                                base_dt = now_dt.replace(hour=f_hm.hour, minute=f_hm.minute, second=0, microsecond=0)
                                if base_dt <= now_dt:
                                    base_dt += timedelta(days=1)
                                resume_dt = base_dt - timedelta(minutes=5)
                                if resume_dt > now_dt:
                                    # 39-6-1-2-2-1-1-1-1. 아직 첫차 5분 전이 안 됐습니다. 계산된 재개 시각까지 위치 조회를 정지합니다.
                                    self.pos_suspend_until[rid] = resume_dt
                                    self.pos_resume_logged.discard(rid)
                                    self.log(f"💤 {rnm}번 차량 없음 → 첫차 {f_tm_str} 5분 전({resume_dt.strftime('%H:%M')})까지 POS 정지")
                                else:
                                    # 39-6-1-2-2-1-1-1-2. 첫차 5분 전이 됐는데도 아직 차량이 없습니다. 1분 후 다시 확인합니다.
                                    self.pos_suspend_until[rid] = now_dt + timedelta(minutes=1)
                                    # 39-6-1-2-2-1-1-1-3.   pos_resume_logged는 유지: 이후 POS 재개 로그가 중복으로 출력되지 않도록 기록을 지우지 않습니다.
                            except Exception:
                                # 39-6-1-2-2-1-1-2. 첫차 시각 파싱에 실패했습니다. 30분 뒤에 다시 확인합니다.
                                self.pos_suspend_until[rid] = now_dt + timedelta(minutes=30)
                                self.pos_resume_logged.discard(rid)
                                self.log(f"💤 {rnm}번 차량 없음 → 첫차 시각 파싱 실패, 30분 후 재시도")
                        else:
                            # 39-6-1-2-2-1-1. 첫차 시각 정보를 가져오지 못했습니다. 30분 뒤에 다시 확인합니다.
                            self.pos_suspend_until[rid] = now_dt + timedelta(minutes=30)
                            self.pos_resume_logged.discard(rid)
                            self.log(f"💤 {rnm}번 차량 없음 → 첫차 정보 없음, 30분 후 재시도")
                        continue  # POS 처리 건너뜀

                    root_pos = pos_result
                    # 39-6-1-2-3. [공유 캐시 저장] 방금 조회한 버스 위치 결과를 임시 저장소에 보관합니다. 정류소2 처리 시 이 데이터를 재사용합니다.
                    if root_pos is not None:
                        self.temp_pos_data[rid] = root_pos

                # 39-6-1-3. [도착 판정] 버스 위치 데이터를 분석하여 버스가 이 정류소에 실제로 도착했는지 판정합니다.
                if root_pos is not None:
                    for bus in root_pos.findall(".//itemList"):
                        # 39-6-1-3-3-1. [정류소 일치 확인] 버스의 현재 위치(stId)가 지금 처리 중인 정류소의 ID와 같으면 도착으로 판정합니다.
                        if bus.findtext("lastStnId") == st_id:
                            veh_no = bus.findtext("plainNo")
                            # 39-6-1-3-3-1-1. [중복 기록 방지] 같은 버스가 5분 안에 다시 기록되지 않도록 막습니다. 출발 후 잠시 같은 정류소에 머무는 상황을 걸러냅니다.
                            if veh_no in self.last_arrival_logs[idx] and (time.time() - self.last_arrival_logs[idx][veh_no] < 300): continue

                            f_time = self.format_datetm(bus.findtext("dataTm"))
                            # 39-6-1-3-3-1-2. [버스 회사 조회] 기록한 버스의 운수 회사 이름을 찾아오고, 공동 배차 정보가 있으면 함께 확인합니다.
                            corp_nm = self.route_corp_map.get(rid, "정보없음")
                            if rnm in self.excel_multi_corp_map and len(self.excel_multi_corp_map[rnm]) >= 2:
                                corp_nm = ", ".join(sorted(list(self.excel_multi_corp_map[rnm])))

                            # 39-6-1-3-3-1-3. [도착 기록 및 자동 저장] 도착 정보를 도착 기록 표에 한 줄 추가하고, 엑셀 자동 저장도 즉시 실행합니다.
                            log_entry = (f_time, rnm, veh_no, corp_nm, "정류소 도착")
                            self.recorded_data.append((f_time, info['nm'], rnm, veh_no, corp_nm))
                            self.last_arrival_logs[idx][veh_no] = time.time()
                            self.perform_auto_save()
                            self.log(f"★ [{info['nm']}] 도착: {rnm} ({veh_no})")
                            # 39-6-1-3-3-1-4. "end"(마지막)로 삽입하면 새 기록이 아래에 쌓여 시간 순서(오름차순)를 유지합니다.
                            # 39-6-1-3-3-1-5. 삽입 후 see()를 호출하면 표가 자동으로 새 항목 위치로 스크롤됩니다.
                            self.root.after(0, lambda r=log_entry, t=self.trees_hist[idx]: (
                                t.insert("", "end", values=r),
                                t.see(t.get_children()[-1])
                            ))
       
    # 40. [엑셀 백그라운드 다운로드 시작] 공동 배차 회사 정보 확인에 쓰이는 서울시 버스 운행 노선 엑셀을 뒷단에서 자동으로 내려받기 시작합니다.
    def start_excel_download_thread(self):
        # 40-1. [백그라운드 스레드 사용] 다운로드가 길어질 수 있으므로 별도 스레드(일꾼)로 실행하여 화면이 멈추지 않도록 합니다.
        threading.Thread(target=self.download_and_load_excel, daemon=True).start()

    # 41. [엑셀 다운로드 실행] 서울 데이터 광장에서 최신 서울시 버스 운행 노선 정보 엑셀 파일을 크롬 자동화로 직접 내려받습니다.
    def download_and_load_excel(self):
        target_url = "https://data.seoul.go.kr/dataList/OA-15066/F/1/datasetView.do" 
        self.log(f"📥 최신 노선정보 확인 중 (Text Search): {target_url}")
        
        download_success = False
        downloaded_filename = None
        driver = None

        try:
            # 41-1-1. [크롬 드라이버 설정] 화면 없이 작동하는 크롬 브라우저(headless)를 준비합니다.
            options = Options()
            options.add_argument("--headless") # 5-20-1-1. 화면에는 안 보이게 숨깁니다.
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--window-size=1920,1080")
            
            # 41-1-2. [봇 감지 우회] 웹사이트가 자동화 프로그램으로 인식하지 않도록 설정합니다.
            options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            options.add_argument("--disable-blink-features=AutomationControlled") 
            options.add_experimental_option("excludeSwitches", ["enable-automation"]) 
            options.add_experimental_option('useAutomationExtension', False)
            
            # 41-1-3. [다운로드 경로 설정] 내려받은 파일이 프로그램 실행 폴더에 저장되도록 크롬 설정을 합니다.
            target_download_dir = os.path.abspath(self.current_dir)
            prefs = {
                "download.default_directory": target_download_dir,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True,
                "profile.default_content_settings.popups": 0
            }
            options.add_experimental_option("prefs", prefs)

            # 41-1-4. [사이트 접속] 크롬 드라이버를 실행하여 서울 데이터 광장 다운로드 페이지로 이동합니다.
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            
            if not os.path.exists(target_download_dir):
                os.makedirs(target_download_dir, exist_ok=True)

            driver.get(target_url) 
            
            # 41-1-5. [페이지 로딩 대기] 버튼이 나타날 때까지 최대 20초 기다립니다.
            wait = WebDriverWait(driver, 20)
            self.log("   ㄴ페이지 분석 중 (파일명 탐색)...")
            
            # 41-1-6. [다운로드 버튼 탐색] 페이지에서 ".xlsx" 텍스트가 포함된 다운로드 링크나 버튼을 찾습니다.
            download_target = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//*[contains(text(), '.xlsx')]")
            ))
            
            found_text = download_target.text.strip()
            self.log(f"   ㄴ다운로드 대상 발견: {found_text}")
            
            # 41-1-7. [중복 파일 정리] 같은 이름의 파일이 이미 있으면 삭제하고 새로 받습니다.
            file_path = os.path.join(self.current_dir, found_text)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    self.log(f"   ℹ 기존 파일 삭제 후 다운로드: {found_text}")
                except Exception as del_err:
                    self.log(f"   ⚠ 기존 파일 삭제 실패 (사용중일 수 있음): {del_err}")

            before_files = set(os.listdir(self.current_dir)) 
            
            # 41-1-8. [다운로드 버튼 클릭] 찾은 버튼을 자동으로 클릭하여 파일 다운로드를 시작합니다.
            driver.execute_script("arguments[0].scrollIntoView(true);", download_target)
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", download_target)
            
            # 41-1-9. [다운로드 완료 대기] 1초마다 폴더를 확인하여 파일이 완전히 저장됐는지 확인합니다.
            found_file = None
            for _ in range(30):
                time.sleep(1)
                after_files = set(os.listdir(self.current_dir))
                new_files = after_files - before_files 
                
                for f in new_files:
                    if f.endswith(".xlsx") and "서울시버스노선기본정보" in f:
                        found_file = f
                        break
                if found_file:
                    break
            
            # 41-1-10. [결과 처리] 다운로드 결과에 따라 성공 또는 실패 메시지를 로그에 기록합니다.
            if found_file:
                self.log(f"✅ 다운로드 성공: {found_file}")
                download_success = True
                downloaded_filename = found_file
            else:
                self.log("⚠ 다운로드 시도했으나 파일 생성 안됨 (시간초과)")

        except Exception as e:
            self.log(f"❌ 다운로드 프로세스 오류: {e}")
            self.log("   (Chrome 브라우저와 인터넷 연결을 확인해주세요)")
        finally:
            if driver:
                driver.quit() # 5-20-9. 로봇을 퇴근시킵니다(인터넷 창 닫기).
        
        # 41-1. [파일 로드 시도] 새로 받은 파일을 읽어옵니다. 없으면 기존 파일 중 가장 최근 것을 찾아 씁니다.
        if download_success and downloaded_filename:
            self.load_excel_routes(specific_filename=downloaded_filename)
        else:
            self.log("ℹ 기존 파일 중 최신 파일을 사용합니다.")
            self.load_excel_routes()

    # 42. [엑셀 데이터 로드] 내려받은 엑셀 파일을 열어 버스 노선번호와 운수사 이름을 메모리에 저장합니다.
    def load_excel_routes(self, specific_filename=None):
        target_file = specific_filename
        
        # 42-1. [파일 자동 탐색] 저장할 파일 이름을 지정하지 않았으면, 폴더에서 이름이 비슷한 가장 최신 파일을 자동으로 찾습니다.
        if not target_file:
            candidates = [f for f in os.listdir(self.current_dir) if "서울시버스노선기본정보" in f and f.endswith(".xlsx")]
            if candidates:
                candidates.sort(reverse=True) 
                target_file = candidates[0]
            else:
                target_file = "서울시버스노선기본정보(20260108).xlsx" 

        path = os.path.join(self.current_dir, target_file)
        
        # 42-2. [파일 존재 확인] 로드할 파일이 실제로 있는지 확인합니다. 없으면 경고 메시지를 남기고 중단합니다.
        if not os.path.exists(path):
            self.log(f"⚠ 로드할 엑셀 파일이 없습니다: {target_file}")
            return

        try:
            # 42-2-1. [표 데이터 파싱] 엑셀 파일을 pandas 라이브러리로 읽어서 노선번호와 업체명 열을 처리합니다.
            df = pd.read_excel(path)
            r_col = [c for c in df.columns if '노선번호' in str(c)]
            c_col = [c for c in df.columns if '업체명' in str(c)]
            
            if r_col and c_col:
                r_col_name = r_col[0]
                c_col_name = c_col[0]
                
                temp_map = {}
                # 42-2-1-1. [노선-회사 매핑] 각 행을 순서대로 읽으면서 "노선번호: 업체명" 형태의 사전(dictionary)을 만듭니다.
                for _, row in df.iterrows():
                    r_name = str(row[r_col_name]).strip()
                    c_name = str(row[c_col_name]).strip()
                    if r_name not in temp_map:
                        temp_map[r_name] = set()
                    temp_map[r_name].add(c_name) 
                
                self.excel_multi_corp_map = temp_map 
                self.log(f"✅ 엑셀 데이터 로드 완료: {target_file}")
            else:
                self.log(f"⚠ 엑셀 형식 오류: '노선번호' 또는 '업체명' 열 미발견")
        except Exception as e:
            self.log(f"❌ 엑셀 로드 중 오류: {e}")

    # 43. [그룹 6] 실시간 모니터링 — 자동/수동 갱신, 모니터링 루프 제어 함수 모음

    # 44. [자동 기록 시작] 버스 도착 자동 기록을 시작합니다. 갱신 주기 검증 → 버튼 비활성화 → 엑셀 파일 생성 → 모니터링 스레드 실행 순으로 진행됩니다.
    def start_monitoring(self):
        # 44-1. [갱신 주기 검증] 갱신 주기 입력창에 유효한 숫자가 있는지 확인합니다. 비어있거나 숫자가 아니면 시작할 수 없습니다.
        if not self.refresh_interval_var.get().strip():
            messagebox.showwarning("입력 누락", "갱신주기를 입력해야 자동 기록을 시작할 수 있습니다.")
            return

        if not self.service_key_var.get().strip(): return 
        
        # 44-2. [입력창 잠금] 자동 기록이 실행되는 동안 갱신 주기 입력창을 수정하지 못하도록 읽기 전용으로 바꿉니다.
        self.entry_refresh_interval.config(state='readonly')
        
        self.is_monitoring = True
        # 44-3.   기록 실행 중에는 정류소를 바꾸지 못하도록 검색 버튼도 비활성화합니다.
        _s_off2 = self.get_btn_style("disabled")
        _s_off2.pop("state", None)
        for _b in self.btn_searches:
            _b.config(state="disabled", **_s_off2)
        
        # 44-4. [엑셀 파일 생성] 오늘 날짜로 이름이 지정된 새 엑셀 파일 경로를 준비하고 자동 저장을 허용합니다.
        filename = f"Bus_Arrival_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.auto_save_path = os.path.join(self.current_dir, filename)
        
        try:
            pd.DataFrame(columns=["데이터시각", "도착정류소명", "노선", "차량번호", "운수사명"]).to_excel(self.auto_save_path, index=False)
            self.can_auto_save = True 
            display_name = os.path.basename(self.auto_save_path)
            self.lbl_auto_save_status.config(text=f"[{display_name}] 파일에 자동 기록 중 ......", fg="#27ae60")
        except Exception as e:
            self.can_auto_save = False
            messagebox.showwarning("오류", f"자동 저장 설정 중 오류 발생: {e}")

        # 44-5. [모니터링 스레드 실행] 별도 백그라운드 스레드를 만들어 main_loop()를 무한 반복 실행합니다.
        self.update_button_states()
        threading.Thread(target=self.main_loop, daemon=True).start()

    # 45. [자동 갱신 루프] 프로그램이 실행되는 동안 설정된 주기마다 자동으로 버스 정보를 갱신하는 무한 반복 함수입니다.
    def main_loop(self):
        next_call = time.time()
        # 45-1. [중지 조건 확인] is_monitoring 플래그가 False가 될 때까지 계속 반복합니다.
        while self.is_monitoring:
            self.refresh_data() # 5-23-1-1. 서버에 물어보고 데이터를 갱신합니다.
            try: interval = int(self.refresh_interval_var.get())
            except: interval = 60
            next_call += interval
            
            # 45-1-1. [분할 대기] 대기를 0.1초씩 나눠서 합니다. 중간에 중지 버튼을 눌러도 즉시 반응할 수 있도록 합니다.
            sleep_time = next_call - time.time()
            if sleep_time < 0: next_call = time.time(); sleep_time = 0
            for _ in range(int(sleep_time * 10)):
                if not self.is_monitoring: break
                time.sleep(0.1)

    # 46. [수동 즉시 갱신] 대기 없이 지금 즉시 버스 정보를 새로고침합니다. 버튼 중복 클릭 방지 처리 포함.
    def manual_refresh(self):
        # 46-1. [맥OS 중복 클릭 방지] 맥OS에서 disabled 상태인 버튼이 클릭되는 버그를 직접 플래그로 차단합니다.
        if CURRENT_OS == "Darwin" and not self._btn_active.get('manual', True):
            return
        # 46-2. [중복 실행 방지] 갱신이 끝나기 전에 버튼을 또 누를 수 없도록 플래그와 버튼 상태를 동시에 비활성화합니다.
        self._btn_active['manual'] = False
        _s_dis = self.get_btn_style("disabled")
        _s_dis.pop("state", None)
        self.btn_manual.config(state="disabled", **_s_dis)
        self.log("🔄 [수동] 데이터 갱신 시작...")

        # 46-3. [스레드로 즉시 갱신] 화면이 멈추지 않도록 별도 스레드에서 refresh_data()를 실행합니다.
        def run_manual():
            self.refresh_data(manual=True)
            self.log("✅ [수동] 데이터 갱신 완료")
            def _re_enable():
                self._btn_active['manual'] = True
                _s_on = self.get_btn_style("normal")
                _s_on.pop("state", None)
                self.btn_manual.config(state="normal", **_s_on)
            self.root.after(0, _re_enable)

        threading.Thread(target=run_manual, daemon=True).start()

    # 47. [전체 갱신 공용 함수] 정류소1과 정류소2 모두의 버스 정보를 순서대로 갱신합니다. 자동/수동 갱신 모두 이 함수를 사용합니다.
    def refresh_data(self, manual=False):
        # 47-1. [노선 변동 주석] 실제 체크는 process_station 내부에서 수행됩니다. ARR1 응답에 새 노선이 나타나면 일시정지 해제.
        #   pos_suspend_until 에 등록된 노선이 ARR1 응답에서 새로 보이면 정지를 해제합니다.
        #   (실제 체크는 process_station 내부에서 수행하며, 여기서는 플래그만 기록합니다.)

        # 47-2. [동시 갱신 방지] _refresh_lock으로 자동 갱신과 수동 갱신이 동시에 실행되지 않도록 막습니다.
                #   두 스레드가 동시에 실행되면 공유 데이터(last_arrival_logs, veh_cache 등)에 중복 기록이나 누락이 생길 수 있습니다.
        #   수동 갱신이 자동 갱신과 겹치면 last_arrival_logs / veh_cache 등의 공유 상태를
        #   두 스레드가 동시에 읽고 쓸 수 있어 중복 기록 또는 누락이 발생할 수 있습니다.
        if not self._refresh_lock.acquire(blocking=False):
            # 47-2-1. 다른 스레드가 이미 갱신 중이면 이번 호출은 건너뜁니다.
            if manual:
                self.log("ℹ 자동 갱신 진행 중이므로 수동 갱신을 건너뜁니다.")
            return

        try:
            # 47-2-2. [두 정류소 순차 처리] 정류소1 처리 후 정류소2를 처리합니다.
            for i in range(2):
                if self.target_st_info[i].get('routes'): 
                    self.process_station(i) 
            
            # 47-2-3. [구제 저장] 갱신 사이클이 끝난 뒤에도 미저장 데이터가 있으면 즉시 재시도합니다.
                    #   perform_auto_save()가 PermissionError 등으로 실패했을 경우 여기서 다시 시도합니다.
            #   process_station() 내부에서 perform_auto_save()를 즉시 호출하지만,
            #   PermissionError나 OS 쓰기 오류로 저장이 실패했다면 recorded_data의 크기가
            #   _saved_record_count보다 크게 됩니다. 이를 감지해 구제 저장을 실행합니다.
            #   최악의 경우에도 누락 데이터는 다음 사이클 종료까지만 지연됩니다.
            if (self.recorded_data
                    and self.auto_save_path
                    and self.can_auto_save
                    and len(self.recorded_data) > self._saved_record_count):
                missed = len(self.recorded_data) - self._saved_record_count
                self.log(f"⚠ 미저장 기록 {missed}건 감지 → 구제 저장 재시도 중...")
                self.perform_auto_save()
                if len(self.recorded_data) == self._saved_record_count:
                    self.log(f"✅ 구제 저장 성공: {missed}건 복구 완료")
                else:
                    self.log(f"❌ 구제 저장도 실패 — 다음 사이클에 재시도됩니다.")

            # 47-2-4. [자동 갱신 로그] 자동 갱신 완료 시에만 로그를 남깁니다. 수동 갱신은 별도 로그가 있으므로 여기선 출력하지 않습니다.
            if not manual:
                self.log("데이터 갱신 완료")
        finally:
            self._refresh_lock.release()

    # 48. [자동 기록 중지] 실행 중인 자동 기록을 멈춥니다. 중지 확인 → 마지막 구제 저장 → 버튼/입력창 복원 순으로 진행됩니다.
    def stop_monitoring(self):
        # 48-1. [중지 확인 팝업] 실수 클릭을 막기 위해 "정말 중지하시겠습니까?" 확인 창을 띄웁니다.
        if not messagebox.askyesno("중지 확인", "정말 중지하시겠습니까?"):
            return
        self.is_monitoring = False
        # 48-2.   중지 후 인증키가 있는 상태라면 정류소 검색 버튼을 다시 활성화합니다.
        if self.key_locked:
            _s_on2 = self.get_btn_style("normal")
            _s_on2.pop("state", None)
            for _b in self.btn_searches:
                _b.config(state="normal", **_s_on2)
        self.log("🛑 자동 기록을 중지합니다.")

        # 48-3. [중지 전 마지막 저장] 모니터링이 완전히 멈추기 전에 미저장 데이터를 마지막으로 한 번 더 강제 저장합니다.
        #   마지막 사이클의 즉시 저장이 실패한 채로 중지하면 데이터가 영구 손실되므로
        #   중지 직전에 한 번 더 강제 저장을 시도합니다.
        if (self.recorded_data
                and self.auto_save_path
                and self.can_auto_save
                and len(self.recorded_data) > self._saved_record_count):
            missed = len(self.recorded_data) - self._saved_record_count
            self.log(f"⚠ 중지 시점 미저장 기록 {missed}건 → 최종 구제 저장 시도 중...")
            self.perform_auto_save()
            if len(self.recorded_data) == self._saved_record_count:
                self.log(f"✅ 최종 구제 저장 성공: {missed}건 복구 완료")
            else:
                self.log(f"❌ 최종 구제 저장 실패 — '다른 이름으로 엑셀 저장' 버튼으로 수동 저장해 주세요.")
        
        # 48-4. [입력창 복원] 갱신 주기 입력창을 다시 수정할 수 있게 열어줍니다.
        self.entry_refresh_interval.config(state='normal')
        self.lbl_auto_save_status.config(text=" ※ 자동 기록 시작 버튼을 작동시키면 도착 기록이 엑셀파일로 자동 저장됩니다.", fg="#e74c3c")
        self.update_button_states()

    # 49. [그룹 7] 데이터 분석 및 기록 — 도착 판정, 로그 기록, 엑셀 저장 함수 모음

    # 50. [도착 기록 삭제] 도착 기록 표의 내용을 전부 지웁니다. 실수 방지를 위해 확인 메시지를 먼저 보여줍니다.
    def clear_history(self, idx):
        # 50-1. [삭제 확인] "기록을 삭제하시겠습니까?" 확인 창을 띄웁니다.
        if messagebox.askyesno("삭제", f"정류소 {idx+1}의 모든 기록을 삭제하시겠습니까?\n기록을 삭제하더라도, 이미 저장된 기록은 삭제되지 않습니다."): 
            tree = self.trees_hist[idx]
            # 50-1-1. [표 비우기] 도착 기록 표의 모든 항목을 순서대로 제거합니다.
            for item in tree.get_children():
                tree.delete(item)
            self.log(f"정류소 {idx+1} 기록을 초기화했습니다.")

    # 51. [시간 형식 변환] 서버에서 받은 "20260101120000" 같은 숫자 덩어리를 "12:00" 형태로 바꿉니다.
    def format_hhmm(self, raw_str):
        if not raw_str or len(raw_str) < 4: return raw_str
        # 51-1. [긴 시간 문자열 처리] 8자리 이상이면 8~11번째 자리(HHMM 위치)를 추출합니다.
        if len(raw_str) >= 14: return f"{raw_str[8:10]}:{raw_str[10:12]}"
        return f"{raw_str[:2]}:{raw_str[2:4]}" 

    # 52. [로그 기록 함수] 시간 도장([HH:MM:SS])을 붙여 화면 하단 로그 창에 메시지를 한 줄 추가합니다. 최대 5,000줄까지만 유지합니다.
    def log(self, msg):
        # 52-1. [스레드 안전 처리] 백그라운드 스레드에서 호출 시 tkinter UI는 메인 스레드에서만 수정해야 합니다. root.after(0, ...)로 메인 스레드에 위임합니다.
        #         tk.Text 조작은 반드시 메인 스레드(root.after)를 통해 실행합니다.
        #         txt_log가 아직 생성되지 않은 경우(초기화 중)는 콘솔 출력으로 대체합니다.
        now = datetime.now().strftime("%H:%M:%S")
        try:
            key = self.service_key_var.get().strip()
            backup_key = self.backup_key_var.get().strip()
        except Exception:
            key, backup_key = "", ""

        # 52-2. [인증키 마스킹] 64자리 인증키가 로그에 그대로 노출되면 보안 위험이 있으므로 ****로 가려서 표시합니다.
        if key and len(key) > 4: msg = msg.replace(key, "********")
        if backup_key and len(backup_key) > 4: msg = msg.replace(backup_key, "********")
        
        line = f"[{now}] {msg}\n"

        # 52-3. [UI 준비 여부 확인] 화면이 다 그려지기 전에 호출될 수 있으므로, 로그 창(txt_log)이 준비됐는지 먼저 확인합니다.
        if not hasattr(self, 'txt_log'):
            # 52-3-1. [초기화 전 콘솔 출력] 화면이 아직 준비되지 않았으면 대신 콘솔(터미널)에 출력합니다.
            print(line, end="")
            return

        def _do_insert():
            try:
                # 52-3-1-1. [쓰기 권한 임시 해제] 로그 창은 평소에 읽기 전용이므로, 글자를 추가할 때만 잠깐 수정 가능 상태로 바꿨다가 다시 잠급니다.
                self.txt_log.config(state="normal")
                self.txt_log.insert(tk.END, line)
                self.txt_log.see(tk.END)
                # 52-3-1-2. [로그 줄 수 제한] 5,000줄을 넘으면 가장 오래된 위쪽 줄부터 지워서 메모리를 절약합니다.
                line_count = int(self.txt_log.index('end-1c').split('.')[0])
                if line_count > 5000:
                    self.txt_log.delete('1.0', '2.0')
                self.txt_log.config(state="disabled")
            except Exception:
                pass  # 5-29-3-1. 창이 닫히는 도중 예외가 발생해도 무시합니다.

        # 52-4. [메인 스레드 예약] root.after(0, _do_insert)로 실제 삽입 작업을 메인 스레드(이벤트 루프)에 안전하게 위임합니다.
        try:
            self.root.after(0, _do_insert)
        except Exception:
            pass

    # 53. [날짜시각 형식 변환] 서버의 "20260101120000" 형태 날짜를 "2026-01-01 12:00:00" 형태로 바꿉니다.
    def format_datetm(self, raw_tm):
        try:
            if raw_tm and len(raw_tm) >= 14: 
                return datetime.strptime(raw_tm, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
            return raw_tm
        except: return raw_tm

    # 54. [엑셀 저장 핵심 로직] 수집한 도착 기록을 날짜별 시트로 나누어 엑셀 파일(.xlsx)에 저장합니다. 완결 파일 자동 저장 기능 포함.
    def _core_excel_save_logic(self, target_path, save_completed=False):
        # 54-1. [완결 파일 저장 조건] save_completed=True이면, 새벽 3시 이전을 이전 영업일로 처리하는 기준으로 완결된 날의 데이터를 별도 파일로도 저장합니다.
        #   영업일(BizDate)이 처음 '완결'되는 순간(새 날짜 시트가 열릴 때)
        #   해당 날짜의 데이터만 Bus_Arrival_Log_YYYYMMDD_completed.xlsx 로 별도 저장합니다.
        try:
            import pandas as pd
            from datetime import datetime, timedelta
            from openpyxl.styles import Font, PatternFill, Alignment

            # 54-1-1. [데이터 변환] recorded_data 리스트를 pandas DataFrame(표)으로 변환합니다.
            cols = ["데이터시각", "도착정류소명", "노선", "차량번호", "운수사명"]
            df = pd.DataFrame(self.recorded_data, columns=cols)
            
            # 54-1-2. [영업일 계산] 새벽 00:00~02:59 사이에 도착한 버스는 실제로는 "어제 밤" 운행이므로 어제 날짜 영업일로 처리합니다.
            def get_biz_date(dt_str):
                try:
                    dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
                    if dt.hour < 3: dt -= timedelta(days=1) 
                    return dt.strftime("%Y-%m-%d")
                except: return "Unknown"
            
            df['BizDate'] = df['데이터시각'].apply(get_biz_date)

            # 54-1-3. [기존 영업일 목록] 현재 수집된 데이터에 포함된 모든 영업일 날짜를 추출합니다.
            biz_dates_in_data = set(df['BizDate'].unique())
            # 54-1-4. [오늘 영업일 계산] 현재 시각 기준으로 오늘의 영업일(새벽 3시 전이면 어제)을 계산합니다.
            now = datetime.now()
            current_biz = (now - timedelta(days=1)).strftime("%Y-%m-%d") if now.hour < 3 else now.strftime("%Y-%m-%d")
            # 54-1-5. [완결 날짜 판별] 데이터에는 있지만 오늘 영업일이 아닌 날짜는 운행이 끝난 "완결된 날"로 간주합니다.
            completed_dates = biz_dates_in_data - {current_biz, "Unknown"}
            
            # 54-1-6. [날짜별 시트 생성] 각 영업일 데이터를 별도 시트로 나누어 한 파일에 저장합니다.
            with pd.ExcelWriter(target_path, engine='openpyxl') as writer:
                for biz_date, group in df.groupby('BizDate'):
                    sheet_data = group.drop(columns=['BizDate'])
                    sheet_data.to_excel(writer, sheet_name=biz_date, index=False)
                    
                    worksheet = writer.sheets[biz_date]
                    
                    # 54-1-6-2-1. [머리글 서식] 첫 행(열 이름 행)을 하늘색 배경에 굵은 글씨로 꾸며 보기 쉽게 합니다.
                    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    header_font = Font(bold=True, size=11)
                    header_alignment = Alignment(horizontal='center', vertical='center')

                    # 54-1-6-2-2. [열 너비 자동 조정] 각 열에서 가장 긴 텍스트에 맞게 열 너비를 자동으로 조정합니다.
                    for idx, col in enumerate(sheet_data.columns):
                        cell = worksheet.cell(row=1, column=idx+1)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment

                        max_data_len = sheet_data[col].astype(str).map(len).max()
                        base_width = max(max_data_len, len(str(col)))
                        
                        # 54-1-6-2-2-1. [특정 열 개별 조정] 정류소 이름 열은 더 넓게, 시간 열은 더 좁게 개별적으로 조정합니다.
                        if col == "데이터시각":
                            adjusted_width = base_width + 1 
                        elif col == "도착정류소명":
                            # 54-1-6-2-2-1-1.   길이 기준보다 넓게 해야 하는 열: 기존 계산값에 1.5배 확장을 적용합니다.
                            adjusted_width = (base_width + 15) * 1.5
                        elif col == "노선":
                            adjusted_width = base_width + 8 
                        elif col == "운수사명":
                            adjusted_width = base_width + 7 
                        else:
                            adjusted_width = base_width + 5 
                        
                        col_letter = worksheet.cell(row=1, column=idx+1).column_letter
                        worksheet.column_dimensions[col_letter].width = adjusted_width

            # 54-1-7. [완결 파일 자동 저장] 특정 영업일 운행이 완전히 끝났을 때, 해당 날의 데이터만 따로 완결 파일(.._completed.xlsx)로 저장합니다.
            #   save_completed=True 이면서, 아직 저장하지 않은 완결 영업일이 있을 때만 저장합니다.
            if save_completed:
                new_completed = completed_dates - self._completed_dates_saved
                for biz_date in sorted(new_completed):
                    # 54-1-7-2-1. [완결 데이터 추출] 완결 처리할 영업일의 데이터만 별도 추출합니다.
                    day_df = df[df['BizDate'] == biz_date].drop(columns=['BizDate'])
                    if day_df.empty:
                        continue
                    # 54-1-7-2-2. [완결 파일 이름] "Bus_Arrival_Log_20260101_completed.xlsx" 형태의 이름으로 저장합니다.
                    safe_date = biz_date.replace("-", "")
                    completed_name = f"Bus_Arrival_Log_{safe_date}_completed.xlsx"
                    completed_path = os.path.join(self.current_dir, completed_name)
                    try:
                        with pd.ExcelWriter(completed_path, engine='openpyxl') as cw:
                            day_df.to_excel(cw, sheet_name=biz_date, index=False)
                            ws = cw.sheets[biz_date]
                            hfill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                            hfont = Font(bold=True, size=11)
                            halign = Alignment(horizontal='center', vertical='center')
                            for ci, col in enumerate(day_df.columns):
                                cell = ws.cell(row=1, column=ci+1)
                                cell.fill = hfill; cell.font = hfont; cell.alignment = halign
                                max_len = day_df[col].astype(str).map(len).max()
                                base_w  = max(max_len, len(str(col)))
                                if col == "도착정류소명":
                                    col_w = (base_w + 15) * 1.5
                                elif col == "데이터시각":
                                    col_w = base_w + 1
                                elif col == "노선":
                                    col_w = base_w + 8
                                elif col == "운수사명":
                                    col_w = base_w + 7
                                else:
                                    col_w = base_w + 5
                                col_letter = ws.cell(row=1, column=ci+1).column_letter
                                ws.column_dimensions[col_letter].width = col_w
                        # 54-1-7-2-2-1. [중복 저장 방지] 같은 날의 완결 파일이 두 번 생기지 않도록 완결 처리된 날짜를 집합에 기록합니다.
                        self._completed_dates_saved.add(biz_date)
                        self.log(f"📁 영업일 완결 파일 저장: {completed_name}")
                    except Exception as ce:
                        self.log(f"⚠ 완결 파일 저장 실패 ({biz_date}): {ce}")

            # 54-1-8. [저장 완료 표시] 저장 성공 시 _saved_record_count를 현재 기록 수로 갱신합니다. 이후 구제 저장 로직이 중복 저장을 시도하지 않도록 합니다.
            #   이 값은 refresh_data() 사이클 종료 시 누락 여부 판단에 사용됩니다.
            self._saved_record_count = len(self.recorded_data)
            return True

        except PermissionError:
            self.log("⚠ 엑셀 파일이 열려 있어 저장을 건너뜁니다. (파일을 닫아주세요)")
            return False
        except Exception as e:
            self.log(f"❌ 엑셀 저장 중 오류 발생: {e}")
            return False
        
    # 55. [다른 이름으로 저장] 사용자가 직접 저장 위치와 파일 이름을 지정하여 엑셀 파일을 저장합니다.
    def save_to_excel(self):
        # 55-1. [빈 데이터 검사] 저장할 도착 기록이 하나도 없으면 경고 메시지를 보여주고 중단합니다.
        if not self.recorded_data:
            messagebox.showwarning("알림", "저장할 데이터가 없습니다.")
            return 
        
        # 55-2. [저장 경로 선택] 파일 탐색기를 열어 저장 위치와 파일 이름을 사용자가 선택하도록 합니다.
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel 파일", "*.xlsx")],
            title="다른 이름으로 저장"
        ) 
        
        if path: 
            # 55-2-1. [파일 저장 실행] 선택한 경로에 실제로 엑셀 파일을 저장합니다.
            if self._core_excel_save_logic(path):
                messagebox.showinfo("저장 완료", f"영업일 기준으로 분류되어 저장되었습니다.\n경로: {path}")
                self.log(f"💾 수동 엑셀 저장 완료: {path}")

    # 56. [자동 백업 저장] 기록이 있을 때마다 프로그램 폴더의 지정 경로에 엑셀 파일을 자동으로 저장합니다.
    def perform_auto_save(self):
        # 56-1. [저장 조건 확인] 저장할 기록이 있고, 자동 저장이 켜져 있고, 저장 경로가 설정되어 있는지 확인합니다.
        if not self.recorded_data or not self.auto_save_path or not self.can_auto_save: 
            return 
        
        # 56-2. [스레드 안전] _save_lock 으로 한 스레드씩만 파일을 씁니다.
        #   자동 모니터링 스레드와 수동 갱신 스레드가 동시에 파일을 덮어써
        #   더 이른 스냅샷이 나중에 기록되면 데이터가 사라지는 버그를 방지합니다.
        #   lock 취득 전에 재진입 여부를 확인해 불필요한 중복 저장을 줄입니다.
        if not self._save_lock.acquire(blocking=False):
            # 56-2-1. 다른 스레드가 저장 중이면 기다리지 않고 바로 건너뜁니다. 현재 저장 중인 스냅샷에 전체 데이터가 포함되므로 손실 없습니다.
            # 56-2-2. 해당 스레드가 완료될 때 recorded_data 의 전체 스냅샷을 포함하므로 데이터 손실은 없습니다.
            return
        try:
            # 56-2-3. [파일 쓰기 실행] _core_excel_save_logic()으로 조용히 파일을 저장합니다. save_completed=True로 전달하여 완결 파일도 함께 처리합니다.
            #         save_completed=True 로 전달해 완결된 영업일 파일도 함께 저장합니다.
            self._core_excel_save_logic(self.auto_save_path, save_completed=True)
        finally:
            self._save_lock.release()

    # 57. [표 갱신] 화면의 표를 완전히 지우고 새 데이터로 다시 채웁니다.
    def refresh_tree(self, tree, data):
        # 57-1. [기존 데이터 삭제] 표의 모든 기존 항목을 삭제합니다.
        for i in tree.get_children(): tree.delete(i)
        # 57-2. [새 데이터 삽입] 새로운 데이터 목록을 순서대로 표에 추가합니다.
        for row in data: tree.insert("", "end", values=row)

# O. [진짜 실행] 여기가 프로그램의 시작 버튼입니다!
if __name__ == "__main__":
    import traceback

    # 1. [오류 기록기] 예상치 못한 오류로 프로그램이 멈출 때 오류 내용을 errorlog.txt 파일에 저장합니다. 실행 파일 또는 .py 파일과 같은 폴더에 저장됩니다.
    #   - --windowed (pyinstaller) 로 빌드된 실행파일에서도 동작합니다.
    #   - 실행 파일(또는 .py) 이 있는 폴더에 errorlog.txt 를 생성합니다.
    def _get_error_log_path():
        """실행 파일 또는 소스 파일이 위치한 폴더에 errorlog.txt 경로를 반환합니다."""
        if getattr(sys, 'frozen', False):
            # 1-2-1. PyInstaller --windowed로 빌드된 실행 파일에서는 sys.executable 경로를 기준으로 로그 파일 위치를 계산합니다.
            base = os.path.dirname(os.path.abspath(sys.executable))
            # 1-2-2.   macOS .app 번들 형태인 경우 Contents/MacOS 디렉토리 위로 3단계 올라가야 실제 저장 폴더에 접근할 수 있습니다.
            if sys.platform == "darwin" and base.endswith("Contents/MacOS"):
                base = os.path.dirname(os.path.dirname(os.path.dirname(base)))
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, "errorlog.txt")

    def _write_error_log(exc_type, exc_value, exc_tb):
        """예외 정보를 errorlog.txt 에 기록하고, 가능하면 messagebox 로도 알립니다."""
        try:
            log_path = _get_error_log_path()
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            tb_text = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"\n{'='*60}\n")
                f.write(f"[{timestamp}] 프로그램 오류 발생\n")
                f.write(f"{'='*60}\n")
                f.write(tb_text)
                f.write("\n")
            # 1-2-3. 오류 발생 시 창이 살아있으면 메시지 박스로도 사용자에게 알립니다.
            try:
                import tkinter.messagebox as _mb
                short_msg = f"{exc_type.__name__}: {exc_value}"
                _mb.showerror(
                    "프로그램 오류",
                    f"예기치 않은 오류가 발생했습니다.\n\n"
                    f"{short_msg}\n\n"
                    f"오류 내용이 다음 파일에 저장되었습니다:\n{log_path}"
                )
            except Exception:
                pass
        except Exception:
            pass  # 로그 기록 자체가 실패해도 프로그램이 이상하게 종료되지 않습니다.

    # 2. [전역 예외 처리기] 어디서 발생했든 처리되지 않은 예외를 가로채어 errorlog.txt에 기록합니다.
    def _global_exception_handler(exc_type, exc_value, exc_tb):
        _write_error_log(exc_type, exc_value, exc_tb)
        sys.__excepthook__(exc_type, exc_value, exc_tb)  # 원래 동작도 유지
    sys.excepthook = _global_exception_handler

    # 3. [오류 출력 저장] --windowed 방식으로 실행파일 빌드 시 콘솔이 없어 stderr가 사라지는 문제를 방지합니다. stderr를 파일로 저장합니다.
    try:
        if sys.stderr is None or getattr(sys, 'frozen', False):
            sys.stderr = open(_get_error_log_path().replace("errorlog.txt", "stderr.txt"),
                              "a", encoding="utf-8")
    except Exception:
        pass

    # 4. [메인 창 생성] tkinter 루트 창을 하나 새로 만듭니다. 이 창이 프로그램의 메인 화면이 됩니다.
    try:
        root = tk.Tk()
        # 4-1. [프로그램 조립] SeoulBusArrivalRecorder 클래스에 메인 창을 넘겨주어 모든 화면과 기능을 만들고 초기화합니다.
        app = SeoulBusArrivalRecorder(root)
        # 4-2. [이벤트 루프 진입] root.mainloop()을 실행하여 창이 열려있는 동안 사용자의 클릭·입력 이벤트를 계속 처리합니다.
        root.mainloop()
    except Exception as e:
        _write_error_log(type(e), e, e.__traceback__)