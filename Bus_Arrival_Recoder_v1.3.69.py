import tkinter as tk # 1. [ì¤€ë¹„ë¬¼] ì»´í“¨í„° í™”ë©´ì— ì°½ì„ ë„ìš°ëŠ” ê¸°ë³¸ ë„êµ¬ìƒìì…ë‹ˆë‹¤.
from tkinter import ttk, messagebox, filedialog # 1-1. [ì¤€ë¹„ë¬¼] ë” ì˜ˆìœ ë²„íŠ¼, ì•Œë¦¼ì°½, íŒŒì¼ ì°¾ê¸° ê¸°ëŠ¥ì„ ê°€ì ¸ì˜µë‹ˆë‹¤.
import threading # 1-2. [ì¤€ë¹„ë¬¼] ì—¬ëŸ¬ ê°€ì§€ ì¼ì„ ë™ì‹œì— ì²˜ë¦¬í•´ì£¼ëŠ” 'ì¼ê¾¼'ì„ ë¶€ë¦…ë‹ˆë‹¤.
import time # 1-3. [ì¤€ë¹„ë¬¼] ì‹œê°„ì„ ì¬ê±°ë‚˜ ì ì‹œ ê¸°ë‹¤ë¦¬ê²Œ í•˜ëŠ” ì‹œê³„ë¥¼ ê°€ì ¸ì˜µë‹ˆë‹¤.
import os # 1-4. [ì¤€ë¹„ë¬¼] ì»´í“¨í„°ì˜ íŒŒì¼ì´ë‚˜ í´ë”ë¥¼ ë‹¤ë£¨ëŠ” ë„êµ¬ë¥¼ ê°€ì ¸ì˜µë‹ˆë‹¤.
import sys # 1-5. [ì¤€ë¹„ë¬¼] í”„ë¡œê·¸ë¨ ì„¤ì •ì´ë‚˜ ì»´í“¨í„° ì •ë³´ë¥¼ í™•ì¸í•˜ëŠ” ë„êµ¬ì…ë‹ˆë‹¤.
from cryptography.fernet import Fernet # 1-6. [ì¤€ë¹„ë¬¼] ê¸€ìë¥¼ ì•”í˜¸ë¡œ ë°”ê¾¸ì–´ ë¹„ë°€ì„ ì§€ì¼œì£¼ëŠ” ë„êµ¬ì…ë‹ˆë‹¤.
MASTER_KEY = b'u7_K-5D4fR9zP2mN8xL1qJ6vH3sB0tG9wE8rT7yU4iA=' 
cipher_suite = Fernet(MASTER_KEY) # 1-7. [ì¤€ë¹„ë¬¼] ì•”í˜¸ë¥¼ í’€ê±°ë‚˜ ì ê·¸ëŠ” ì—´ì‡ ë¥¼ ë§Œë“­ë‹ˆë‹¤.
import webbrowser # 1-8. [ì¤€ë¹„ë¬¼] ì¸í„°ë„· ì‚¬ì´íŠ¸ë¥¼ ì—´ì–´ì£¼ëŠ” ë„êµ¬ì…ë‹ˆë‹¤.
import platform # 1-9. [ì¤€ë¹„ë¬¼] ì§€ê¸ˆ ì“°ëŠ” ì»´í“¨í„°ê°€ ì–´ë–¤ ì¢…ë¥˜ì¸ì§€ ì•Œë ¤ì£¼ëŠ” ë„êµ¬ì…ë‹ˆë‹¤.
from datetime import datetime, timedelta # 1-10. [ì¤€ë¹„ë¬¼] ì˜¤ëŠ˜ ë‚ ì§œì™€ í˜„ì¬ ì‹œê°„ì„ ê³„ì‚°í•˜ëŠ” ë„êµ¬ì…ë‹ˆë‹¤.
import xml.etree.ElementTree as ET # 1-11. [ì¤€ë¹„ë¬¼] ë³µì¡í•œ ì¸í„°ë„· ë¬¸ì„œë¥¼ ì½ê¸° ì¢‹ê²Œ ì •ë¦¬í•´ì£¼ëŠ” ë„êµ¬ì…ë‹ˆë‹¤.
from urllib.parse import unquote # 1-12. [ì¤€ë¹„ë¬¼] ì¸í„°ë„· ì£¼ì†Œì— ë“¤ì–´ìˆëŠ” íŠ¹ìˆ˜ ë¬¸ìë¥¼ ê¸€ìë¡œ ë°”ê¿”ì¤ë‹ˆë‹¤.

# 2. [ì£¼ë³€ ê²€ì‚¬] ì§€ê¸ˆ ë‚´ ì»´í“¨í„° ìƒíƒœê°€ ì–´ë–¤ì§€ í™•ì¸í•´ë´…ë‹ˆë‹¤.
# 2-1. ìœˆë„ìš°ì¸ì§€ ë§¥ì¸ì§€ í™•ì¸í•´ì„œ ì´ë¦„ì„ ê¸°ì–µí•´ë‘¡ë‹ˆë‹¤.
CURRENT_OS = platform.system() 
# 2-2. [ì„ íƒ]ë§¥os ì•± ì‹¤í–‰ ê¶Œí•œ(Gatekeeper) ê²©ë¦¬ ì†ì„± ì œê±° í„°ë¯¸ë„ ëª…ë ¹ì–´ xattr -d com.apple.quarantine /ê²½ë¡œ/to/your/app_file(í˜¹ì€ ë“œë˜ê·¸)

# 3. [ì˜ˆìœ ê¸€ì”¨] í™”ë©´ì— ë‚˜ì˜¬ ê¸€ìë“¤ì˜ ê¸€ê¼´ê³¼ í¬ê¸°ë¥¼ ë¯¸ë¦¬ ì •í•©ë‹ˆë‹¤.
if CURRENT_OS == "Windows":
    # 3-1. ìœˆë„ìš° ì»´í“¨í„°ìš© ê¸€ê¼´ê³¼ í¬ê¸° ì„¤ì •ì…ë‹ˆë‹¤.
    FONT_MAIN = "ë§‘ì€ ê³ ë”•" 
    FONT_SUB = "ë‹ì›€"       
    FONT_MONO = "Consolas" 
    # 3-1-1. ê¸€ì í¬ê¸°ë“¤ì„ ë¯¸ë¦¬ ì •í•´ë‘¡ë‹ˆë‹¤.
    SZ_L = 19  
    SZ_M = 11
    SZ_S = 9
    SZ_XS = 8
    SZ_XXS = 7
else: 
    # 3-2. ë§¥(Mac) ì»´í“¨í„°ìš© ê¸€ê¼´ê³¼ í¬ê¸° ì„¤ì •ì…ë‹ˆë‹¤.
    FONT_MAIN = "AppleGothic"         
    FONT_SUB = "Apple SD Gothic Neo" 
    FONT_MONO = "Menlo"              
    # 3-2-1. ë§¥ì€ ê¸€ìê°€ ì‘ê²Œ ë³´ì—¬ì„œ 1.4ë°° ì •ë„ í‚¤ì›Œì¤ë‹ˆë‹¤.
    # 3-2-1. macOS ë²„íŠ¼ í¬ê¸°ê°€ Windowsë³´ë‹¤ í¬ê²Œ ë³´ì´ëŠ” ë¬¸ì œë¥¼ ë°©ì§€í•˜ê¸° ìœ„í•´
    #        ë°°ìœ¨ì„ 1.4â†’1.15 ìˆ˜ì¤€ìœ¼ë¡œ ë‚®ì¶¥ë‹ˆë‹¤.
    SZ_L = 19
    SZ_M = 11
    SZ_S = 9
    SZ_XS = 8
    SZ_XXS = 7

# 4. [ì°½ê³  ê²€ì‚¬] í”„ë¡œê·¸ë¨ì„ ëŒë¦´ ë•Œ í•„ìš”í•œ íŠ¹ìˆ˜ ë„êµ¬ë“¤ì´ ë‹¤ ìˆëŠ”ì§€ í™•ì¸í•©ë‹ˆë‹¤.
try:
    # 4-1. ì¸í„°ë„· ì—°ê²°ê³¼ ì—‘ì…€ ì‘ì—…ì— ê¼­ í•„ìš”í•œ ì™¸ë¶€ ë„êµ¬ë“¤ì„ ë¶ˆëŸ¬ì˜µë‹ˆë‹¤.
    import requests # ì¸í„°ë„· ì„œë²„ì— ì§ˆë¬¸ì„ ë³´ë‚´ëŠ” ì¼ê¾¼
    import pandas as pd # í‘œ í˜•íƒœì˜ ë°ì´í„°ë¥¼ ì—‘ì…€ì²˜ëŸ¼ ë‹¤ë£¨ëŠ” ë„êµ¬
    from selenium import webdriver # ì¸í„°ë„· ì°½ì„ ìë™ìœ¼ë¡œ ì¡°ì¢…í•˜ëŠ” ë¡œë´‡
    from selenium.webdriver.chrome.service import Service 
    from selenium.webdriver.chrome.options import Options 
    from selenium.webdriver.common.by import By 
    from selenium.webdriver.support.ui import WebDriverWait 
    from selenium.webdriver.support import expected_conditions as EC 
    from webdriver_manager.chrome import ChromeDriverManager 
except ImportError as e:
    # 4-2. [ì˜¤ë¥˜ ëŒ€ì²˜] í•„ìš”í•œ ë„êµ¬ê°€ ì—†ìœ¼ë©´ ì•ˆë‚´ ì°½ì„ ë„ìš°ê³  í”„ë¡œê·¸ë¨ì„ ì¢…ë£Œí•©ë‹ˆë‹¤.
    root = tk.Tk() 
    root.withdraw() 
    msg = f"í•„ìˆ˜ ë¼ì´ë¸ŒëŸ¬ë¦¬ê°€ ì—†ìŠµë‹ˆë‹¤.\nì•„ë˜ ëª…ë ¹ì–´ë¥¼ í„°ë¯¸ë„(CMD)ì— ì‹¤í–‰í•´ ì£¼ì„¸ìš”.\n\npip install requests pandas openpyxl cryptography selenium webdriver-manager\n\nì—ëŸ¬ ë‚´ìš©: {e}"
    messagebox.showerror("ì‹¤í–‰ ì˜¤ë¥˜", msg) 
    exit() 

# 5. [ê¸°ê³„ ì„¤ê³„ë„] ì„œìš¸ ë²„ìŠ¤ ë„ì°©ì„ ê¸°ë¡í•˜ëŠ” í”„ë¡œê·¸ë¨ì˜ ì „ì²´ ê¸°ëŠ¥ ëª¨ìŒ
class SeoulBusArrivalRecorder:

    # 1ê·¸ë£¹ : ì´ˆê¸°ë‹¨ê³„(Initialization)

    # 5-1. [íƒ„ìƒ] í”„ë¡œê·¸ë¨ì´ ì²˜ìŒ ì¼œì§ˆ ë•Œ ì‹¤í–‰ë˜ëŠ” ì¤€ë¹„ ê³¼ì •
    def __init__(self, root):
        # 5-1-1. í™”ë©´ì˜ ê¸°ë³¸ ì •ë³´ë“¤ì„ ì„¤ì •í•©ë‹ˆë‹¤.
        self.root = root 
        self.root.title("ì„œìš¸ë²„ìŠ¤ ì •ë¥˜ì†Œ ë“€ì–¼ ë„ì°©ê¸°ë¡ í”„ë¡œê·¸ë¨ v1.3.69") 
        self.root.geometry("1200x800") 
        # 5-1-1-1. ì°½ì´ ë„ˆë¬´ ì‘ì•„ì§€ë©´ í™”ë©´ì´ ê¹¨ì§€ë¯€ë¡œ ìµœì†Œ í¬ê¸°ë¥¼ ì •í•©ë‹ˆë‹¤.
        self.root.minsize(960, 400) 
        
        # 5-1-2. ë””ìì¸ í…Œë§ˆë¥¼ ì„¤ì •í•©ë‹ˆë‹¤.
        self.style = ttk.Style() 
        try:
            self.style.theme_use('clam') 
        except:
            pass

        # 5-1-3. í”„ë¡œê·¸ë¨ì´ ëŒì•„ê°€ëŠ” ìœ„ì¹˜(í´ë”)ë¥¼ ì •í™•íˆ ì°¾ì•„ëƒ…ë‹ˆë‹¤.
        if getattr(sys, 'frozen', False):
            # 5-1-3-1. ì™„ì„±ëœ ì‹¤í–‰ íŒŒì¼(.exe) ìƒíƒœì¼ ë•Œ í´ë” ìœ„ì¹˜ ì°¾ê¸°
            self.resource_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(sys.executable)))
            base_exe_path = os.path.abspath(sys.executable)
            if CURRENT_OS == "Darwin": 
                self.current_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(base_exe_path))))
            else:
                self.current_dir = os.path.dirname(base_exe_path)
        else:
            # 5-1-3-2. ì†ŒìŠ¤ ì½”ë“œ(.py) ìƒíƒœë¡œ ì‹¤í–‰ ì¤‘ì¼ ë•Œ í´ë” ìœ„ì¹˜ ì°¾ê¸°
            self.resource_dir = os.path.dirname(os.path.abspath(__file__))
            self.current_dir = self.resource_dir

        # 5-1-4. í•„ìš”í•œ íŒŒì¼ë“¤ì˜ ìœ„ì¹˜ë¥¼ ë¯¸ë¦¬ ì•½ì†í•´ë‘¡ë‹ˆë‹¤.
        self.key_file_path = os.path.join(self.current_dir, "key.cfg")
        icon_path_png = os.path.join(self.resource_dir, "icon.png")
        icon_path_ico = os.path.join(self.resource_dir, "icon.ico")
        icon_path_icns = os.path.join(self.resource_dir, "icon.icns")

        # 5-1-5. ìœˆë„ìš° ì»´í“¨í„°ë¼ë©´ ì•„ì´ì½˜ì´ ì˜ˆì˜ê²Œ ë³´ì´ë„ë¡ íŠ¹ë³„ ì„¤ì •ì„ í•©ë‹ˆë‹¤.
        if CURRENT_OS == "Windows":
            import ctypes
            try:
                myappid = 'Bus_Arrival_Recoder(v1.3.49)' 
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
            except: pass

        # 5-1-6. í”„ë¡œê·¸ë¨ì˜ ì•„ì´ì½˜ ì´ë¯¸ì§€ë¥¼ ì°½ì— ì…í™ë‹ˆë‹¤.
        try:
            if os.path.exists(icon_path_png):
                img = tk.PhotoImage(file=icon_path_png)
                self.root.iconphoto(True, img) 
            if CURRENT_OS == "Windows" and os.path.exists(icon_path_ico):
                self.root.iconbitmap(icon_path_ico)
        except Exception as e:
            print(f"ì•„ì´ì½˜ ì„¤ì • ì‹¤íŒ¨: {e}")

        # 5-1-7. í™”ë©´ ì œì¼ ìœ„ìª½ì— ë©”ë‰´ ë°” ê³µê°„ì„ ë§Œë“­ë‹ˆë‹¤.
        self.top_container = tk.Frame(self.root)
        self.top_container.pack(side="top", fill="x")

        # 5-1-8. [ê¸°ì–µ ì €ì¥ì†Œ] í”„ë¡œê·¸ë¨ì´ ëŒì•„ê°€ëŠ” ë™ì•ˆ ê¸°ì–µí•´ì•¼ í•  ì†Œì¤‘í•œ ì •ë³´ë“¤
        self.service_key_var = tk.StringVar(value="") # ë©”ì¸ ì—´ì‡ (ì¸ì¦í‚¤)
        self.backup_key_var = tk.StringVar(value="")  # ë¹„ìƒìš© ì—´ì‡ (ë°±ì—…í‚¤)
        self.key_locked = False # í‚¤ ì…ë ¥ì°½ ì ê¸ˆ ìƒíƒœ

        # 5-1-8-1. ê²€ìƒ‰ ë²„íŠ¼ê³¼ í™•ì •ëœ ì—´ì‡ ë“¤ì„ ê´€ë¦¬í•  ê³µê°„ì…ë‹ˆë‹¤.
        self.btn_searches = [] 
        self.final_main_key = ""
        self.final_backup_key = ""

        # 5-1-8-2. ë²„ìŠ¤ ì •ë¥˜ì†Œ ì •ë³´ì™€ ê°ì‹œ ìƒíƒœë¥¼ ì €ì¥í•©ë‹ˆë‹¤.
        self.ars_ids = [tk.StringVar(), tk.StringVar()] # ì •ë¥˜ì†Œ ë²ˆí˜¸ 2ê°œ
        self.refresh_interval_var = tk.StringVar(value="60") # ê°±ì‹  ì‹œê°„ (ê¸°ë³¸ 60ì´ˆ)
        self.is_monitoring = False # ì§€ê¸ˆ ê°ì‹œ ì¤‘ì¸ì§€ ì•„ë‹Œì§€ (ON/OFF)
        
        # 5-1-8-3. ì—‘ì…€ ì €ì¥ ìœ„ì¹˜ì™€ ìë™ ì €ì¥ í—ˆìš© ì—¬ë¶€ì…ë‹ˆë‹¤.
        self.auto_save_path = None # ìë™ ì €ì¥í•  ì—‘ì…€ íŒŒì¼ ìœ„ì¹˜
        self.can_auto_save = True # ì €ì¥ì„ í•´ë„ ë˜ëŠ”ì§€ í™•ì¸í•˜ëŠ” ìŠ¤ìœ„ì¹˜
        
        # 5-1-8-4. ë¬¸ì œê°€ ìƒê²¼ì„ ë•Œ ì“¸ ë¹„ìƒìš© API ì‚¬ìš© ìƒíƒœì…ë‹ˆë‹¤.
        self.use_fallback_api = False # ë¹„ìƒìš© API 1ë²ˆ ì‚¬ìš© ì—¬ë¶€
        self.use_fallback_pos_api = False # ë¹„ìƒìš© API 2ë²ˆ ì‚¬ìš© ì—¬ë¶€
        
        # 5-1-8-5. ì„ íƒëœ ì •ë¥˜ì†Œ ì •ë³´ì™€ ë²„ìŠ¤ íšŒì‚¬ ì´ë¦„ ë“±ì„ ëª¨ì•„ë‘¡ë‹ˆë‹¤.
        self.target_st_info = [{}, {}] # ì„ íƒí•œ ë²„ìŠ¤ ì •ë³´ ì €ì¥ì†Œ
        self.recorded_data = [] # ê¸°ë¡ëœ ë‚´ìš©ì„ ì ì‹œ ë‹´ì•„ë‘ëŠ” ë°”êµ¬ë‹ˆ
        self.last_arrival_logs = [{}, {}] # ì¤‘ë³µ ê¸°ë¡ì„ ë§‰ê¸° ìœ„í•œ ë©”ëª¨ì¥
        self.route_corp_map = {} # ë²„ìŠ¤ íšŒì‚¬ ì´ë¦„ ì €ì¥ì†Œ
        self.rid_to_rnm = {} # ë²„ìŠ¤ IDë¥¼ ë²ˆí˜¸ë¡œ ë°”ê¿”ì£¼ëŠ” ì‚¬ì „
        self.excel_multi_corp_map = {} # ì—‘ì…€ì—ì„œ ì½ì–´ì˜¨ íšŒì‚¬ ì •ë³´
        self.temp_pos_data = {} # POS ë°ì´í„°ë¥¼ ì„ì‹œë¡œ ì €ì¥í•˜ëŠ” ê³³
        # 5-1-8-6. ì°¨ëŸ‰ë²ˆí˜¸ ìºì‹œ: ARR1/ARR2 ì‘ë‹µì—ì„œ ë°›ì€ ì°¨ëŸ‰ë²ˆí˜¸ë¥¼ ì €ì¥í•´ UID í´ë°±ì— í™œìš©í•©ë‹ˆë‹¤.
        #   {(st_id, rid): (plainNo1, plainNo2)}  â† ì •ë¥˜ì†Œ+ë…¸ì„  ì¡°í•© í‚¤
        self.veh_cache = {}
        # 5-1-8-7. POS1 ì¼ì‹œì •ì§€ í…Œì´ë¸”: ì°¨ëŸ‰ ì—†ìŒ í™•ì¸ ì‹œ ì²«ì°¨ ì‹œê°ê¹Œì§€ í˜¸ì¶œì„ ê±´ë„ˆëœë‹ˆë‹¤.
        #   {rid: datetime}  â€” ì •ì§€ í•´ì œ ì‹œê° (ì´ ì‹œê°ì´ ë˜ë©´ ë‹¤ì‹œ í˜¸ì¶œ í—ˆìš©)
        self.pos_suspend_until = {}
        self.pos_resume_logged = set()  # POS ì¬ê°œ ë¡œê·¸ ì¤‘ë³µ ë°©ì§€: í•œ ë²ˆë§Œ ê¸°ë¡
        # 5-1-8-8. ë§ˆì§€ë§‰ ë‚ ì§œ ì²´í¬: ìì • ì´í›„ suspend ì´ˆê¸°í™”ì— ì‚¬ìš©í•©ë‹ˆë‹¤.
        self._last_date = datetime.now().date()
        # 5-1-8-9. SLST(getStaionByRoute) ìºì‹œ: ë…¸ì„  ID â†’ ì „ì²´ ì •ë¥˜ì†Œ ëª©ë¡.
        #   on_station_select ì—ì„œ ë…¸ì„ ë‹¹ 1íšŒë§Œ í˜¸ì¶œí•˜ê³  confirm_selection ì—ì„œ ì¬ì‚¬ìš©í•©ë‹ˆë‹¤.
        #   ì •ë¥˜ì†Œ ì„ íƒì´ ë°”ë€” ë•Œë§ˆë‹¤ ì´ˆê¸°í™”ë©ë‹ˆë‹¤.
        #   {busRouteId: root_element}
        self._strt_cache = {}
        # 5-1-8-10-1. SLST ord ìºì‹œ: on_station_select ì—ì„œ ë¯¸ë¦¬ êµ¬í•´ë‘” (ars_id, seq) ë¥¼ ì €ì¥í•©ë‹ˆë‹¤.
        #   {busRouteId: ord_value}  â€” ì •ë¥˜ì†Œê°€ ë°”ë€” ë•Œë§ˆë‹¤ ì´ˆê¸°í™”ë©ë‹ˆë‹¤.
        self._strt_ord_cache = {}
        # 5-1-8-10. ìë™ ì™„ê²° íŒŒì¼ ì €ì¥ìš©: ì´ë¯¸ ì™„ê²° íŒŒì¼ì„ ì €ì¥í•œ ì˜ì—…ì¼ì„ ê¸°ì–µí•©ë‹ˆë‹¤.
        #   {"YYYY-MM-DD"} í˜•íƒœì˜ ì§‘í•©ìœ¼ë¡œ ì¤‘ë³µ ì €ì¥ì„ ë°©ì§€í•©ë‹ˆë‹¤.
        self._completed_dates_saved = set()

        # 5-1-9. APIë¥¼ ëª‡ ë²ˆ í˜¸ì¶œí–ˆëŠ”ì§€ ì„¸ëŠ” í†µê³„ ìˆ«ìíŒì…ë‹ˆë‹¤.
        # 5-1-9-1. í•©ì‚° í†µê³„(ë©”ì¸í‚¤ + ë°±ì—…í‚¤ í•©ê³„) â€” ë©”ì¸ í™”ë©´ ìƒë‹¨ì— í‘œì‹œë©ë‹ˆë‹¤.
        self.api_stats = {
            "ARR1": 0, "ARR2": 0, "SINF": 0, "POS1": 0, "POS2": 0, "SCNM": 0, "SCID": 0, "RINF": 0, "SLST": 0, "VLD": 0
        }
        # 5-1-9-2. ë©”ì¸/ë°±ì—… í‚¤ë³„ ê°œë³„ í†µê³„ â€” API ìƒì„¸ ì°½ì—ì„œ êµ¬ë¶„ í‘œì‹œë©ë‹ˆë‹¤.
        self.api_stats_by_key = {
            "main": {k: 0 for k in self.api_stats},
            "back": {k: 0 for k in self.api_stats},
        }
        
        # 5-1-10. ì¸í„°ë„·ì„ í†µí•´ ì •ë³´ë¥¼ ê°€ì ¸ì˜¬ ì£¼ì†Œ ëª©ë¡ì…ë‹ˆë‹¤.
        self.api_urls = {
            "ARR1": "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRoute",
            "ARR2": "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRouteAll",
            "POS1": "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRtid",
            "SINF": "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
            "POS2": "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRouteSt",
            "SCNM": "http://ws.bus.go.kr/api/rest/stationinfo/getStationByName",
            "SCID": "http://ws.bus.go.kr/api/rest/stationinfo/getRouteByStation",
            "RINF": "http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo",
            "SLST": "http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute",
            "VLD": "http://ws.bus.go.kr/api/rest/busRouteInfo/getBusRouteList  (ì¸ì¦í‚¤ ê²€ì¦)"
        }
        
        self.stats_win = None # í†µê³„ ì°½ì„ ì €ì¥í•  ë¹ˆ ê³µê°„

        # 5-1-11. [ìµœì¢… ì‹œì‘] ì €ì¥ëœ ì—´ì‡ ë¥¼ ë¶ˆëŸ¬ì˜¤ê³  í™”ë©´ì„ ê·¸ë¦° ë’¤ ì •ë³´ë¥¼ ê°€ì ¸ì˜µë‹ˆë‹¤.
        self.load_saved_key() # ì €ì¥ëœ ì¸ì¦í‚¤ê°€ ìˆìœ¼ë©´ ë¶ˆëŸ¬ì˜µë‹ˆë‹¤.
        self.setup_ui() 
        self.root.after(500, self.start_excel_download_thread) 

    # 2ê·¸ë£¹ : ë©”ì¸ UI (Main UI)

    # 5-2. [ê·¸ë¦¼ ê·¸ë¦¬ê¸°] í™”ë©´ì˜ ì „ì²´ì ì¸ ë ˆì´ì•„ì›ƒê³¼ ë²„íŠ¼ë“¤ì„ ë°°ì¹˜í•˜ëŠ” í•¨ìˆ˜
    def setup_ui(self):
        # 5-2-1. ë§¥ ì»´í“¨í„°ë¼ë©´ ê¸€ììƒ‰ì„ ì•½ê°„ ì¡°ì ˆí•©ë‹ˆë‹¤.
        mac_lbl_opts = {"fg": "#2d3436"} if CURRENT_OS == "Darwin" else {}

        # 5-2-2. í™”ë©´ ì œì¼ ìƒë‹¨ì— í”„ë¡œê·¸ë¨ ì•ˆë‚´ ë¬¸êµ¬ë¥¼ ë„£ìŠµë‹ˆë‹¤.
        frame_status = tk.Frame(self.top_container, pady=2, bg="#ecf0f1")
        frame_status.pack(fill="x", side="top")
        
        # 5-2-2-1. ì™¼ìª½ ì˜ì—­ ì•ˆë‚´ ê¸€ì”¨ë“¤
        frame_left_status = tk.Frame(frame_status, bg="#ecf0f1")
        frame_left_status.pack(side="left", padx=10, pady=2)
        self.lbl_info_container = tk.Frame(frame_left_status, bg="#ecf0f1")
        self.lbl_info_container.pack(anchor="w", pady=(2, 0))

        tk.Label(self.lbl_info_container, text="ì •ë¥˜ì†Œì™€ ë…¸ì„ ì„ ì„ íƒí•˜ê³  ", font=(FONT_MAIN, SZ_L, "bold"), fg="#2d3436", bg="#ecf0f1").pack(side="left")
        tk.Label(self.lbl_info_container, text="ìë™ ê¸°ë¡ ì‹œì‘", font=(FONT_MAIN, SZ_L, "bold"), fg="#0984e3", bg="#ecf0f1").pack(side="left")
        tk.Label(self.lbl_info_container, text=" ë²„íŠ¼ì„ ëˆ„ë¥´ì„¸ìš”.", font=(FONT_MAIN, SZ_L, "bold"), fg="#2d3436", bg="#ecf0f1").pack(side="left")
        
        # 5-2-2-2. ì˜¤ë¥¸ìª½ ì˜ì—­ ì œì‘ì ì •ë³´ ë° ì‚¬ì´íŠ¸ ë§í¬ë“¤
        frame_right_status = tk.Frame(frame_status, bg="#ecf0f1")
        frame_right_status.pack(side="right", padx=10, pady=2)
        frame_right_top = tk.Frame(frame_right_status, bg="#ecf0f1")
        frame_right_top.pack(side="top", anchor="e")
        tk.Label(frame_right_top, text=" [ ë§Œë“ ì´ : ë°• êµ­ í™˜ (", font=(FONT_SUB, SZ_XS), fg="#7f8c8d", bg="#ecf0f1").pack(side="left", padx=(5, 0))
        lbl_email = tk.Label(frame_right_top, text="ggoyong2@naver.com", font=(FONT_SUB, SZ_XS, "underline"), fg="#2980b9", bg="#ecf0f1", cursor="hand2")
        lbl_email.pack(side="left"); lbl_email.bind("<Button-1>", lambda e: self.send_email_link("ggoyong2@naver.com"))
        tk.Label(frame_right_top, text=") ]", font=(FONT_SUB, SZ_XS), fg="#7f8c8d", bg="#ecf0f1").pack(side="left")
        
        frame_sources = tk.Frame(frame_right_status, bg="#ecf0f1")
        frame_sources.pack(side="top", anchor="e")
        lbl_link1 = tk.Label(frame_sources, text="ê³µê³µë°ì´í„°í¬í„¸ Open API (https://www.data.go.kr)", font=(FONT_SUB, SZ_XS), fg="#006400", bg="#ecf0f1", cursor="hand2")
        lbl_link1.pack(side="top", anchor="e"); lbl_link1.bind("<Button-1>", lambda e: self.open_link("https://www.data.go.kr"))
        lbl_link2 = tk.Label(frame_sources, text="ì„œìš¸ì‹œ ë²„ìŠ¤ìš´í–‰ë…¸ì„  ì •ë³´ (https://data.seoul.go.kr)", font=(FONT_SUB, SZ_XS), fg="#000080", bg="#ecf0f1", cursor="hand2")
        lbl_link2.pack(side="top", anchor="e"); lbl_link2.bind("<Button-1>", lambda e: self.open_link("https://data.seoul.go.kr/dataList/OA-15066/F/1/datasetView.do"))

        # 5-2-3. ì£¼ìš” ì¡°ì¢… ë²„íŠ¼ë“¤ì´ ëª¨ì—¬ìˆëŠ” íŒ¨ë„ì„ ë§Œë“­ë‹ˆë‹¤.
        self.frame_ctrl_master = tk.Frame(self.top_container, pady=4, bg="#f1f2f6")
        self.frame_ctrl_master.pack(fill="x", side="top")

        frame_main_content = tk.Frame(self.frame_ctrl_master, bg="#f1f2f6")
        frame_main_content.pack(side="left", fill="x", expand=True)

        PAD = 5  # ë²„íŠ¼Â·ìœ„ì ¯ ì‚¬ì´ ê³µí†µ ê°„ê²©

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # 1í–‰: ì¢Œ=ìë™ì €ì¥ ì•ˆë‚´ë¬¸êµ¬  /  ìš°=(ê°±ì‹ ì£¼ê¸° + ìë™ê¸°ë¡ì‹œì‘ + ìˆ˜ë™ê°±ì‹  + ë‹¤ë¥¸ì´ë¦„ì €ì¥)
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        frame_row1 = tk.Frame(frame_main_content, bg="#f1f2f6")
        frame_row1.pack(fill="x", side="top", pady=(2, 1))

        # 1í–‰ ì¢Œì¸¡: ìë™ ì €ì¥ ì•ˆë‚´ ë¬¸êµ¬
        self.lbl_auto_save_status = tk.Label(
            frame_row1,
            text="ìë™ ê¸°ë¡ ì‹œì‘ ë²„íŠ¼ì„ ì‘ë™ì‹œí‚¤ë©´ ë„ì°© ê¸°ë¡ì´ ì—‘ì…€íŒŒì¼ë¡œ ìë™ ì €ì¥ë©ë‹ˆë‹¤.",
            font=(FONT_SUB, SZ_XS, "bold"), fg="#e74c3c", bg="#f1f2f6"
        )
        self.lbl_auto_save_status.pack(side="left", padx=(5, 0))

        # 1í–‰ ìš°ì¸¡ ê·¸ë£¹: ê°±ì‹ ì£¼ê¸° | ìë™ê¸°ë¡ì‹œì‘ | ìˆ˜ë™ê°±ì‹  | ë‹¤ë¥¸ì´ë¦„ìœ¼ë¡œì—‘ì…€ì €ì¥
        right_r1 = tk.Frame(frame_row1, bg="#f1f2f6")
        right_r1.pack(side="right", padx=(0, 5))

        # ê°±ì‹ ì£¼ê¸° ë ˆì´ë¸” + ì…ë ¥ì¹¸ (ê°€ì¥ ì™¼ìª½)
        tk.Label(
            right_r1, text="ê°±ì‹ ì£¼ê¸°(ì´ˆ):", bg="#f1f2f6",
            font=(FONT_MAIN, SZ_S, "bold"), **mac_lbl_opts
        ).pack(side="left", padx=(PAD * 2, 2))
        entry_ival_opts = {
            "bg": "white", "fg": "black", "insertbackground": "black",
            "font": (FONT_MONO, SZ_S), "readonlybackground": "#f0f0f0"
        }
        self.entry_refresh_interval = tk.Entry(
            right_r1, textvariable=self.refresh_interval_var,
            width=4, **entry_ival_opts
        )
        self.entry_refresh_interval.pack(side="left", padx=(0, PAD))

        # ìë™ ê¸°ë¡ ì‹œì‘/ì¤‘ì§€ í† ê¸€
        # BTN_H: ì´ í–‰ì˜ ê¸°ì¤€ ë†’ì´(í”½ì…€)ì…ë‹ˆë‹¤. btn_save_excelì˜ ì‘ì€ í°íŠ¸ê°€
        #        ë²„íŠ¼ì„ ì¶•ì†Œì‹œí‚¤ì§€ ì•Šë„ë¡ ëª¨ë“  ë²„íŠ¼ì— ë™ì¼í•œ heightë¥¼ ì§€ì •í•©ë‹ˆë‹¤.
        BTN_H = 1  # tk.Button height ë‹¨ìœ„(ì¤„ ìˆ˜). 1ì¤„ ê³ ì •ìœ¼ë¡œ ë†’ì´ë¥¼ í†µì¼í•©ë‹ˆë‹¤.
        self.btn_toggle = tk.Button(
            right_r1, text="ìë™ ê¸°ë¡ ì‹œì‘",
            command=self._on_toggle_monitoring, width=11, height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_toggle.pack(side="left", padx=(0, PAD), ipady=4)

        # ìˆ˜ë™ ê°±ì‹ 
        self.btn_manual = tk.Button(
            right_r1, text="ìˆ˜ë™ ê°±ì‹ ",
            command=self.manual_refresh, width=7, height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_manual.pack(side="left", padx=(0, PAD), ipady=4)

        # ë‹¤ë¥¸ ì´ë¦„ìœ¼ë¡œ ì—‘ì…€ ì €ì¥ â€” 1ì¤„, ìƒë‹¨ ë²„íŠ¼ë“¤ê³¼ ë™ì¼í•œ í°íŠ¸Â·ë†’ì´
        self.btn_save_excel = tk.Button(
            right_r1, text="ë‹¤ë¥¸ ì´ë¦„ìœ¼ë¡œ ì—‘ì…€ ì €ì¥",
            command=self.save_to_excel, height=BTN_H,
            **self.get_btn_style("normal")
        )
        self.btn_save_excel.pack(side="left", padx=(0, 8), ipady=4)

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # 2í–‰: ì¢Œ=(ì¸ì¦í‚¤ ì…ë ¥ì°½ + ì¸ì¦í‚¤ë²„íŠ¼)  /  ìš°=(APIí˜„í™©ë²„íŠ¼ + APIì¹´ìš´íŠ¸í‘œ)
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        frame_row2 = tk.Frame(frame_main_content, bg="#f1f2f6")
        frame_row2.pack(fill="x", side="top", pady=(1, 2))

        # 2í–‰ ì¢Œì¸¡: ì¸ì¦í‚¤ ì…ë ¥ ì˜ì—­ + ì¸ì¦í‚¤ ë²„íŠ¼
        left_r2 = tk.Frame(frame_row2, bg="#f1f2f6")
        left_r2.pack(side="left", padx=(5, 0), fill="y")

        entry_key_opts = {
            "show": "*", "bg": "white", "fg": "black",
            "insertbackground": "black", "font": (FONT_MONO, SZ_XXS),
            "readonlybackground": "#f0f0f0"
        }
        key_input_area = tk.Frame(left_r2, bg="#f1f2f6")
        key_input_area.pack(side="left", fill="y")

        main_key_row = tk.Frame(key_input_area, bg="#f1f2f6")
        main_key_row.pack(fill="x", pady=1)
        tk.Label(main_key_row, text="ë©”ì¸ì¸ì¦í‚¤ :", font=(FONT_MAIN, SZ_S),
                 bg="#f1f2f6", **mac_lbl_opts).pack(side="left")
        self.entry_service_key = tk.Entry(
            main_key_row, textvariable=self.service_key_var,
            width=68, **entry_key_opts)
        self.entry_service_key.pack(side="left", padx=2)

        backup_key_row = tk.Frame(key_input_area, bg="#f1f2f6")
        backup_key_row.pack(fill="x", pady=1)
        tk.Label(backup_key_row, text="ë°±ì—…ì¸ì¦í‚¤ :", font=(FONT_MAIN, SZ_S),
                 bg="#f1f2f6", **mac_lbl_opts).pack(side="left")
        self.entry_backup_key = tk.Entry(
            backup_key_row, textvariable=self.backup_key_var,
            width=68, **entry_key_opts)
        self.entry_backup_key.pack(side="left", padx=2)

        # ì¸ì¦í‚¤ í™•ì¸/ì ê¸ˆ ë²„íŠ¼ â€” "ì¸ì¦í‚¤ ì…ë ¥": normal(ë…¹ìƒ‰) / "ì¸ì¦í‚¤ ë³€ê²½": outline(í…Œë‘ë¦¬)
        self.btn_key_manage = tk.Button(
            left_r2, text="ì¸ì¦í‚¤\nì…ë ¥",
            command=self.toggle_key_lock, width=7,
            **self.get_btn_style("normal")
        )
        self.btn_key_manage.pack(side="left", padx=(6, 0), fill="y")

        # 2í–‰ ìš°ì¸¡: API í˜¸ì¶œ í˜„í™© ë²„íŠ¼ + í†µê³„ ìˆ«ìíŒ
        right_r2 = tk.Frame(frame_row2, bg="#f1f2f6")
        right_r2.pack(side="right", padx=(0, 5), fill="y")

        self.btn_api_stats = tk.Button(
            right_r2, text="API\ní˜¸ì¶œ í˜„í™©",
            command=self.open_api_stats_window,
            width=7, **self.get_btn_style("normal")
        )
        self.btn_api_stats.pack(side="left", padx=(0, 4), fill="y")

        self.api_stats_container = tk.Frame(right_r2, bg="#f1f2f6")
        self.api_stats_container.pack(side="left", fill="y", anchor="center")
        self.stat_value_labels = {}
        stat_layout = [
            ["ARR1", "ARR2", "SINF", "POS1", "POS2"],
            ["SCNM", "SCID", "RINF", "SLST", "VLD"]
        ]
        for r, row_keys in enumerate(stat_layout):
            for c, key in enumerate(row_keys):
                tk.Label(
                    self.api_stats_container,
                    text=key, font=(FONT_MONO, SZ_XS, "bold"),
                    fg="#636e72", bg="#f1f2f6", width=5, anchor="w"
                ).grid(row=r, column=c*2, padx=(6, 0), sticky="w")
                val_lbl = tk.Label(
                    self.api_stats_container,
                    text="0", font=(FONT_MONO, SZ_XS, "bold"),
                    fg="#2d3436", bg="#f1f2f6", width=4, anchor="w"
                )
                val_lbl.grid(row=r, column=c*2+1, padx=(0, 4), sticky="w")
                self.stat_value_labels[key] = val_lbl

        # 5-2-6. ì´ˆê¸° í™”ë©´ ìƒíƒœë¥¼ ì—…ë°ì´íŠ¸í•©ë‹ˆë‹¤.
        self.update_api_counter_ui()
        self.update_button_states()

        # 5-2-5-1. ì¸ì¦í‚¤ ê¸¸ì´ ë³€í™”ë¥¼ ê°ì§€í•´ ì¸ì¦í‚¤ ì…ë ¥ ë²„íŠ¼ì˜ í™œì„±í™”ë¥¼ ì œì–´í•©ë‹ˆë‹¤.
        #   ë©”ì¸í‚¤ê°€ ì •í™•íˆ 64ê¸€ìì¼ ë•Œë§Œ ë²„íŠ¼ì´ í™œì„±í™”ë©ë‹ˆë‹¤.
        self.service_key_var.trace_add("write", lambda *_: self._check_key_btn_state())
        self.backup_key_var.trace_add("write",  lambda *_: self._check_key_btn_state())
        self._check_key_btn_state()  # ì´ˆê¸° ë¡œë“œ ì‹œ í•œ ë²ˆ ì¦‰ì‹œ ì‹¤í–‰

        # 5-2-7. ì—´ì‡ ê°€ ì ê²¨ ìˆë‹¤ë©´ ì…ë ¥ì¹¸ì„ ì½ê¸° ì „ìš©ìœ¼ë¡œ ë°”ê¿‰ë‹ˆë‹¤.
        if self.key_locked:
            self.entry_service_key.config(state='readonly')
            self.entry_backup_key.config(state='readonly')
            _ok = self.get_btn_style("outline")
            _ok.pop("state", None)
            self.btn_key_manage.config(text="ì¸ì¦í‚¤\në³€ê²½", **_ok)

        # 5-2-8. í™”ë©´ ì¤‘ê°„ê³¼ ì•„ë˜ìª½ì˜ ìœ„ì•„ë˜ í¬ê¸°ë¥¼ ì¡°ì ˆí•  ìˆ˜ ìˆëŠ” ê³µê°„ì„ ë§Œë“­ë‹ˆë‹¤.
        self.super_paned = tk.PanedWindow(self.root, orient=tk.VERTICAL, sashrelief=tk.RAISED, sashwidth=6)
        self.super_paned.pack(fill="both", expand=True)
        self.main_paned = tk.PanedWindow(self.super_paned, orient=tk.VERTICAL, sashrelief=tk.RAISED, sashwidth=6)
        self.setup_trees()
        self.super_paned.add(self.main_paned, stretch="always", minsize=220)

        # 5-2-9. í™”ë©´ ì œì¼ ì•„ë˜ìª½ì— ì‘ì—… ì¼ì§€(ë¡œê·¸)ë¥¼ ì ëŠ” ê¹Œë§Œ ì°½ì„ ë§Œë“­ë‹ˆë‹¤.
        log_outer_frame = tk.Frame(self.super_paned, padx=10, pady=5)
        log_scroll_frame = tk.Frame(log_outer_frame); log_scroll_frame.pack(fill="both", expand=True)
        self.txt_log = tk.Text(log_scroll_frame, height=10, bg="#2d3436", fg="#dfe6e9",
                               font=(FONT_MONO, SZ_S), state="disabled")  # ì½ê¸°ì „ìš©: íƒ€ì´í•‘ìœ¼ë¡œ ìˆ˜ì • ë¶ˆê°€
        log_scrollbar = ttk.Scrollbar(log_scroll_frame, orient="vertical", command=self.txt_log.yview)
        self.txt_log.configure(yscrollcommand=log_scrollbar.set)
        self.txt_log.pack(side="left", fill="both", expand=True); log_scrollbar.pack(side="right", fill="y")
        self.super_paned.add(log_outer_frame, minsize=60, height=180)

    # 5-3. [ë²„ìŠ¤ ì •ë³´íŒ ë§Œë“¤ê¸°] ì‹¤ì‹œê°„ ë²„ìŠ¤ ìœ„ì¹˜ì™€ ë„ì°© ê¸°ë¡ì„ ë³´ì—¬ì£¼ëŠ” í‘œ ìƒì„± í•¨ìˆ˜
    def setup_trees(self):
        # 5-3-1. í™”ë©´ ìƒë‹¨ì— 2ê°œì˜ ì‹¤ì‹œê°„ í˜„í™© í‘œë¥¼ ë°°ì¹˜í•©ë‹ˆë‹¤.
        frame_rt_container = tk.Frame(self.main_paned)
        rt_inner = tk.Frame(frame_rt_container); rt_inner.pack(fill="both", expand=True)
        rt_inner.grid_columnconfigure(0, weight=1); rt_inner.grid_columnconfigure(1, weight=1)
        rt_inner.grid_rowconfigure(0, weight=1)

        self.btn_searches = [] 
        self.trees_rt = []; self.lbl_st_names = [] 
        
        # 5-3-1-1. ì™¼ìª½(0)ê³¼ ì˜¤ë¥¸ìª½(1)ì— ê°ê° í‘œë¥¼ í•˜ë‚˜ì”© ê·¸ë¦½ë‹ˆë‹¤.
        for i in range(2): 
            f = tk.Frame(rt_inner); f.grid(row=0, column=i, sticky="nsew", padx=2)
            header = tk.Frame(f); header.pack(fill="x", pady=2)
            inner_header = tk.Frame(header); inner_header.pack(anchor="center")
            
            # 5-3-1-2. ì •ë¥˜ì†Œ ì œëª© ë¼ë²¨ê³¼ ARS-ID ë²ˆí˜¸ ì…ë ¥ì¹¸ì„ ë§Œë“­ë‹ˆë‹¤.
            lbl = tk.Label(inner_header, text=f"[ì •ë¥˜ì†Œ {i+1}] ì‹¤ì‹œê°„ í˜„í™©", fg="#0984e3", font=(FONT_SUB, SZ_M, "bold"))
            lbl.pack(side="left"); self.lbl_st_names.append(lbl)
            tk.Entry(inner_header, textvariable=self.ars_ids[i], width=10, state="readonly", readonlybackground="#f0f0f0", fg="#2d3436", font=(FONT_MONO, SZ_M, "bold")).pack(side="left", padx=5)
            
            # 5-3-1-3. ë²„ìŠ¤ ì •ë¥˜ì†Œë¥¼ ì°¾ê¸° ìœ„í•œ ê²€ìƒ‰ ë²„íŠ¼ì„ ë§Œë“­ë‹ˆë‹¤.
            # ë¹„í™œì„± ìƒíƒœ(disabled) ì´ˆê¸°ê°’: disabled ìŠ¤íƒ€ì¼ë¡œ íšŒìƒ‰ ë°°ê²½ í‘œì‹œ
            _s_search = self.get_btn_style("disabled")
            _s_search.pop("state", None)
            btn_search = tk.Button(inner_header, text="ê²€ìƒ‰",
                                  command=lambda idx=i: self.open_search_window(idx),
                                  width=5, state="disabled", **_s_search)
            btn_search.pack(side="left", padx=(4, 0), ipady=1)
            self.btn_searches.append(btn_search) 

            # 5-3-1-4. ì‹¤ì‹œê°„ ë„ì°© ì •ë³´ í‘œ(Treeview)ë¥¼ ì„¤ì •í•©ë‹ˆë‹¤.
            tree_frame = tk.Frame(f); tree_frame.pack(fill="both", expand=True)
            cols = ("route", "bus1_no", "bus1_msg", "bus2_no", "bus2_msg")
            tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
            for col in cols: tree.heading(col, text=self.get_col_name(col))
            tree.column("route", width=69); tree.column("bus1_no", width=86); tree.column("bus1_msg", width=130)
            tree.column("bus2_no", width=86); tree.column("bus2_msg", width=129)
            
            # 5-3-1-5. ìŠ¤í¬ë¡¤ë°”ë¥¼ í‘œì— ë¶™ì…ë‹ˆë‹¤.
            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=vsb.set); tree.pack(side="left", fill="both", expand=True); vsb.pack(side="right", fill="y")
            self.trees_rt.append(tree)
        self.main_paned.add(frame_rt_container, minsize=100) 

        # 5-3-2. í™”ë©´ í•˜ë‹¨ì— 2ê°œì˜ ë„ì°© ê¸°ë¡ í‘œë¥¼ ë°°ì¹˜í•©ë‹ˆë‹¤.
        frame_hist_container = tk.Frame(self.main_paned)
        hist_inner = tk.Frame(frame_hist_container); hist_inner.pack(fill="both", expand=True)
        hist_inner.grid_columnconfigure(0, weight=1); hist_inner.grid_columnconfigure(1, weight=1)
        hist_inner.grid_rowconfigure(0, weight=1)
        
        self.trees_hist = []; self.lbl_hist_titles = []
        for i in range(2):
            f = tk.Frame(hist_inner); f.grid(row=0, column=i, sticky="nsew", padx=2)
            header = tk.Frame(f); header.pack(fill="x", pady=2)
            inner_header = tk.Frame(header); inner_header.pack(anchor="center")
            
            # 5-3-2-1. ë„ì°© ê¸°ë¡íŒ ì œëª©ê³¼ ê¸°ë¡ ì‚­ì œ ë²„íŠ¼ì„ ë§Œë“­ë‹ˆë‹¤.
            lbl = tk.Label(inner_header, text=f"[ì •ë¥˜ì†Œ {i+1}] ë„ì°© ê¸°ë¡", fg="#d63031", font=(FONT_SUB, SZ_M, "bold"))
            lbl.pack(side="left"); self.lbl_hist_titles.append(lbl)
            # ê¸°ë¡ ì‚­ì œ ë²„íŠ¼: normal ìŠ¤íƒ€ì¼, í°íŠ¸ í¬ê¸°ëŠ” SZ_XS(í•œ ë‹¨ê³„ ì‘ê²Œ)
            _s_del = self.get_btn_style("normal", font_size=SZ_XS)
            tk.Button(inner_header, text="ê¸°ë¡ ì‚­ì œ",
                      command=lambda idx=i: self.clear_history(idx),
                      **_s_del).pack(side="left", padx=10, ipady=1)

            # 5-3-2-2. ë„ì°© ê¸°ë¡ í‘œ(Treeview)ë¥¼ ì„¤ì •í•©ë‹ˆë‹¤.
            tree_frame = tk.Frame(f); tree_frame.pack(fill="both", expand=True)
            cols = ("data_time", "route", "veh_no", "corp", "status")
            tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
            for col in cols: tree.heading(col, text=self.get_col_name(col)) 
            tree.column("data_time", width=135); tree.column("route", width=63); tree.column("veh_no", width=94)
            tree.column("corp", width=135); tree.column("status", width=73)
            # 5-3-2-3. ë„ì°© ê¸°ë¡ í‘œì— ìŠ¤í¬ë¡¤ë°”ë¥¼ ë¶™ì…ë‹ˆë‹¤.
            
            # ìŠ¤í¬ë¡¤ë°” ë¶™ì´ê¸°
            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=vsb.set); tree.pack(side="left", fill="both", expand=True); vsb.pack(side="right", fill="y")
            self.trees_hist.append(tree)
        self.main_paned.add(frame_hist_container, minsize=100)

    # 3ê·¸ë£¹ : UI ë³´ì¡° í•¨ìˆ˜ (UI Auxiliary function)

    # 5-4. [ë²„íŠ¼ ì˜·ì…íˆê¸°] ì»´í“¨í„° ì¢…ë¥˜ì™€ ë²„íŠ¼ ìƒíƒœì— ë”°ë¼ ëª¨ì–‘ì„ ì •í•´ì£¼ëŠ” í•¨ìˆ˜
    # 5-4-0. [ë²„íŠ¼ ë””ìì¸ ê°€ì´ë“œ]
    #   btn_type="normal"  : ì´ˆë¡ ê·¸ë¼ë°ì´ì…˜ ë°°ê²½ + í°ìƒ‰ ê¸€ì + ì…ì²´ í…Œë‘ë¦¬  â€” ì¼ë°˜ í™œì„± ë²„íŠ¼
    #   btn_type="outline" : í° ë°°ê²½ + ì´ˆë¡ í…Œë‘ë¦¬ + ì´ˆë¡ ê¸€ì               â€” ì¸ì¦í‚¤ë³€ê²½ ë²„íŠ¼
    #   btn_type="pressed" : ì§„í•œ ì´ˆë¡ ë°°ê²½(ëˆŒë¦¼ íš¨ê³¼)                        â€” ëˆŒë¦° ìƒíƒœ
    #   btn_type="disabled": ì—°íšŒìƒ‰ ë°°ê²½(#e8e8e8) + íšŒìƒ‰ ê¸€ì                â€” ë¹„í™œì„± ë²„íŠ¼
    #   macOSì—ì„œëŠ” ë°°ê²½ìƒ‰ì´ ì œí•œì ì´ë¯€ë¡œ ë³„ë„ ë¶„ê¸°ë¡œ ì²˜ë¦¬í•©ë‹ˆë‹¤.
    def get_btn_style(self, btn_type="normal", font_size=None):
        # 5-4-0-1. í•˜ìœ„ í˜¸í™˜: êµ¬ë²„ì „ì—ì„œ theme_color ë¬¸ìì—´ì„ ë„˜ê¸¸ ê²½ìš° ìë™ ë§¤í•‘í•©ë‹ˆë‹¤.
        _color_map = {
            "#28a745": "normal", "#007bff": "normal", "#8e44ad": "normal",
            "#f1c40f": "normal", "#2ecc71": "normal",
            "#d63031": "outline", "#7f8c8d": "outline",
        }
        if btn_type.startswith("#"):
            btn_type = _color_map.get(btn_type, "normal")

        target_sz = font_size if font_size else SZ_S

        # 5-4-1. ìƒ‰ìƒ íŒ”ë ˆíŠ¸ ì •ì˜
        C_GREEN        = "#2ecc71"   # ê¸°ë³¸ ì´ˆë¡ ë°°ê²½
        C_GREEN_DARK   = "#27ae60"   # ì§„í•œ ì´ˆë¡ (hover / active)
        C_GREEN_DEEPER = "#1e8449"   # ë” ì§„í•œ ì´ˆë¡ (pressed / í…Œë‘ë¦¬ í•˜ë‹¨ ê·¸ë¦¼ì)
        C_GREEN_PALE   = "#eafaf1"   # outline ë²„íŠ¼ ë°°ê²½ (ê±°ì˜ í°ìƒ‰)
        C_GRAY_BG      = "#e8e8e8"   # disabled ë°°ê²½
        C_GRAY_BORDER  = "#c0c0c0"   # disabled í…Œë‘ë¦¬
        C_GRAY_FG      = "#a0a0a0"   # disabled ê¸€ì
        C_WHITE        = "#ffffff"
        CURSOR_HAND    = "pointinghand" if CURRENT_OS == "Darwin" else "hand2"

        if CURRENT_OS == "Darwin":
            # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
            # macOS: Aqua ãƒ†ãƒ¼ãƒã§ tk.Button ã® bg/activebackground ã¯
            #        ã‚·ã‚¹ãƒ†ãƒ ã«ä¸Šæ›¸ãã•ã‚Œã¾ã™ãŒã€ä»¥ä¸‹ã®å±æ€§ã¯æœ‰åŠ¹ã§ã™:
            #   â€¢ fg / activeforeground  â†’ æ–‡å­—è‰²
            #   â€¢ font                   â†’ ãƒ•ã‚©ãƒ³ãƒˆ
            #   â€¢ highlightbackground / highlightthickness â†’ å¤–æ ã®è‰²ãƒ»å¤ªã•
            #   â€¢ relief / bd            â†’ ç«‹ä½“æ„Ÿã®å½¢çŠ¶
            #   â€¢ padx / pady            â†’ å†…å´ä½™ç™½ (ãƒœã‚¿ãƒ³ã®å¤§ãã•ã«å½±éŸ¿)
            #   â€¢ overrelief             â†’ ãƒ›ãƒãƒ¼æ™‚ã®å½¢çŠ¶å¤‰åŒ–
            #   â€¢ cursor                 â†’ ã‚«ãƒ¼ã‚½ãƒ«å½¢çŠ¶
            #
            # macOS: Aqua í…Œë§ˆì—ì„œ tk.Buttonì˜ bg/activebackground ëŠ”
            #        ì‹œìŠ¤í…œì´ ë®ì–´ì”ë‹ˆë‹¤. ê·¸ëŸ¬ë‚˜ ì•„ë˜ í•­ëª©ì€ ìœ íš¨í•©ë‹ˆë‹¤:
            #   fg/activeforeground(ê¸€ììƒ‰), highlightbackground(ì™¸ê³½ í…Œë‘ë¦¬),
            #   relief/bd(ì…ì²´ê°), padx/pady(ë‚´ë¶€ ì—¬ë°±), overrelief(í˜¸ë²„ í˜•ìƒ),
            #   cursor(ë§ˆìš°ìŠ¤ ëª¨ì–‘)
            # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
            if btn_type == "disabled":
                return {
                    "fg": C_GRAY_FG,
                    "font": (FONT_MAIN, target_sz, "normal"),
                    "cursor": "arrow",
                    "highlightthickness": 1,
                    "highlightbackground": C_GRAY_BORDER,
                    "relief": "groove",
                    "bd": 1,
                    "padx": 6,
                    "pady": 3,
                    "state": "normal",
                }
            elif btn_type == "outline":
                return {
                    "fg": C_GREEN_DARK,
                    "activeforeground": C_GREEN_DEEPER,
                    "font": (FONT_MAIN, target_sz, "bold"),
                    "cursor": CURSOR_HAND,
                    "highlightthickness": 2,
                    "highlightbackground": C_GREEN_DARK,
                    "relief": "ridge",
                    "bd": 2,
                    "padx": 6,
                    "pady": 3,
                    "overrelief": "solid",
                }
            else:  # normal
                return {
                    "fg": C_GREEN_DEEPER,
                    "activeforeground": C_GREEN_DARK,
                    "font": (FONT_MAIN, target_sz, "bold"),
                    "cursor": CURSOR_HAND,
                    "highlightthickness": 2,
                    "highlightbackground": C_GREEN,
                    "relief": "raised",
                    "bd": 2,
                    "padx": 6,
                    "pady": 3,
                    "overrelief": "solid",
                }
        else:
            # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
            # Windows: ë°°ê²½/ê¸€ììƒ‰ + ì…ì²´ í…Œë‘ë¦¬ë¡œ ê·¸ë˜í”½ ë²„íŠ¼ íš¨ê³¼ë¥¼ ëƒ…ë‹ˆë‹¤.
            #   - relief="raised"  + ì´ˆë¡ ë°°ê²½  â†’ ë³¼ë¡í•œ ì…ì²´ ë²„íŠ¼
            #   - highlightbackground ë¡œ ì™¸ê³½ í…Œë‘ë¦¬ ìƒ‰ ì§€ì •
            #   - activebackground  ë¡œ hover ìƒ‰ìƒ ì§€ì •
            # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
            if btn_type == "disabled":
                return {
                    "bg": C_GRAY_BG,
                    "fg": C_GRAY_FG,
                    "font": (FONT_MAIN, target_sz, "normal"),
                    "relief": "groove",
                    "bd": 2,
                    "highlightbackground": C_GRAY_BORDER,
                    "highlightthickness": 1,
                    "cursor": "arrow",
                    "state": "disabled",
                    "disabledforeground": C_GRAY_FG,
                }
            elif btn_type == "outline":
                return {
                    "bg": C_GREEN_PALE,
                    "fg": C_GREEN_DARK,
                    "activebackground": C_GREEN_PALE,
                    "activeforeground": C_GREEN_DEEPER,
                    "font": (FONT_MAIN, target_sz, "bold"),
                    "relief": "ridge",
                    "bd": 2,
                    "highlightbackground": C_GREEN_DARK,
                    "highlightthickness": 1,
                    "cursor": CURSOR_HAND,
                    "overrelief": "solid",
                }
            else:  # normal
                return {
                    "bg": C_GREEN,
                    "fg": C_WHITE,
                    "activebackground": C_GREEN_DARK,
                    "activeforeground": C_WHITE,
                    "font": (FONT_MAIN, target_sz, "bold"),
                    "relief": "raised",
                    "bd": 2,
                    "highlightbackground": C_GREEN_DEEPER,
                    "highlightthickness": 1,
                    "cursor": CURSOR_HAND,
                    "overrelief": "solid",
                }
        
    # 5-5. [ì´ë¦„í‘œ ë²ˆì—­] ì˜ì–´ë¡œ ëœ ë°ì´í„° ì´ë¦„ì„ í•œê¸€ë¡œ ì•Œê¸° ì‰½ê²Œ ë°”ê¿”ì£¼ëŠ” í•¨ìˆ˜
    def get_col_name(self, code):
        mapping = {"route": "ë…¸ì„ ", "bus1_no": "1ë²ˆì°¨ëŸ‰", "bus1_msg": "ë„ì°©ì •ë³´", "bus2_no": "2ë²ˆì°¨ëŸ‰", "bus2_msg": "ë„ì°©ì •ë³´", "data_time": "ë°ì´í„° ì‹œê°", "veh_no": "ì°¨ëŸ‰ë²ˆí˜¸", "corp": "ìš´ìˆ˜ì‚¬ëª…", "status": "ìƒíƒœ"}
        return mapping.get(code, code)
    
    # 5-6. [ë§í¬ ì—´ê¸°] ì¸í„°ë„· ì£¼ì†Œë¥¼ ì›¹ ë¸Œë¼ìš°ì €ë¡œ ì—´ì–´ì£¼ëŠ” í•¨ìˆ˜
    def open_link(self, url):
        webbrowser.open_new(url)
        
    # 5-7. [ë©”ì¼ ë³´ë‚´ê¸°] í´ë¦­í•˜ë©´ ë°”ë¡œ ì´ë©”ì¼ì„ ë³´ë‚¼ ìˆ˜ ìˆê²Œ ì°½ì„ ë„ìš°ëŠ” í•¨ìˆ˜
    def send_email_link(self, email):
        webbrowser.open_new(f"mailto:{email}")

    # 5-8. [ì¹¸ ê³ ì •] í‘œì˜ ì¹¸ ë„ˆë¹„ë¥¼ ì‹¤ìˆ˜ë¡œ ë°”ê¾¸ì§€ ëª»í•˜ê²Œ ê³ ì •í•˜ëŠ” í•¨ìˆ˜
    def prevent_column_resize(self, event):
        if event.widget.identify_region(event.x, event.y) == "separator":
            return "break" 

    # 5-9. [í†µê³„ ì°½] APIë¥¼ ëª‡ ë²ˆ í˜¸ì¶œí–ˆëŠ”ì§€ ìƒì„¸íˆ ë³´ì—¬ì£¼ëŠ” ìƒˆ ì°½ ì—´ê¸° í•¨ìˆ˜
    def open_api_stats_window(self):
        # 5-9-1. ì°½ì´ ì´ë¯¸ ë–  ìˆë‹¤ë©´ ê·¸ ì°½ì„ ë§¨ ì•ìœ¼ë¡œ ê°€ì ¸ì˜µë‹ˆë‹¤.
        if self.stats_win is not None and self.stats_win.winfo_exists():
            self.stats_win.lift()
            self.stats_win.focus_force()
            return

        # 5-9-2. ìƒˆë¡œìš´ ì°½ì„ ë§Œë“­ë‹ˆë‹¤.
        self.stats_win = tk.Toplevel(self.root) 
        self.stats_win.title("API í˜¸ì¶œ ìƒì„¸ í˜„í™©")
        self.stats_win.geometry("680x400")
        self.stats_win.minsize(680, 380)
        self.stats_win.transient(self.root) 
        self.stats_win.protocol("WM_DELETE_WINDOW", self.on_stats_win_close) 
        
        tk.Label(self.stats_win, text="API í˜¸ì¶œ ìƒì„¸ í†µê³„  (ë©”ì¸í‚¤ / ë°±ì—…í‚¤ / í•©ê³„ ìˆœ)", font=(FONT_MAIN, SZ_M, "bold"), pady=10).pack(side="top", fill="x")
        
        # 5-9-3. í†µê³„ ì •ë³´ë¥¼ ë³´ì—¬ì¤„ í‘œë¥¼ ë§Œë“­ë‹ˆë‹¤.
        #        ì•½ì¹­ / URL / ë©”ì¸í‚¤ / ë°±ì—…í‚¤ / í•©ê³„ 5ê°œ ì—´ë¡œ êµ¬ì„±í•©ë‹ˆë‹¤.
        # 5-9-3-1. ì—´ ìˆœì„œ: ì•½ì¹­ / URL / ë©”ì¸í‚¤ / ë°±ì—…í‚¤ / í•©ê³„ (ì‚¬ìš©ì ìš”ì²­ ìˆœì„œ)
        cols = ("abbr", "url", "main", "back", "total")
        tree = ttk.Treeview(self.stats_win, columns=cols, show="headings") 
        tree.heading("abbr",  text="ì•½ì¹­")
        tree.heading("url",   text="API ì‹¤ì œ í˜¸ì¶œ ì£¼ì†Œ")
        tree.heading("main",  text="ë©”ì¸í‚¤")
        tree.heading("back",  text="ë°±ì—…í‚¤")
        tree.heading("total", text="í•©ê³„")
        tree.column("abbr",  width=60,  anchor="center", stretch=False)
        tree.column("url",   width=360, anchor="w",      stretch=True)
        tree.column("main",  width=60,  anchor="center", stretch=False)
        tree.column("back",  width=60,  anchor="center", stretch=False)
        tree.column("total", width=60,  anchor="center", stretch=False)
        tree.bind('<Button-1>', self.prevent_column_resize)
        
        scrollbar = ttk.Scrollbar(self.stats_win, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10)
        
        # 5-9-4. 1ì´ˆë§ˆë‹¤ ìˆ«ìê°€ ë°”ë€Œë„ë¡ ê³„ì† ì‹¤í–‰ë˜ëŠ” í•¨ìˆ˜ì…ë‹ˆë‹¤.
        def update_stats_loop():
            if not self.stats_win or not self.stats_win.winfo_exists():
                return
            for item in tree.get_children():
                tree.delete(item)
            total_all = 0
            total_main = 0
            total_back = 0
            for key, count in self.api_stats.items():
                url = self.api_urls.get(key, "-")
                m_cnt = self.api_stats_by_key["main"].get(key, 0)
                b_cnt = self.api_stats_by_key["back"].get(key, 0)
                # 5-9-4-1. ì—´ ìˆœì„œ: ì•½ì¹­ / URL / ë©”ì¸í‚¤ / ë°±ì—…í‚¤ / í•©ê³„
                tree.insert("", "end", values=(key, url, f"{m_cnt}", f"{b_cnt}", f"{count}"))
                total_all  += count
                total_main += m_cnt
                total_back += b_cnt
            # 5-9-4-2. ë§¨ ì•„ë˜ì— í•©ê³„ í–‰ì„ ì¶”ê°€í•©ë‹ˆë‹¤.
            tree.insert("", "end", values=("í•©ê³„", " - ", f"{total_main}", f"{total_back}", f"{total_all}"))
            self.stats_win.after(1000, update_stats_loop) 

        update_stats_loop()

    # 5-10. [ìˆ˜ì²© ë‹«ê¸°] í†µê³„ ì°½ì„ ë‹«ì„ ë•Œ ë©”ëª¨ë¦¬ì—ì„œ ì§€ì›Œì£¼ëŠ” í•¨ìˆ˜
    def on_stats_win_close(self):
        if self.stats_win:
            self.stats_win.destroy()
            self.stats_win = None

    # 5-11. [ìŠ¤ìœ„ì¹˜ ì¡°ì ˆ] ìƒí™©(ê°ì‹œ ì¤‘ì¸ì§€ ë“±)ì— ë”°ë¼ ë²„íŠ¼ë“¤ì„ ì¼œê±°ë‚˜ ë„ëŠ” í•¨ìˆ˜
    def update_button_states(self):
        # 5-11-1. ì •ë¥˜ì†Œ ì •ë³´ê°€ ìˆëŠ”ì§€ ë¨¼ì € í™•ì¸í•©ë‹ˆë‹¤.
        has_info = (bool(self.target_st_info[0].get('routes')) or
                    bool(self.target_st_info[1].get('routes')))

        # 5-11-2. ë²„íŠ¼ì˜ ì‘ë™ ì—¬ë¶€ë¥¼ ì„¤ì •í•˜ëŠ” ë„ìš°ë¯¸ í•¨ìˆ˜
        def apply_btn(btn, btn_type, is_active, cmd, text=None):
            # 5-11-2-1. ë¹„í™œì„± ìƒíƒœë¼ë©´ disabled ìŠ¤íƒ€ì¼ë¡œ ë®ì–´ì”ë‹ˆë‹¤.
            effective_type = btn_type if is_active else "disabled"
            style = self.get_btn_style(effective_type)
            if text:
                style["text"] = text
            # 5-11-2-2. state í‚¤ë¥¼ styleì—ì„œ ì œê±° â€” btn.config()ì— state= ë¥¼ ë³„ë„ë¡œ ë„˜ê¸°ë¯€ë¡œ
            #           style ë”•ì…”ë„ˆë¦¬ì— state ê°€ ê°™ì´ ìˆìœ¼ë©´ "multiple values" TypeError ë°œìƒ.
            style.pop("state", None)
            if CURRENT_OS == "Darwin":
                btn.config(state="normal", command=(cmd if is_active else None), **style)
            else:
                btn.config(state=("normal" if is_active else "disabled"),
                           command=(cmd if is_active else lambda: None), **style)

        # 5-11-3. ê°ì‹œ ì¤‘ì¼ ë•Œ â†’ í†µí•© ë²„íŠ¼ì„ "ì¤‘ì§€"(outline), ìˆ˜ë™ ê°±ì‹  í™œì„±í™”
        if self.is_monitoring:
            apply_btn(self.btn_toggle, "outline", True,
                      self.stop_monitoring, text="ì¤‘ì§€")
            apply_btn(self.btn_manual, "normal", True, self.manual_refresh)
        # 5-11-4. ê°ì‹œ ì¤‘ì´ ì•„ë‹ ë•Œ â†’ í†µí•© ë²„íŠ¼ì„ "ìë™ ê¸°ë¡ ì‹œì‘"
        else:
            apply_btn(self.btn_toggle, "normal", has_info,
                      self.start_monitoring, text="ìë™ ê¸°ë¡ ì‹œì‘")
            apply_btn(self.btn_manual, "normal", has_info, self.manual_refresh)

    # 5-11-5. [í†µí•© í† ê¸€] ì‹œì‘/ì¤‘ì§€ ë²„íŠ¼ í´ë¦­ ì‹œ í˜„ì¬ ìƒíƒœì— ë”°ë¼ ë¶„ê¸°í•©ë‹ˆë‹¤.
    def _on_toggle_monitoring(self):
        if self.is_monitoring:
            self.stop_monitoring()
        else:
            self.start_monitoring()

    # 5-11-6. [ì¸ì¦í‚¤ ê¸¸ì´ ê°ì‹œ] ì¸ì¦í‚¤ ì…ë ¥ ë²„íŠ¼ í™œì„±í™” ì¡°ê±´ì„ ê²€ì‚¬í•©ë‹ˆë‹¤.
    #   í™œì„±í™” ì¡°ê±´ (ë‘ ì¡°ê±´ ëª¨ë‘ ì¶©ì¡±í•´ì•¼ í•¨):
    #     â‘  ë©”ì¸í‚¤: ì •í™•íˆ 64ê¸€ì
    #     â‘¡ ë°±ì—…í‚¤: ë¹„ì–´ìˆê±°ë‚˜(0ì) ì •í™•íˆ 64ê¸€ì  â† ê·¸ ì™¸ ê¸¸ì´ë©´ ë¹„í™œì„±
    #   - key_locked=True(ì ê¸ˆ ìƒíƒœ)ì´ë©´ "ì¸ì¦í‚¤ ë³€ê²½" ë²„íŠ¼ì´ë¯€ë¡œ í•­ìƒ í™œì„±í™”í•©ë‹ˆë‹¤.
    def _check_key_btn_state(self):
        if self.key_locked:
            # ì ê¸ˆ ìƒíƒœì—ì„œëŠ” "ì¸ì¦í‚¤ ë³€ê²½" ë²„íŠ¼ì´ í•­ìƒ ëˆŒë¦´ ìˆ˜ ìˆì–´ì•¼ í•©ë‹ˆë‹¤.
            return
        main_len = len(self.service_key_var.get().strip())
        back_len = len(self.backup_key_var.get().strip())
        # ë©”ì¸í‚¤ 64ì AND (ë°±ì—…í‚¤ 0ì OR ë°±ì—…í‚¤ 64ì)
        is_valid = (main_len == 64) and (back_len == 0 or back_len == 64)
        if is_valid:
            _s = self.get_btn_style("normal")
            _s.pop("state", None)
            self.btn_key_manage.config(state="normal", **_s)
        else:
            _s = self.get_btn_style("disabled")
            _s.pop("state", None)
            self.btn_key_manage.config(state="disabled", **_s)


    # 5-12. [ìˆ«ìíŒ ì—…ë°ì´íŠ¸] ë©”ì¸ í™”ë©´ì˜ API í˜¸ì¶œ íšŸìˆ˜ë¥¼ ì‹¤ì‹œê°„ìœ¼ë¡œ ê°±ì‹ í•˜ëŠ” í•¨ìˆ˜
    def update_api_counter_ui(self):
        # 5-12-1. ì €ì¥ëœ ìˆ«ìíŒë“¤ì„ ëŒë©´ì„œ ìµœì‹  ìˆ«ìë¡œ ê¸€ìë¥¼ ë°”ê¿‰ë‹ˆë‹¤.
        for key, val_lbl in self.stat_value_labels.items():
            if key in self.api_stats:
                val_lbl.config(text=str(self.api_stats[key]))

    # 4ê·¸ë£¹ : ë³´ì•ˆ ë° ì¸ì¦ ê´€ë¦¬ (Security & Auth)

    # 5-13. [ì—´ì‡  ë¶ˆëŸ¬ì˜¤ê¸°] í”„ë¡œê·¸ë¨ì´ ì¼œì§ˆ ë•Œ ì €ì¥í•´ë‘” ë¹„ë°€ ì—´ì‡ (ì¸ì¦í‚¤)ë¥¼ êº¼ë‚´ì„œ í™•ì¸í•˜ëŠ” ë‹¨ê³„
    def load_saved_key(self):
        # 5-13-1. ì—´ì‡ ê°€ ì €ì¥ëœ íŒŒì¼ì´ ìˆëŠ”ì§€ í™•ì¸í•©ë‹ˆë‹¤.
        if os.path.exists(self.key_file_path):
            try:
                # 5-13-2. íŒŒì¼ì„ ì—´ì–´ì„œ ì•”í˜¸í™”ëœ ë‚´ìš©ì„ ì½ì–´ì˜µë‹ˆë‹¤.
                with open(self.key_file_path, "rb") as f:
                    encrypted_data = f.read()
                    if encrypted_data:
                        # 5-13-2-1. [ë¹„ë°€ í’€ê¸°] ì•”í˜¸í™”ëœ ë‚´ìš©ì„ ë‹¤ì‹œ ê¸€ìë¡œ ë°”ê¿‰ë‹ˆë‹¤.
                        decrypted_data = cipher_suite.decrypt(encrypted_data).decode('utf-8')
                        main_k, back_k = "", ""
                        # 5-13-2-2. ë©”ì¸ ì—´ì‡ ì™€ ë°±ì—… ì—´ì‡ ê°€ ê°™ì´ ìˆë‹¤ë©´ ë‚˜ëˆ„ì–´ì„œ ì €ì¥í•©ë‹ˆë‹¤.
                        if "|" in decrypted_data:
                            main_k, back_k = decrypted_data.split("|", 1)
                        else:
                            main_k = decrypted_data
                        
                        # 5-13-2-3. í™”ë©´ì— ìˆëŠ” ì…ë ¥ì°½ì— ë¶ˆëŸ¬ì˜¨ ì—´ì‡ ë“¤ì„ ë„£ì–´ì¤ë‹ˆë‹¤.
                        self.service_key_var.set(main_k)
                        self.backup_key_var.set(back_k)
                        
                        # 5-13-2-4. [ìë™ ê²€ì‚¬] ë¶ˆëŸ¬ì˜¨ ì—´ì‡ ê°€ ì§„ì§œì¸ì§€ ë°”ë¡œ í™•ì¸ì„ ì‹œì‘í•©ë‹ˆë‹¤.
                        if main_k.strip() or back_k.strip():
                            # 5-13-2-4-1. ë¡œê·¸ì°½ì´ ì¤€ë¹„ë  ë•Œê¹Œì§€ ì•„ì£¼ ì ì‹œ(0.1ì´ˆ) ê¸°ë‹¤ë ¸ë‹¤ê°€ ë©”ì‹œì§€ë¥¼ ë„ì›ë‹ˆë‹¤.
                            self.root.after(100, lambda: self.log("ğŸš€ ì €ì¥ëœ ì¸ì¦í‚¤ë¥¼ ë¶ˆëŸ¬ì™€ ìœ íš¨ì„±ì„ ê²€ì‚¬í•©ë‹ˆë‹¤..."))
                            self.root.after(150, self.toggle_key_lock)

            except Exception as e:
                # 5-13-3. ë§Œì•½ ì—´ì‡ ë¥¼ ì½ë‹¤ê°€ ì‹¤ìˆ˜ë¥¼ í•˜ë©´ ë¬´ì—‡ì´ í‹€ë ¸ëŠ”ì§€ ì•Œë ¤ì¤ë‹ˆë‹¤.
                print(f"ì¸ì¦í‚¤ ë³µí˜¸í™” ì‹¤íŒ¨: {e}")

    # 5-14. [ì—´ì‡  ì €ì¥í•˜ê¸°] ì…ë ¥í•œ ì†Œì¤‘í•œ ì—´ì‡ (ì¸ì¦í‚¤)ë¥¼ ìŠì–´ë²„ë¦¬ì§€ ì•Šê²Œ íŒŒì¼ë¡œ ì ê¶ˆë‘ëŠ” í•¨ìˆ˜
    def save_key_to_file(self):
        # 5-14-1. ì…ë ¥ì°½ì— ì íŒ ë©”ì¸ ì—´ì‡ ì™€ ë°±ì—… ì—´ì‡ ë¥¼ ê°€ì ¸ì˜µë‹ˆë‹¤.
        main_k = self.service_key_var.get().strip()
        back_k = self.backup_key_var.get().strip()
        
        # 5-14-2. ë‘ ê°œì˜ ì—´ì‡ ë¥¼ í•˜ë‚˜ë¡œ ë¬¶ì–´ì¤ë‹ˆë‹¤.
        combined_data = f"{main_k}|{back_k}"
        
        try:
            # 5-14-3. [ë¹„ë°€ ì ê¸ˆ] ë‚¨ë“¤ì´ ëª» ë³´ê²Œ ì•”í˜¸ë¡œ ë°”ê¾¸ì–´ íŒŒì¼ì— ê¸°ë¡í•©ë‹ˆë‹¤.
            encrypted_data = cipher_suite.encrypt(combined_data.encode('utf-8'))
            with open(self.key_file_path, "wb") as f:
                f.write(encrypted_data)
            self.log("ğŸ”’ ë©”ì¸/ë°±ì—… ì¸ì¦í‚¤ê°€ ì•”í˜¸í™”ë˜ì–´ ì•ˆì „í•˜ê²Œ ì €ì¥ë˜ì—ˆìŠµë‹ˆë‹¤.")
        except Exception as e:
            # 5-14-4. ì €ì¥í•˜ë‹¤ê°€ ë¬¸ì œê°€ ìƒê¸°ë©´ ë¡œê·¸ì°½ì— ì•Œë¦½ë‹ˆë‹¤.
            self.log(f"âš  ì¸ì¦í‚¤ ì•”í˜¸í™” ì €ì¥ ì‹¤íŒ¨: {e}")

    # 5-15. [ì—´ì‡  í™•ì¸ ë° ì ê¸ˆ] ì…ë ¥í•œ ì—´ì‡ ê°€ ì§„ì§œì¸ì§€ ë‚˜ë¼ ì„œë²„ì— ë¬¼ì–´ë³´ê³  ì°½ì„ ì ê·¸ëŠ” í•¨ìˆ˜
    def toggle_key_lock(self):
        import xml.etree.ElementTree as ET
        from urllib.parse import unquote
        import requests

        # 5-15-1. ì•„ì§ ì—´ì‡ ê°€ í™•ì¸ë˜ì§€ ì•Šì•„ ì—´ë ¤ìˆëŠ” ìƒíƒœë¼ë©´ ê²€ì‚¬ë¥¼ ì‹œì‘í•©ë‹ˆë‹¤.
        if not self.key_locked:
            # 5-15-1-1. ì…ë ¥ì°½ì— ì íŒ ê¸€ìë“¤ì„ ê°€ì ¸ì˜µë‹ˆë‹¤.
            main_input = self.service_key_var.get().strip()
            back_input = self.backup_key_var.get().strip()

            # 5-15-1-2. ë©”ì¸ ì—´ì‡ ê°€ ì—†ìœ¼ë©´ ê²½ê³ ì°½ì„ ë„ìš°ê³  ë©ˆì¶¥ë‹ˆë‹¤.
            if not main_input:
                messagebox.showwarning("ì•Œë¦¼", "ë©”ì¸ ì¸ì¦í‚¤ë¥¼ ì…ë ¥í•´ì£¼ì„¸ìš”.")
                return

            self.log("ğŸ”‘ ì¸ì¦í‚¤ ìœ íš¨ì„± ê²€ì‚¬ ë° ì ìš© ì¤‘...")
            
            # 5-15-1-3. ê²€ì‚¬í•  ì—´ì‡ ë“¤ì„ ëª©ë¡ìœ¼ë¡œ ë§Œë“­ë‹ˆë‹¤.
            keys_to_test = [("ë©”ì¸", main_input, self.service_key_var)]
            if back_input:
                keys_to_test.append(("ë°±ì—…", back_input, self.backup_key_var))

            valid_main, valid_back = "", ""
            any_success = False

            # 5-15-1-4. [ì§„ì§œì¸ì§€ í™•ì¸] ëª©ë¡ì— ìˆëŠ” ì—´ì‡ ë“¤ì„ í•˜ë‚˜ì”© API ì„œë²„ì— ë¬¼ì–´ë´…ë‹ˆë‹¤.
            for name, key_val, var_obj in keys_to_test:
                test_url = "http://ws.bus.go.kr/api/rest/busRouteInfo/getBusRouteList"
                try:
                    # 5-15-1-4-1. ì„œë²„ì— ì•„ì£¼ ì§§ì€ ì§ˆë¬¸ì„ ë˜ì ¸ë´…ë‹ˆë‹¤.
                    r = requests.get(test_url, params={'serviceKey': unquote(key_val), 'strSrch': '12'}, timeout=5)
                    # 5-15-1-4-2. ì„œë²„ê°€ 'OK(0)'ë¼ê³  ëŒ€ë‹µí•˜ë©´ ì§„ì§œ ì—´ì‡ ì…ë‹ˆë‹¤.
                    if "<headerCd>0</headerCd>" in r.text:
                        self.log(f"âœ… {name} ì¸ì¦í‚¤: ì •ìƒ ì‘ë™ í™•ì¸")
                        if name == "ë©”ì¸": valid_main = key_val
                        else: valid_back = key_val
                        any_success = True
                    else:
                        # 5-15-1-4-3. ê°€ì§œê±°ë‚˜ ë§ê°€ì§„ ì—´ì‡ ë©´ ì…ë ¥ì°½ì—ì„œ ì§€ì›Œë²„ë¦½ë‹ˆë‹¤.
                        var_obj.set("") 
                        self.log(f"âŒ {name} ì¸ì¦í‚¤: ì‘ë™ ë¶ˆê°€í•˜ì—¬ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤.")
                except Exception as e:
                    # 5-15-1-4-4. ì¸í„°ë„· ì—°ê²° ë¬¸ì œ ë“±ìœ¼ë¡œ í™•ì¸ì´ ì•ˆ ë˜ë©´ ì‚­ì œ ì²˜ë¦¬í•©ë‹ˆë‹¤.
                    var_obj.set("")
                    self.log(f"âŒ {name} í†µì‹  ì—ëŸ¬ë¡œ ì‚­ì œ ì²˜ë¦¬ë˜ì—ˆìŠµë‹ˆë‹¤.")

            # 5-15-1-5. [ì„±ê³µ ì‹œ ì²˜ë¦¬] í•˜ë‚˜ë¼ë„ ì§„ì§œ ì—´ì‡ ê°€ ìˆë‹¤ë©´ í”„ë¡œê·¸ë¨ì„ í™œì„±í™”í•©ë‹ˆë‹¤.
            if any_success:
                # 5-15-1-5-1. ì„±ê³µí•œ ì—´ì‡ ë“¤ì„ ê¸°ì–µí•˜ê³  ì ê¸ˆ ìƒíƒœë¡œ ë°”ê¿‰ë‹ˆë‹¤.
                self.final_main_key = valid_main
                self.final_backup_key = valid_back
                self.key_locked = True
                
                # 5-15-1-5-2. ì‹¤ìˆ˜ë¡œ ì§€ìš°ì§€ ëª»í•˜ê²Œ ì…ë ¥ì°½ì„ íšŒìƒ‰(ì½ê¸°ì „ìš©)ìœ¼ë¡œ ì ê¸‰ë‹ˆë‹¤.
                self.entry_service_key.config(state='readonly', readonlybackground='#f0f0f0')
                self.entry_backup_key.config(state='readonly', readonlybackground='#f0f0f0')
                # ì ê¸ˆ ì™„ë£Œ â†’ ë²„íŠ¼ì„ outline(í…Œë‘ë¦¬) ìŠ¤íƒ€ì¼ë¡œ ë³€ê²½
                _outline = self.get_btn_style("outline")
                _outline.pop("state", None)
                self.btn_key_manage.config(text="ì¸ì¦í‚¤\në³€ê²½", **_outline)
                
                # 5-15-1-5-3. ì´ì œ ë²„ìŠ¤ë¥¼ ì°¾ì„ ìˆ˜ ìˆë„ë¡ ê²€ìƒ‰ ë²„íŠ¼ë“¤ì„ ì¼­ë‹ˆë‹¤.
                _s_on = self.get_btn_style("normal")
                _s_on.pop("state", None)
                for btn in self.btn_searches:
                    btn.config(state="normal", **_s_on)
                
                # 5-15-1-5-4. VLD í†µê³„ ìˆ«ìë¥¼ ì˜¬ë¦¬ê³  ì—´ì‡ ë¥¼ íŒŒì¼ì— ì €ì¥í•©ë‹ˆë‹¤.
                #             ì„±ê³µí•œ ê° í‚¤(ë©”ì¸/ë°±ì—…)ë³„ë¡œ ê°œë³„ ì¹´ìš´í„°ë„ ì˜¬ë¦½ë‹ˆë‹¤.
                if valid_main:
                    self.api_stats['VLD'] += 1
                    self.api_stats_by_key["main"]['VLD'] = self.api_stats_by_key["main"].get('VLD', 0) + 1
                if valid_back:
                    self.api_stats['VLD'] += 1
                    self.api_stats_by_key["back"]['VLD'] = self.api_stats_by_key["back"].get('VLD', 0) + 1
                self.root.after(0, self.update_api_counter_ui)
                self.save_key_to_file()
                self.log("ğŸ”’ ì •ìƒ ì¸ì¦í‚¤ê°€ í™•ì •ë˜ì—ˆìŠµë‹ˆë‹¤. ì´ì œ ê²€ìƒ‰ì´ ê°€ëŠ¥í•©ë‹ˆë‹¤.")
            else:
                # 5-15-1-6. ì§„ì§œ ì—´ì‡ ê°€ í•˜ë‚˜ë„ ì—†ìœ¼ë©´ ìŠ¬í”ˆ ì•Œë¦¼ì°½ì„ ë„ì›ë‹ˆë‹¤.
                messagebox.showerror("ì¸ì¦ ì‹¤íŒ¨", "ì‚¬ìš© ê°€ëŠ¥í•œ ì¸ì¦í‚¤ê°€ ì—†ìŠµë‹ˆë‹¤.")

        # 5-15-2. ì´ë¯¸ ì ê²¨ìˆëŠ” ìƒíƒœì—ì„œ ë²„íŠ¼ì„ ëˆ„ë¥´ë©´ ì ê¸ˆì„ í’€ê³  ìˆ˜ì • ëª¨ë“œë¡œ ë°”ê¿‰ë‹ˆë‹¤.
        else:
            self.key_locked = False
            # 5-15-2-1. ì…ë ¥ì°½ì„ ë‹¤ì‹œ í•˜ì–€ìƒ‰ìœ¼ë¡œ ë°”ê¿”ì„œ ê¸€ìë¥¼ ì“¸ ìˆ˜ ìˆê²Œ í•©ë‹ˆë‹¤.
            self.entry_service_key.config(state='normal', background='white')
            self.entry_backup_key.config(state='normal', background='white')
            # ì ê¸ˆ í•´ì œ â†’ ë²„íŠ¼ì„ normal(ë…¹ìƒ‰) ìŠ¤íƒ€ì¼ë¡œ ë³µì›
            _normal = self.get_btn_style("normal")
            _normal.pop("state", None)
            self.btn_key_manage.config(text="ì¸ì¦í‚¤\nì…ë ¥", **_normal)
            
            # 5-15-2-2. ì—´ì‡ ë¥¼ ê³ ì¹˜ëŠ” ë™ì•ˆì—ëŠ” ê²€ìƒ‰ ë²„íŠ¼ì„ ëª» ëˆ„ë¥´ê²Œ ë•ë‹ˆë‹¤.
            _s_off = self.get_btn_style("disabled")
            _s_off.pop("state", None)
            for btn in self.btn_searches:
                btn.config(state="disabled", **_s_off)
            
            # 5-15-2-3. ì ê¸ˆ í•´ì œ ì¦‰ì‹œ 64ì ê²€ì¦ì„ ë‹¤ì‹œ ì ìš©í•©ë‹ˆë‹¤.
            self._check_key_btn_state()
            self.log("ğŸ”“ ì¸ì¦í‚¤ ìˆ˜ì • ëª¨ë“œì…ë‹ˆë‹¤. (ê²€ì¦ ì „ê¹Œì§€ ê²€ìƒ‰ ë²„íŠ¼ ì ê¸ˆ)")

    # 5ê·¸ë£¹: ë°ì´í„° ìˆ˜ì§‘ ë° í†µì‹  (Fetching & API)

    # 5-16. [ì¸í„°ë„· ëŒ€í™”í•˜ê¸°] ë‚˜ë¼ ì„œë²„ì— ë²„ìŠ¤ ì •ë³´ë¥¼ ë¬¼ì–´ë³´ê³  ë‹µë³€ì„ ë°›ì•„ì˜¤ëŠ” í•µì‹¬ í•¨ìˆ˜
    def fetch_api(self, url, params, api_type=None):
        from urllib.parse import unquote
        # 5-16-1. ìš°ë¦¬ê°€ ê°€ì§„ ë©”ì¸ ì—´ì‡ ì™€ ë°±ì—… ì—´ì‡ ë¥¼ ì¤€ë¹„í•©ë‹ˆë‹¤.
        keys = [self.service_key_var.get().strip(), self.backup_key_var.get().strip()]
        
        # 5-16-2. ì–´ë–¤ ì •ë³´ë¥¼ ë¬¼ì–´ë³¼ì§€ì— ë”°ë¼ ì¸í„°ë„· ì£¼ì†Œë¥¼ ë¯¸ë¦¬ ì •í•©ë‹ˆë‹¤. API íƒ€ì…ë³„ ì£¼ì†Œ ì„¤ì • (ARR1->SINF / POS1->POS2 ìˆœì„œ)
        p_url, f_url = url, url
        if api_type == "ARR1":
            p_url = "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRoute"
            f_url = "http://ws.bus.go.kr/api/rest/arrive/getArrInfoByRouteAll"
        elif api_type == "POS2":
            p_url = "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRtid"
            f_url = "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRouteSt"

        # 5-16-3. ë©”ì¸ ì£¼ì†Œì™€ ë¹„ìƒ ì£¼ì†Œë¥¼ ì„ì–´ì„œ ì´ 4ë²ˆì˜ ì‹œë„ ê³„íšì„ ì„¸ì›ë‹ˆë‹¤.
        attempts = [
            (p_url, "ë©”ì¸", keys[0]), (p_url, "ë°±ì—…", keys[1]), 
            (f_url, "ë©”ì¸", keys[0]), (f_url, "ë°±ì—…", keys[1])
        ]

        # 5-16-4. ê³„íšì— ë”°ë¼ ì„œë²„ì— ì§ˆë¬¸ì„ ë˜ì§‘ë‹ˆë‹¤.
        for step, (curr_url, name, key) in enumerate(attempts):
            if not key: continue
            try:
                # 5-16-4-1. ì—´ì‡ ë¥¼ ê½‚ê³  ì„œë²„ì— ì—°ê²°í•©ë‹ˆë‹¤.
                params['serviceKey'] = unquote(key)
                resp = requests.get(curr_url, params=params, timeout=10)
                
                # 5-16-4-2. ì¼ì¼í˜¸ì¶œëŸ‰ì´ ì´ˆê³¼ë˜ì–´ ì„œë²„ê°€ ê±°ì ˆí•˜ë©´ ë‹¤ìŒ ì‹œë„ë¡œ ë„˜ì–´ê°‘ë‹ˆë‹¤.
                if "LIMITED NUMBER OF SERVICE REQUESTS" in resp.text or "<headerCd>22</headerCd>" in resp.text:
                    continue 

                # 5-16-4-3. ì—´ì‡ ê°€ ì˜ëª»ë˜ì—ˆë‹¤ê³  í•˜ë©´ ë‹¤ìŒ ì—´ì‡ ë¡œ ì‹œë„í•©ë‹ˆë‹¤.
                if "SERVICE KEY IS NOT REGISTERED" in resp.text or "UNREGISTERED_KEY" in resp.text:
                    continue

                # 5-16-4-4. ì„œë²„ê°€ ëŒ€ë‹µì„ ì˜ í•´ì£¼ì—ˆë‹¤ë©´ ë‚´ìš©ì„ ê¼¼ê¼¼íˆ ì½ì–´ë´…ë‹ˆë‹¤.
                if resp.status_code == 200:
                    root = ET.fromstring(resp.text)
                    header_cd = root.findtext(".//headerCd")
                    err_msg = root.findtext('.//headerMsg') or ""
                    
                    # 5-16-4-4-1. ì •ë³´ë¥¼ ì°¾ì•˜ê±°ë‚˜, ì •ë³´ê°€ ì—†ë‹¤ëŠ” í™•ì‹¤í•œ ëŒ€ë‹µì„ ë“¤ì—ˆì„ ë•Œë§Œ íšŸìˆ˜ë¥¼ ì…‰ë‹ˆë‹¤.
                    is_no_result = ("ê²°ê³¼ê°€ ì—†ìŠµë‹ˆë‹¤" in err_msg) or ("NODATA" in err_msg)
                    
                    if header_cd == "0" or is_no_result:
                        # 5-16-4-4-1-1. ì–´ë–¤ ì§ˆë¬¸ì„ í–ˆëŠ”ì§€ ì¢…ë¥˜ë³„ë¡œ í†µê³„ ìˆ«ìë¥¼ ì˜¬ë¦½ë‹ˆë‹¤.
                        success_url = curr_url
                        key_type = "VLD"  # ê¸°ë³¸ê°’: ì•Œ ìˆ˜ ì—†ëŠ” í˜¸ì¶œ
                        if "getArrInfoByRoute" in success_url and "All" not in success_url: key_type = "ARR1"
                        elif "getArrInfoByRouteAll" in success_url: key_type = "ARR2"
                        elif "getBusPosByRtid" in success_url: key_type = "POS1"
                        elif "getStationByUid" in success_url: key_type = "SINF"
                        elif "getBusPosByRouteSt" in success_url: key_type = "POS2"
                        elif "getStationByName" in success_url: key_type = "SCNM"
                        elif "getRouteByStation" in success_url: key_type = "SCID"
                        elif "getRouteInfo" in success_url: key_type = "RINF"
                        elif "getStaionByRoute" in success_url: key_type = "SLST"
                        
                        # 5-16-4-4-1-1-1. í•©ì‚° ì¹´ìš´í„°ë¥¼ ì˜¬ë¦½ë‹ˆë‹¤.
                        self.api_stats[key_type] += 1
                        # 5-16-4-4-1-1-2. ì–´ë–¤ í‚¤(ë©”ì¸/ë°±ì—…)ê°€ ì„±ê³µí–ˆëŠ”ì§€ ê°œë³„ ì¹´ìš´í„°ë„ ì˜¬ë¦½ë‹ˆë‹¤.
                        key_slot = "main" if name == "ë©”ì¸" else "back"
                        self.api_stats_by_key[key_slot][key_type] = \
                            self.api_stats_by_key[key_slot].get(key_type, 0) + 1
                        self.root.after(0, self.update_api_counter_ui)

                        # 5-16-4-4-1-2. ì§„ì§œ ë°ì´í„°ê°€ ë“¤ì–´ìˆìœ¼ë©´ ë‚´ìš©ì„ ì „ë‹¬í•˜ê³ , ì—†ìœ¼ë©´ ì´ìœ ë¥¼ ì ìŠµë‹ˆë‹¤.
                        if header_cd == "0":
                            return root
                        else:
                            # 5-16-4-4-1-2-1. ë²„ìŠ¤ ìœ„ì¹˜ë¥¼ ë¬¼ì–´ë´¤ëŠ”ë° ë²„ìŠ¤ê°€ í•œ ëŒ€ë„ ì—†ìœ¼ë©´ ë¡œê·¸ë¥¼ ë‚¨ê¸°ê³ 
                            #   ì²«ì°¨ ì‹œê°„(f_tm)ì„ í•¨ê»˜ ë°˜í™˜í•©ë‹ˆë‹¤. â€” í˜¸ì¶œìëŠ” (None, f_tm) í˜•íƒœë¥¼ ì²˜ë¦¬í•´ì•¼ í•©ë‹ˆë‹¤.
                            if any(u in success_url for u in ["getBusPosByRtid", "getBusPosByRouteSt"]):
                                rid_param = params.get('busRouteId', '')
                                rnm_log = self.rid_to_rnm.get(rid_param, rid_param)
                                self.log(f"{rnm_log}ë²ˆì€ ìš´í–‰ì¤‘ì¸ ì°¨ëŸ‰ì´ ì—†ìŠµë‹ˆë‹¤.")
                                # ì²«ì°¨ ì‹œê°„ì„ ë…¸ì„ ì •ë³´ API(getRouteInfo)ì—ì„œ ê°€ì ¸ì˜µë‹ˆë‹¤.
                                f_tm = self._fetch_route_first_time(rid_param)
                                return ('NO_BUS', f_tm)
                            return None
                        
            except Exception:
                # 5-16-4-5. ì¤‘ê°„ì— ì‹¤ìˆ˜ê°€ ìƒê¸°ë©´ ë‹¤ìŒ ë°©ë²•ìœ¼ë¡œ ë‹¤ì‹œ ì‹œë„í•©ë‹ˆë‹¤.
                continue
        return None
    
    # 5-16b. [ì²«ì°¨ ì‹œê° ì¡°íšŒ] í•´ë‹¹ ë…¸ì„ ì˜ ì²« ë²„ìŠ¤ ì‹œê°(f_tm)ì„ getRouteInfo APIë¡œ ì¡°íšŒí•©ë‹ˆë‹¤.
    def _fetch_route_first_time(self, rid):
        """getRouteInfoì—ì„œ firstBusTm í•„ë“œë¥¼ ê°€ì ¸ì˜µë‹ˆë‹¤ (HHmm ë˜ëŠ” HHmmss í˜•ì‹).
        ì‹¤íŒ¨ ë˜ëŠ” ì—†ëŠ” ì˜¤ì „ ì‹œë¦¬ì¦ˆ ë…¸ì„ ì€ None ë°˜í™˜."""
        if not rid:
            return None
        try:
            root_ri = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo",
                {'busRouteId': rid}
            )
            if root_ri is None or isinstance(root_ri, tuple):
                return None
            f_tm_raw = root_ri.findtext(".//firstBusTm") or ""
            if f_tm_raw:
                return self.format_hhmm(f_tm_raw)  # HH:MM í˜•ì‹
        except Exception:
            pass
        return None

    # 5-17. [ê²€ìƒ‰ ì°½ ì—´ê¸°] ë²„ìŠ¤ ì •ë¥˜ì†Œë¥¼ ì°¾ê³  ê¸°ë¡í•  ë²„ìŠ¤ë¥¼ ê³ ë¥´ëŠ” íŒì—… ì°½ í•¨ìˆ˜
    def open_search_window(self, target_idx):
        # 5-17-1. ì—´ì‡ ê°€ ì—†ìœ¼ë©´ ê²€ìƒ‰ì„ í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤.
        if not self.service_key_var.get().strip(): 
            messagebox.showwarning("ì•Œë¦¼", "ì¸ì¦í‚¤ë¥¼ ë¨¼ì € ì…ë ¥í•´ì£¼ì„¸ìš”."); return
        
        # 5-17-2. ê²€ìƒ‰ìš© ìƒˆ ì°½ì„ ì˜ˆì˜ê²Œ ë§Œë“­ë‹ˆë‹¤.
        search_win = tk.Toplevel(self.root)
        search_win.title(f"ì •ë¥˜ì†Œ {target_idx+1} ê²€ìƒ‰ ë° ë…¸ì„  ì„ íƒ")
        search_win.geometry("1000x850")
        search_win.grab_set() # 5-17-2-1. ì´ ì°½ì„ ë‹«ê¸° ì „ê¹Œì§€ëŠ” ë©”ì¸ ì°½ì„ ëª» ê±´ë“¤ê²Œ í•©ë‹ˆë‹¤.
        search_win.minsize(500, 500)
        
        # 5-17-3. ê¸€ìë¥¼ ì…ë ¥í•  ê²€ìƒ‰ì°½ì„ ë§Œë“­ë‹ˆë‹¤.
        frame_search = tk.Frame(search_win, pady=10)
        frame_search.pack(fill="x")
        tk.Label(frame_search, text="ì •ë¥˜ì†Œëª…/ID:", font=(FONT_MAIN, SZ_S)).pack(side="left", padx=(15,5))
        search_ent = tk.Entry(frame_search, width=30)
        search_ent.pack(side="left", padx=5)
        search_ent.focus_set() # 5-17-3-1. ì°½ì´ ì—´ë¦¬ìë§ˆì ë°”ë¡œ ê¸€ìë¥¼ ì“¸ ìˆ˜ ìˆê²Œ í•©ë‹ˆë‹¤.
        
        # 5-17-4. ì°¾ì€ ì •ë¥˜ì†Œë“¤ì„ ëª©ë¡ìœ¼ë¡œ ë³´ì—¬ì£¼ëŠ” í‘œë¥¼ ë§Œë“­ë‹ˆë‹¤.
        tree_st_frame = tk.Frame(search_win)
        tree_st_frame.pack(fill="x", padx=15, pady=5)
        self.style.configure("Search.Treeview.Heading", background="#dfe6e9", font=(FONT_SUB, SZ_S, "bold"))
        st_scroll = ttk.Scrollbar(tree_st_frame, orient="vertical")
        tree_st = ttk.Treeview(tree_st_frame, columns=("name", "arsid"), show="headings", height=5, style="Search.Treeview", yscrollcommand=st_scroll.set)
        st_scroll.config(command=tree_st.yview)
        tree_st.heading("name", text="ì •ë¥˜ì†Œëª…")
        tree_st.heading("arsid", text="ARS-ID")
        tree_st.column("name", width=659, stretch=True)
        tree_st.column("arsid", width=141, stretch=True)
        tree_st.bind('<Button-1>', self.prevent_column_resize)
        tree_st.pack(side="left", fill="both", expand=True)
        st_scroll.pack(side="right", fill="y")

        # 5-17-5. ë²„ìŠ¤ ë…¸ì„ ì„ ê³ ë¥¼ ë•Œ ì“¸ ë²„íŠ¼ë“¤ì„ ë°°ì¹˜í•©ë‹ˆë‹¤.
        btn_action_frame = tk.Frame(search_win)
        btn_action_frame.pack(fill="x", padx=15, pady=5)
        tk.Label(btn_action_frame, text="â–¼ ë…¸ì„  ì„ íƒ", font=(FONT_SUB, SZ_S, "bold"), fg="#2980b9").pack(side="left")

        self.current_route_data = [] 
        self.sort_state = {'key': None, 'reverse': False} 
        type_order = {"ê°„ì„ ": 0, "ì§€ì„ ": 1, "ê´‘ì—­": 2, "ê¸°íƒ€": 3, "ê²½ê¸°": 4, "ì¸ì²œ": 5} 

        # 5-17-6. ë²„ìŠ¤ ëª©ë¡ì„ ë‹¤ì‹œ ê·¸ë¦¬ëŠ” ë„ìš°ë¯¸ ê¸°ëŠ¥
        def update_route_tree_view(): 
            render_route_list()

        # 5-17-7. ëª©ë¡ì— ìˆëŠ” ëª¨ë“  ë²„ìŠ¤ë¥¼ í•œ ë²ˆì— ê³ ë¥´ê±°ë‚˜ ì·¨ì†Œí•˜ëŠ” ê¸°ëŠ¥
        def set_all_check(status): 
            for item in self.current_route_data: 
                item['checked'].set(status)
            update_route_tree_view()

        tk.Button(btn_action_frame, text="ì „ì²´ í•´ì œ", command=lambda: set_all_check(False), font=(FONT_SUB, SZ_XS)).pack(side="right", padx=2)
        tk.Button(btn_action_frame, text="ì „ì²´ ì„ íƒ", command=lambda: set_all_check(True), font=(FONT_SUB, SZ_XS)).pack(side="right", padx=2)

        # 5-17-8. ì„ íƒí•œ ì •ë¥˜ì†Œë¥¼ ì§€ë‚˜ê°€ëŠ” ë²„ìŠ¤ë“¤ì˜ ìƒì„¸ ì •ë³´ë¥¼ ë³´ì—¬ì¤„ í° í‘œë¥¼ ë§Œë“­ë‹ˆë‹¤.
        frame_route_list = tk.Frame(search_win, bd=1, relief="sunken")
        frame_route_list.pack(fill="both", expand=True, padx=15, pady=5)
        route_scroll = ttk.Scrollbar(frame_route_list, orient="vertical")
        route_cols = ("status", "rnm", "rtype", "path", "term", "f_tm", "l_tm", "st_cnt")
        tree_route = ttk.Treeview(frame_route_list, columns=route_cols, show="headings", 
                                  yscrollcommand=route_scroll.set, style="Search.Treeview", 
                                  selectmode="none") 
        route_scroll.config(command=tree_route.yview)
        
        # 5-17-8-1. ê³ ë¥¸ ë²„ìŠ¤ëŠ” ì—°ë‘ìƒ‰, ì•ˆ ê³ ë¥¸ ë²„ìŠ¤ëŠ” í•˜ì–€ìƒ‰ìœ¼ë¡œ ì¹ í•©ë‹ˆë‹¤.
        tree_route.tag_configure("selected", background="#d4edda", foreground="#155724") 
        tree_route.tag_configure("unselected", background="white", foreground="black") 
        
        # 5-17-8-2. í‘œì˜ ì œëª©ë“¤ì„ ì„¤ì •í•©ë‹ˆë‹¤. (ì œëª©ì„ ëˆ„ë¥´ë©´ ìˆœì„œëŒ€ë¡œ ì •ë ¬ë©ë‹ˆë‹¤.)
        tree_route.heading("status", text="ì„ íƒì—¬ë¶€")
        tree_route.column("status", width=42, anchor="center", stretch=True)
        tree_route.heading("rnm", text="ë…¸ì„ ë²ˆí˜¸", command=lambda: sort_data('rnm'))
        tree_route.column("rnm", width=117, anchor="center", stretch=True)
        tree_route.heading("rtype", text="ìœ í˜•", command=lambda: sort_data('rtype'))
        tree_route.column("rtype", width=50, anchor="center", stretch=True)
        tree_route.heading("path", text="ê¸°ì â†”ì¢…ì ")
        tree_route.column("path", width=347, anchor="center", stretch=True)
        tree_route.heading("term", text="ë°°ì°¨ê°„ê²©")
        tree_route.column("term", width=50, anchor="center", stretch=True)
        tree_route.heading("f_tm", text="ì²«ì°¨ì‹œê°")
        tree_route.column("f_tm", width=67, anchor="center", stretch=True)
        tree_route.heading("l_tm", text="ë§‰ì°¨ì‹œê°")
        tree_route.column("l_tm", width=67, anchor="center", stretch=True)
        tree_route.heading("st_cnt", text="ì •ë¥˜ì¥ìˆ˜")
        tree_route.column("st_cnt", width=60, anchor="center", stretch=True)
        
        tree_route.bind('<Button-1>', self.prevent_column_resize)
        tree_route.pack(side="left", fill="both", expand=True)
        route_scroll.pack(side="right", fill="y")

        # 5-17-9. [ì§„ì§œ ê²€ìƒ‰] ë‚´ê°€ ì“´ ê¸€ìë¡œ ì •ë¥˜ì†Œë¥¼ ì°¾ì•„ë‚´ëŠ” ë˜‘ë˜‘í•œ í•¨ìˆ˜
        def perform_search(event=None):
            keyword = search_ent.get().strip()
            if not keyword: return 
            
            # 5-17-9-1. ì˜ˆì „ì— ì°¾ì•˜ë˜ ëª©ë¡ì€ ê¹¨ë—í•˜ê²Œ ì§€ì›ë‹ˆë‹¤.
            for i in tree_st.get_children():
                tree_st.delete(i) 
            
            root_res = None
            is_digit_search = keyword.isdigit() and len(keyword) >= 3 

            # 5-17-9-2. ìˆ«ìë§Œ ì¼ë‹¤ë©´ ì •ë¥˜ì†Œ ë²ˆí˜¸(ARS-ID)ë¡œ ë¨¼ì € ì°¾ì•„ë´…ë‹ˆë‹¤.
            if is_digit_search:
                root_res = self.fetch_api("http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid", {'arsId': keyword})
            
            # 5-17-9-3. ë²ˆí˜¸ë¡œ ëª» ì°¾ì•˜ê±°ë‚˜ ê¸€ìë¥¼ ì¼ë‹¤ë©´ ì •ë¥˜ì†Œ ì´ë¦„ìœ¼ë¡œ ë‹¤ì‹œ ì°¾ì•„ë´…ë‹ˆë‹¤.
            if root_res is None or not root_res.findall(".//itemList"):
                if is_digit_search:
                    pass  # self.log(f"â„¹ ARS-ID({keyword}) ê²€ìƒ‰ ê²°ê³¼ ì—†ìŒ -> ëª…ì¹­ ê²€ìƒ‰ìœ¼ë¡œ ì¬ì‹œë„í•©ë‹ˆë‹¤.")
                root_res = self.fetch_api("http://ws.bus.go.kr/api/rest/stationinfo/getStationByName", {'stSrch': keyword})
            
            # 5-17-9-4. ì°¾ì€ ê²°ê³¼ê°€ ìˆë‹¤ë©´ ëª©ë¡ì— í•˜ë‚˜ì”© ì˜¬ë ¤ì¤ë‹ˆë‹¤.
            insert_count = 0
            seen_stations = set()
            if root_res is not None:
                items = root_res.findall(".//itemList")
                for item in items:
                    name = item.findtext("stNm")
                    ars_id = item.findtext("arsId")
                    if ars_id and ars_id != "0": 
                        key = (name, str(ars_id))
                        if key not in seen_stations:
                            seen_stations.add(key)
                            tree_st.insert("", "end", values=key) 
                            insert_count += 1
            
            # 5-17-9-5. ì•„ë¬´ê²ƒë„ ëª» ì°¾ì•˜ë‹¤ë©´ ì—†ë‹¤ê³  ì•Œë ¤ì¤ë‹ˆë‹¤.
            if insert_count == 0:
                tree_st.insert("", "end", values=("ê²€ìƒ‰ ê²°ê³¼ê°€ ì—†ìŠµë‹ˆë‹¤.", ""))

        search_ent.bind("<Return>", perform_search) # 5-17-9-6. ì—”í„°í‚¤ë¥¼ ì³ë„ ê²€ìƒ‰ì´ ë©ë‹ˆë‹¤.
        tk.Button(frame_search, text="ê²€ìƒ‰", command=perform_search, font=(FONT_SUB, SZ_XS)).pack(side="left", padx=5)

        # 5-17-10. [ì •ë ¬í•˜ê¸°] ëª©ë¡ì„ ê°€ë‚˜ë‹¤ìˆœì´ë‚˜ ë²„ìŠ¤ ì¢…ë¥˜ë³„ë¡œ ì •ë¦¬í•˜ëŠ” í•¨ìˆ˜
        def sort_data(key):
            if self.sort_state['key'] == key:
                self.sort_state['reverse'] = not self.sort_state['reverse'] 
            else:
                self.sort_state['key'] = key
                self.sort_state['reverse'] = False
            
            if key == 'rtype':
                self.current_route_data.sort(key=lambda x: (type_order.get(x['rtype'], 99), x['rnm']), reverse=self.sort_state['reverse'])
            else:
                self.current_route_data.sort(key=lambda x: x[key], reverse=self.sort_state['reverse'])
            render_route_list()

        # 5-17-11. [ë³´ì—¬ì£¼ê¸°] ì°¾ì€ ë²„ìŠ¤ ì •ë³´ë“¤ì„ í‘œì— ì‹¤ì œë¡œ ì±„ì›Œë„£ëŠ” í•¨ìˆ˜
        def render_route_list():
            for item in tree_route.get_children():
                tree_route.delete(item)
            for item in self.current_route_data:
                status_text = "V" if item['checked'].get() else "" 
                tag = "selected" if item['checked'].get() else "unselected"
                vals = [status_text] + item['data_vals']
                tree_route.insert("", "end", iid=item['rid'], values=vals, tags=(tag,))

        # 5-17-12. [ë²„ìŠ¤ ê³ ë¥´ê¸°] ëª©ë¡ì—ì„œ ë²„ìŠ¤ë¥¼ í´ë¦­í•˜ë©´ ì²´í¬ë°•ìŠ¤ê°€ ì¼œì¡Œë‹¤ êº¼ì¡Œë‹¤ í•˜ëŠ” í•¨ìˆ˜
        def on_route_click(event):
            if tree_route.identify_region(event.x, event.y) == "separator":
                return "break"
                
            item_id = tree_route.identify_row(event.y)
            if not item_id: return
            for item in self.current_route_data:
                if item['rid'] == item_id:
                    new_state = not item['checked'].get() 
                    item['checked'].set(new_state)
                    cur_values = list(tree_route.item(item_id, "values"))
                    cur_values[0] = "V" if new_state else ""
                    tree_route.item(item_id, values=cur_values, tags=("selected" if new_state else "unselected",))
                    break
        tree_route.bind("<Button-1>", on_route_click)

        # 5-17-13. [ë²„ìŠ¤ ë¶ˆëŸ¬ì˜¤ê¸°] ì •ë¥˜ì†Œë¥¼ ê³ ë¥´ë©´ ê·¸ê³³ì— ì •ì°¨í•˜ëŠ” ëª¨ë“  ë²„ìŠ¤ë¥¼ ì„œë²„ì—ì„œ ê°€ì ¸ì˜¤ëŠ” í•¨ìˆ˜
        def on_station_select(event):
            selected = tree_st.selection()
            if not selected: return
            ars_id = str(tree_st.item(selected[0])['values'][1]).zfill(5) 
            if not ars_id or ars_id == "": return
            
            # 5-17-13-1. ì˜ˆì „ì— ê³¨ëë˜ ë²„ìŠ¤ê°€ ìˆë‹¤ë©´ ê¸°ì–µí•´ë‘¡ë‹ˆë‹¤.
            saved_checked_routes = set()
            if self.ars_ids[target_idx].get().strip() == ars_id and self.target_st_info[target_idx].get('routes'):
                saved_checked_routes = {r[0] for r in self.target_st_info[target_idx]['routes']}
            
            self.current_route_data.clear()
            # 5-17-13-1-1. ì •ë¥˜ì†Œê°€ ë°”ë€Œë©´ SLSTÂ·ord ìºì‹œë¥¼ ì´ˆê¸°í™”í•©ë‹ˆë‹¤.
            #   ìƒˆ ì •ë¥˜ì†Œì˜ ars_idì— ë§ëŠ” ord ê°’ì´ ë‹¬ë¼ì§ˆ ìˆ˜ ìˆê¸° ë•Œë¬¸ì…ë‹ˆë‹¤.
            self._strt_cache.clear()
            self._strt_ord_cache.clear()
            # 5-17-13-2. ì´ ì •ë¥˜ì†Œì— ì–´ë–¤ ë²„ìŠ¤ê°€ ì˜¤ëŠ”ì§€ ì„œë²„ì— ë¬¼ì–´ë´…ë‹ˆë‹¤.
            root_route = self.fetch_api("http://ws.bus.go.kr/api/rest/stationinfo/getRouteByStation", {'arsId': ars_id})
            if root_route is not None:
                type_map = {"1":"ê³µí•­", "2":"ë§ˆì„", "3":"ê°„ì„ ", "4":"ì§€ì„ ", "5":"ìˆœí™˜", "6":"ê´‘ì—­", "7":"ì¸ì²œ", "8":"ê²½ê¸°", "9":"íì§€", "0":"ê³µìš©"}
                for it in root_route.findall(".//itemList"):
                    rid, rnm = it.findtext("busRouteId"), it.findtext("busRouteNm")
                    # 5-17-13-3. ê° ë²„ìŠ¤ì˜ ë°°ì°¨ê°„ê²©ì´ë‚˜ ì²«ì°¨ ì‹œê°„ì„ ë” ìì„¸íˆ ì•Œì•„ë´…ë‹ˆë‹¤.
                    root_info = self.fetch_api("http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo", {'busRouteId': rid})
                    rtype, path, term, f_tm, l_tm, st_cnt_str = "-", "-", "-", "-", "-", "-"
                    st_cnt_val = 100
                    if root_info is not None:
                        info = root_info.find(".//itemList")
                        if info is not None:
                            rtype = type_map.get(info.findtext("routeType"), "ê¸°íƒ€")
                            path = f"{info.findtext('stStationNm')}â†”{info.findtext('edStationNm')}"
                            term = info.findtext("term") + "ë¶„"; f_tm = self.format_hhmm(info.findtext("firstBusTm", "-"))
                            l_tm = self.format_hhmm(info.findtext("lastBusTm", "-"))
                            self.route_corp_map[rid] = info.findtext("corpNm", "ì •ë³´ì—†ìŒ") 
                            # 5-17-13-4. ì´ ë²„ìŠ¤ê°€ ëª‡ ê°œì˜ ì •ë¥˜ì¥ì„ ê±°ì³ ê°€ëŠ”ì§€ ì•Œì•„ë´…ë‹ˆë‹¤.
                            #   ë™ì‹œì— ì´ ì •ë¥˜ì†Œì˜ ìˆœë²ˆ(ord)ë„ ìºì‹œí•´ ë‘ì–´
                            #   confirm_selection ì—ì„œ ì¬í˜¸ì¶œ ì—†ì´ ì¬ì‚¬ìš©í•©ë‹ˆë‹¤.
                            root_st_list = self.fetch_api("http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute", {'busRouteId': rid})
                            if root_st_list is not None:
                                st_list_items = root_st_list.findall(".//itemList")
                                st_cnt_val = len(st_list_items)
                                st_cnt_str = f"{st_cnt_val}ê°œ"
                                # 5-17-13-4-1. ì‘ë‹µ ì „ì²´ë¥¼ ìºì‹œì— ë³´ê´€í•©ë‹ˆë‹¤.
                                self._strt_cache[rid] = root_st_list
                                # 5-17-13-4-2. í˜„ì¬ ì •ë¥˜ì†Œì˜ ìˆœë²ˆ(seq)ì„ ë¯¸ë¦¬ ì¶”ì¶œí•´ ord ìºì‹œì— ì €ì¥í•©ë‹ˆë‹¤.
                                for _si in st_list_items:
                                    _item_ars = str(_si.findtext("arsId") or "").zfill(5)
                                    if _item_ars == ars_id:
                                        _ord_val = _si.findtext("seq") or _si.findtext("staOrd") or ""
                                        if _ord_val:
                                            self._strt_ord_cache[rid] = _ord_val
                                        break
                    
                    self.rid_to_rnm[rid] = rnm
                    is_checked = (rid in saved_checked_routes) if saved_checked_routes else True 
                    self.current_route_data.append({
                        'rid': rid, 'rnm': rnm, 'rtype': rtype, 'st_cnt': st_cnt_val,
                        'data_vals': [rnm, rtype, path, term, f_tm, l_tm, st_cnt_str],
                        'checked': tk.BooleanVar(value=is_checked)
                    })
            
            # 5-17-13-5. ë³´ê¸° ì¢‹ê²Œ ë²„ìŠ¤ ì¢…ë¥˜ë³„ë¡œ ì •ë ¬í•´ì„œ ë³´ì—¬ì¤ë‹ˆë‹¤.
            self.sort_state = {'key': 'rtype', 'reverse': False}
            self.current_route_data.sort(key=lambda x: (type_order.get(x['rtype'], 99), x['rnm']))
            render_route_list()
        
        tree_st.bind("<<TreeviewSelect>>", on_station_select) 

        # 5-17-14. ì´ë¯¸ ì…ë ¥ëœ ì •ë¥˜ì†Œê°€ ìˆë‹¤ë©´ ì°½ì´ ì—´ë¦´ ë•Œ ë°”ë¡œ ë³´ì—¬ì¤ë‹ˆë‹¤.
        # 5-17-14-1. selection_set ì´ <<TreeviewSelect>> ì´ë²¤íŠ¸ë¥¼ ìë™ ë°œí™”í•˜ë¯€ë¡œ
        #   on_station_select ë¥¼ ì§ì ‘ ë‹¤ì‹œ í˜¸ì¶œí•˜ì§€ ì•ŠìŠµë‹ˆë‹¤.
        #   (ì´ì „ ë²„ì „ì—ì„œ ë‘ ë²ˆ í˜¸ì¶œí•´ SLST ê°€ 2Në²ˆ ë°œìƒí•˜ë˜ ë¬¸ì œë¥¼ ìˆ˜ì •í–ˆìŠµë‹ˆë‹¤.)
        if self.ars_ids[target_idx].get().strip():
            item_id = tree_st.insert("", "end", values=(self.target_st_info[target_idx].get('nm', ''), self.ars_ids[target_idx].get()))
            tree_st.selection_set(item_id)
            # on_station_select(None)  â† ì œê±°: selection_set ì´ ì´ë²¤íŠ¸ë¥¼ ìë™ ë°œí™”í•©ë‹ˆë‹¤.

        # 5-17-15. [ê²°ì •] ê³ ë¥¸ ë²„ìŠ¤ë“¤ì„ ì§„ì§œë¡œ í”„ë¡œê·¸ë¨ì— ì ìš©í•˜ëŠ” í•¨ìˆ˜
        def confirm_selection():
            # 5-17-15-1. ì²´í¬ëœ ë²„ìŠ¤ë“¤ë§Œ ê³¨ë¼ëƒ…ë‹ˆë‹¤.
            chosen = [(item['rid'], item['rnm'], item['st_cnt']) for item in self.current_route_data if item['checked'].get()]
            selected_st = tree_st.selection()
            if not selected_st or not chosen: return
            st_name, ars_id = tree_st.item(selected_st[0])['values'][0], str(tree_st.item(selected_st[0])['values'][1]).zfill(5)
            
            # 5-17-15-2. ë©”ì¸ í™”ë©´ì˜ ì •ë¥˜ì†Œ ì´ë¦„ì„ ë°”ê¿‰ë‹ˆë‹¤.
            self.ars_ids[target_idx].set(ars_id)
            self.lbl_st_names[target_idx].config(text=f"[{st_name}] ì‹¤ì‹œê°„ í˜„í™©")
            self.lbl_hist_titles[target_idx].config(text=f"[{st_name}] ë„ì°© ê¸°ë¡")
            
            # 5-17-15-3. ì •ë¥˜ì†Œì˜ ê³ ìœ  ë²ˆí˜¸(stId)ë¥¼ ì•Œì•„ë‚´ì–´ ì €ì¥í•©ë‹ˆë‹¤.
            root_st_uid = self.fetch_api("http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid", {'arsId': ars_id})
            st_id = root_st_uid.findtext(".//stId") if (root_st_uid is not None and not isinstance(root_st_uid, tuple)) else ""

            # 5-17-15-4. ê° ë…¸ì„ ì˜ ì´ ì •ë¥˜ì†Œ ìˆœë²ˆ(ord)ì„ ë¯¸ë¦¬ ì•Œì•„ë‘¡ë‹ˆë‹¤.
            #   getArrInfoByRoute í˜¸ì¶œ ì‹œ stId + busRouteId + ord ê°€ ëª¨ë‘ í•„ìš”í•˜ê¸° ë•Œë¬¸ì…ë‹ˆë‹¤.
            #   getStaionByRoute ì‘ë‹µì˜ ê° itemListì—ëŠ” arsId ì™€ seq(ìˆœë²ˆ) í•„ë“œê°€ ìˆìŠµë‹ˆë‹¤.
            #
            # 5-17-15-4-1. [SLST ìºì‹œ ì „ëµ]
            #   on_station_select ì—ì„œ ê° ë…¸ì„ ì˜ getStaionByRoute ê²°ê³¼ë¥¼ _strt_cache ì—,
            #   í•´ë‹¹ ì •ë¥˜ì†Œì˜ seq(ìˆœë²ˆ)ë¥¼ _strt_ord_cache ì— ë¯¸ë¦¬ ì €ì¥í•´ ë‘ì—ˆìŠµë‹ˆë‹¤.
            #   confirm_selection ì—ì„œëŠ” ì´ ìºì‹œë¥¼ ê·¸ëŒ€ë¡œ ì‚¬ìš©í•˜ë¯€ë¡œ ì¶”ê°€ API í˜¸ì¶œì´ ì—†ìŠµë‹ˆë‹¤.
            #   â†’ ì •ë¥˜ì†Œ 1ê°œÂ·ë…¸ì„  Nê°œ ì„ íƒ ì‹œ SLST í˜¸ì¶œì€ Në²ˆ(on_station_select)ì—ì„œ ëë‚©ë‹ˆë‹¤.
            #
            # 5-17-15-4-2. ì •ë¥˜ì†Œê°€ ë‹¤ì‹œ ì„ íƒë  ë•ŒëŠ” suspendÂ·veh_cacheë„ ì •ë¦¬í•©ë‹ˆë‹¤.
            if hasattr(self, 'target_st_info') and self.target_st_info[target_idx].get('routes'):
                old_rids = {r[0] for r in self.target_st_info[target_idx]['routes']}
                new_rids  = {r[0] for r, *_ in chosen}  # type: ignore[misc]
                new_rids  = {item[0] for item in chosen}
                gone = old_rids - new_rids
                for rid in gone:
                    self.pos_suspend_until.pop(rid, None)
                    # 5-17-15-4-2-1. ì‚¬ë¼ì§„ ë…¸ì„ ì˜ ì°¨ëŸ‰ë²ˆí˜¸ ìºì‹œë„ ì œê±°í•©ë‹ˆë‹¤.
                    old_st_id = self.target_st_info[target_idx].get('id', '')
                    self.veh_cache.pop((old_st_id, rid), None)

            ord_map = {}  # {busRouteId: ord}
            for rid, rnm, _ in chosen:
                # 5-17-15-4-3. on_station_select ì—ì„œ ë¯¸ë¦¬ ì±„ìš´ ord ìºì‹œë¥¼ ìš°ì„  ì‚¬ìš©í•©ë‹ˆë‹¤.
                if rid in self._strt_ord_cache:
                    ord_map[rid] = self._strt_ord_cache[rid]
                    # 5-17-15-4-3-1. SLST API ì¬í˜¸ì¶œ ì—†ì´ ìºì‹œ ê°’ìœ¼ë¡œ ì²˜ë¦¬í•©ë‹ˆë‹¤.
                elif rid in self._strt_cache:
                    # 5-17-15-4-3-2. ord ìºì‹œëŠ” ì—†ì§€ë§Œ SLST ì‘ë‹µì€ ìˆëŠ” ê²½ìš° ì§ì ‘ íƒìƒ‰í•©ë‹ˆë‹¤.
                    root_strt = self._strt_cache[rid]
                    for st_item in root_strt.findall(".//itemList"):
                        item_ars = str(st_item.findtext("arsId") or "").zfill(5)
                        if item_ars == ars_id:
                            ord_val = st_item.findtext("seq") or st_item.findtext("staOrd") or ""
                            if ord_val:
                                ord_map[rid] = ord_val
                            break
                else:
                    # 5-17-15-4-3-3. ìºì‹œê°€ ì „í˜€ ì—†ëŠ” ì˜ˆì™¸ì  ìƒí™©ì—ë§Œ APIë¥¼ í˜¸ì¶œí•©ë‹ˆë‹¤.
                    #   (ì •ìƒì ìœ¼ë¡œëŠ” on_station_select ì—ì„œ ì´ë¯¸ ì±„ì›Œì ¸ ìˆì–´ì•¼ í•©ë‹ˆë‹¤.)
                    self.log(f"âš  SLST ìºì‹œ ëˆ„ë½: {rnm}ë²ˆ â€” ì¬í˜¸ì¶œí•©ë‹ˆë‹¤.")
                    root_strt = self.fetch_api(
                        "http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute",
                        {'busRouteId': rid}
                    )
                    if root_strt is not None and not isinstance(root_strt, tuple):
                        for st_item in root_strt.findall(".//itemList"):
                            item_ars = str(st_item.findtext("arsId") or "").zfill(5)
                            if item_ars == ars_id:
                                ord_val = st_item.findtext("seq") or st_item.findtext("staOrd") or ""
                                if ord_val:
                                    ord_map[rid] = ord_val
                                break

            # 5-17-15-5. ì •ë¥˜ì†Œ ì„ íƒì´ ì™„ë£Œë˜ë©´ SLSTÂ·ord ìºì‹œë¥¼ ë¹„ì›Œì„œ ë‹¤ìŒ ì„ íƒ ë•Œ ìƒˆë¡œ ë°›ìŠµë‹ˆë‹¤.
            self._strt_cache.clear()
            self._strt_ord_cache.clear()
            
            self.target_st_info[target_idx] = {'id': st_id, 'nm': st_name, 'routes': chosen, 'ord_map': ord_map}
            self.log(f"ì •ë¥˜ì†Œ {target_idx+1} ì„¤ì • ì™„ë£Œ: {st_name} (ë…¸ì„  ì„ íƒ ì™„ë£Œ : ì´ {len(ord_map)}ê°œ ë…¸ì„  ì¤‘, {len(chosen)}ê°œ ë…¸ì„  ì„ íƒ)")
            search_win.destroy() 
            self.update_button_states() 
            
        # 5-17-16. ìµœì¢… í™•ì¸ ë²„íŠ¼ì„ ì°½ ì•„ë˜ì— ë§Œë“­ë‹ˆë‹¤.
        tk.Button(search_win, text="ì„ íƒí•œ ë…¸ì„ ë“¤ë¡œ ì ìš©", command=confirm_selection, 
                  height=2,
                  **self.get_btn_style("#27ae60", font_size=SZ_M) 
        ).pack(fill="x", padx=15, pady=10)

    # 5-18. [ì •ë³´ ê°€ì ¸ì˜¤ê¸°] ê³ ë¥¸ ì •ë¥˜ì†Œì˜ ë²„ìŠ¤ ë„ì°© ì‹œê°„ê³¼ ìœ„ì¹˜ë¥¼ ì‹¤ì œë¡œ í™•ì¸í•˜ëŠ” í•¨ìˆ˜
    def process_station(self, idx):
        # 5-18-0. [ë‚ ì§œ ì „í™˜ ì²´í¬] ìì •ì„ ë„˜ìœ¼ë©´ POS ì •ì§€ í…Œì´ë¸”ì„ ì´ˆê¸°í™”í•©ë‹ˆë‹¤.
        today = datetime.now().date()
        if today != self._last_date:
            self._last_date = today
            self.pos_suspend_until.clear()
            # veh_cacheëŠ” (st_id, rid) ê¸°ë°˜ìœ¼ë¡œ ë‚ ì§œì™€ ë¬´ê´€í•˜ê²Œ ìœ íš¨í•˜ë¯€ë¡œ ìì •ì— ì´ˆê¸°í™”í•˜ì§€ ì•ŠìŠµë‹ˆë‹¤.
            self.log("ğŸ“… ë‚ ì§œê°€ ë°”ë€Œì—ˆìŠµë‹ˆë‹¤. POS1 ì¼ì‹œì •ì§€ ê¸°ë¡ì„ ì´ˆê¸°í™”í•©ë‹ˆë‹¤.")

        # 5-18-1. ì–´ë–¤ ì •ë¥˜ì†Œë¥¼ í™•ì¸í•´ì•¼ í•˜ëŠ”ì§€ ì •ë³´ë¥¼ ê°€ì ¸ì˜µë‹ˆë‹¤.
        info, ars_id = self.target_st_info[idx], self.ars_ids[idx].get().strip()
        st_id, rt_rows = info.get('id', ''), []

        # 5-18-2. [ì‹¤ì‹œê°„ ì‹œê°„] ê° ë²„ìŠ¤ê°€ ëª‡ ë¶„ ë’¤ì— ì˜¤ëŠ”ì§€ ì„œë²„ì— ë¬¼ì–´ë´…ë‹ˆë‹¤.
        # â”€â”€ ARR1(ì£¼) ì´ˆê³¼ ëŒ€ë¹„: SINF(ë³´ì¡°)ëŠ” ì •ë¥˜ì†Œë‹¹ 1íšŒë§Œ í˜¸ì¶œí•˜ëŠ” ìºì‹œ â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        uid_cache = None  # None=ë¯¸í˜¸ì¶œ / {}=í˜¸ì¶œí–ˆìœ¼ë‚˜ ê²°ê³¼ì—†ìŒ / {rid:item}=ì •ìƒìˆ˜ì‹ 
        ord_map = info.get('ord_map', {})  # {busRouteId: ì •ë¥˜ì†Œìˆœë²ˆ} â€” confirm_selectionì—ì„œ ë¯¸ë¦¬ ì €ì¥

        def get_uid_cache():
            """SINF APIë¥¼ ì •ë¥˜ì†Œë‹¹ ìµœëŒ€ 1íšŒë§Œ í˜¸ì¶œí•´ ë‘ ê°€ì§€ í‚¤ë¡œ ì¡°íšŒ ê°€ëŠ¥í•œ ë”•ì…”ë„ˆë¦¬ë¥¼ ë°˜í™˜í•©ë‹ˆë‹¤.
            ê¸°ë³¸ í‚¤: busRouteId (ìˆ«ìí˜• ID)
            ë³´ì¡° í‚¤: rtNm ë˜ëŠ” busRouteAbrv (ë…¸ì„  ë²ˆí˜¸ ë¬¸ìì—´) â† busRouteIdê°€ ì—†ê±°ë‚˜ 0ì¸ ê²½ìš° ëŒ€ë¹„
            ë°˜í™˜ê°’: {'by_rid': {busRouteId: item}, 'by_nm': {rtNm: item}}
            """
            nonlocal uid_cache
            if uid_cache is not None:
                return uid_cache
            root_uid = self.fetch_api(
                "http://ws.bus.go.kr/api/rest/stationinfo/getStationByUid",
                {'arsId': ars_id}
            )
            if root_uid is None or isinstance(root_uid, tuple):
                uid_cache = {'by_rid': {}, 'by_nm': {}}
                return uid_cache
            by_rid, by_nm = {}, {}
            for item in root_uid.findall(".//itemList"):
                rid_key = item.findtext("busRouteId") or ""
                nm_key  = item.findtext("busRouteAbrv") or item.findtext("rtNm") or ""
                # rid í‚¤ ë“±ë¡: ë¹„ì–´ìˆê±°ë‚˜ "0"ì´ ì•„ë‹Œ ê²½ìš°ë§Œ
                if rid_key and rid_key not in by_rid:
                    by_rid[rid_key] = item
                # ë…¸ì„ ëª… í‚¤ ë“±ë¡: í•­ìƒ ë³´ì¡° ìˆ˜ë‹¨ìœ¼ë¡œ ì €ì¥
                if nm_key and nm_key not in by_nm:
                    by_nm[nm_key] = item
            uid_cache = {'by_rid': by_rid, 'by_nm': by_nm}
            self.log(f"â„¹ ARR1 ì´ˆê³¼ â†’ SINF ë³´ì¡° í˜¸ì¶œ ì™„ë£Œ (ì •ë¥˜ì†Œ {idx+1}, {len(by_rid)}ê°œ ë…¸ì„  ìˆ˜ì‹ )")
            return uid_cache

        # ë…¸ì„ ë§ˆë‹¤ ARR1 ìš°ì„  ì‹œë„, ì‹¤íŒ¨ ì‹œ SINF ìºì‹œì—ì„œ í•´ë‹¹ ë…¸ì„  í•­ëª©ë§Œ ì¶”ì¶œ
        for rid, rnm, _ in info.get('routes', []):
            row = None

            # 5-18-2-1. [ì£¼ API] getArrInfoByRoute â€” stId + busRouteId + ord ë¡œ í•´ë‹¹ ì •ë¥˜ì†Œ 1ê°œ í–‰ë§Œ ì¡°íšŒ
            ord_val = ord_map.get(rid, '0')
            arr_params = {'stId': st_id, 'busRouteId': rid, 'ord': ord_val}
            root_arr = self.fetch_api("", arr_params, api_type="ARR1")

            if root_arr is not None and not isinstance(root_arr, tuple):
                # 5-18-2-1-0. [ë…¸ì„  ì¬ì¶œí˜„] ARR1 ì‘ë‹µì— ì‹¤ì œ ìš´í–‰ ë²„ìŠ¤ê°€ ìˆì„ ë•Œë§Œ POS ì •ì§€ í•´ì œ
                #   íŒë‹¨ ê¸°ì¤€ (ë‘˜ ì¤‘ í•˜ë‚˜ë¼ë„ í•´ë‹¹í•˜ë©´ ìš´í–‰ ì¤‘):
                #     â‘  plainNo1/plainNo2 ì— ì‹¤ì œ ì°¨ëŸ‰ë²ˆí˜¸(ìˆ«ì+ì˜ë¬¸, "-"/"" ì•„ë‹˜)ê°€ ì¡´ì¬
                #     â‘¡ arrmsg1/arrmsg2 ê°€ ìš´í–‰ ì—†ìŒ ë©”ì‹œì§€("ìš´í–‰ì •ë³´ì—†ìŒ","ìš´í–‰ì¢…ë£Œ","ì¶œë°œëŒ€ê¸°","-","") ì´ì™¸ì˜ ê°’
                #   â†’ XML ì‘ë‹µ ìì²´ê°€ ì„±ê³µì´ë”ë¼ë„ ì°¨ëŸ‰ ë°ì´í„° ì—†ìœ¼ë©´ ì •ì§€ ìœ ì§€
                NO_BUS_MSGS = {"ìš´í–‰ì •ë³´ì—†ìŒ", "ìš´í–‰ì¢…ë£Œ", "ì¶œë°œëŒ€ê¸°", "-", ""}
                arr_items = root_arr.findall(".//itemList")
                has_active_bus = any(
                    (item.findtext("plainNo1") or "").strip() not in ("", "-")
                    or (item.findtext("plainNo2") or "").strip() not in ("", "-")
                    or (item.findtext("arrmsg1") or "").strip() not in NO_BUS_MSGS
                    or (item.findtext("arrmsg2") or "").strip() not in NO_BUS_MSGS
                    for item in arr_items
                )
                if rid in self.pos_suspend_until:
                    if has_active_bus:
                        del self.pos_suspend_until[rid]
                        self.log(f"ğŸ”„ {rnm}ë²ˆ ìš´í–‰ ì¬ê°œ í™•ì¸ â†’ POS ì •ì§€ í•´ì œ")
                    # else: ì•„ì§ ìš´í–‰ ì •ë³´ ì—†ìŒ â†’ ì •ì§€ ìœ ì§€ (ë¶ˆí•„ìš”í•œ ë¡œê·¸ ì—†ìŒ)
                for item in root_arr.findall(".//itemList"):
                    res_ars  = str(item.findtext("arsId") or "").zfill(5)
                    res_ord  = str(item.findtext("staOrd") or "")

                    # â‘  ì£¼ API(getArrInfoByRoute) ì‘ë‹µ: í•­ìƒ í•´ë‹¹ ì •ë¥˜ì†Œ 1í–‰ â†’ arsId ë¬´ì¡°ê±´ ìˆ˜ìš©
                    # â‘¡ ARR2 í´ë°± ì‘ë‹µ: arsId ì¼ì¹˜ í–‰ ìš°ì„ , arsId=0ì¸ ê²½ìš° staOrdë¡œ 2ì°¨ ê²€ì¦
                    #    - arsIdê°€ ì •í™•íˆ ì¼ì¹˜í•˜ë©´ OK
                    #    - arsId=0("00000")ì´ê³  ord_valì´ ì¼ì¹˜í•˜ë©´ ì¸ì²œ/ê²½ê¸° ë²„ìŠ¤ í•´ë‹¹ ì •ë¥˜ì†Œ í–‰ OK
                    #    - arsId=0ì´ê³  ordë¥¼ ëª¨ë¥¼ ë•Œ(ord_val='0')ëŠ” ì²« ë²ˆì§¸ 0í–‰ ìˆ˜ìš©(ê¸°ì¡´ ë™ì‘ ìœ ì§€)
                    arsid_match = (res_ars == ars_id)
                    arsid_zero  = (not res_ars.strip("0"))   # "00000" ë“± 0ìœ¼ë¡œë§Œ êµ¬ì„±
                    ord_match   = (ord_val != '0' and res_ord == ord_val)

                    if arsid_match or (arsid_zero and (ord_match or ord_val == '0')):
                        p1 = item.findtext("plainNo1") or "-"
                        p2 = item.findtext("plainNo2") or "-"
                        # 5-18-2-1-1. [ì°¨ëŸ‰ë²ˆí˜¸ ìºì‹œ] ARR1/ARR2ì—ì„œ ì–»ì€ ì‹¤ì œ ë²ˆí˜¸íŒì„ ì €ì¥í•©ë‹ˆë‹¤.
                        cache_key = (st_id, rid)
                        if p1 not in ("-", "", None) or p2 not in ("-", "", None):
                            self.veh_cache[cache_key] = (p1, p2)
                        row = (
                            item.findtext("rtNm") or rnm,
                            p1,
                            item.findtext("arrmsg1") or "-",
                            p2,
                            item.findtext("arrmsg2") or "-"
                        )
                        break
            elif root_arr is None:
                # 5-18-2-2. [ë³´ì¡°] ARR ì „ë¶€ ì‹¤íŒ¨(íšŸìˆ˜ ì´ˆê³¼) â†’ SINF ìºì‹œì—ì„œ í•´ë‹¹ ë…¸ì„  í•­ëª© ì¶”ì¶œ
                uid_data = get_uid_cache()
                # busRouteId(ìˆ«ìí˜•) ìš°ì„  ì¡°íšŒ, ì—†ìœ¼ë©´ ë…¸ì„ ëª…(ë¬¸ìì—´)ìœ¼ë¡œ ì¬ì‹œë„ (DeprecationWarning ìˆ˜ì •: ëª…ì‹œì  None ë¹„êµ)
                uid_item = uid_data['by_rid'].get(rid)
                if uid_item is None:
                    uid_item = uid_data['by_nm'].get(rnm)

                if uid_item is not None:
                    route_display = (
                        uid_item.findtext("busRouteAbrv")
                        or uid_item.findtext("rtNm")
                        or rnm
                    )
                    # 5-18-2-2-1. [ì°¨ëŸ‰ë²ˆí˜¸ ìºì‹œ ìš°ì„  í™œìš©] ARRì—ì„œ ë°›ì€ ë²ˆí˜¸íŒì´ ìˆìœ¼ë©´ vehId ëŒ€ì‹  ì‚¬ìš©í•©ë‹ˆë‹¤.
                    cache_key = (st_id, rid)
                    cached = self.veh_cache.get(cache_key)
                    if cached:
                        p1_display, p2_display = cached
                    else:
                        # ìºì‹œ ì—†ìŒ: 9ìë¦¬ vehIdë¥¼ í‘œì‹œí•©ë‹ˆë‹¤ (ê°œì„  ì—¬ì§€ê°€ ìˆìŠµë‹ˆë‹¤).
                        p1_display = uid_item.findtext("vehId1") or "-"
                        p2_display = uid_item.findtext("vehId2") or "-"
                    row = (
                        route_display,
                        p1_display,
                        uid_item.findtext("arrmsg1") or "-",
                        p2_display,
                        uid_item.findtext("arrmsg2") or "-"
                    )

            if row:
                rt_rows.append(row)

        # 5-18-2-3. ì•Œì•„ë‚¸ ì‹œê°„ ì •ë³´ë¥¼ ì‹¤ì‹œê°„ í‘œì— ë³´ì—¬ì¤ë‹ˆë‹¤.
        self.root.after(0, lambda d=rt_rows, t=self.trees_rt[idx]: self.refresh_tree(t, d))

        # 5-18-3. [ìŠ¤ë§ˆíŠ¸ ë„ì°© ì—¬ë¶€] ì¤‘ë³µ ë…¸ì„ ì€ API í˜¸ì¶œì„ ì•„ë¼ê³  ì •ë³´ë¥¼ ê³µìœ í•©ë‹ˆë‹¤.
        if st_id:
            # 5-18-3-1. ì²« ë²ˆì§¸ ì •ë¥˜ì†Œ(idx=0)ë¥¼ ì‹œì‘í•  ë•ŒëŠ” ì„ì‹œ ì €ì¥ì†Œë¥¼ ë¹„ì›ë‹ˆë‹¤.
            if idx == 0: self.temp_pos_data = {}

            for rid, rnm, st_cnt in info.get('routes', []):
                root_pos = None

                # 5-18-3-1-1. [POS ì •ì§€ í™•ì¸] ì²«ì°¨ ì‹œê° ì´ì „ì´ë©´ í˜¸ì¶œ ê±´ë„ˆëœë‹ˆë‹¤.
                now_dt = datetime.now()
                if rid in self.pos_suspend_until:
                    resume_dt = self.pos_suspend_until[rid]
                    if now_dt < resume_dt:
                        # ì•„ì§ ì¬ê°œ ì‹œê° ì „ â†’ POS í˜¸ì¶œ ìƒëµ
                        continue
                    else:
                        # ì¬ê°œ ì‹œê° ë„ë‹¬ â†’ ì •ì§€ í•´ì œ (ìµœì´ˆ 1íšŒë§Œ ë¡œê·¸ ì¶œë ¥)
                        del self.pos_suspend_until[rid]
                        if rid not in self.pos_resume_logged:
                            self.log(f"â° {rnm}ë²ˆ POS í˜¸ì¶œ ì¬ê°œ (ì²«ì°¨ 5ë¶„ ì „ ë„ë‹¬)")
                            self.pos_resume_logged.add(rid)

                # 5-18-3-2. [ìŠ¤ë§ˆíŠ¸ íŒë‹¨] ì´ë¯¸ ë‹¤ë¥¸ ì •ë¥˜ì†Œì—ì„œ ì´ ë…¸ì„ ì˜ ìœ„ì¹˜ ì •ë³´ë¥¼ ê°€ì ¸ì™”ëŠ”ì§€ í™•ì¸í•©ë‹ˆë‹¤.
                if rid in self.temp_pos_data:
                    # 5-18-3-2-1. ì´ë¯¸ ìˆë‹¤ë©´ ì„œë²„ì— ë¬»ì§€ ì•Šê³  ì €ì¥ëœ ì •ë³´ë¥¼ êº¼ë‚´ ì”ë‹ˆë‹¤.
                    root_pos = self.temp_pos_data[rid]
                else:
                    # 5-18-3-2-2. ì €ì¥ëœ ì •ë³´ê°€ ì—†ë‹¤ë©´ ì„œë²„ì— ì§ì ‘ ì¢Œí‘œë¥¼ ë¬¼ì–´ë´…ë‹ˆë‹¤.
                    pos_result = self.fetch_api("", {'busRouteId': rid, 'startOrd': '1', 'endOrd': str(st_cnt)}, api_type="POS2")

                    if isinstance(pos_result, tuple) and pos_result[0] == 'NO_BUS':
                        # 5-18-3-2-2-1. [ìš´í–‰ ì—†ìŒ] ì²«ì°¨ ì‹œê°ì„ ì¡°íšŒí•˜ì—¬ 5ë¶„ ì „ ì¬ê°œ ì‹œê°ì„ ë“±ë¡í•©ë‹ˆë‹¤.
                        _, f_tm_str = pos_result
                        if f_tm_str:
                            try:
                                f_hm = datetime.strptime(f_tm_str, "%H:%M")
                                # ë‚´ì¼ ë‚ ì§œë¡œ ì²«ì°¨ ì‹œê° ì¡°í•© (ì²«ì°¨ê°€ ë‹¤ìŒ ë‚ ì¸ ê²½ìš° ëŒ€ë¹„)
                                base_dt = now_dt.replace(hour=f_hm.hour, minute=f_hm.minute, second=0, microsecond=0)
                                if base_dt <= now_dt:
                                    base_dt += timedelta(days=1)
                                resume_dt = base_dt - timedelta(minutes=5)
                                if resume_dt > now_dt:
                                    # ì•„ì§ ì¬ê°œ ì‹œê°ì´ ì˜¤ì§€ ì•ŠìŒ â†’ ì¬ê°œ ì‹œê°ê¹Œì§€ ì •ì§€
                                    self.pos_suspend_until[rid] = resume_dt
                                    self.pos_resume_logged.discard(rid)
                                    self.log(f"ğŸ’¤ {rnm}ë²ˆ ì°¨ëŸ‰ ì—†ìŒ â†’ ì²«ì°¨ {f_tm_str} 5ë¶„ ì „({resume_dt.strftime('%H:%M')})ê¹Œì§€ POS ì •ì§€")
                                else:
                                    # ì´ë¯¸ 5ë¶„ ì „ ì°½ ì•ˆì— ìˆì§€ë§Œ ì•„ì§ ì°¨ëŸ‰ ì—†ìŒ â†’ 1ë¶„ í›„ ì¬ì‹œë„ (ë¡œê·¸ ì—†ìŒ)
                                    self.pos_suspend_until[rid] = now_dt + timedelta(minutes=1)
                                    # pos_resume_loggedëŠ” ìœ ì§€ (ì¤‘ë³µ ë¡œê·¸ ë°©ì§€)
                            except Exception:
                                # íŒŒì‹± ì‹¤íŒ¨ì‹œ 30ë¶„ í›„ ì¬ì‹œë„
                                self.pos_suspend_until[rid] = now_dt + timedelta(minutes=30)
                                self.pos_resume_logged.discard(rid)
                                self.log(f"ğŸ’¤ {rnm}ë²ˆ ì°¨ëŸ‰ ì—†ìŒ â†’ ì²«ì°¨ ì‹œê° íŒŒì‹± ì‹¤íŒ¨, 30ë¶„ í›„ ì¬ì‹œë„")
                        else:
                            # ì²«ì°¨ ì‹œê° ì—†ìŒ: 30ë¶„ í›„ ì¬ì‹œë„
                            self.pos_suspend_until[rid] = now_dt + timedelta(minutes=30)
                            self.pos_resume_logged.discard(rid)
                            self.log(f"ğŸ’¤ {rnm}ë²ˆ ì°¨ëŸ‰ ì—†ìŒ â†’ ì²«ì°¨ ì •ë³´ ì—†ìŒ, 30ë¶„ í›„ ì¬ì‹œë„")
                        continue  # POS ì²˜ë¦¬ ê±´ë„ˆëœ€

                    root_pos = pos_result
                    # 5-18-3-2-3. ê°€ì ¸ì˜¨ ì •ë³´ëŠ” ë‹¤ìŒ ì •ë¥˜ì†Œë¥¼ ìœ„í•´ ì„ì‹œ ì €ì¥ì†Œì— ë³´ê´€í•©ë‹ˆë‹¤.
                    if root_pos is not None:
                        self.temp_pos_data[rid] = root_pos

                # 5-18-3-3. ê°€ì ¸ì˜¨(ë˜ëŠ” ê³µìœ ë°›ì€) ìœ„ì¹˜ ì •ë³´ë¡œ ë„ì°© ì—¬ë¶€ë¥¼ íŒì •í•©ë‹ˆë‹¤.
                if root_pos is not None:
                    for bus in root_pos.findall(".//itemList"):
                        # 5-18-3-3-1. ë²„ìŠ¤ì˜ í˜„ì¬ ìœ„ì¹˜ê°€ ì§€ê¸ˆ ì²˜ë¦¬ ì¤‘ì¸ ì •ë¥˜ì†Œ(st_id)ì¸ì§€ í™•ì¸í•©ë‹ˆë‹¤.
                        if bus.findtext("lastStnId") == st_id:
                            veh_no = bus.findtext("plainNo")
                            # 5-18-3-3-2. ë°©ê¸ˆ ê¸°ë¡í•œ ë²„ìŠ¤ëŠ” 5ë¶„ ë™ì•ˆ ë‹¤ì‹œ ê¸°ë¡í•˜ì§€ ì•ŠìŠµë‹ˆë‹¤.
                            if veh_no in self.last_arrival_logs[idx] and (time.time() - self.last_arrival_logs[idx][veh_no] < 300): continue

                            f_time = self.format_datetm(bus.findtext("dataTm"))
                            # 5-18-3-3-3. ë²„ìŠ¤ íšŒì‚¬ ì´ë¦„ì„ ì°¾ì•„ë³´ê³  ê³µë™ë°°ì°¨ ì •ë³´ë¥¼ í™•ì¸í•©ë‹ˆë‹¤.
                            corp_nm = self.route_corp_map.get(rid, "ì •ë³´ì—†ìŒ")
                            if rnm in self.excel_multi_corp_map and len(self.excel_multi_corp_map[rnm]) >= 2:
                                corp_nm = ", ".join(sorted(list(self.excel_multi_corp_map[rnm])))

                            # 5-18-3-3-4. ë„ì°© ê¸°ë¡íŒì— í•œ ì¤„ ì ê³  ì—‘ì…€ì—ë„ ìë™ìœ¼ë¡œ ì €ì¥í•©ë‹ˆë‹¤.
                            log_entry = (f_time, rnm, veh_no, corp_nm, "ì •ë¥˜ì†Œ ë„ì°©")
                            self.recorded_data.append((f_time, info['nm'], rnm, veh_no, corp_nm, "ì •ë¥˜ì†Œ ë„ì°©"))
                            self.last_arrival_logs[idx][veh_no] = time.time()
                            self.perform_auto_save()
                            self.log(f"â˜… ì •ë¥˜ì†Œ {idx+1} ë„ì°©: {rnm} ({veh_no})")
                            self.root.after(0, lambda r=log_entry, t=self.trees_hist[idx]: t.insert("", 0, values=r))
       
    # 5-19. [ê³µë™ë°°ì°¨í™•ì¸ìš© ì—‘ì…€ íŒŒì¼ ë¶€ë¥´ê¸°] ë²„ìŠ¤ íšŒì‚¬ ì •ë³´ë¥¼ ë‹´ì€ ì—‘ì…€ì„ ë’·ë‹¨ì—ì„œ ëª°ë˜ ë°›ì•„ì˜¤ê¸° ì‹œì‘í•˜ëŠ” í•¨ìˆ˜
    def start_excel_download_thread(self):
        # 5-19-1. í”„ë¡œê·¸ë¨ì´ ë©ˆì¶”ì§€ ì•Šê²Œ ìŠ¤ë ˆë“œ(ë³„ë„ì˜ ì¼ê¾¼)ë¥¼ ì¨ì„œ ë‹¤ìš´ë¡œë“œí•©ë‹ˆë‹¤.
        threading.Thread(target=self.download_and_load_excel, daemon=True).start()

    # 5-20. [ê³µë™ë°°ì°¨í™•ì¸ìš© ì—‘ì…€ íŒŒì¼ ë‹¤ìš´ë¡œë“œ] ì„œìš¸ ë°ì´í„° ê´‘ì¥ì—ì„œ ìµœì‹  ë²„ìŠ¤ ì •ë³´ ì—‘ì…€ì„ ì§ì ‘ ë°›ì•„ì˜¤ëŠ” í•¨ìˆ˜
    def download_and_load_excel(self):
        target_url = "https://data.seoul.go.kr/dataList/OA-15066/F/1/datasetView.do" 
        self.log(f"ğŸ“¥ ìµœì‹  ë…¸ì„ ì •ë³´ í™•ì¸ ì¤‘ (Text Search): {target_url}")
        
        download_success = False
        downloaded_filename = None
        driver = None

        try:
            # 5-20-1. ì¸í„°ë„· ë¡œë´‡(í¬ë¡¬ ë“œë¼ì´ë²„) ì„¤ì •ì„ í•©ë‹ˆë‹¤.
            options = Options()
            options.add_argument("--headless") # 5-20-1-1. í™”ë©´ì—ëŠ” ì•ˆ ë³´ì´ê²Œ ìˆ¨ê¹ë‹ˆë‹¤.
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--window-size=1920,1080")
            
            # 5-20-1-2. ë¡œë´‡ì´ ì•„ë‹ˆë¼ ì§„ì§œ ì‚¬ëŒì¸ ê²ƒì²˜ëŸ¼ ê¾¸ë°‰ë‹ˆë‹¤.
            options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            options.add_argument("--disable-blink-features=AutomationControlled") 
            options.add_experimental_option("excludeSwitches", ["enable-automation"]) 
            options.add_experimental_option('useAutomationExtension', False)
            
            # 5-20-1-3. íŒŒì¼ì„ ì €ì¥í•  ìš°ë¦¬ ì»´í“¨í„°ì˜ í´ë”ë¥¼ ì •í•©ë‹ˆë‹¤.
            target_download_dir = os.path.abspath(self.current_dir)
            prefs = {
                "download.default_directory": target_download_dir,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True,
                "profile.default_content_settings.popups": 0
            }
            options.add_experimental_option("prefs", prefs)

            # 5-20-2. í¬ë¡¬ ë¡œë´‡ì„ ì‹¤í–‰í•´ì„œ ì‚¬ì´íŠ¸ë¡œ ë³´ëƒ…ë‹ˆë‹¤.
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            
            if not os.path.exists(target_download_dir):
                os.makedirs(target_download_dir, exist_ok=True)

            driver.get(target_url) 
            
            # 5-20-3. ì‚¬ì´íŠ¸ê°€ ë‹¤ ì—´ë¦´ ë•Œê¹Œì§€ ìµœëŒ€ 20ì´ˆ ë™ì•ˆ ê¸°ë‹¤ë ¤ì¤ë‹ˆë‹¤.
            wait = WebDriverWait(driver, 20)
            self.log("   ã„´í˜ì´ì§€ ë¶„ì„ ì¤‘ (íŒŒì¼ëª… íƒìƒ‰)...")
            
            # 5-20-4. '.xlsx'ë¼ê³  ì¨ì§„ ì—‘ì…€ íŒŒì¼ ë‹¤ìš´ë¡œë“œ ë‹¨ì¶”ë¥¼ ì°¾ìŠµë‹ˆë‹¤.
            download_target = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//*[contains(text(), '.xlsx')]")
            ))
            
            found_text = download_target.text.strip()
            self.log(f"   ã„´ë‹¤ìš´ë¡œë“œ ëŒ€ìƒ ë°œê²¬: {found_text}")
            
            # 5-20-5. ì´ë¯¸ ê°™ì€ íŒŒì¼ì´ ìˆë‹¤ë©´ ì§€ìš°ê³  ìƒˆë¡œ ë°›ê¸°ë¡œ í•©ë‹ˆë‹¤.
            file_path = os.path.join(self.current_dir, found_text)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    self.log(f"   â„¹ ê¸°ì¡´ íŒŒì¼ ì‚­ì œ í›„ ë‹¤ìš´ë¡œë“œ: {found_text}")
                except Exception as del_err:
                    self.log(f"   âš  ê¸°ì¡´ íŒŒì¼ ì‚­ì œ ì‹¤íŒ¨ (ì‚¬ìš©ì¤‘ì¼ ìˆ˜ ìˆìŒ): {del_err}")

            before_files = set(os.listdir(self.current_dir)) 
            
            # 5-20-6. ë§ˆìš°ìŠ¤ë¡œ ë‹¤ìš´ë¡œë“œ ë²„íŠ¼ì„ í´ë¦­í•©ë‹ˆë‹¤.
            driver.execute_script("arguments[0].scrollIntoView(true);", download_target)
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", download_target)
            
            # 5-20-7. íŒŒì¼ì´ ë‹¤ ë°›ì•„ì§ˆ ë•Œê¹Œì§€ 1ì´ˆë§ˆë‹¤ í´ë”ë¥¼ í™•ì¸í•©ë‹ˆë‹¤.
            found_file = None
            for _ in range(30):
                time.sleep(1)
                after_files = set(os.listdir(self.current_dir))
                new_files = after_files - before_files 
                
                for f in new_files:
                    if f.endswith(".xlsx") and "ì„œìš¸ì‹œë²„ìŠ¤ë…¸ì„ ê¸°ë³¸ì •ë³´" in f:
                        found_file = f
                        break
                if found_file:
                    break
            
            # 5-20-8. ì„±ê³µí•˜ë©´ ì•Œë ¤ì£¼ê³  ì‹¤íŒ¨í•˜ë©´ ì´ìœ ë¥¼ ì ìŠµë‹ˆë‹¤.
            if found_file:
                self.log(f"âœ… ë‹¤ìš´ë¡œë“œ ì„±ê³µ: {found_file}")
                download_success = True
                downloaded_filename = found_file
            else:
                self.log("âš  ë‹¤ìš´ë¡œë“œ ì‹œë„í–ˆìœ¼ë‚˜ íŒŒì¼ ìƒì„± ì•ˆë¨ (ì‹œê°„ì´ˆê³¼)")

        except Exception as e:
            self.log(f"âŒ ë‹¤ìš´ë¡œë“œ í”„ë¡œì„¸ìŠ¤ ì˜¤ë¥˜: {e}")
            self.log("   (Chrome ë¸Œë¼ìš°ì €ì™€ ì¸í„°ë„· ì—°ê²°ì„ í™•ì¸í•´ì£¼ì„¸ìš”)")
        finally:
            if driver:
                driver.quit() # 5-20-9. ë¡œë´‡ì„ í‡´ê·¼ì‹œí‚µë‹ˆë‹¤(ì¸í„°ë„· ì°½ ë‹«ê¸°).
        
        # 5-20-10. ìƒˆë¡œ ë°›ì€ íŒŒì¼ì„ ì—´ê³ , ì—†ìœ¼ë©´ ì˜›ë‚  íŒŒì¼ì´ë¼ë„ ì°¾ì•„ë´…ë‹ˆë‹¤.
        if download_success and downloaded_filename:
            self.load_excel_routes(specific_filename=downloaded_filename)
        else:
            self.log("â„¹ ê¸°ì¡´ íŒŒì¼ ì¤‘ ìµœì‹  íŒŒì¼ì„ ì‚¬ìš©í•©ë‹ˆë‹¤.")
            self.load_excel_routes()

    # 5-21. [ê³µë™ë°°ì°¨í™•ì¸ìš© ì—‘ì…€ ì½ê¸°] ë°›ì•„ì˜¨ ì—‘ì…€ íŒŒì¼ì„ ì—´ì–´ì„œ ë²„ìŠ¤ ë…¸ì„ ê³¼ íšŒì‚¬ ì´ë¦„ì„ ì •ë¦¬í•˜ëŠ” í•¨ìˆ˜
    def load_excel_routes(self, specific_filename=None):
        target_file = specific_filename
        
        # 5-21-1. ì–´ë–¤ íŒŒì¼ì¸ì§€ ëª¨ë¥´ë©´ ì´ë¦„ì´ ë¹„ìŠ·í•œ ê²ƒ ì¤‘ì— ê°€ì¥ ìµœê·¼ ê²ƒì„ ê³ ë¦…ë‹ˆë‹¤.
        if not target_file:
            candidates = [f for f in os.listdir(self.current_dir) if "ì„œìš¸ì‹œë²„ìŠ¤ë…¸ì„ ê¸°ë³¸ì •ë³´" in f and f.endswith(".xlsx")]
            if candidates:
                candidates.sort(reverse=True) 
                target_file = candidates[0]
            else:
                target_file = "ì„œìš¸ì‹œë²„ìŠ¤ë…¸ì„ ê¸°ë³¸ì •ë³´(20260108).xlsx" 

        path = os.path.join(self.current_dir, target_file)
        
        # 5-21-2. íŒŒì¼ì´ ì§„ì§œ ìˆëŠ”ì§€ í™•ì¸í•©ë‹ˆë‹¤.
        if not os.path.exists(path):
            self.log(f"âš  ë¡œë“œí•  ì—‘ì…€ íŒŒì¼ì´ ì—†ìŠµë‹ˆë‹¤: {target_file}")
            return

        try:
            # 5-21-3. ì—‘ì…€ì„ í‘œ ëª¨ì–‘ìœ¼ë¡œ ì½ì–´ì„œ 'ë…¸ì„ ë²ˆí˜¸'ì™€ 'ì—…ì²´ëª…'ì„ ê¸°ì–µí•©ë‹ˆë‹¤.
            df = pd.read_excel(path)
            r_col = [c for c in df.columns if 'ë…¸ì„ ë²ˆí˜¸' in str(c)]
            c_col = [c for c in df.columns if 'ì—…ì²´ëª…' in str(c)]
            
            if r_col and c_col:
                r_col_name = r_col[0]
                c_col_name = c_col[0]
                
                temp_map = {}
                # 5-21-3-1. í‘œë¥¼ í•œ ì¤„ì”© ì½ìœ¼ë©´ì„œ ë²„ìŠ¤ ë²ˆí˜¸ì— íšŒì‚¬ë¥¼ ì§ì§€ì–´ì¤ë‹ˆë‹¤.
                for _, row in df.iterrows():
                    r_name = str(row[r_col_name]).strip()
                    c_name = str(row[c_col_name]).strip()
                    if r_name not in temp_map:
                        temp_map[r_name] = set()
                    temp_map[r_name].add(c_name) 
                
                self.excel_multi_corp_map = temp_map 
                self.log(f"âœ… ì—‘ì…€ ë°ì´í„° ë¡œë“œ ì™„ë£Œ: {target_file}")
            else:
                self.log(f"âš  ì—‘ì…€ í˜•ì‹ ì˜¤ë¥˜: 'ë…¸ì„ ë²ˆí˜¸' ë˜ëŠ” 'ì—…ì²´ëª…' ì—´ ë¯¸ë°œê²¬")
        except Exception as e:
            self.log(f"âŒ ì—‘ì…€ ë¡œë“œ ì¤‘ ì˜¤ë¥˜: {e}")

    # 6ê·¸ë£¹ : ì‹¤ì‹œê°„ ëª¨ë‹ˆí„°ë§ ë¡œì§ (Monitoring)        

    # 5-22. [ê¸°ë¡ ì—”ì§„ ì¼œê¸°] ê°ì‹œì™€ ê¸°ë¡ì„ ì‹¤ì œë¡œ ì‹œì‘í•˜ëŠ” í•¨ìˆ˜
    def start_monitoring(self):
        # 5-22-1. ëª‡ ì´ˆë§ˆë‹¤ í™•ì¸í• ì§€ ì•ˆ ì ì—ˆë‹¤ë©´ ì‹œì‘í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤.
        if not self.refresh_interval_var.get().strip():
            messagebox.showwarning("ì…ë ¥ ëˆ„ë½", "ê°±ì‹ ì£¼ê¸°ë¥¼ ì…ë ¥í•´ì•¼ ìë™ ê¸°ë¡ì„ ì‹œì‘í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤.")
            return

        if not self.service_key_var.get().strip(): return 
        
        # 5-22-2. ê¸°ë¡ ì¤‘ì—ëŠ” ì£¼ê¸° ì…ë ¥ì°½ì„ ëª» ê±´ë“œë¦¬ê²Œ ë§‰ìŠµë‹ˆë‹¤.
        self.entry_refresh_interval.config(state='readonly')
        
        self.save_key_to_file()
        self.is_monitoring = True
        # ìë™ ê¸°ë¡ ì¤‘ì—ëŠ” ì •ë¥˜ì†Œ ê²€ìƒ‰ ë²„íŠ¼ì„ ë¹„í™œì„±í™”í•©ë‹ˆë‹¤.
        _s_off2 = self.get_btn_style("disabled")
        _s_off2.pop("state", None)
        for _b in self.btn_searches:
            _b.config(state="disabled", **_s_off2)
        
        # 5-22-3. ì˜¤ëŠ˜ ê¸°ë¡ì„ ì €ì¥í•  ìƒˆ ì—‘ì…€ íŒŒì¼ì„ í•˜ë‚˜ ë§Œë“­ë‹ˆë‹¤.
        filename = f"Bus_Arrival_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.auto_save_path = os.path.join(self.current_dir, filename)
        
        try:
            pd.DataFrame(columns=["ë°ì´í„°ì‹œê°", "ì •ë¥˜ì†Œëª…", "ë…¸ì„ ", "ì°¨ëŸ‰ë²ˆí˜¸", "ìš´ìˆ˜ì‚¬ëª…", "ìƒíƒœ"]).to_excel(self.auto_save_path, index=False)
            self.can_auto_save = True 
            display_name = os.path.basename(self.auto_save_path)
            self.lbl_auto_save_status.config(text=f"[{display_name}] íŒŒì¼ì— ìë™ ê¸°ë¡ ì¤‘ ......", fg="#27ae60")
        except Exception as e:
            self.can_auto_save = False
            messagebox.showwarning("ì˜¤ë¥˜", f"ìë™ ì €ì¥ ì„¤ì • ì¤‘ ì˜¤ë¥˜ ë°œìƒ: {e}")

        # 5-22-4. ì¼ê¾¼(ìŠ¤ë ˆë“œ)ì„ ë¶ˆëŸ¬ì„œ ë¬´í•œ ë£¨í”„(main_loop)ë¥¼ ëŒë¦½ë‹ˆë‹¤.
        self.update_button_states()
        threading.Thread(target=self.main_loop, daemon=True).start()

    # 5-23. [ë¬´í•œ ë°˜ë³µ] ì •í•´ì§„ ì‹œê°„ë§ˆë‹¤ ê³„ì†í•´ì„œ ì •ë³´ë¥¼ ìƒˆë¡œê³ ì¹¨í•˜ëŠ” í•¨ìˆ˜
    def main_loop(self):
        next_call = time.time()
        # 5-23-1. ìš°ë¦¬ê°€ ë©ˆì¶”ë¼ê³  í•  ë•Œê¹Œì§€ ê³„ì† ë°˜ë³µí•©ë‹ˆë‹¤.
        while self.is_monitoring:
            self.refresh_data() # 5-23-1-1. ì„œë²„ì— ë¬¼ì–´ë³´ê³  ë°ì´í„°ë¥¼ ê°±ì‹ í•©ë‹ˆë‹¤.
            try: interval = int(self.refresh_interval_var.get())
            except: interval = 60
            next_call += interval
            
            # 5-23-2. ë‹¤ìŒ ê°±ì‹  ë•Œê¹Œì§€ ê¸°ë‹¤ë¦½ë‹ˆë‹¤. (ì¤‘ê°„ì— ë©ˆì¶¤ì„ ëˆ„ë¥¼ ìˆ˜ ìˆê²Œ 0.1ì´ˆì”© ë‚˜ëˆ ì„œ ì‰½ë‹ˆë‹¤.)
            sleep_time = next_call - time.time()
            if sleep_time < 0: next_call = time.time(); sleep_time = 0
            for _ in range(int(sleep_time * 10)):
                if not self.is_monitoring: break
                time.sleep(0.1)

    # 5-24. [ì§€ê¸ˆ ë°”ë¡œ] ê¸°ë‹¤ë¦¬ì§€ ì•Šê³  ë‹¹ì¥ ì •ë³´ë¥¼ ìƒˆë¡œ ê°€ì ¸ì˜¤ëŠ” í•¨ìˆ˜
    def manual_refresh(self):
        self.btn_manual.config(state="disabled")
        self.log("ğŸ”„ [ìˆ˜ë™] ë°ì´í„° ê°±ì‹  ì‹œì‘...")

        # 5-24-1. ë³„ë„ì˜ ì¼ê¾¼ì„ ì¨ì„œ ì§€ê¸ˆ ë°”ë¡œ ì •ë³´ë¥¼ ê°±ì‹ í•©ë‹ˆë‹¤.
        def run_manual():
            self.refresh_data(manual=True) 
            self.log("âœ… [ìˆ˜ë™] ë°ì´í„° ê°±ì‹  ì™„ë£Œ")
            self.root.after(0, lambda: self.btn_manual.config(state="normal"))

        threading.Thread(target=run_manual, daemon=True).start()

    # 5-25. [ëª¨ë‘ ìƒˆë¡œê³ ì¹¨] ëª¨ë“  ì •ë¥˜ì†Œì˜ ì •ë³´ë¥¼ í•œêº¼ë²ˆì— ê°±ì‹ í•˜ëŠ” ê³µìš© í•¨ìˆ˜
    def refresh_data(self, manual=False):
        # 5-25-0. [ë…¸ì„  ë³€ë™ ì²´í¬] ARR1 ì‘ë‹µì—ì„œ ê¸°ì¡´ì— ì—†ë˜ ë…¸ì„ ì´ ë‚˜íƒ€ë‚˜ê±°ë‚˜ ì‚¬ë¼ì§„ ê²½ìš°ë¥¼ ê°ì§€í•©ë‹ˆë‹¤.
        #   pos_suspend_until ì— ë“±ë¡ëœ ë…¸ì„ ì´ ARR1 ì‘ë‹µì—ì„œ ìƒˆë¡œ ë³´ì´ë©´ ì •ì§€ë¥¼ í•´ì œí•©ë‹ˆë‹¤.
        #   (ì‹¤ì œ ì²´í¬ëŠ” process_station ë‚´ë¶€ì—ì„œ ìˆ˜í–‰í•˜ë©°, ì—¬ê¸°ì„œëŠ” í”Œë˜ê·¸ë§Œ ê¸°ë¡í•©ë‹ˆë‹¤.)

        # 5-25-1. ì •ë¥˜ì†Œ 1ë²ˆê³¼ 2ë²ˆì„ ì°¨ë¡€ëŒ€ë¡œ í™•ì¸í•©ë‹ˆë‹¤.
        for i in range(2):
            if self.target_st_info[i].get('routes'): 
                self.process_station(i) 
        
        # 5-25-2. ìë™ìœ¼ë¡œ ê°±ì‹ ë  ë•Œë§Œ ë¡œê·¸ë¥¼ ë‚¨ê¹ë‹ˆë‹¤. (ë„ˆë¬´ ë§ì´ ë‚¨ìœ¼ë©´ ì§€ì €ë¶„í•˜ë‹ˆê¹Œìš”.)
        if not manual:
            self.log("ë°ì´í„° ê°±ì‹  ì™„ë£Œ")

    # 5-26. [ê¸°ë¡ ë©ˆì¶¤] í•˜ë˜ ì¼ì„ ë©ˆì¶”ê³  ì‰¬ëŠ” ë‹¨ê³„ë¡œ ëŒì•„ê°€ëŠ” í•¨ìˆ˜
    def stop_monitoring(self):
        self.is_monitoring = False
        # ìë™ ê¸°ë¡ ì¤‘ì§€ í›„ í‚¤ê°€ ì ê¸´ ìƒíƒœë©´ ê²€ìƒ‰ ë²„íŠ¼ì„ ë‹¤ì‹œ í™œì„±í™”í•©ë‹ˆë‹¤.
        if self.key_locked:
            _s_on2 = self.get_btn_style("normal")
            _s_on2.pop("state", None)
            for _b in self.btn_searches:
                _b.config(state="normal", **_s_on2)
        self.log("ğŸ›‘ ìë™ ê¸°ë¡ì„ ì¤‘ì§€í•©ë‹ˆë‹¤.")
        
        # 5-26-1. ì´ì œ ì£¼ê¸°ë¥¼ ë‹¤ì‹œ ê³ ì¹  ìˆ˜ ìˆê²Œ ì…ë ¥ì°½ì„ ì—½ë‹ˆë‹¤.
        self.entry_refresh_interval.config(state='normal')
        self.lbl_auto_save_status.config(text=" â€» ìë™ ê¸°ë¡ ì‹œì‘ ë²„íŠ¼ì„ ì‘ë™ì‹œí‚¤ë©´ ë„ì°© ê¸°ë¡ì´ ì—‘ì…€íŒŒì¼ë¡œ ìë™ ì €ì¥ë©ë‹ˆë‹¤.", fg="#e74c3c")
        self.update_button_states()

    # 7ê·¸ë£¹ : ë°ì´í„° ë¶„ì„ ë° ê¸°ë¡, ì¶œë ¥ (Analysis, Log, Export)

    # 5-27. [ê¸°ë¡ ì§€ìš°ê¸°] í‘œì— ìŒ“ì¸ ë„ì°© ê¸°ë¡ë“¤ì„ ì‹¹ ì²­ì†Œí•˜ëŠ” í•¨ìˆ˜
    def clear_history(self, idx):
        # 5-27-1. í˜¹ì‹œ ì‹¤ìˆ˜ë¡œ ëˆ„ë¥¸ ê±¸ ìˆ˜ë„ ìˆìœ¼ë‹ˆ í•œ ë²ˆ ë” ë¬¼ì–´ë´…ë‹ˆë‹¤.
        if messagebox.askyesno("ì‚­ì œ", f"ì •ë¥˜ì†Œ {idx+1}ì˜ ëª¨ë“  ê¸°ë¡ì„ ì‚­ì œí•˜ì‹œê² ìŠµë‹ˆê¹Œ?\nê¸°ë¡ì„ ì‚­ì œí•˜ë”ë¼ë„, ì´ë¯¸ ì €ì¥ëœ ê¸°ë¡ì€ ì‚­ì œë˜ì§€ ì•ŠìŠµë‹ˆë‹¤."): 
            tree = self.trees_hist[idx]
            # 5-27-2. í‘œì˜ ëª¨ë“  ì¤„ì„ í•˜ë‚˜ì”© ì§€ì›ë‹ˆë‹¤.
            for item in tree.get_children():
                tree.delete(item)
            self.log(f"ì •ë¥˜ì†Œ {idx+1} ê¸°ë¡ì„ ì´ˆê¸°í™”í–ˆìŠµë‹ˆë‹¤.")

    # 5-28. [ì‹œê°„ ì˜ˆì˜ê²Œ] ìˆ«ìë§Œ ìˆëŠ” ì‹œê°„ì„ ì‚¬ëŒì´ ì½ê¸° ì¢‹ê²Œ(12:34) ë°”ê¿”ì£¼ëŠ” í•¨ìˆ˜
    def format_hhmm(self, raw_str):
        if not raw_str or len(raw_str) < 4: return raw_str
        # 5-28-1. ê¸´ ì‹œê°„ ë°ì´í„°ë©´ ì¤‘ê°„ì— ì‹œê°„ë§Œ ì™ ë½‘ì•„ëƒ…ë‹ˆë‹¤.
        if len(raw_str) >= 14: return f"{raw_str[8:10]}:{raw_str[10:12]}"
        return f"{raw_str[:2]}:{raw_str[2:4]}" 

    # 5-29. [ì¼ì§€ ì ê¸°] í™”ë©´ í•˜ë‹¨ ë¡œê·¸ì°½ì— ë©”ì‹œì§€ë¥¼ ë‚¨ê¸°ëŠ” í•¨ìˆ˜ (ìµœëŒ€ 5,000ì¤„ê¹Œì§€ë§Œ ê¸°ì–µí•´ìš”)
    def log(self, msg):
        # 5-29-0. [ë§¥OS ìŠ¤ë ˆë“œ ì•ˆì „] ë°±ê·¸ë¼ìš´ë“œ ìŠ¤ë ˆë“œì—ì„œ í˜¸ì¶œë  ìˆ˜ ìˆìœ¼ë¯€ë¡œ
        #         tk.Text ì¡°ì‘ì€ ë°˜ë“œì‹œ ë©”ì¸ ìŠ¤ë ˆë“œ(root.after)ë¥¼ í†µí•´ ì‹¤í–‰í•©ë‹ˆë‹¤.
        #         txt_logê°€ ì•„ì§ ìƒì„±ë˜ì§€ ì•Šì€ ê²½ìš°(ì´ˆê¸°í™” ì¤‘)ëŠ” ì½˜ì†” ì¶œë ¥ìœ¼ë¡œ ëŒ€ì²´í•©ë‹ˆë‹¤.
        now = datetime.now().strftime("%H:%M:%S")
        try:
            key = self.service_key_var.get().strip()
            backup_key = self.backup_key_var.get().strip()
        except Exception:
            key, backup_key = "", ""

        # 5-29-1. ë¹„ë°€ë²ˆí˜¸ê°€ ë¡œê·¸ì— ë³´ì´ë©´ ì•ˆ ë˜ë‹ˆ ë³„í‘œ(****)ë¡œ ê°€ë ¤ì¤ë‹ˆë‹¤.
        if key and len(key) > 4: msg = msg.replace(key, "********")
        if backup_key and len(backup_key) > 4: msg = msg.replace(backup_key, "********")
        
        line = f"[{now}] {msg}\n"

        # 5-29-2. txt_logê°€ ì¤€ë¹„ëëŠ”ì§€ í™•ì¸í•˜ê³  ë©”ì¸ ìŠ¤ë ˆë“œë¡œ UIë¥¼ ì—…ë°ì´íŠ¸í•©ë‹ˆë‹¤.
        if not hasattr(self, 'txt_log'):
            # 5-29-2-1. UI ì´ˆê¸°í™” ì „ì—ëŠ” ì½˜ì†”ì— ì¶œë ¥í•©ë‹ˆë‹¤(ë§¥ í„°ë¯¸ë„ì—ì„œë„ í™•ì¸ ê°€ëŠ¥).
            print(line, end="")
            return

        def _do_insert():
            try:
                # 5-29-2-2. ì½ê¸°ì „ìš©(disabled) ìœ„ì ¯ì— ì“°ë ¤ë©´ ì ì‹œ NORMALë¡œ ì—´ì—ˆë‹¤ê°€ ë‹¤ì‹œ ë‹«ìŠµë‹ˆë‹¤.
                self.txt_log.config(state="normal")
                self.txt_log.insert(tk.END, line)
                self.txt_log.see(tk.END)
                # 5-29-3. ë¡œê·¸ 5,000ì¤„ì´ ë„˜ì–´ê°€ë©´ ê°€ì¥ ì˜¤ë˜ëœ(ì œì¼ ìœ„) ì¤„ë¶€í„° ì§€ì›ë‹ˆë‹¤.
                line_count = int(self.txt_log.index('end-1c').split('.')[0])
                if line_count > 5000:
                    self.txt_log.delete('1.0', '2.0')
                self.txt_log.config(state="disabled")
            except Exception:
                pass  # 5-29-3-1. ì°½ì´ ë‹«íˆëŠ” ë„ì¤‘ ì˜ˆì™¸ê°€ ë°œìƒí•´ë„ ë¬´ì‹œí•©ë‹ˆë‹¤.

        # 5-29-4. ë©”ì¸ ìŠ¤ë ˆë“œì—ì„œ ì‹¤í–‰ë˜ë„ë¡ ì˜ˆì•½í•©ë‹ˆë‹¤.
        try:
            self.root.after(0, _do_insert)
        except Exception:
            pass

    # 5-30. [ë‚ ì§œ ì˜ˆì˜ê²Œ] ìˆ«ì ë©ì–´ë¦¬ ë‚ ì§œë¥¼ ë‹¬ë ¥ì²˜ëŸ¼(2026-01-01 12:00:00) ë°”ê¾¸ëŠ” í•¨ìˆ˜
    def format_datetm(self, raw_tm):
        try:
            if raw_tm and len(raw_tm) >= 14: 
                return datetime.strptime(raw_tm, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
            return raw_tm
        except: return raw_tm

    # 5-31. [ì—‘ì…€ ë§Œë“¤ê¸°] ê¸°ë¡í•œ ë°ì´í„°ë“¤ì„ ë©‹ì§„ ì—‘ì…€ íŒŒì¼ë¡œ ì™„ì„±í•˜ëŠ” í•µì‹¬ í•¨ìˆ˜
    def _core_excel_save_logic(self, target_path, save_completed=False):
        # 5-31-0. save_completed=True ì¼ ë•Œ:
        #   ì˜ì—…ì¼(BizDate)ì´ ì²˜ìŒ 'ì™„ê²°'ë˜ëŠ” ìˆœê°„(ìƒˆ ë‚ ì§œ ì‹œíŠ¸ê°€ ì—´ë¦´ ë•Œ)
        #   í•´ë‹¹ ë‚ ì§œì˜ ë°ì´í„°ë§Œ Bus_Arrival_Log_YYYYMMDD_completed.xlsx ë¡œ ë³„ë„ ì €ì¥í•©ë‹ˆë‹¤.
        try:
            import pandas as pd
            from datetime import datetime, timedelta
            from openpyxl.styles import Font, PatternFill, Alignment

            # 5-31-1. ìš°ë¦¬ê°€ ëª¨ì€ ê¸°ë¡ë“¤ì„ ì—‘ì…€ í‘œ ëª¨ì–‘ìœ¼ë¡œ ì •ë¦¬í•©ë‹ˆë‹¤.
            cols = ["ë°ì´í„°ì‹œê°", "ì •ë¥˜ì†Œëª…", "ë…¸ì„ ", "ì°¨ëŸ‰ë²ˆí˜¸", "ìš´ìˆ˜ì‚¬ëª…", "ìƒíƒœ"]
            df = pd.DataFrame(self.recorded_data, columns=cols)
            
            # 5-31-2. [ì˜ì—…ì¼ ê³„ì‚°] ìƒˆë²½ 3ì‹œ ì „ì— ë“¤ì–´ì˜¨ ë²„ìŠ¤ëŠ” 'ì–´ì œ' ë²„ìŠ¤ë¡œ ì³ì¤ë‹ˆë‹¤.
            def get_biz_date(dt_str):
                try:
                    dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
                    if dt.hour < 3: dt -= timedelta(days=1) 
                    return dt.strftime("%Y-%m-%d")
                except: return "Unknown"
            
            df['BizDate'] = df['ë°ì´í„°ì‹œê°'].apply(get_biz_date)

            # 5-31-2-1. í˜„ì¬ ì¡´ì¬í•˜ëŠ” ì˜ì—…ì¼ ëª©ë¡ì„ êµ¬í•©ë‹ˆë‹¤.
            biz_dates_in_data = set(df['BizDate'].unique())
            # 5-31-2-2. ì˜¤ëŠ˜(í˜„ì¬ ì§„í–‰ ì¤‘ì¸) ì˜ì—…ì¼ì„ ê³„ì‚°í•©ë‹ˆë‹¤.
            now = datetime.now()
            current_biz = (now - timedelta(days=1)).strftime("%Y-%m-%d") if now.hour < 3 else now.strftime("%Y-%m-%d")
            # 5-31-2-3. ì™„ê²°ëœ ë‚ ì§œ: ë°ì´í„°ì— ìˆìœ¼ë‚˜ í˜„ì¬ ì˜ì—…ì¼ì´ ì•„ë‹Œ ê³¼ê±° ë‚ ì§œì…ë‹ˆë‹¤.
            completed_dates = biz_dates_in_data - {current_biz, "Unknown"}
            
            # 5-31-3. ì—‘ì…€ì„ ì €ì¥í•˜ë©´ì„œ ë‚ ì§œë³„ë¡œ ì‹œíŠ¸(ì¢…ì´)ë¥¼ ë‚˜ëˆ„ì–´ ë‹´ìŠµë‹ˆë‹¤.
            with pd.ExcelWriter(target_path, engine='openpyxl') as writer:
                for biz_date, group in df.groupby('BizDate'):
                    sheet_data = group.drop(columns=['BizDate'])
                    sheet_data.to_excel(writer, sheet_name=biz_date, index=False)
                    
                    worksheet = writer.sheets[biz_date]
                    
                    # 5-31-3-1. ì²« ë²ˆì§¸ ì¤„(ì œëª©ì¤„)ì„ í•˜ëŠ˜ìƒ‰ìœ¼ë¡œ ì¹ í•˜ê³  ê¸€ì”¨ë¥¼ êµµê²Œ í•©ë‹ˆë‹¤.
                    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    header_font = Font(bold=True, size=11)
                    header_alignment = Alignment(horizontal='center', vertical='center')

                    # 5-31-3-2. ê° ì¹¸ì˜ ë„ˆë¹„ë¥¼ ê¸€ì ê¸¸ì´ì— ë§ì¶°ì„œ ë³´ê¸° ì¢‹ê²Œ ì¡°ì ˆí•©ë‹ˆë‹¤.
                    for idx, col in enumerate(sheet_data.columns):
                        cell = worksheet.cell(row=1, column=idx+1)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment

                        max_data_len = sheet_data[col].astype(str).map(len).max()
                        base_width = max(max_data_len, len(str(col)))
                        
                        # 5-31-3-2-1. ì •ë¥˜ì†Œ ì´ë¦„ì²˜ëŸ¼ ê¸´ ê¸€ìëŠ” ë” ë„“ê²Œ, ì‹œê°ì€ ì¢ê²Œ ê°œë³„ ì¡°ì ˆí•©ë‹ˆë‹¤.
                        if col == "ë°ì´í„°ì‹œê°":
                            adjusted_width = base_width + 1 
                        elif col == "ì •ë¥˜ì†Œëª…":
                            adjusted_width = base_width + 15
                        elif col == "ë…¸ì„ ":
                            adjusted_width = base_width + 8 
                        elif col == "ìš´ìˆ˜ì‚¬ëª…":
                            adjusted_width = base_width + 7 
                        else:
                            adjusted_width = base_width + 5 
                        
                        col_letter = worksheet.cell(row=1, column=idx+1).column_letter
                        worksheet.column_dimensions[col_letter].width = adjusted_width

            # 5-31-4. [ì™„ê²° íŒŒì¼ ìë™ ì €ì¥]
            #   save_completed=True ì´ë©´ì„œ, ì•„ì§ ì €ì¥í•˜ì§€ ì•Šì€ ì™„ê²° ì˜ì—…ì¼ì´ ìˆì„ ë•Œë§Œ ì €ì¥í•©ë‹ˆë‹¤.
            if save_completed:
                new_completed = completed_dates - self._completed_dates_saved
                for biz_date in sorted(new_completed):
                    # 5-31-4-1. í•´ë‹¹ ì˜ì—…ì¼ ë°ì´í„°ë§Œ ì¶”ì¶œí•©ë‹ˆë‹¤.
                    day_df = df[df['BizDate'] == biz_date].drop(columns=['BizDate'])
                    if day_df.empty:
                        continue
                    # 5-31-4-2. íŒŒì¼ëª…: Bus_Arrival_Log_YYYYMMDD_completed.xlsx
                    safe_date = biz_date.replace("-", "")
                    completed_name = f"Bus_Arrival_Log_{safe_date}_completed.xlsx"
                    completed_path = os.path.join(self.current_dir, completed_name)
                    try:
                        with pd.ExcelWriter(completed_path, engine='openpyxl') as cw:
                            day_df.to_excel(cw, sheet_name=biz_date, index=False)
                            ws = cw.sheets[biz_date]
                            hfill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                            hfont = Font(bold=True, size=11)
                            halign = Alignment(horizontal='center', vertical='center')
                            for ci, col in enumerate(day_df.columns):
                                cell = ws.cell(row=1, column=ci+1)
                                cell.fill = hfill; cell.font = hfont; cell.alignment = halign
                                max_len = day_df[col].astype(str).map(len).max()
                                col_letter = ws.cell(row=1, column=ci+1).column_letter
                                ws.column_dimensions[col_letter].width = max(max_len, len(str(col))) + 5
                        # 5-31-4-3. ì €ì¥ ì„±ê³µ ì‹œ ì¤‘ë³µ ì €ì¥ ë°©ì§€ ì§‘í•©ì— ì¶”ê°€í•©ë‹ˆë‹¤.
                        self._completed_dates_saved.add(biz_date)
                        self.log(f"ğŸ“ ì˜ì—…ì¼ ì™„ê²° íŒŒì¼ ì €ì¥: {completed_name}")
                    except Exception as ce:
                        self.log(f"âš  ì™„ê²° íŒŒì¼ ì €ì¥ ì‹¤íŒ¨ ({biz_date}): {ce}")

            return True

        except PermissionError:
            self.log("âš  ì—‘ì…€ íŒŒì¼ì´ ì—´ë ¤ ìˆì–´ ì €ì¥ì„ ê±´ë„ˆëœë‹ˆë‹¤. (íŒŒì¼ì„ ë‹«ì•„ì£¼ì„¸ìš”)")
            return False
        except Exception as e:
            self.log(f"âŒ ì—‘ì…€ ì €ì¥ ì¤‘ ì˜¤ë¥˜ ë°œìƒ: {e}")
            return False
        
    # 5-32. [ìˆ˜ë™ ì €ì¥] ìš°ë¦¬ê°€ ì§ì ‘ ë²„íŠ¼ì„ ëˆŒëŸ¬ì„œ ì›í•˜ëŠ” ì´ë¦„ìœ¼ë¡œ ì—‘ì…€ì„ ì €ì¥í•˜ëŠ” í•¨ìˆ˜
    def save_to_excel(self):
        # 5-32-1. ê¸°ë¡ì´ í•˜ë‚˜ë„ ì—†ìœ¼ë©´ ì €ì¥í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤.
        if not self.recorded_data:
            messagebox.showwarning("ì•Œë¦¼", "ì €ì¥í•  ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
            return 
        
        # 5-32-2. ì–´ë””ì— ì–´ë–¤ ì´ë¦„ìœ¼ë¡œ ì €ì¥í• ì§€ ë¬¼ì–´ë³´ëŠ” ì°½ì„ ë„ì›ë‹ˆë‹¤.
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel íŒŒì¼", "*.xlsx")],
            title="ë‹¤ë¥¸ ì´ë¦„ìœ¼ë¡œ ì €ì¥"
        ) 
        
        if path: 
            # 5-32-3. ì§„ì§œë¡œ íŒŒì¼ì„ ë§Œë“­ë‹ˆë‹¤.
            if self._core_excel_save_logic(path):
                messagebox.showinfo("ì €ì¥ ì™„ë£Œ", f"ì˜ì—…ì¼ ê¸°ì¤€ìœ¼ë¡œ ë¶„ë¥˜ë˜ì–´ ì €ì¥ë˜ì—ˆìŠµë‹ˆë‹¤.\nê²½ë¡œ: {path}")
                self.log(f"ğŸ’¾ ìˆ˜ë™ ì—‘ì…€ ì €ì¥ ì™„ë£Œ: {path}")

    # 5-33. [ìë™ ì €ì¥] í”„ë¡œê·¸ë¨ì´ ì •í•´ì§„ ì¥ì†Œì— ì•Œì•„ì„œ ì²™ì²™ ê¸°ë¡ì„ ë°±ì—…í•˜ëŠ” í•¨ìˆ˜
    def perform_auto_save(self):
        # 5-33-1. ì €ì¥í•  ê²Œ ìˆëŠ”ì§€, ìë™ ì €ì¥ì´ ì¼œì ¸ ìˆëŠ”ì§€ í™•ì¸í•©ë‹ˆë‹¤.
        if not self.recorded_data or not self.auto_save_path or not self.can_auto_save: 
            return 
        
        # 5-33-2. ì†Œë¦¬ ì—†ì´ ì¡°ìš©íˆ íŒŒì¼ì— ê¸°ë¡í•©ë‹ˆë‹¤.
        #         save_completed=True ë¡œ ì „ë‹¬í•´ ì™„ê²°ëœ ì˜ì—…ì¼ íŒŒì¼ë„ í•¨ê»˜ ì €ì¥í•©ë‹ˆë‹¤.
        if self._core_excel_save_logic(self.auto_save_path, save_completed=True):
            pass

    # 5-34. [í‘œ ìƒˆë¡œê³ ì¹¨] í‘œì˜ ë‚´ìš©ì„ ë‹¤ ì§€ìš°ê³  ìƒˆë¡œìš´ ë²„ìŠ¤ ì •ë³´ë¡œ ì±„ìš°ëŠ” í•¨ìˆ˜
    def refresh_tree(self, tree, data):
        # 5-34-1. ì˜›ë‚  ì •ë³´ëŠ” ì‹¹ ì§€ì›ë‹ˆë‹¤.
        for i in tree.get_children(): tree.delete(i)
        # 5-34-2. ìƒˆë¡œ ê°€ì ¸ì˜¨ ì •ë³´ë¥¼ í•œ ì¤„ì”© ë„£ìŠµë‹ˆë‹¤.
        for row in data: tree.insert("", "end", values=row)

# 6. [ì§„ì§œ ì‹¤í–‰] ì—¬ê¸°ê°€ í”„ë¡œê·¸ë¨ì˜ ì‹œì‘ ë²„íŠ¼ì…ë‹ˆë‹¤!
if __name__ == "__main__":
    import traceback

    # 6-0. [ì˜¤ë¥˜ ê¸°ë¡ê¸°] í”„ë¡œê·¸ë¨ì´ ì˜ˆê¸°ì¹˜ ì•Šê²Œ ì¢…ë£Œë  ë•Œ ì˜¤ë¥˜ ë‚´ìš©ì„ errorlog.txt ì— ì €ì¥í•©ë‹ˆë‹¤.
    #   - --windowed (pyinstaller) ë¡œ ë¹Œë“œëœ ì‹¤í–‰íŒŒì¼ì—ì„œë„ ë™ì‘í•©ë‹ˆë‹¤.
    #   - ì‹¤í–‰ íŒŒì¼(ë˜ëŠ” .py) ì´ ìˆëŠ” í´ë”ì— errorlog.txt ë¥¼ ìƒì„±í•©ë‹ˆë‹¤.
    def _get_error_log_path():
        """ì‹¤í–‰ íŒŒì¼ ë˜ëŠ” ì†ŒìŠ¤ íŒŒì¼ì´ ìœ„ì¹˜í•œ í´ë”ì— errorlog.txt ê²½ë¡œë¥¼ ë°˜í™˜í•©ë‹ˆë‹¤."""
        if getattr(sys, 'frozen', False):
            # PyInstaller --windowed ë¹Œë“œ: sys.executable ìœ„ì¹˜
            base = os.path.dirname(os.path.abspath(sys.executable))
            # macOS .app ë²ˆë“¤ì¼ ê²½ìš° Contents/MacOS ìœ„ë¡œ 3ë‹¨ê³„ ì˜¬ë¼ê°‘ë‹ˆë‹¤.
            if sys.platform == "darwin" and base.endswith("Contents/MacOS"):
                base = os.path.dirname(os.path.dirname(os.path.dirname(base)))
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, "errorlog.txt")

    def _write_error_log(exc_type, exc_value, exc_tb):
        """ì˜ˆì™¸ ì •ë³´ë¥¼ errorlog.txt ì— ê¸°ë¡í•˜ê³ , ê°€ëŠ¥í•˜ë©´ messagebox ë¡œë„ ì•Œë¦½ë‹ˆë‹¤."""
        try:
            log_path = _get_error_log_path()
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            tb_text = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"\n{'='*60}\n")
                f.write(f"[{timestamp}] í”„ë¡œê·¸ë¨ ì˜¤ë¥˜ ë°œìƒ\n")
                f.write(f"{'='*60}\n")
                f.write(tb_text)
                f.write("\n")
            # GUI ë©”ì‹œì§€ë°•ìŠ¤ë¡œë„ ì•Œë¦½ë‹ˆë‹¤ (ì°½ì´ ì‚´ì•„ìˆì„ ë•Œ)
            try:
                import tkinter.messagebox as _mb
                short_msg = f"{exc_type.__name__}: {exc_value}"
                _mb.showerror(
                    "í”„ë¡œê·¸ë¨ ì˜¤ë¥˜",
                    f"ì˜ˆê¸°ì¹˜ ì•Šì€ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤.\n\n"
                    f"{short_msg}\n\n"
                    f"ì˜¤ë¥˜ ë‚´ìš©ì´ ë‹¤ìŒ íŒŒì¼ì— ì €ì¥ë˜ì—ˆìŠµë‹ˆë‹¤:\n{log_path}"
                )
            except Exception:
                pass
        except Exception:
            pass  # ë¡œê·¸ ê¸°ë¡ ìì²´ê°€ ì‹¤íŒ¨í•´ë„ í”„ë¡œê·¸ë¨ì´ ì´ìƒí•˜ê²Œ ì¢…ë£Œë˜ì§€ ì•ŠìŠµë‹ˆë‹¤.

    # 6-0-1. ì²˜ë¦¬ë˜ì§€ ì•Šì€ ì˜ˆì™¸ë¥¼ ê°€ë¡œì±„ì„œ errorlog.txt ì— ê¸°ë¡í•©ë‹ˆë‹¤.
    def _global_exception_handler(exc_type, exc_value, exc_tb):
        _write_error_log(exc_type, exc_value, exc_tb)
        sys.__excepthook__(exc_type, exc_value, exc_tb)  # ì›ë˜ ë™ì‘ë„ ìœ ì§€
    sys.excepthook = _global_exception_handler

    # 6-0-2. stderr ë¥¼ íŒŒì¼ë¡œ ë¦¬ë‹¤ì´ë ‰íŠ¸ (--windowed ë¹Œë“œ ì‹œ stderr ê°€ ì—†ì–´ ì˜¤ë¥˜ ë°œìƒ ë°©ì§€)
    try:
        if sys.stderr is None or getattr(sys, 'frozen', False):
            sys.stderr = open(_get_error_log_path().replace("errorlog.txt", "stderr.txt"),
                              "a", encoding="utf-8")
    except Exception:
        pass

    # 6-1. ìœˆë„ìš° ì°½ì„ í•˜ë‚˜ ìƒˆë¡œ ë§Œë“­ë‹ˆë‹¤.
    try:
        root = tk.Tk()
        # 6-2. ìš°ë¦¬ê°€ ë§Œë“  'ë²„ìŠ¤ ê¸°ë¡ê¸°' ì„¤ê³„ë„ëŒ€ë¡œ í”„ë¡œê·¸ë¨ì„ ì¡°ë¦½í•©ë‹ˆë‹¤.
        app = SeoulBusArrivalRecorder(root)
        # 6-3. ì°½ì´ êº¼ì§€ì§€ ì•Šê³  ê³„ì† ìš°ë¦¬ë¥¼ ê¸°ë‹¤ë¦¬ê²Œ ë§Œë“­ë‹ˆë‹¤.
        root.mainloop()
    except Exception as e:
        _write_error_log(type(e), e, e.__traceback__)

   